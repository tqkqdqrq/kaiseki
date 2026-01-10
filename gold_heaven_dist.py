# 沖ドキGOLD 天国中当選G数分布スクリプト
import pandas as pd
from collections import Counter

# 設定
INPUT_FILE = r"C:\Users\ilove\Desktop\解析\沖ドキＧＯＬＤ(本)２ - Sheet1.csv"
HEAVEN_THRESHOLD = 32  # 天国閾値

print("=== 沖ドキGOLD 天国中当選G数分布 ===")
print(f"天国閾値: {HEAVEN_THRESHOLD}G")
print()

# データ読み込み
print("データ読み込み中...")
df = pd.read_csv(INPUT_FILE, encoding='utf-8')
print(f"  読み込み完了: {len(df)}行")

# カラム名を設定（ヘッダーがずれているので位置で指定）
df.columns = ['URL', '台番号', '機種名', '日付', '大当たり', 'スタート', '出玉', '種別', '時間']

# 数値に変換
df['大当たり'] = pd.to_numeric(df['大当たり'], errors='coerce')
df['スタート'] = pd.to_numeric(df['スタート'], errors='coerce')

# ソート（URL、台番号、日付、時間順）
df = df.sort_values(['URL', '台番号', '日付', '時間']).reset_index(drop=True)

# 連チャンIDを生成
print("連チャンID生成中...")
chain_ids = []
chain_positions = []  # 連チャン内の位置（1=初当たり、2=2連目...）

current_chain_id = 0
current_chain_pos = 0
prev_url = None
prev_dai = None
prev_date = None

for idx, row in df.iterrows():
    url = row['URL']
    dai = row['台番号']
    date = row['日付']
    hit_num = row['大当たり']
    start_g = row['スタート']
    
    # URL+台番号+日付が変わったらリセット
    if url != prev_url or dai != prev_dai or date != prev_date:
        prev_url = url
        prev_dai = dai
        prev_date = date
        current_chain_id += 1
        current_chain_pos = 1
    else:
        # 同じ台+日付内
        if hit_num == 1 or pd.isna(hit_num):
            # 初当たり = 新しい連チャン開始
            current_chain_id += 1
            current_chain_pos = 1
        elif hit_num == 0:
            # RB単発（連チャンリセット）→ ただし32G以内なら天国中の可能性
            if start_g <= HEAVEN_THRESHOLD:
                # 天国中のRB、同じチェーン
                current_chain_pos += 1
            else:
                # 天国抜け、新しいチェーン扱い
                current_chain_id += 1
                current_chain_pos = 1
        else:
            # 連チャン継続 (hit_num >= 2)
            current_chain_pos = hit_num
    
    chain_ids.append(current_chain_id)
    chain_positions.append(current_chain_pos)

df['Chain_ID'] = chain_ids
df['Chain_Pos'] = chain_positions

print(f"  連チャン数: {df['Chain_ID'].nunique()}")

# 天国判定（3連以上の連チャン）
chain_lengths = df.groupby('Chain_ID')['Chain_Pos'].max()
heaven_chains = chain_lengths[chain_lengths >= 3].index

print(f"  天国チェーン数（3連以上）: {len(heaven_chains)}")

# 天国チェーン中の2連目以降（32G以内）を抽出
heaven_hits = df[
    (df['Chain_ID'].isin(heaven_chains)) & 
    (df['Chain_Pos'] >= 2) & 
    (df['スタート'] <= HEAVEN_THRESHOLD)
]

print(f"  天国中当選数: {len(heaven_hits)}")
print()

# G数分布を集計
g_counts = Counter(heaven_hits['スタート'].astype(int))
total_hits = len(heaven_hits)

# 0Gを1Gに合算
count_0g = g_counts.get(0, 0)
if count_0g > 0:
    g_counts[1] = g_counts.get(1, 0) + count_0g
    g_counts[0] = 0
    print(f"※ 0Gデータ {count_0g}件 を1Gに合算しました")
    print()

# 結果表示
print("【天国中 当選G数分布】")
print("-" * 40)
print(f"{'G数':<8} {'回数':<10} {'割合':<10}")
print("-" * 40)

for g in range(1, HEAVEN_THRESHOLD + 1):
    count = g_counts.get(g, 0)
    rate = count / total_hits * 100 if total_hits > 0 else 0
    print(f"{g}G{'':<5} {count:<10} {rate:.2f}%")

print("-" * 40)
print(f"{'合計':<8} {total_hits:<10} {'100.00%':<10}")

# 平均G数（0Gを除く）
valid_hits = heaven_hits[heaven_hits['スタート'] > 0]
avg_g = sum(valid_hits['スタート']) / len(valid_hits) if len(valid_hits) > 0 else 0
print(f"\n平均当選G数: {avg_g:.2f}G (0G除く)")

# Excelに出力
print("\nExcel出力中...")
from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()
ws = wb.active
ws.title = "天国中当選分布"

# タイトル
ws['A1'] = "【沖ドキGOLD 天国中当選G数分布】"
ws['A1'].font = Font(bold=True, size=14)
ws['A2'] = f"天国閾値: {HEAVEN_THRESHOLD}G、3連以上チェーンの2連目以降"

# ヘッダー
ws['A4'] = "当選G数"
ws['B4'] = "発生回数"
ws['C4'] = "割合"
for col in ['A', 'B', 'C']:
    ws[f'{col}4'].font = Font(bold=True)

# データ
row = 5

for g in range(1, HEAVEN_THRESHOLD + 1):
    count = g_counts.get(g, 0)
    rate = count / total_hits if total_hits > 0 else 0
    ws.cell(row=row, column=1, value=f"{g}G")
    ws.cell(row=row, column=2, value=count)
    ws.cell(row=row, column=3, value=rate)
    ws.cell(row=row, column=3).number_format = '0.00%'
    row += 1

# 合計行
ws.cell(row=row, column=1, value="合計")
ws.cell(row=row, column=1).font = Font(bold=True)
ws.cell(row=row, column=2, value=total_hits)
ws.cell(row=row, column=2).font = Font(bold=True)
ws.cell(row=row, column=3, value=1.0)
ws.cell(row=row, column=3).number_format = '0.00%'
ws.cell(row=row, column=3).font = Font(bold=True)

# 平均G数
row += 1
ws.cell(row=row, column=1, value="平均G数")
ws.cell(row=row, column=1).font = Font(bold=True)
ws.cell(row=row, column=2, value=avg_g)
ws.cell(row=row, column=2).number_format = '0.0'
ws.cell(row=row, column=2).font = Font(bold=True)

# 列幅調整
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 12

# 保存
OUTPUT_FILE = r"C:\Users\ilove\Desktop\解析\gold_heaven_distribution.xlsx"
wb.save(OUTPUT_FILE)
print(f"出力完了: {OUTPUT_FILE}")
