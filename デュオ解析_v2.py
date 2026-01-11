import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import pickle
from tqdm import tqdm

# ===== 設定 =====
INPUT_FILE = r'C:\Users\ilove\Desktop\解析\jyogai20251223_duo2_azukun.csv'
OUTPUT_FILE = r'C:\Users\ilove\Desktop\解析\duo解析\duo_analysis_' + datetime.now().strftime('%Y%m%d_%H%M%S') + '.xlsx'
CACHE_FILE = r'C:\Users\ilove\Desktop\解析\chains_cache.pkl'

# 設定
HEAVEN_THRESHOLD = 35  # 天国連チャンとみなすG数
COIN_HOLD = 25.3  # コイン持ち (G/50枚)
BB_GAMES = 51          # BB獲得G数
RB_GAMES = 21          # RB獲得G数


def load_data(filepath):
    """azukunフォーマットCSVを読み込み"""
    try:
        df = pd.read_csv(filepath, encoding='utf-8-sig')
    except:
        df = pd.read_csv(filepath, encoding='cp932')
    
    # 元の順序を保持
    df['Original_Order'] = range(len(df))
    
    # ID列を保持したまま、内部処理用に分解
    if 'ID' not in df.columns:
        # ID列がない場合は作成を試みる（既存のロジックで対応できない場合はエラーになるが、今回はazukun.csvを使うのでOKなはず）
        pass

    id_parts = df['ID'].str.rsplit('_', n=2, expand=True)
    df['Hall_Name'] = id_parts[0]
    df['Machine_No'] = pd.to_numeric(id_parts[1], errors='coerce').fillna(0).astype(int)
    df['Date'] = id_parts[2].str.replace(r'(\d{4})(\d{2})(\d{2})', r'\1-\2-\3', regex=True)
    
    df['Start'] = pd.to_numeric(df['Start'], errors='coerce').fillna(0).astype(int)
    df['Dedama'] = pd.to_numeric(df['Dedama'], errors='coerce').fillna(0).astype(int)
    
    # Count列作成（時刻順）
    df_sorted = df.sort_values(by=['Hall_Name', 'Date', 'Machine_No', 'Time'])
    df_sorted['Count'] = df_sorted.groupby(['Hall_Name', 'Date', 'Machine_No']).cumcount() + 1
    
    # 元の順序に戻す
    df = df_sorted.sort_values(by='Original_Order').reset_index(drop=True)
    
    return df


def analyze_chains(df):
    """連チャン判定と天国移行率を計算（元のCSV順序で処理）"""
    # 元の順序でソート
    df = df.sort_values(by='Original_Order').reset_index(drop=True)

    all_chains = []
    chain_id = 0
    chain_number_per_id = {}  # {ID: 連チャン回数}

    # DataFrameに Chain_ID, Chain_Position 列を追加
    df['Chain_ID'] = 0
    df['Chain_Position'] = 0

    # 現在のチェーン情報
    current_id = None
    current_chain_start_idx = 0
    current_chain_hits = [] # インデックスのリスト
    prev_chain_len = 0

    for idx, row in tqdm(df.iterrows(), total=len(df), desc="Analyzing Chains"):
        row_id = row['ID']
        start_g = row['Start']

        # 新しい台または天国終了
        if current_id != row_id or start_g > HEAVEN_THRESHOLD:
            # 前のチェーンを保存
            if current_chain_hits:
                chain_id += 1
                chain_len = len(current_chain_hits)
                is_heaven = chain_len >= 2

                # Chain_ID, Positionを設定
                for pos, hit_idx in enumerate(current_chain_hits, start=1):
                    df.at[hit_idx, 'Chain_ID'] = chain_id
                    df.at[hit_idx, 'Chain_Position'] = pos

                # チェーンの差枚計算
                first_start = df.at[current_chain_hits[0], 'Start']
                last_dedama = df.at[current_chain_hits[-1], 'Dedama']
                raw_dedama = df.loc[current_chain_hits, 'Dedama'].sum()  # 出玉合計（生データ）
                total_invest = df.loc[current_chain_hits, 'Start'].sum() / COIN_HOLD * 50
                first_invest = first_start / COIN_HOLD * 50
                # 天国中投資（初当り以外の消費）= 全投資 - 初当り投資
                heaven_invest = total_invest - first_invest
                # 純増枚数 = 出玉合計 - 天国中消費（初当り投資は含まない）
                total_dedama = raw_dedama - heaven_invest
                # 差枚 = 出玉合計 - 全投資（初当り含む）
                net_diff = raw_dedama - total_invest
                # 特殊判定 = 純増 - 最終出玉（最後の出玉を除いた天国中の純増）
                special_judge = total_dedama - last_dedama
                
                # 当選G数のリストを取得
                hit_games = df.loc[current_chain_hits, 'Start'].tolist()
                # Dedamaのリストを取得（ドキハナ計算用）
                hit_dedamas = df.loc[current_chain_hits, 'Dedama'].tolist()
                # Statusのリストを取得（BB/RB判定用）
                hit_statuses = df.loc[current_chain_hits, 'Status'].tolist()

                # 台ごとの連チャン番号をカウント
                machine_id = df.at[current_chain_hits[0], 'ID']
                if machine_id not in chain_number_per_id:
                    chain_number_per_id[machine_id] = 0
                chain_number_per_id[machine_id] += 1

                # 統計用に記録
                all_chains.append({
                    'ID': machine_id,
                    'Chain_ID': chain_id,
                    'Chain_Number': chain_number_per_id[machine_id],
                    'Chain_Length': chain_len,
                    'Hit_Games': hit_games,
                    'Hit_Dedamas': hit_dedamas,
                    'Hit_Statuses': hit_statuses,
                    'Is_Heaven': is_heaven,
                    'First_G': df.at[current_chain_hits[0], 'Start'],
                    'Through_Before': 0,
                    'Prev_Chain_Length': prev_chain_len,
                    'Raw_Dedama': raw_dedama,       # 出玉合計（生データ）
                    'Total_Dedama': total_dedama,   # 純増枚数 = 出玉 - 天国中消費
                    'Total_Invest': total_invest,
                    'Heaven_Invest': heaven_invest,
                    'Net_Diff': net_diff,
                    'Special_Judge': special_judge,
                    'Prev_Special_Judge': 0.0
                })

                if is_heaven:
                    prev_chain_len = chain_len
                else:
                    pass # 連チャンでない場合は前回連チャン数を維持しない（スルー扱いなので、前回の「天国連チャン」数は保持すべきだが、ロジック上は後で計算し直す）
            
            # 新しいチェーン開始
            if current_id != row_id:
                prev_chain_len = 0
            
            current_id = row_id
            current_chain_hits = [idx]
        else:
            # 天国継続
            current_chain_hits.append(idx)
    
    # 最後のチェーンを保存
    if current_chain_hits:
        chain_id += 1
        chain_len = len(current_chain_hits)
        is_heaven = chain_len >= 2

        for pos, hit_idx in enumerate(current_chain_hits, start=1):
            df.at[hit_idx, 'Chain_ID'] = chain_id
            df.at[hit_idx, 'Chain_Position'] = pos

        # チェーンの差枚計算
        first_start = df.at[current_chain_hits[0], 'Start']
        last_dedama = df.at[current_chain_hits[-1], 'Dedama']
        raw_dedama = df.loc[current_chain_hits, 'Dedama'].sum()  # 出玉合計（生データ）
        total_invest = df.loc[current_chain_hits, 'Start'].sum() / COIN_HOLD * 50
        first_invest = first_start / COIN_HOLD * 50
        heaven_invest = total_invest - first_invest
        # 純増枚数 = 出玉合計 - 天国中消費（初当り投資は含まない）
        total_dedama = raw_dedama - heaven_invest
        # 差枚 = 出玉合計 - 全投資（初当り含む）
        net_diff = raw_dedama - total_invest
        # 特殊判定 = 純増 - 最終出玉
        special_judge = total_dedama - last_dedama
        
        # 当選G数のリストを取得
        hit_games = df.loc[current_chain_hits, 'Start'].tolist()
        # Dedamaのリストを取得
        hit_dedamas = df.loc[current_chain_hits, 'Dedama'].tolist()
        # Statusのリストを取得
        hit_statuses = df.loc[current_chain_hits, 'Status'].tolist()

        # 台ごとの連チャン番号をカウント
        machine_id = df.at[current_chain_hits[0], 'ID']
        if machine_id not in chain_number_per_id:
            chain_number_per_id[machine_id] = 0
        chain_number_per_id[machine_id] += 1

        all_chains.append({
            'ID': machine_id,
            'Chain_ID': chain_id,
            'Chain_Number': chain_number_per_id[machine_id],
            'Chain_Length': chain_len,
            'Hit_Games': hit_games,
            'Hit_Dedamas': hit_dedamas,
            'Hit_Statuses': hit_statuses,
            'Is_Heaven': is_heaven,
            'First_G': df.at[current_chain_hits[0], 'Start'],
            'Through_Before': 0,
            'Prev_Chain_Length': prev_chain_len,
            'Raw_Dedama': raw_dedama,       # 出玉合計（生データ）
            'Total_Dedama': total_dedama,   # 純増枚数 = 出玉 - 天国中消費
            'Total_Invest': total_invest,
            'Heaven_Invest': heaven_invest,
            'Net_Diff': net_diff,
            'Special_Judge': special_judge,
            'Prev_Special_Judge': 0.0
        })

    chains_df = pd.DataFrame(all_chains)

    # Through_Before, Prev_Chain_Length, Prev_Special_Judgeを正確に計算
    if not chains_df.empty:
        chains_df['Through_Before'] = 0
        chains_df['Prev_Chain_Length'] = 0
        chains_df['Prev_Special_Judge'] = 0.0  # float型に変更
        chains_df['Daily_Balance_Before'] = 0.0    # 当日差枚（チェーン開始前の累計）
        chains_df['Max_Daily_Balance_Before'] = 0.0  # 当日差枚最大（チェーン開始前までの最大値）
        chains_df['Is_Reset'] = chains_df['Chain_Number'] == 1  # リセット判定（台内初回）

        prev_id = None
        through_count = 0
        prev_heaven_len = 0
        prev_special = 0.0
        cumulative_balance = 0.0   # 累計差枚（前回チェーン終了時点）
        max_balance = 0.0          # 最大差枚（前回チェーン終了時点までの最大）

        for idx in chains_df.index:
            current_id = chains_df.at[idx, 'ID']
            is_heaven = chains_df.at[idx, 'Is_Heaven']
            chain_len = chains_df.at[idx, 'Chain_Length']
            special_judge = chains_df.at[idx, 'Special_Judge']
            net_diff = chains_df.at[idx, 'Net_Diff']  # このチェーンの差枚

            if current_id != prev_id:
                through_count = 0
                prev_heaven_len = 0
                prev_special = 0.0
                cumulative_balance = 0.0
                max_balance = 0.0

            # チェーン開始前の値を記録（このチェーンの差枚を加算する前）
            chains_df.at[idx, 'Through_Before'] = through_count
            chains_df.at[idx, 'Prev_Chain_Length'] = prev_heaven_len
            chains_df.at[idx, 'Prev_Special_Judge'] = prev_special
            chains_df.at[idx, 'Daily_Balance_Before'] = cumulative_balance
            chains_df.at[idx, 'Max_Daily_Balance_Before'] = max_balance

            # このチェーンの差枚を加算（次のチェーンのために）
            cumulative_balance += net_diff
            # 最大値を更新
            if cumulative_balance > max_balance:
                max_balance = cumulative_balance

            if is_heaven:
                through_count = 0
                prev_heaven_len = chain_len
            else:
                through_count += 1

            # 前回の特殊判定を保存（天国チェーンの場合のみ更新、天国以外は直近の天国の値を保持）
            if is_heaven:
                prev_special = special_judge

            prev_id = current_id

    return df, chains_df


def calculate_heaven_rate(chains_df):
    """スルー回数別の天国移行率を計算"""
    results = []
    
    # 0〜10スルー
    for thr in range(11):
        subset = chains_df[chains_df['Through_Before'] == thr]
        total = len(subset)
        heaven = len(subset[subset['Is_Heaven'] == True])
        rate = heaven / total if total > 0 else 0
        
        results.append({
            'スルー回数': f'{thr}スルー',
            '天国移行数': heaven,
            'サンプル数': total,
            '天国移行率': rate
        })
    
    # 11スルー以上
    subset = chains_df[chains_df['Through_Before'] >= 11]
    total = len(subset)
    heaven = len(subset[subset['Is_Heaven'] == True])
    rate = heaven / total if total > 0 else 0
    
    results.append({
        'スルー回数': '11スルー以上',
        '天国移行数': heaven,
        'サンプル数': total,
        '天国移行率': rate
    })
    
    # 合計
    total_all = len(chains_df)
    heaven_all = len(chains_df[chains_df['Is_Heaven'] == True])
    rate_all = heaven_all / total_all if total_all > 0 else 0
    
    results.append({
        'スルー回数': '合計',
        '天国移行数': heaven_all,
        'サンプル数': total_all,
        '天国移行率': rate_all
    })
    
    return pd.DataFrame(results)


def calculate_heaven_rate_by_chain(chains_df):
    """前回連チャン長別の天国移行率を計算"""
    results = []
    
    # 前回連チャン2連〜19連
    for chain_len in range(2, 20):
        subset = chains_df[chains_df['Prev_Chain_Length'] == chain_len]
        total = len(subset)
        heaven = len(subset[subset['Is_Heaven'] == True])
        rate = heaven / total if total > 0 else 0
        
        results.append({
            '前回連チャン': f'{chain_len}連後',
            '天国移行数': heaven,
            'サンプル数': total,
            '天国移行率': rate
        })
    
    # 20連以上
    subset = chains_df[chains_df['Prev_Chain_Length'] >= 20]
    total = len(subset)
    heaven = len(subset[subset['Is_Heaven'] == True])
    rate = heaven / total if total > 0 else 0
    
    results.append({
        '前回連チャン': '20連以上後',
        '天国移行数': heaven,
        'サンプル数': total,
        '天国移行率': rate
    })
    
    # 合計 (前回天国があったもののみ)
    subset_all = chains_df[chains_df['Prev_Chain_Length'] >= 2]
    total_all = len(subset_all)
    heaven_all = len(subset_all[subset_all['Is_Heaven'] == True])
    rate_all = heaven_all / total_all if total_all > 0 else 0
    
    results.append({
        '前回連チャン': '合計',
        '天国移行数': heaven_all,
        'サンプル数': total_all,
        '天国移行率': rate_all
    })
    
    return pd.DataFrame(results)


def calculate_heaven_rate_crosstab(chains_df):
    """前回連チャン長 × スルー回数 のクロス集計（天国移行率とサンプル数を分離）+ 特殊判定分析を統合"""
    results = []

    # 朝イチを除外（Chain_Number > 1のみ）
    chains_with_prev = chains_df[chains_df['Chain_Number'] > 1].copy()

    # ===== 前回連チャン長別 - 天国移行率 =====
    results.append({'': '【前回連チャン長別 - 天国移行率】'})

    for chain_len in range(2, 20):
        rate_row = {'': f'{chain_len}連後'}
        for thr in range(11):
            subset = chains_with_prev[(chains_with_prev['Prev_Chain_Length'] == chain_len) &
                                      (chains_with_prev['Through_Before'] == thr)]
            total = len(subset)
            heaven = len(subset[subset['Is_Heaven'] == True])
            rate = heaven / total if total > 0 else None
            rate_row[f'{thr}スルー'] = rate

        # 合計列には天国移行率を入れる
        total_all = len(chains_with_prev[chains_with_prev['Prev_Chain_Length'] == chain_len])
        heaven_all = len(chains_with_prev[(chains_with_prev['Prev_Chain_Length'] == chain_len) &
                                          (chains_with_prev['Is_Heaven'] == True)])
        rate_all = heaven_all / total_all if total_all > 0 else None
        rate_row['合計'] = rate_all

        results.append(rate_row)

    # 20連以上
    rate_row = {'': '20連以上後'}
    for thr in range(11):
        subset = chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 20) &
                                  (chains_with_prev['Through_Before'] == thr)]
        total = len(subset)
        heaven = len(subset[subset['Is_Heaven'] == True])
        rate = heaven / total if total > 0 else None
        rate_row[f'{thr}スルー'] = rate
    total_all = len(chains_with_prev[chains_with_prev['Prev_Chain_Length'] >= 20])
    heaven_all = len(chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 20) &
                                      (chains_with_prev['Is_Heaven'] == True)])
    rate_all = heaven_all / total_all if total_all > 0 else None
    rate_row['合計'] = rate_all
    results.append(rate_row)

    # 3～9連後（まとめ）
    rate_row = {'': '3～9連後'}
    for thr in range(11):
        subset = chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 3) &
                                  (chains_with_prev['Prev_Chain_Length'] <= 9) &
                                  (chains_with_prev['Through_Before'] == thr)]
        total = len(subset)
        heaven = len(subset[subset['Is_Heaven'] == True])
        rate = heaven / total if total > 0 else None
        rate_row[f'{thr}スルー'] = rate
    total_all = len(chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 3) & (chains_with_prev['Prev_Chain_Length'] <= 9)])
    heaven_all = len(chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 3) & (chains_with_prev['Prev_Chain_Length'] <= 9) &
                                      (chains_with_prev['Is_Heaven'] == True)])
    rate_all = heaven_all / total_all if total_all > 0 else None
    rate_row['合計'] = rate_all
    results.append(rate_row)

    # 10連以上後（まとめ）
    rate_row = {'': '10連以上後'}
    for thr in range(11):
        subset = chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 10) &
                                  (chains_with_prev['Through_Before'] == thr)]
        total = len(subset)
        heaven = len(subset[subset['Is_Heaven'] == True])
        rate = heaven / total if total > 0 else None
        rate_row[f'{thr}スルー'] = rate
    total_all = len(chains_with_prev[chains_with_prev['Prev_Chain_Length'] >= 10])
    heaven_all = len(chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 10) &
                                      (chains_with_prev['Is_Heaven'] == True)])
    rate_all = heaven_all / total_all if total_all > 0 else None
    rate_row['合計'] = rate_all
    results.append(rate_row)

    # 空行
    results.append({})

    # ===== 前回連チャン長別 - サンプル数 =====
    results.append({'': '【前回連チャン長別 - サンプル数】'})

    for chain_len in range(2, 20):
        count_row = {'': f'{chain_len}連後'}
        for thr in range(11):
            subset = chains_with_prev[(chains_with_prev['Prev_Chain_Length'] == chain_len) &
                                      (chains_with_prev['Through_Before'] == thr)]
            count_row[f'{thr}スルー'] = len(subset)

        total_all = len(chains_with_prev[chains_with_prev['Prev_Chain_Length'] == chain_len])
        count_row['合計'] = total_all

        results.append(count_row)

    # 20連以上
    count_row = {'': '20連以上後'}
    for thr in range(11):
        subset = chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 20) &
                                  (chains_with_prev['Through_Before'] == thr)]
        count_row[f'{thr}スルー'] = len(subset)
    total_all = len(chains_with_prev[chains_with_prev['Prev_Chain_Length'] >= 20])
    count_row['合計'] = total_all
    results.append(count_row)

    # 3～9連後（まとめ）
    count_row = {'': '3～9連後'}
    for thr in range(11):
        subset = chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 3) &
                                  (chains_with_prev['Prev_Chain_Length'] <= 9) &
                                  (chains_with_prev['Through_Before'] == thr)]
        count_row[f'{thr}スルー'] = len(subset)
    total_all = len(chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 3) & (chains_with_prev['Prev_Chain_Length'] <= 9)])
    count_row['合計'] = total_all
    results.append(count_row)

    # 10連以上後（まとめ）
    count_row = {'': '10連以上後'}
    for thr in range(11):
        subset = chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 10) &
                                  (chains_with_prev['Through_Before'] == thr)]
        count_row[f'{thr}スルー'] = len(subset)
    total_all = len(chains_with_prev[chains_with_prev['Prev_Chain_Length'] >= 10])
    count_row['合計'] = total_all
    results.append(count_row)

    # 空行
    results.append({})

    # ===== 前回特殊判定別 - 天国移行率 =====
    results.append({'': '【前回特殊判定別 - 天国移行率】'})

    # 100枚刻みのレンジ定義
    spec_ranges = []
    for start in range(0, 2001, 100):
        end = start + 99
        spec_ranges.append((f"{start}～{end}", start, end))
    spec_ranges.append(("2001以上", 2001, None))

    for label, low, high in spec_ranges:
        rate_row = {'': label}

        if high is None:
            range_subset = chains_with_prev[chains_with_prev['Prev_Special_Judge'] >= low]
        else:
            range_subset = chains_with_prev[(chains_with_prev['Prev_Special_Judge'] >= low) &
                                            (chains_with_prev['Prev_Special_Judge'] <= high)]

        for thr in range(11):
            subset = range_subset[range_subset['Through_Before'] == thr]
            total = len(subset)
            heaven = len(subset[subset['Is_Heaven'] == True])
            rate = heaven / total if total > 0 else None
            rate_row[f'{thr}スルー'] = rate

        # 合計列には天国移行率を入れる
        total_all = len(range_subset)
        heaven_all = len(range_subset[range_subset['Is_Heaven'] == True])
        rate_all = heaven_all / total_all if total_all > 0 else None
        rate_row['合計'] = rate_all

        results.append(rate_row)

    # 空行
    results.append({})

    # ===== 前回特殊判定別 - サンプル数 =====
    results.append({'': '【前回特殊判定別 - サンプル数】'})

    for label, low, high in spec_ranges:
        count_row = {'': label}

        if high is None:
            range_subset = chains_with_prev[chains_with_prev['Prev_Special_Judge'] >= low]
        else:
            range_subset = chains_with_prev[(chains_with_prev['Prev_Special_Judge'] >= low) &
                                            (chains_with_prev['Prev_Special_Judge'] <= high)]

        for thr in range(11):
            subset = range_subset[range_subset['Through_Before'] == thr]
            count_row[f'{thr}スルー'] = len(subset)

        count_row['合計'] = len(range_subset)

        
        results.append(rate_row)
        results.append(count_row)
        
    return pd.DataFrame(results)


def calculate_3chain_rule_analysis(chains_df):
    """3連以上を天国、2連を隠れスルーとする分析（率とサンプル数を別列で出力）"""
    # 分析用データを収集
    # 構造: {'Category': [{'2Chain_Loc': ..., 'Through': ..., 'Is_True_Heaven': ...}]}
    analysis_data = {
        '全体': [],
        '2連目1G_RB': [],
        '2連目1G_BB': [],
        '2連目[2-5,32]': [],
        '2連目[6-12]': [],
        '2連目[13-21]': [],
        '2連目[22-31]': [],
        '2連目BB_[2-5,32]': [],
        '2連目BB_[6-12]': [],
        '2連目BB_[13-21]': [],
        '2連目BB_[22-31]': [],
        '2連目RB_[2-5,32]': [],
        '2連目RB_[6-12]': [],
        '2連目RB_[13-21]': [],
        '2連目RB_[22-31]': [],
    }
    
    # 変数初期化
    prev_id = None
    hits_since_true_heaven = 0  # 3連以上基準のスルー回数
    first_2chain_idx = None     # 最初の2連が発生したスルー回数 (None: 未発生)
    prev_true_heaven_valid = False 
    
    # カテゴリ判定用のステート
    current_2chain_feature = None 
    
    trackers = {}
    for cat in analysis_data.keys():
        trackers[cat] = {
            'hits_since': 0,
            'first_2chain_idx': None,
            'active': False, # このカテゴリのデータとして有効か（3連天国後スタート）
            'ignore': False # このカテゴリの条件に合わない「最初の2連」が発生した場合、以降天国まで無視
        }

    # DataFrameに追加するためのリスト初期化
    col_r3_through = []
    col_r3_end = []
    col_r3_heaven = []
    # カテゴリごとのLocation列
    cols_loc = {cat: [] for cat in analysis_data}
    
    for idx, row in chains_df.iterrows():
        row_id = row['ID']
        chain_len = row['Chain_Length']
        hit_games = row.get('Hit_Games', [])
        hit_statuses = row.get('Hit_Statuses', [])
        
        # 台が変わったらリセット
        if row_id != prev_id:
            for cat in trackers:
                trackers[cat]['hits_since'] = 0
                trackers[cat]['first_2chain_idx'] = None
                trackers[cat]['active'] = False
                trackers[cat]['ignore'] = False
            
            if chain_len >= 3:
                for cat in trackers:
                    trackers[cat]['active'] = True
            
            # リセット時は全て空/無効値を入れるが、DataFrameの行数と合わせるため追加が必要
            # ただし、ループは全ての行に対して回るので、ここで追加する値が「その行の値」になる
            # リセット直後の行（=新しい台の1行目）の処理は下流で行うので、ここではcontinueしない方が良いが
            # prev_id logic matches flow.
            # ここでは状態リセットのみ。
            
            # --- ここで continue すると append がスキップされるので注意 ---
            # 元のコードでは continue していた。
            # 今回はリストに必ず値を入れる必要がある（行ズレ防止）。
            # リセット時の行も処理対象となる（e.g. 朝イチHeavenなど）
            prev_id = row_id # ID更新
            
            # 既存ロジックでは continue して次の行へ行っていた（active判定等は次ループで）。
            # しかしリスト追加は必須。
            # ここは「新しい台の1行目」のコンテキストで処理を続行させる。
            # activeチェック等をこの後で行う。
        
        # 前回の天国が3連以上でなければスキップ -> activeフラグで管理
        # 元コード: continueしていた。
        # 今回: continueするとリストが埋まらない。
        # -> 「対象外」として埋める。
        
        if not trackers['全体']['active']:
            # activeでない期間（前任が単発天国等）
            col_r3_through.append(None)
            col_r3_end.append(None)
            col_r3_heaven.append(None)
            for cat in analysis_data:
                cols_loc[cat].append('非アクティブ')
                
            # 次回アクティブ化判定
            if chain_len >= 3:
                for cat in trackers:
                    trackers[cat]['active'] = True
                    trackers[cat]['hits_since'] = 0
                    trackers[cat]['first_2chain_idx'] = None
                    trackers[cat]['ignore'] = False
            prev_id = row_id
            continue
            
        # ここからアクティブな処理
        is_true_heaven = (chain_len >= 3)
        
        # 今回のチェーン情報の解析
        current_2chain_details = {}
        is_valid_2chain = False
        
        if chain_len == 2:
            second_hit_g = hit_games[1] if len(hit_games) > 1 else 0
            if second_hit_g >= 34:
                is_valid_2chain = False
            else:
                is_valid_2chain = True
                status_2 = hit_statuses[1] if len(hit_statuses) > 1 else ''
                is_rb = 'RB' in status_2
                is_bb = 'BB' in status_2
                g = second_hit_g
                rng = None
                if (2 <= g <= 5) or (g == 32): rng = '[2-5,32]'
                elif 6 <= g <= 12: rng = '[6-12]'
                elif 13 <= g <= 21: rng = '[13-21]'
                elif 22 <= g <= 31: rng = '[22-31]'
                current_2chain_details = {
                    'is_1g': (g == 1),
                    'is_rb': is_rb,
                    'is_bb': is_bb,
                    'range': rng
                }

        # 共通値
        current_through = trackers['全体']['hits_since']
        
        # End Through計算 (Excel数式用: Start <= T < End)
        # 天国の場合は Start == T (End不要だが、Startを入れておく)
        if is_true_heaven:
            end_through = current_through
        else:
            # 天国でない場合、Chain Length分だけスルー回数を消費する
            end_through = current_through + chain_len

        col_r3_through.append(current_through)
        col_r3_end.append(end_through)
        col_r3_heaven.append(is_true_heaven)
        
        # 各カテゴリごとに処理
        for cat in analysis_data:
            tr = trackers[cat]
            
            # --- まず状態更新（2連検出を先に行う）---
            if is_true_heaven:
                # 天国到達 -> リセットは後で（ラベル決定後）
                pass
            else:
                # まだ天国でない（スルー中）
                if not tr['ignore']:
                    if is_valid_2chain and tr['first_2chain_idx'] is None:
                        # まだ2連未発生状態で、今回2連が発生した
                        # この2連が「このカテゴリの条件」を満たすか？
                        match = True
                        det = current_2chain_details
                        if cat == '全体': match = True
                        elif cat == '2連目1G_RB': match = (det['is_rb'] and det['is_1g'])
                        elif cat == '2連目1G_BB': match = (det['is_bb'] and det['is_1g'])
                        elif cat == '2連目[2-5,32]': match = (det['range'] == '[2-5,32]')
                        elif cat == '2連目[6-12]': match = (det['range'] == '[6-12]')
                        elif cat == '2連目[13-21]': match = (det['range'] == '[13-21]')
                        elif cat == '2連目[22-31]': match = (det['range'] == '[22-31]')
                        elif cat == '2連目BB_[2-5,32]': match = (det['is_bb'] and det['range'] == '[2-5,32]')
                        elif cat == '2連目BB_[6-12]': match = (det['is_bb'] and det['range'] == '[6-12]')
                        elif cat == '2連目BB_[13-21]': match = (det['is_bb'] and det['range'] == '[13-21]')
                        elif cat == '2連目BB_[22-31]': match = (det['is_bb'] and det['range'] == '[22-31]')
                        elif cat == '2連目RB_[2-5,32]': match = (det['is_rb'] and det['range'] == '[2-5,32]')
                        elif cat == '2連目RB_[6-12]': match = (det['is_rb'] and det['range'] == '[6-12]')
                        elif cat == '2連目RB_[13-21]': match = (det['is_rb'] and det['range'] == '[13-21]')
                        elif cat == '2連目RB_[22-31]': match = (det['is_rb'] and det['range'] == '[22-31]')
                        
                        if match:
                            tr['first_2chain_idx'] = tr['hits_since']
                        else:
                            tr['ignore'] = True
            
            # Locationラベル決定（2連検出後なので、2連発生時はそのスルー目ラベルになる）
            if tr['ignore']:
                loc_label = '対象外'
            elif not tr['active']:
                loc_label = '非アクティブ'
            elif tr['first_2chain_idx'] is None:
                loc_label = '2連なし'
            else:
                loc_label = f"{tr['first_2chain_idx']}スルー目"
            
            cols_loc[cat].append(loc_label)

            # --- データ記録（集計用リストへ） ---
            if not tr['ignore']:
                analysis_data[cat].append({
                    '2Chain_Loc': loc_label,
                    'Through': tr['hits_since'],
                    'Is_True_Heaven': is_true_heaven
                })
            
            # --- 状態更新（スルー回数加算とリセット） ---
            if is_true_heaven:
                tr['hits_since'] = 0
                tr['first_2chain_idx'] = None
                tr['ignore'] = False
            else:
                if not tr['ignore']:
                            
                    # スルー回数加算
                    # ユーザー要望: 2連（失敗）は2スルー分としてカウント（3連以上基準のためボーナス回数でカウント）
                    # 単発なら+1、2連なら+2
                    tr['hits_since'] += chain_len

        prev_id = row_id
    
    # DataFrameに列を追加
    chains_df['R3_Through'] = col_r3_through
    chains_df['R3_End'] = col_r3_end
    chains_df['R3_Heaven'] = col_r3_heaven
    for cat, vals in cols_loc.items():
        chains_df[f'Loc_{cat}'] = vals
        
    # 集計とクロス集計作成
    results = []
    loc_order = ['2連なし'] + [f'{i}スルー目' for i in range(11)] + ['11スルー以上']
    
    # カテゴリ順序
    cat_order = [
         '全体', 
         '2連目1G_RB', '2連目1G_BB',
         '2連目[2-5,32]', '2連目[6-12]', '2連目[13-21]', '2連目[22-31]',
         '2連目BB_[2-5,32]', '2連目BB_[6-12]', '2連目BB_[13-21]', '2連目BB_[22-31]',
         '2連目RB_[2-5,32]', '2連目RB_[6-12]', '2連目RB_[13-21]', '2連目RB_[22-31]'
    ]

    for cat in cat_order:
        data_list = analysis_data[cat]
        if not data_list: continue
        
        df_cat = pd.DataFrame(data_list)
        
        # --- 1. 天国移行率 ---
        results.append({'2連発生位置': f'【{cat} - 天国移行率】'})
        
        for loc in loc_order:
            row_subset = df_cat[df_cat['2Chain_Loc'] == loc] if not df_cat.empty else pd.DataFrame()
            # 空でもスキップしない（行構造を固定するため）
            
            row_data = {'2連発生位置': loc}
            for thr in range(11):
                subset = row_subset[row_subset['Through'] == thr]
                total = len(subset)
                success = len(subset[subset['Is_True_Heaven'] == True])
                rate = success / total if total > 0 else None
                row_data[f'{thr}スルー'] = rate
            
            # 11以降
            subset_over = row_subset[row_subset['Through'] >= 11]
            total = len(subset_over)
            success = len(subset_over[subset_over['Is_True_Heaven'] == True])
            rate = success / total if total > 0 else None
            row_data['11スルー以上'] = rate
            
            # 合計
            total_all = len(row_subset)
            success_all = len(row_subset[row_subset['Is_True_Heaven'] == True])
            rate_all = success_all / total_all if total_all > 0 else None
            row_data['合計'] = rate_all
            
            results.append(row_data)

        # --- 2. 天国移行数（分子） ---
        results.append({'2連発生位置': f'【{cat} - 天国移行数】'})
        
        for loc in loc_order:
            row_subset = df_cat[df_cat['2Chain_Loc'] == loc] if not df_cat.empty else pd.DataFrame()
            # 空でもスキップしない
            
            row_data = {'2連発生位置': loc}
            for thr in range(11):
                subset = row_subset[row_subset['Through'] == thr]
                success = len(subset[subset['Is_True_Heaven'] == True])
                row_data[f'{thr}スルー'] = success
            
            subset_over = row_subset[row_subset['Through'] >= 11]
            success = len(subset_over[subset_over['Is_True_Heaven'] == True])
            row_data['11スルー以上'] = success
            
            # 合計
            success_all = len(row_subset[row_subset['Is_True_Heaven'] == True])
            row_data['合計'] = success_all
            
            results.append(row_data)

        # --- 3. サンプル数（分母） ---
        results.append({'2連発生位置': f'【{cat} - サンプル数】'})
        for loc in loc_order:
            row_subset = df_cat[df_cat['2Chain_Loc'] == loc] if not df_cat.empty else pd.DataFrame()
            # 空でもスキップしない
            
            row_data = {'2連発生位置': loc}
            for thr in range(11):
                subset = row_subset[row_subset['Through'] == thr]
                row_data[f'{thr}スルー'] = len(subset)
            
            subset_over = row_subset[row_subset['Through'] >= 11]
            row_data['11スルー以上'] = len(subset_over)
            
            row_data['合計'] = len(row_subset)
            
            results.append(row_data)

        # 空行
        results.append({})
        
    return pd.DataFrame(results)


def calculate_dokihana_analysis(chains_df):
    """
    ドキハナチャンス分析
    - 天国連チャン中に34Gまたは35Gで当たりが発生（Position>=2）
    - その後も連チャンが続く（合計3連以上）
    - chains_dfにドキハナ判定列を追加
    """
    # ドキハナ列を初期化
    chains_df['Dokihana_Success'] = False
    chains_df['Dokihana_Start_Pos'] = 0
    chains_df['Dokihana_Chain_Len'] = 0
    chains_df['Dokihana_Dedama'] = 0
    
    for idx in chains_df.index:
        chain_len = chains_df.at[idx, 'Chain_Length']
        is_heaven = chains_df.at[idx, 'Is_Heaven']
        hit_games = chains_df.at[idx, 'Hit_Games']
        
        # ドキハナ判定: 天国連チャン中にPosition>=2で34か35Gがあり、その後2連以上続く
        if is_heaven and chain_len >= 3 and len(hit_games) >= 2:
            hit_dedamas = chains_df.at[idx, 'Hit_Dedamas']
            
            # Position>=2で34か35Gを探す
            for pos, g in enumerate(hit_games[1:], start=2):  # Position 2から
                if g in [34, 35]:
                    dokihana_len = chain_len - pos + 1
                    # ドキハナ連チャン数が2以上の場合のみ成功
                    if dokihana_len >= 2:
                        chains_df.at[idx, 'Dokihana_Success'] = True
                        chains_df.at[idx, 'Dokihana_Start_Pos'] = pos
                        chains_df.at[idx, 'Dokihana_Chain_Len'] = dokihana_len
                        
                        # ドキハナ開始位置以降の純増を計算
                        # pos-1はインデックス（0始まり）
                        dokihana_dedamas = sum(hit_dedamas[pos-1:])  # 出玉合計
                        dokihana_games = hit_games[pos-1:]  # 当選G数リスト
                        # ドキハナ区間の天国消費（初回除く = pos以降の当選G数合計）
                        dokihana_invest = sum(dokihana_games[1:]) / COIN_HOLD * 50 if len(dokihana_games) > 1 else 0
                        # 純増 = 出玉 - 天国消費
                        dokihana_net = dokihana_dedamas - dokihana_invest
                        chains_df.at[idx, 'Dokihana_Dedama'] = int(dokihana_net)
                    break
    
    # 統計計算（コンソール表示用）
    dokihana_success_df = chains_df[chains_df['Dokihana_Success'] == True]
    
    stats = {
        'total_count': len(dokihana_success_df),
        'avg_chain_len': dokihana_success_df['Dokihana_Chain_Len'].mean() if len(dokihana_success_df) > 0 else 0,
        'avg_dedama': dokihana_success_df['Dokihana_Dedama'].mean() if len(dokihana_success_df) > 0 else 0,
        'total_dedama': dokihana_success_df['Dokihana_Dedama'].sum() if len(dokihana_success_df) > 0 else 0
    }
    
    return chains_df, stats


def write_excel(df, heaven_rate_df, heaven_by_chain_df, crosstab_df, rule3_df, chains_df):
    """エクセルファイルに出力"""
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        # Dataシート - 基本データ出力
        data_cols = ['ID', 'Status', 'Start', 'Dedama', 'Time', 'Chain_ID', 'Chain_Position']
        df_out = df[data_cols].copy()
        df_out.columns = ['ID', 'Status', 'Start', 'Dedama', 'Time', '連チャンID', '連チャン位置']
        df_out.to_excel(writer, sheet_name='Data', index=False)
        
        ws_data = writer.sheets['Data']
        max_r = ws_data.max_row
        
        # 追加列ヘッダー (H列から)
        extra_headers = ['朝イチ連', '投資詳細(枚)', '差枚詳細(枚)', '当日差枚', '有利区間G数']
        for i, h in enumerate(extra_headers):
            ws_data.cell(row=1, column=8 + i, value=h)
        
        # 各行に数式を追加
        for r in range(2, max_r + 1):
            prev_r = r - 1
            
            # H: 朝イチ連 (column 8) - 台切り替え判定
            # IDが同じなら連チャン位置1でも朝イチ継続
            if r == 2:
                ws_data.cell(row=r, column=8, value=True)
            else:
                f_same_id = f"A{r}=A{prev_r}"
                f_morn = f"=IF(G{r}=1, IF({f_same_id}, H{prev_r}, TRUE), H{prev_r})"
                ws_data.cell(row=r, column=8, value=f_morn)
            
            # I: 投資詳細(枚) (column 9) = Start / COIN_HOLD * 50
            ws_data.cell(row=r, column=9, value=f"=C{r}/{COIN_HOLD}*50")
            
            # J: 差枚詳細(枚) (column 10) = Dedama - 投資
            ws_data.cell(row=r, column=10, value=f"=D{r}-I{r}")
            
            # K: 当日差枚 (column 11) - 前回終了時点の累積差枚
            if r == 2:
                ws_data.cell(row=r, column=11, value=0)
            else:
                f_same_id = f"A{r}=A{prev_r}"
                f_daily = f"=IF({f_same_id}, K{prev_r}+J{prev_r}, 0)"
                ws_data.cell(row=r, column=11, value=f_daily)
            
            # L: 有利区間G数 (column 12) - 当選時点の有利区間G数 (BB=51G, RB=21G)
            f_bonus_g = f'IF(ISNUMBER(SEARCH("BB", B{r})), 51, IF(ISNUMBER(SEARCH("RB", B{r})), 21, 0))'
            f_current_g = f"C{r}+{f_bonus_g}"
            
            # 有利区間リセット判定をどうするか？暫定で「連チャン位置1」でリセットとするか、あるいは累積し続けるか
            # ここではシンプルに「累積」としているが、実際には有利区間切れでリセットが必要。
            # ただし仕様が複雑なため、単純累積で実装
            if r == 2:
                ws_data.cell(row=r, column=12, value=f"=C{r}") # 初回はStartのみ
            else:
                f_same_id = f"A{r}=A{prev_r}"
                f_yuuri = f"=IF({f_same_id}, L{prev_r}+{f_current_g}, {f_current_g})"
                ws_data.cell(row=r, column=12, value=f_yuuri)
        
        # 天国移行率シート（スルー回数別）
        heaven_rate_df.to_excel(writer, sheet_name='天国移行率_スルー', index=False)
        ws = writer.sheets['天国移行率_スルー']
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
            for cell in row:
                cell.number_format = '0.0%'
        ws.column_dimensions['A'].width = 15
        
        # 天国移行率シート（前回連チャン別）
        heaven_by_chain_df.to_excel(writer, sheet_name='天国移行率_前回連', index=False)
        ws2 = writer.sheets['天国移行率_前回連']
        for row in ws2.iter_rows(min_row=2, max_col=ws2.max_column):
            for cell in row:
                cell.number_format = '0.0%'
        ws2.column_dimensions['A'].width = 15
        
        # クロス集計シート（天国移行率とサンプル数交互 + 特殊判定分析統合）
        crosstab_df.to_excel(writer, sheet_name='クロス集計', index=False)
        ws3 = writer.sheets['クロス集計']

        # パーセント表示（天国移行率セクションのみ）
        current_section = None

        for row_idx in range(2, ws3.max_row + 1):
            a_cell_value = ws3.cell(row=row_idx, column=1).value

            # セクションヘッダーを検出
            if a_cell_value and isinstance(a_cell_value, str) and '【' in a_cell_value:
                current_section = a_cell_value
                continue

            # 空行をスキップ
            if not a_cell_value:
                continue

            # 天国移行率セクションのみパーセント表示
            if current_section and '天国移行率' in current_section:
                for col in range(2, 13):  # B-L列
                    cell = ws3.cell(row=row_idx, column=col)
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.0%'

        ws3.column_dimensions['A'].width = 25  # セクション名が長いので幅を広げる

        # 3連基準2連スルー分析シート
        rule3_df.to_excel(writer, sheet_name='3連基準2連スルー分析', index=False)
        ws4 = writer.sheets['3連基準2連スルー分析']
        
        # パーセント表示＆数式適用（天国移行率セクションのみ）
        current_section = None


        
        # ChainDataシート (日本語ヘッダー作成)
        rename_dict = {
            'ID': 'ID',
            'Chain_ID': '連チャンID',
            'Chain_Number': '台内連チャン番号',
            'Chain_Length': '連チャン長',
            'Is_Heaven': '天国判定',
            'First_G': '初当G',
            'Through_Before': 'スルー回数',
            'Prev_Chain_Length': '前回連チャン長',
            'Raw_Dedama': '出玉合計',
            'Total_Dedama': '純増枚数',
            'Total_Invest': '投資枚数',
            'Heaven_Invest': '天国中消費',
            'Net_Diff': '差枚',
            'Special_Judge': '特殊判定',
            'Prev_Special_Judge': '前回特殊判定',
            'R3_Through': '3連基準スルー',
            'R3_End': '3連基準End',
            'R3_Heaven': '3連基準天国',
            'Dokihana_Success': 'ドキハナ成功',
            'Dokihana_Start_Pos': 'ドキハナ開始位置',
            'Dokihana_Chain_Len': 'ドキハナ連チャン数',
            'Dokihana_Dedama': 'ドキハナ獲得枚数',
            'Daily_Balance_Before': '当日差枚_開始前',
            'Max_Daily_Balance_Before': '当日差枚最大_開始前',
            'Is_Reset': 'リセット判定'
        }
        # Loc列もリネーム
        for col in chains_df.columns:
            if col.startswith('Loc_'):
                cat_name = col.replace('Loc_', '')
                rename_dict[col] = f'判定_{cat_name}'
                
        chains_jp = chains_df.rename(columns=rename_dict)
        chains_jp.to_excel(writer, sheet_name='ChainData', index=False)
        ws_chain = writer.sheets['ChainData']
        ws_chain.column_dimensions['A'].width = 35
        
        # ChainDataの列マッピング取得
        col_map = {}
        from openpyxl.utils import get_column_letter
        for col in range(1, ws_chain.max_column + 1):
            val = ws_chain.cell(1, col).value
            if val:
                col_map[val] = get_column_letter(col)
        
        # 参照用カラムレター
        col_r3_thr = col_map.get('3連基準スルー')
        col_r3_end = col_map.get('3連基準End')
        col_r3_hvn = col_map.get('3連基準天国')
        col_len = col_map.get('連チャン長')
        
        # 3連基準2連スルー分析シート（数式適用）
        current_section = None
        current_cat = None
        
        # データ構造に基づくオフセット
        OFFSET_SUCCESS = 14
        OFFSET_SAMPLE = 28

        for row_idx in range(2, ws4.max_row + 1):
            a_cell_value = ws4.cell(row=row_idx, column=1).value

            # セクションヘッダー検出
            if a_cell_value and isinstance(a_cell_value, str) and '【' in a_cell_value:
                current_section = a_cell_value
                # カテゴリ名抽出: "【全体 - 天国移行率】" -> "全体"
                part1 = a_cell_value.split(' - ')[0] # "【全体"
                current_cat = part1.replace('【', '')
                continue

            if not a_cell_value: continue
            
            # --- 数式適用 ---
            loc_label = a_cell_value # "2連なし", "0スルー目", etc.
            
            # カテゴリ判定列
            col_judge = col_map.get(f'判定_{current_cat}')
            if not col_judge: continue # マッピングエラー回避

            # 1. 天国移行率セクション (=Success/Sample, 0や空は[-])
            if current_section and '天国移行率' in current_section:
                for col in range(2, 15): 
                    cell_rate = ws4.cell(row=row_idx, column=col)
                    cell_success = ws4.cell(row=row_idx + OFFSET_SUCCESS, column=col)
                    cell_sample = ws4.cell(row=row_idx + OFFSET_SAMPLE, column=col)
                    addr_success = cell_success.coordinate
                    addr_sample = cell_sample.coordinate
                    # サンプル0 or 成功0 → [-]
                    formula = f'=IF(OR({addr_sample}=0, {addr_sample}=""), "[-]", IF({addr_success}=0, "[-]", {addr_success}/{addr_sample}))'
                    cell_rate.value = formula
                    cell_rate.number_format = '0.0%'

            # 2. 天国移行数セクション (COUNTIFS TrueHeaven=True)
            elif current_section and '天国移行数' in current_section:
                for col_idx, col_header in enumerate(rule3_df.columns):
                    if col_header == '2連発生位置': continue
                    col_excel = col_idx + 1 # 1-based index (A=1, B=2...)
                    cell = ws4.cell(row=row_idx, column=col_excel)
                    
                    # 条件作成
                    base_cond = f"ChainData!${col_judge}:${col_judge}, \"{loc_label}\", ChainData!${col_r3_hvn}:${col_r3_hvn}, TRUE"
                    
                    formula = None
                    if 'スルー' in col_header and '以上' not in col_header:
                        thr_val = col_header.replace('スルー', '')
                        formula = f'=COUNTIFS({base_cond}, ChainData!${col_r3_thr}:${col_r3_thr}, {thr_val})'
                    elif '11スルー以上' in col_header:
                        formula = f'=COUNTIFS({base_cond}, ChainData!${col_r3_thr}:${col_r3_thr}, ">=11")'
                    elif col_header == '合計':
                        formula = f'=COUNTIFS({base_cond})'
                    
                    if formula:
                        # 0件は[-]表示
                        expr = formula[1:] # =を除く
                        cell.value = f'=IF({expr}=0, "[-]", {expr})'

            # 3. サンプル数セクション
            # ロジック: 天国でない場合(Start<=T<End) + 天国の場合(Start=T)
            # つまり、2連(非天国)なら StartとStart+1 の2つのスルー回数でサンプルとなる
            elif current_section and 'サンプル数' in current_section:
                for col_idx, col_header in enumerate(rule3_df.columns):
                    if col_header == '2連発生位置': continue
                    col_excel = col_idx + 1
                    cell = ws4.cell(row=row_idx, column=col_excel)
                    
                    # 共通: Judge=Loc
                    cond_loc = f"ChainData!${col_judge}:${col_judge}, \"{loc_label}\""
                    
                    formula = None
                    if 'スルー' in col_header and '以上' not in col_header:
                        thr_val = col_header.replace('スルー', '')
                        # Non-Heaven: Start<=T, End>T
                        cnt_non_hvn = f"COUNTIFS({cond_loc}, ChainData!${col_r3_hvn}:${col_r3_hvn}, FALSE, ChainData!${col_r3_thr}:${col_r3_thr}, \"<=\"&{thr_val}, ChainData!${col_r3_end}:${col_r3_end}, \">\"&{thr_val})"
                        # Heaven: Start=T
                        cnt_hvn = f"COUNTIFS({cond_loc}, ChainData!${col_r3_hvn}:${col_r3_hvn}, TRUE, ChainData!${col_r3_thr}:${col_r3_thr}, {thr_val})"
                        formula = f'={cnt_non_hvn} + {cnt_hvn}'
                        
                    elif '11スルー以上' in col_header:
                        # Non-Heaven: End > 11 (covers 11+)
                        cnt_non_hvn = f"COUNTIFS({cond_loc}, ChainData!${col_r3_hvn}:${col_r3_hvn}, FALSE, ChainData!${col_r3_end}:${col_r3_end}, \">11\")"
                        # Heaven: Start >= 11
                        cnt_hvn = f"COUNTIFS({cond_loc}, ChainData!${col_r3_hvn}:${col_r3_hvn}, TRUE, ChainData!${col_r3_thr}:${col_r3_thr}, \">=11\")"
                        formula = f'={cnt_non_hvn} + {cnt_hvn}'
                        
                    elif col_header == '合計':
                        # 合計サンプル数 = 非天国の総ボーナス数(ChainLength合計) + 天国回数(Chain数)
                        # SUMIFS(Length, NonHeaven) + COUNTIFS(Heaven)
                        sum_non_hvn = f"SUMIFS(ChainData!${col_len}:${col_len}, {cond_loc}, ChainData!${col_r3_hvn}:${col_r3_hvn}, FALSE)"
                        cnt_hvn = f"COUNTIFS({cond_loc}, ChainData!${col_r3_hvn}:${col_r3_hvn}, TRUE)"
                        formula = f'={sum_non_hvn} + {cnt_hvn}'
                        
                    if formula:
                        # 0件は[-]表示
                        expr = formula[1:] # =を除く
                        cell.value = f'=IF({expr}=0, "[-]", {expr})'

        ws4.column_dimensions['A'].width = 25

        # ========== P列以降: 新形式ピボットテーブル ==========
        # 行: カテゴリ（状況）
        # 列: スルー回数 (0〜10, 11以上) - B〜M列形式
        # 各2連発生スルー目ごとにセクション
        
        pivot_start_col = 16  # P列 = 16
        
        # スタイル定義
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # 青
        header_font = Font(bold=True, color='FFFFFF')  # 白文字
        label_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')  # 薄青
        label_font = Font(bold=True)
        row_fill_odd = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # 白
        row_fill_even = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # 薄灰色
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # カテゴリ順序 (行)
        categories = [
            '全体',
            '2連目1G_RB', '2連目1G_BB',
            '2連目[2-5,32]', '2連目[6-12]', '2連目[13-21]', '2連目[22-31]',
            '2連目BB_[2-5,32]', '2連目BB_[6-12]', '2連目BB_[13-21]', '2連目BB_[22-31]',
            '2連目RB_[2-5,32]', '2連目RB_[6-12]', '2連目RB_[13-21]', '2連目RB_[22-31]'
        ]
        
        # 2連発生スルー目のリスト
        loc_list = ['2連なし'] + [f'{i}スルー目' for i in range(11)] + ['11スルー以上']
        
        # 2連発生基準のスルー回数（0〜9の10列）
        relative_through_count = 10
        
        current_row = 1
        
        for loc_idx, loc_label in enumerate(loc_list):
            section_start_row = current_row
            
            # セクションヘッダー行
            for c in range(pivot_start_col, pivot_start_col + relative_through_count + 1):
                cell = ws4.cell(row=current_row, column=c)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
            ws4.cell(row=current_row, column=pivot_start_col, value=f'【{loc_label}】')
            current_row += 1
            
            # 2連発生位置からオフセットを計算
            if loc_label == '2連なし':
                base_offset = 0
            elif loc_label == '11スルー以上':
                base_offset = 11 + 2
            else:
                loc_num = int(loc_label.replace('スルー目', ''))
                base_offset = loc_num + 2
            
            # ラベル行1: 3連後基準スルー
            for c in range(pivot_start_col, pivot_start_col + relative_through_count + 1):
                cell = ws4.cell(row=current_row, column=c)
                cell.fill = label_fill
                cell.font = label_font
                cell.border = thin_border
            ws4.cell(row=current_row, column=pivot_start_col, value='3連後基準スルー')
            for rel_idx in range(relative_through_count):
                col = pivot_start_col + 1 + rel_idx
                if loc_label == '2連なし':
                    ws4.cell(row=current_row, column=col, value=f'{rel_idx}スルー')
                else:
                    absolute_through = base_offset + rel_idx
                    ws4.cell(row=current_row, column=col, value=f'{absolute_through}スルー')
            current_row += 1
            
            # ラベル行2: 2連発生基準スルー
            for c in range(pivot_start_col, pivot_start_col + relative_through_count + 1):
                cell = ws4.cell(row=current_row, column=c)
                cell.fill = label_fill
                cell.font = label_font
                cell.border = thin_border
            ws4.cell(row=current_row, column=pivot_start_col, value='2連発生基準スルー')
            for rel_idx in range(relative_through_count):
                col = pivot_start_col + 1 + rel_idx
                ws4.cell(row=current_row, column=col, value=f'{rel_idx}スルー')
            current_row += 1
            
            # 各カテゴリの行
            for cat_idx, cat in enumerate(categories):
                row = current_row
                row_fill = row_fill_even if cat_idx % 2 == 0 else row_fill_odd
                
                # 行ラベル (カテゴリ名)
                cell = ws4.cell(row=row, column=pivot_start_col, value=cat)
                cell.fill = row_fill
                cell.border = thin_border
                
                col_judge = col_map.get(f'判定_{cat}')
                if not col_judge:
                    current_row += 1
                    continue
                
                cond_loc = f'ChainData!${col_judge}:${col_judge}, "{loc_label}"'
                
                # 各スルー回数の列（2連発生基準で0～9スルー）
                for rel_idx in range(relative_through_count):
                    col = pivot_start_col + 1 + rel_idx
                    
                    if loc_label == '2連なし':
                        thr_val = rel_idx
                    else:
                        thr_val = base_offset + rel_idx
                    
                    cond_thr = f'ChainData!${col_r3_thr}:${col_r3_thr}, {thr_val}'
                    cnt_hvn = f'COUNTIFS({cond_loc}, ChainData!${col_r3_hvn}:${col_r3_hvn}, TRUE, {cond_thr})'
                    cnt_sample_non = f'COUNTIFS({cond_loc}, ChainData!${col_r3_hvn}:${col_r3_hvn}, FALSE, ChainData!${col_r3_thr}:${col_r3_thr}, "<="&{thr_val}, ChainData!${col_r3_end}:${col_r3_end}, ">"&{thr_val})'
                    cnt_sample_hvn = f'COUNTIFS({cond_loc}, ChainData!${col_r3_hvn}:${col_r3_hvn}, TRUE, {cond_thr})'
                    sample_expr = f'{cnt_sample_non}+{cnt_sample_hvn}'
                    
                    # 天国移行率 = 分子 / 分母 (0の場合は[-])
                    formula = f'=IF(OR({sample_expr}=0, {cnt_hvn}=0), "[-]", {cnt_hvn}/({sample_expr}))'
                    cell = ws4.cell(row=row, column=col)
                    cell.value = formula
                    cell.number_format = '0.0%'
                    cell.fill = row_fill
                    cell.border = thin_border
                
                current_row += 1
            
            # セクション間の空行
            current_row += 1
        
        # P列幅調整
        ws4.column_dimensions[get_column_letter(pivot_start_col)].width = 20

        # ========== AD列以降: サンプル数テーブル ==========
        sample_start_col = 30  # AD列 = 30
        
        # サンプル数テーブル用のスタイル（緑系）
        sample_header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')  # 緑
        sample_label_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')  # 薄緑
        
        current_row = 1
        
        for loc_idx, loc_label in enumerate(loc_list):
            # セクションヘッダー行
            for c in range(sample_start_col, sample_start_col + relative_through_count + 1):
                cell = ws4.cell(row=current_row, column=c)
                cell.fill = sample_header_fill
                cell.font = header_font
                cell.border = thin_border
            ws4.cell(row=current_row, column=sample_start_col, value=f'【{loc_label}】サンプル数')
            current_row += 1
            
            # 2連発生位置からオフセットを計算
            if loc_label == '2連なし':
                base_offset = 0
            elif loc_label == '11スルー以上':
                base_offset = 11 + 2
            else:
                loc_num = int(loc_label.replace('スルー目', ''))
                base_offset = loc_num + 2
            
            # ラベル行1: 3連後基準スルー
            for c in range(sample_start_col, sample_start_col + relative_through_count + 1):
                cell = ws4.cell(row=current_row, column=c)
                cell.fill = sample_label_fill
                cell.font = label_font
                cell.border = thin_border
            ws4.cell(row=current_row, column=sample_start_col, value='3連後基準スルー')
            for rel_idx in range(relative_through_count):
                col = sample_start_col + 1 + rel_idx
                if loc_label == '2連なし':
                    ws4.cell(row=current_row, column=col, value=f'{rel_idx}スルー')
                else:
                    absolute_through = base_offset + rel_idx
                    ws4.cell(row=current_row, column=col, value=f'{absolute_through}スルー')
            current_row += 1
            
            # ラベル行2: 2連発生基準スルー
            for c in range(sample_start_col, sample_start_col + relative_through_count + 1):
                cell = ws4.cell(row=current_row, column=c)
                cell.fill = sample_label_fill
                cell.font = label_font
                cell.border = thin_border
            ws4.cell(row=current_row, column=sample_start_col, value='2連発生基準スルー')
            for rel_idx in range(relative_through_count):
                col = sample_start_col + 1 + rel_idx
                ws4.cell(row=current_row, column=col, value=f'{rel_idx}スルー')
            current_row += 1
            
            # 各カテゴリの行
            for cat_idx, cat in enumerate(categories):
                row = current_row
                row_fill = row_fill_even if cat_idx % 2 == 0 else row_fill_odd
                
                # 行ラベル (カテゴリ名)
                cell = ws4.cell(row=row, column=sample_start_col, value=cat)
                cell.fill = row_fill
                cell.border = thin_border
                
                col_judge = col_map.get(f'判定_{cat}')
                if not col_judge:
                    current_row += 1
                    continue
                
                cond_loc = f'ChainData!${col_judge}:${col_judge}, "{loc_label}"'
                
                # 各スルー回数の列（2連発生基準で0～9スルー）
                for rel_idx in range(relative_through_count):
                    col = sample_start_col + 1 + rel_idx
                    
                    if loc_label == '2連なし':
                        thr_val = rel_idx
                    else:
                        thr_val = base_offset + rel_idx
                    
                    cnt_sample_non = f'COUNTIFS({cond_loc}, ChainData!${col_r3_hvn}:${col_r3_hvn}, FALSE, ChainData!${col_r3_thr}:${col_r3_thr}, "<="&{thr_val}, ChainData!${col_r3_end}:${col_r3_end}, ">"&{thr_val})'
                    cnt_sample_hvn = f'COUNTIFS({cond_loc}, ChainData!${col_r3_hvn}:${col_r3_hvn}, TRUE, ChainData!${col_r3_thr}:${col_r3_thr}, {thr_val})'
                    sample_expr = f'{cnt_sample_non}+{cnt_sample_hvn}'
                    
                    # サンプル数 (0の場合は[-])
                    formula = f'=IF({sample_expr}=0, "[-]", {sample_expr})'
                    cell = ws4.cell(row=row, column=col)
                    cell.value = formula
                    cell.fill = row_fill
                    cell.border = thin_border
                
                current_row += 1
            
            # セクション間の空行
            current_row += 1
        
        # AD列幅調整
        ws4.column_dimensions[get_column_letter(sample_start_col)].width = 20

        # ========== ドキハナチャンス分析シート（式計算ベース） ==========
        ws_doki = writer.book.create_sheet("DokihanaAnalysis")
        
        # タイトル
        ws_doki['A1'] = "【ドキハナチャンス分析】"
        ws_doki['A1'].font = Font(bold=True, size=14)
        ws_doki['A2'] = "34Gまたは35Gで発生する特別な天国連チャンの分析"
        
        # ChainDataの行数とドキハナ列参照を取得
        chain_max_row = ws_chain.max_row
        col_doki_success = col_map.get('ドキハナ成功')
        col_doki_len = col_map.get('ドキハナ連チャン数')
        col_doki_dedama = col_map.get('ドキハナ獲得枚数')
        col_doki_start = col_map.get('ドキハナ開始位置')
        col_max_balance = col_map.get('当日差枚最大_開始前')
        
        # ChainDataの範囲参照
        ref_success = f"ChainData!${col_doki_success}$2:${col_doki_success}${chain_max_row}"
        ref_doki_len = f"ChainData!${col_doki_len}$2:${col_doki_len}${chain_max_row}"
        ref_doki_dedama = f"ChainData!${col_doki_dedama}$2:${col_doki_dedama}${chain_max_row}"
        ref_doki_start = f"ChainData!${col_doki_start}$2:${col_doki_start}${chain_max_row}"
        ref_max_balance = f"ChainData!${col_max_balance}$2:${col_max_balance}${chain_max_row}"
        
        # ========== 基本統計 ==========
        ws_doki['A4'] = "■ 基本統計"
        ws_doki['A4'].font = Font(bold=True)
        
        ws_doki['A5'] = "ドキハナ発生回数"
        ws_doki['B5'] = f"=COUNTIF({ref_success}, TRUE)"
        
        ws_doki['A6'] = "平均ドキハナ連チャン数"
        ws_doki['B6'] = f'=IFERROR(AVERAGEIF({ref_success}, TRUE, {ref_doki_len}), "-")'
        ws_doki['B6'].number_format = '0.00'
        
        ws_doki['A7'] = "平均ドキハナ獲得枚数"
        ws_doki['B7'] = f'=IFERROR(AVERAGEIF({ref_success}, TRUE, {ref_doki_dedama})-32/{COIN_HOLD}*50, "-")'
        ws_doki['B7'].number_format = '#,##0'
        
        ws_doki['A8'] = "合計ドキハナ獲得枚数"
        ws_doki['B8'] = f"=SUMIF({ref_success}, TRUE, {ref_doki_dedama})"
        ws_doki['B8'].number_format = '#,##0'
        
        # ========== ドキハナ開始位置別 分析（全体） ==========
        ws_doki['A10'] = "■ ドキハナ開始位置別 分析（全体）"
        ws_doki['A10'].font = Font(bold=True)
        
        ws_doki['A11'] = "開始位置"
        ws_doki['B11'] = "発生回数"
        ws_doki['C11'] = "割合"
        ws_doki['D11'] = "平均ドキハナ連数"
        ws_doki['E11'] = "平均獲得枚数"
        
        row = 12
        for pos in range(2, 21):
            ws_doki.cell(row=row, column=1, value=f"{pos}連目")
            ws_doki.cell(row=row, column=2, value=f"=COUNTIFS({ref_success}, TRUE, {ref_doki_start}, {pos})")
            ws_doki.cell(row=row, column=3, value=f"=IFERROR(B{row}/$B$5, 0)")
            ws_doki.cell(row=row, column=3).number_format = '0.0%'
            ws_doki.cell(row=row, column=4, value=f'=IFERROR(AVERAGEIFS({ref_doki_len}, {ref_success}, TRUE, {ref_doki_start}, {pos}), "-")')
            ws_doki.cell(row=row, column=4).number_format = '0.00'
            ws_doki.cell(row=row, column=5, value=f'=IFERROR(AVERAGEIFS({ref_doki_dedama}, {ref_success}, TRUE, {ref_doki_start}, {pos})-32/{COIN_HOLD}*50, "-")')
            ws_doki.cell(row=row, column=5).number_format = '#,##0'
            row += 1
        
        # 21連目以降
        ws_doki.cell(row=row, column=1, value="21連目以降")
        ws_doki.cell(row=row, column=2, value=f'=COUNTIFS({ref_success}, TRUE, {ref_doki_start}, ">=21")')
        ws_doki.cell(row=row, column=3, value=f"=IFERROR(B{row}/$B$5, 0)")
        ws_doki.cell(row=row, column=3).number_format = '0.0%'
        ws_doki.cell(row=row, column=4, value=f'=IFERROR(AVERAGEIFS({ref_doki_len}, {ref_success}, TRUE, {ref_doki_start}, ">=21"), "-")')
        ws_doki.cell(row=row, column=4).number_format = '0.00'
        ws_doki.cell(row=row, column=5, value=f'=IFERROR(AVERAGEIFS({ref_doki_dedama}, {ref_success}, TRUE, {ref_doki_start}, ">=21")-32/{COIN_HOLD}*50, "-")')
        ws_doki.cell(row=row, column=5).number_format = '#,##0'
        
        # ========== ドキハナ開始位置別 分析（当日差枚最大<1 = 一度もプラスになっていない） ==========
        ws_doki['A35'] = "■ ドキハナ開始位置別 分析（当日差枚が一度も1以上になっていない）"
        ws_doki['A35'].font = Font(bold=True)
        ws_doki['A36'] = "※ その時点までの当日差枚最大が1未満の場合のみカウント"
        
        ws_doki['A37'] = "開始位置"
        ws_doki['B37'] = "発生回数"
        ws_doki['C37'] = "割合"
        ws_doki['D37'] = "平均ドキハナ連数"
        ws_doki['E37'] = "平均獲得枚数"
        
        # 当日差枚条件付きの発生回数合計（割合計算用）
        ws_doki['G35'] = "条件付き合計"
        ws_doki['G36'] = f'=COUNTIFS({ref_success}, TRUE, {ref_max_balance}, "<1")'
        
        row = 38
        for pos in range(2, 21):
            ws_doki.cell(row=row, column=1, value=f"{pos}連目")
            # 条件: ドキハナ成功=TRUE AND 開始位置=pos AND 当日差枚最大<1
            ws_doki.cell(row=row, column=2, value=f'=COUNTIFS({ref_success}, TRUE, {ref_doki_start}, {pos}, {ref_max_balance}, "<1")')
            ws_doki.cell(row=row, column=3, value=f"=IFERROR(B{row}/$G$36, 0)")
            ws_doki.cell(row=row, column=3).number_format = '0.0%'
            ws_doki.cell(row=row, column=4, value=f'=IFERROR(AVERAGEIFS({ref_doki_len}, {ref_success}, TRUE, {ref_doki_start}, {pos}, {ref_max_balance}, "<1"), "-")')
            ws_doki.cell(row=row, column=4).number_format = '0.00'
            ws_doki.cell(row=row, column=5, value=f'=IFERROR(AVERAGEIFS({ref_doki_dedama}, {ref_success}, TRUE, {ref_doki_start}, {pos}, {ref_max_balance}, "<1")-32/{COIN_HOLD}*50, "-")')
            ws_doki.cell(row=row, column=5).number_format = '#,##0'
            row += 1
        
        # 21連目以降
        ws_doki.cell(row=row, column=1, value="21連目以降")
        ws_doki.cell(row=row, column=2, value=f'=COUNTIFS({ref_success}, TRUE, {ref_doki_start}, ">=21", {ref_max_balance}, "<1")')
        ws_doki.cell(row=row, column=3, value=f"=IFERROR(B{row}/$G$36, 0)")
        ws_doki.cell(row=row, column=3).number_format = '0.0%'
        ws_doki.cell(row=row, column=4, value=f'=IFERROR(AVERAGEIFS({ref_doki_len}, {ref_success}, TRUE, {ref_doki_start}, ">=21", {ref_max_balance}, "<1"), "-")')
        ws_doki.cell(row=row, column=4).number_format = '0.00'
        ws_doki.cell(row=row, column=5, value=f'=IFERROR(AVERAGEIFS({ref_doki_dedama}, {ref_success}, TRUE, {ref_doki_start}, ">=21", {ref_max_balance}, "<1")-32/{COIN_HOLD}*50, "-")')
        ws_doki.cell(row=row, column=5).number_format = '#,##0'
        
        # ========== スルー回数別 ドキハナ分析（条件付き） ==========
        # ChainDataのスルー回数列を参照
        col_through = col_map.get('スルー回数')
        ref_through = f"ChainData!${col_through}$2:${col_through}${chain_max_row}"
        
        # スタイル定義
        doki_header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # 青
        doki_header_font = Font(bold=True, color='FFFFFF')  # 白文字
        doki_row_fill_odd = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # 白
        doki_row_fill_even = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # 薄灰色
        doki_total_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')  # 薄青
        doki_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        ws_doki['A62'] = "■ スルー回数別 ドキハナ分析（当日プラス未経験 & 6連目以内開始）"
        ws_doki['A62'].font = Font(bold=True, size=12)
        ws_doki['A63'] = "※ 当日差枚が一度も1以上になっていない & ドキハナ開始位置が6連目以内"
        
        # ヘッダー行（64行目）
        headers = ["スルー回数", "発生回数", "平均ドキハナ連数", "平均獲得枚数"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_doki.cell(row=64, column=col_idx, value=header)
            cell.fill = doki_header_fill
            cell.font = doki_header_font
            cell.border = doki_border
            cell.alignment = Alignment(horizontal='center')
        
        row = 65
        for thr in range(10):  # 0～9スルー
            row_fill = doki_row_fill_even if thr % 2 == 0 else doki_row_fill_odd
            
            # スルー回数ラベル
            cell = ws_doki.cell(row=row, column=1, value=f"{thr}スルー")
            cell.fill = row_fill
            cell.border = doki_border
            
            # 発生回数
            cell = ws_doki.cell(row=row, column=2, value=f'=COUNTIFS({ref_success}, TRUE, {ref_through}, {thr}, {ref_max_balance}, "<1", {ref_doki_start}, "<=6")')
            cell.fill = row_fill
            cell.border = doki_border
            
            # 平均連チャン数
            cell = ws_doki.cell(row=row, column=3, value=f'=IFERROR(AVERAGEIFS({ref_doki_len}, {ref_success}, TRUE, {ref_through}, {thr}, {ref_max_balance}, "<1", {ref_doki_start}, "<=6"), "-")')
            cell.number_format = '0.00'
            cell.fill = row_fill
            cell.border = doki_border
            
            # 平均獲得枚数
            cell = ws_doki.cell(row=row, column=4, value=f'=IFERROR(AVERAGEIFS({ref_doki_dedama}, {ref_success}, TRUE, {ref_through}, {thr}, {ref_max_balance}, "<1", {ref_doki_start}, "<=6")-32/{COIN_HOLD}*50, "-")')
            cell.number_format = '#,##0'
            cell.fill = row_fill
            cell.border = doki_border
            
            row += 1
        
        # 10スルー以上
        row_fill = doki_row_fill_even
        cell = ws_doki.cell(row=row, column=1, value="10スルー以上")
        cell.fill = row_fill
        cell.border = doki_border
        
        cell = ws_doki.cell(row=row, column=2, value=f'=COUNTIFS({ref_success}, TRUE, {ref_through}, ">=10", {ref_max_balance}, "<1", {ref_doki_start}, "<=6")')
        cell.fill = row_fill
        cell.border = doki_border
        
        cell = ws_doki.cell(row=row, column=3, value=f'=IFERROR(AVERAGEIFS({ref_doki_len}, {ref_success}, TRUE, {ref_through}, ">=10", {ref_max_balance}, "<1", {ref_doki_start}, "<=6"), "-")')
        cell.number_format = '0.00'
        cell.fill = row_fill
        cell.border = doki_border
        
        cell = ws_doki.cell(row=row, column=4, value=f'=IFERROR(AVERAGEIFS({ref_doki_dedama}, {ref_success}, TRUE, {ref_through}, ">=10", {ref_max_balance}, "<1", {ref_doki_start}, "<=6")-32/{COIN_HOLD}*50, "-")')
        cell.number_format = '#,##0'
        cell.fill = row_fill
        cell.border = doki_border
        row += 1
        
        # 合計行（薄青背景）
        cell = ws_doki.cell(row=row, column=1, value="合計")
        cell.fill = doki_total_fill
        cell.border = doki_border
        cell.font = Font(bold=True)
        
        cell = ws_doki.cell(row=row, column=2, value=f'=COUNTIFS({ref_success}, TRUE, {ref_max_balance}, "<1", {ref_doki_start}, "<=6")')
        cell.fill = doki_total_fill
        cell.border = doki_border
        cell.font = Font(bold=True)
        
        cell = ws_doki.cell(row=row, column=3, value=f'=IFERROR(AVERAGEIFS({ref_doki_len}, {ref_success}, TRUE, {ref_max_balance}, "<1", {ref_doki_start}, "<=6"), "-")')
        cell.number_format = '0.00'
        cell.fill = doki_total_fill
        cell.border = doki_border
        cell.font = Font(bold=True)
        
        cell = ws_doki.cell(row=row, column=4, value=f'=IFERROR(AVERAGEIFS({ref_doki_dedama}, {ref_success}, TRUE, {ref_max_balance}, "<1", {ref_doki_start}, "<=6")-32/{COIN_HOLD}*50, "-")')
        cell.number_format = '#,##0'
        cell.fill = doki_total_fill
        cell.border = doki_border
        cell.font = Font(bold=True)
        
        # 列幅調整
        ws_doki.column_dimensions['A'].width = 35
        ws_doki.column_dimensions['B'].width = 15
        ws_doki.column_dimensions['C'].width = 18
        ws_doki.column_dimensions['D'].width = 18
        ws_doki.column_dimensions['E'].width = 18

        # ========== 天国当選G数分布シート ==========
        ws_dist = writer.book.create_sheet("HeavenHitDistribution")
        
        ws_dist['A1'] = "【3連以上天国チェーン 当選G数分布】"
        ws_dist['A1'].font = Font(bold=True, size=14)
        ws_dist['A2'] = "※ 3連以上の天国チェーンにおけるPosition 2以降（天国連チャン中）の当選G数分布"
        
        # 3連以上のchainからPosition 2以降のhit_gamesを収集
        all_heaven_hits = []
        for idx in chains_df.index:
            chain_len = chains_df.at[idx, 'Chain_Length']
            is_heaven = chains_df.at[idx, 'Is_Heaven']
            hit_games = chains_df.at[idx, 'Hit_Games']
            
            if is_heaven and chain_len >= 3:
                # Position 2以降（インデックス1以降）の当選G数を収集
                for g in hit_games[1:]:
                    if g <= 35:
                        all_heaven_hits.append(g)
        
        # 分布を集計
        from collections import Counter
        hit_counter = Counter(all_heaven_hits)
        total_hits = len(all_heaven_hits)
        
        # ヘッダー
        ws_dist['A4'] = "当選G数"
        ws_dist['B4'] = "発生回数"
        ws_dist['C4'] = "割合"
        ws_dist['A4'].font = Font(bold=True)
        ws_dist['B4'].font = Font(bold=True)
        ws_dist['C4'].font = Font(bold=True)
        
        row = 5
        for g in range(1, 36):  # 1G〜35G
            ws_dist.cell(row=row, column=1, value=f"{g}G")
            count = hit_counter.get(g, 0)
            ws_dist.cell(row=row, column=2, value=count)
            rate = count / total_hits if total_hits > 0 else 0
            ws_dist.cell(row=row, column=3, value=rate)
            ws_dist.cell(row=row, column=3).number_format = '0.00%'
            row += 1
        
        # 合計行
        ws_dist.cell(row=row, column=1, value="合計")
        ws_dist.cell(row=row, column=1).font = Font(bold=True)
        ws_dist.cell(row=row, column=2, value=total_hits)
        ws_dist.cell(row=row, column=2).font = Font(bold=True)
        ws_dist.cell(row=row, column=3, value=1.0 if total_hits > 0 else 0)
        ws_dist.cell(row=row, column=3).number_format = '0.00%'
        ws_dist.cell(row=row, column=3).font = Font(bold=True)
        
        # 平均G数
        avg_g = sum(all_heaven_hits) / total_hits if total_hits > 0 else 0
        ws_dist.cell(row=row+1, column=1, value="平均G数")
        ws_dist.cell(row=row+1, column=1).font = Font(bold=True)
        ws_dist.cell(row=row+1, column=2, value=avg_g)
        ws_dist.cell(row=row+1, column=2).number_format = '0.0'
        ws_dist.cell(row=row+1, column=2).font = Font(bold=True)
        
        # 列幅調整
        ws_dist.column_dimensions['A'].width = 12
        ws_dist.column_dimensions['B'].width = 12
        ws_dist.column_dimensions['C'].width = 12

        # ========== 天井到達率分析シート ==========
        ws_ceil = writer.book.create_sheet("CeilingAnalysis")
        
        ws_ceil['A1'] = "【天井到達率分析】"
        ws_ceil['A1'].font = Font(bold=True, size=14)
        ws_ceil['A2'] = "※ サンプル数=そのG以上で当選した総数、天井到達=800G以上で当選、到達率=天井到達/サンプル"
        
        # Is_Reset列を参照
        col_reset = col_map.get('リセット判定')
        col_through = col_map.get('スルー回数')
        col_first_g = col_map.get('初当G')
        
        ref_reset = f"ChainData!${col_reset}$2:${col_reset}${chain_max_row}"
        ref_through = f"ChainData!${col_through}$2:${col_through}${chain_max_row}"
        ref_first_g = f"ChainData!${col_first_g}$2:${col_first_g}${chain_max_row}"
        
        # セクション1: リセット0スルー (A-D列)
        ws_ceil['A4'] = "■ リセット時0スルー"
        ws_ceil['A4'].font = Font(bold=True, size=12)
        
        ws_ceil['A5'] = "ゲーム数"
        ws_ceil['B5'] = "サンプル数"
        ws_ceil['C5'] = "天井到達"
        ws_ceil['D5'] = "到達率"
        for col in range(1, 5):
            ws_ceil.cell(row=5, column=col).font = Font(bold=True)
        
        # セクション2: リセット0スルー以外 (F-I列)
        ws_ceil['F4'] = "■ リセット0スルー以外"
        ws_ceil['F4'].font = Font(bold=True, size=12)
        
        ws_ceil['F5'] = "ゲーム数"
        ws_ceil['G5'] = "サンプル数"
        ws_ceil['H5'] = "天井到達"
        ws_ceil['I5'] = "到達率"
        for col in range(6, 10):
            ws_ceil.cell(row=5, column=col).font = Font(bold=True)
        
        # データ行（両セクション同時に出力）
        row = 6
        for g in range(0, 801, 5):
            # セクション1: リセット0スルー (A-D列)
            ws_ceil.cell(row=row, column=1, value=f"{g}G～")
            ws_ceil.cell(row=row, column=2, value=f'=COUNTIFS({ref_reset}, TRUE, {ref_through}, 0, {ref_first_g}, ">={g}")')
            ws_ceil.cell(row=row, column=3, value=f'=COUNTIFS({ref_reset}, TRUE, {ref_through}, 0, {ref_first_g}, ">=800")')
            ws_ceil.cell(row=row, column=4, value=f'=IFERROR(C{row}/B{row}, "-")')
            ws_ceil.cell(row=row, column=4).number_format = '0.00%'
            
            # セクション2: リセット0スルー以外 (F-I列)
            ws_ceil.cell(row=row, column=6, value=f"{g}G～")
            ws_ceil.cell(row=row, column=7, value=f'=COUNTIFS({ref_first_g}, ">={g}")-COUNTIFS({ref_reset}, TRUE, {ref_through}, 0, {ref_first_g}, ">={g}")')
            ws_ceil.cell(row=row, column=8, value=f'=COUNTIFS({ref_first_g}, ">=800")-COUNTIFS({ref_reset}, TRUE, {ref_through}, 0, {ref_first_g}, ">=800")')
            ws_ceil.cell(row=row, column=9, value=f'=IFERROR(H{row}/G{row}, "-")')
            ws_ceil.cell(row=row, column=9).number_format = '0.00%'
            
            row += 1
        
        # 列幅調整
        ws_ceil.column_dimensions['A'].width = 10
        ws_ceil.column_dimensions['B'].width = 12
        ws_ceil.column_dimensions['C'].width = 10
        ws_ceil.column_dimensions['D'].width = 10
        ws_ceil.column_dimensions['E'].width = 2  # 空き列
        ws_ceil.column_dimensions['F'].width = 10
        ws_ceil.column_dimensions['G'].width = 12
        ws_ceil.column_dimensions['H'].width = 10
        ws_ceil.column_dimensions['I'].width = 10

        # ========== 全体出玉率の出力（CeilingAnalysisシート右側） ==========
        # 除外なしで全データから算出
        # IN枚数 = (通常G * 3) + (純増 / 4.1 * 3)  ※純増4.1枚/Gと仮定してボーナス消化G数を逆算
        sum_first_g = chains_df['First_G'].sum()
        sum_ty = chains_df['Total_Dedama'].sum()
        sum_net_diff = chains_df['Net_Diff'].sum()
        
        total_in = sum_first_g * 3 + (sum_ty / 4.1 * 3)
        payout_rate = (total_in + sum_net_diff) / total_in if total_in > 0 else 0
        
        # 出力
        ws_ceil['K2'] = "【全体出玉率】(全データ)"
        ws_ceil['K2'].font = Font(bold=True)
        
        ws_ceil['K3'] = "サンプル数"
        ws_ceil['L3'] = len(chains_df)
        
        ws_ceil['K4'] = "総IN枚数"
        ws_ceil['L4'] = total_in
        ws_ceil['L4'].number_format = '#,##0'
        
        ws_ceil['K5'] = "総差枚"
        ws_ceil['L5'] = sum_net_diff
        ws_ceil['L5'].number_format = '#,##0'
        
        ws_ceil['K6'] = "出玉率"
        ws_ceil['L6'] = payout_rate
        ws_ceil['L6'].number_format = '0.00%'
        
        # 設定1調整係数: M6 = 97.2% / L6 (全体出玉率)
        ws_ceil['M5'] = "設定1調整係数"
        ws_ceil['M5'].font = Font(bold=True)
        ws_ceil['M6'] = f'=0.972/L6'
        ws_ceil['M6'].number_format = '0.000'
        
        ws_ceil.column_dimensions['K'].width = 15
        ws_ceil.column_dimensions['L'].width = 15
        ws_ceil.column_dimensions['M'].width = 15

        # ========== 天国チェーン中 実質純増の計算 ==========
        # 対象: 3連以上の天国チェーン (Is_Heaven=True, Chain_Length>=3)
        heaven_chains = chains_df[(chains_df['Is_Heaven'] == True) & (chains_df['Chain_Length'] >= 3)]
        
        total_heaven_dedama = 0
        total_heaven_wait_g = 0
        total_heaven_bonus_g = 0
        
        for idx in heaven_chains.index:
            hit_games = heaven_chains.at[idx, 'Hit_Games']
            total_dedama = heaven_chains.at[idx, 'Total_Dedama']
            
            # 天国中待機G (2連目以降、35G以下のみ)
            wait_g = sum(g for g in hit_games[1:] if g <= 35)
            
            # 推定ボーナス消化G (純増4枚/Gと仮定)
            bonus_g = total_dedama / 4.0 if total_dedama > 0 else 0
            
            total_heaven_dedama += total_dedama
            total_heaven_wait_g += wait_g
            total_heaven_bonus_g += bonus_g
        
        total_heaven_g = total_heaven_wait_g + total_heaven_bonus_g
        real_dedama_per_g = total_heaven_dedama / total_heaven_g if total_heaven_g > 0 else 0
        
        # 出力
        ws_ceil['K8'] = "【天国中 実質純増】"
        ws_ceil['K8'].font = Font(bold=True)
        
        ws_ceil['K9'] = "天国チェーン数"
        ws_ceil['L9'] = len(heaven_chains)
        
        ws_ceil['K10'] = "天国中待機G計"
        ws_ceil['L10'] = total_heaven_wait_g
        ws_ceil['L10'].number_format = '#,##0'
        
        ws_ceil['K11'] = "ボーナスG計(推定)"
        ws_ceil['L11'] = total_heaven_bonus_g
        ws_ceil['L11'].number_format = '#,##0'
        
        ws_ceil['K12'] = "純増計"
        ws_ceil['L12'] = total_heaven_dedama
        ws_ceil['L12'].number_format = '#,##0'
        
        ws_ceil['K13'] = "実質純増/G"
        ws_ceil['L13'] = real_dedama_per_g
        ws_ceil['L13'].number_format = '0.00'


        # 天国移行率シート（スルー回数別） - 出力しない
        # heaven_rate_df.to_excel(writer, sheet_name='天国移行率_スルー', index=False)
        
        # 天国移行率シート（前回連チャン別） - 出力しない
        # heaven_by_chain_df.to_excel(writer, sheet_name='天国移行率_前回連', index=False)
        
        # クロス集計シート（48行目以降は出力しない -> ヘッダー1行+46行=47行）
        crosstab_df = crosstab_df.iloc[:46]
        crosstab_df.to_excel(writer, sheet_name='クロス集計', index=False)
        ws3 = writer.sheets['クロス集計']

        # パーセント表示（天国移行率セクションのみ）
        current_section = None

        for row_idx in range(2, ws3.max_row + 1):
            a_cell_value = ws3.cell(row=row_idx, column=1).value

            # セクションヘッダーを検出
            if a_cell_value and isinstance(a_cell_value, str) and '【' in a_cell_value:
                current_section = a_cell_value
                continue

            # 空行をスキップ
            if not a_cell_value:
                continue

            # 天国移行率セクションのみパーセント表示
            if current_section and '天国移行率' in current_section:
                for col in range(2, 13):  # B-L列
                    cell = ws3.cell(row=row_idx, column=col)
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.0%'

        ws3.column_dimensions['A'].width = 25  # セクション名が長いので幅を広げる

        # 3連基準2連スルー分析シート
        rule3_df.to_excel(writer, sheet_name='3連基準2連スルー分析', index=False)
        ws4 = writer.sheets['3連基準2連スルー分析']
        
        # パーセント表示＆数式適用（天国移行率セクションのみ）
        current_section = None



        
        # ChainDataシート (日本語ヘッダー作成)
        rename_dict = {
            'ID': 'ID',
            'Chain_ID': '連チャンID',
            'Chain_Number': '台内連チャン番号',
            'Chain_Length': '連チャン長',
            'Is_Heaven': '天国判定',
            'First_G': '初当G',
            'Through_Before': 'スルー回数',
            'Prev_Chain_Length': '前回連チャン長',
            'Raw_Dedama': '出玉合計',
            'Total_Dedama': '純増枚数',
            'Total_Invest': '投資枚数',
            'Heaven_Invest': '天国中消費',
            'Net_Diff': '差枚',
            'Special_Judge': '特殊判定',
            'Prev_Special_Judge': '前回特殊判定',
            'R3_Through': '3連基準スルー',
            'R3_End': '3連基準End',
            'R3_Heaven': '3連基準天国',
            'Dokihana_Success': 'ドキハナ成功',
            'Dokihana_Start_Pos': 'ドキハナ開始位置',
            'Dokihana_Chain_Len': 'ドキハナ連チャン数',
            'Dokihana_Dedama': 'ドキハナ獲得枚数',
            'Daily_Balance_Before': '当日差枚_開始前',
            'Max_Daily_Balance_Before': '当日差枚最大_開始前',
            'Is_Reset': 'リセット判定'
        }
        # Loc列もリネーム
        for col in chains_df.columns:
            if col.startswith('Loc_'):
                cat_name = col.replace('Loc_', '')
                rename_dict[col] = f'判定_{cat_name}'
                
        chains_jp = chains_df.rename(columns=rename_dict)
        chains_jp.to_excel(writer, sheet_name='ChainData', index=False)
        ws_chain = writer.sheets['ChainData']
        ws_chain.column_dimensions['A'].width = 35
        
        # ChainDataの列マッピング取得
        col_map = {}
        from openpyxl.utils import get_column_letter
        for col in range(1, ws_chain.max_column + 1):
            val = ws_chain.cell(1, col).value
            if val:
                col_map[val] = get_column_letter(col)
        
        # 参照用カラムレター
        col_r3_thr = col_map.get('3連基準スルー')
        col_r3_end = col_map.get('3連基準End')
        col_r3_hvn = col_map.get('3連基準天国')
        col_len = col_map.get('連チャン長')
        
        # 3連基準2連スルー分析シート（数式適用）
        # ... (中略) ...

        col_total_dedama = col_map.get('純増枚数')
        col_max_balance = col_map.get('当日差枚最大_開始前')  # T列
        col_prev_special = col_map.get('前回特殊判定')  # R列

        # 共通の列参照
        ref_r3_hvn_full = f"ChainData!${col_r3_hvn}:${col_r3_hvn}"
        ref_r3_thr_full = f"ChainData!${col_r3_thr}:${col_r3_thr}"
        ref_first_g_full = f"ChainData!${col_first_g}:${col_first_g}"
        ref_total_dedama_full = f"ChainData!${col_total_dedama}:${col_total_dedama}"
        ref_max_balance_full = f"ChainData!${col_max_balance}:${col_max_balance}"  # 差枚条件用
        ref_prev_special_full = f"ChainData!${col_prev_special}:${col_prev_special}"  # 前回特殊判定条件用
        
        # カテゴリリスト
        categories = [
            '2連目1G_RB', '2連目1G_BB',
            '2連目[2-5,32]', '2連目[6-12]', '2連目[13-21]', '2連目[22-31]',
            '2連目BB_[2-5,32]', '2連目BB_[6-12]', '2連目BB_[13-21]', '2連目BB_[22-31]',
            '2連目RB_[2-5,32]', '2連目RB_[6-12]', '2連目RB_[13-21]', '2連目RB_[22-31]'
        ]

        # スタイル定義
        ev_header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # 青
        ev_header_font = Font(bold=True, color='FFFFFF')  # 白文字
        ev_row_fill_odd = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # 白
        ev_row_fill_even = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # 薄灰色
        ev_cat_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid') # オレンジ（カテゴリヘッダー）
        ev_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # ========== 期待値表シート（0〜8スルー目2連発生） ==========
        # Loop for Sheets (0スルー目 〜 8スルー目)
        # ループ: 9つのシートを作成
        for sheet_n in tqdm(range(9), desc="Creating Sheets"):
            sheet_name = f"{sheet_n}スルー目2連発生期待値表一覧"
            ws_ev = writer.book.create_sheet(sheet_name)
            
            # スタイル定義（シートごとに再定義した方が安全だが、同じオブジェクト使い回しでもOK。ここでは参照のみなので外で定義しておけばよいが、念のため）
            # (外で定義済み)

            # 各テーブルの行数: ヘッダー2行 + データ153行 (35G〜800G, 5G刻み) + 空白2行 = 157行
            rows_per_table = 159
            
            # Loop for Tables within Sheet (2連発生0スルー目 〜 9スルー目)
            for k in range(10):
                judge_val_num = k
                judge_val = f"{k}スルー目"
                thr_val = k + sheet_n + 2
                title_text = f"3連基準{thr_val}スルー、2連発生{judge_val}"
                
                # 行オフセット計算 (シート内での位置)
                row_offset = k * rows_per_table
                
                # タイトル
                ws_ev.cell(row=1 + row_offset, column=1, value=f"【{sheet_name}】{title_text}")
                ws_ev.cell(row=1 + row_offset, column=1).font = Font(bold=True, size=16)
                ws_ev.cell(row=2 + row_offset, column=1, value=f"※ 条件: 各パターンの判定={judge_val} かつ 3連基準{thr_val}スルー(発生時)")
                
                current_col = 1
            
                for cat in categories:
                    # カテゴリごとの判定列を取得
                    col_judge = col_map.get(f'判定_{cat}')
                    if not col_judge:
                        continue
                    
                    # 列幅調整 (7列分)
                    col_letter = get_column_letter(current_col)
                    ws_ev.column_dimensions[get_column_letter(current_col)].width = 10
                    ws_ev.column_dimensions[get_column_letter(current_col+1)].width = 12
                    ws_ev.column_dimensions[get_column_letter(current_col+2)].width = 12
                    ws_ev.column_dimensions[get_column_letter(current_col+3)].width = 12
                    ws_ev.column_dimensions[get_column_letter(current_col+4)].width = 10
                    ws_ev.column_dimensions[get_column_letter(current_col+5)].width = 10
                    ws_ev.column_dimensions[get_column_letter(current_col+6)].width = 12
                        
                    ref_judge_full = f"ChainData!${col_judge}:${col_judge}"
                    
                    # カテゴリヘッダー
                    cell = ws_ev.cell(row=4 + row_offset, column=current_col, value=f"■ {cat}")
                    cell.font = Font(bold=True, size=12)
                    cell.fill = ev_cat_fill
                    cell.border = ev_border
                    # 横方向に結合（スタイル適用のためループで塗る）
                    for c in range(current_col+1, current_col+7):
                        cell = ws_ev.cell(row=4 + row_offset, column=c)
                        cell.fill = ev_cat_fill
                        cell.border = ev_border
                    
                    # テーブルヘッダー
                    headers = ["開始G", "初当たり", "TY", "5枚等価", "出玉率", "サンプル数", "調整TY"]
                    for i, header in enumerate(headers):
                        cell = ws_ev.cell(row=5 + row_offset, column=current_col + i, value=header)
                        cell.fill = ev_header_fill
                        cell.font = ev_header_font
                        cell.border = ev_border
                        cell.alignment = Alignment(horizontal='center')
                    
                    # データ行
                    data_start_row = 6 + row_offset
                    for i, start_g in enumerate(range(35, 801, 5)):
                        row_idx = data_start_row + i
                        
                        # 交互の背景色
                        offset = (start_g - 35) // 5
                        row_fill = ev_row_fill_even if offset % 2 == 0 else ev_row_fill_odd
                        
                        # 列参照用のヘルパー
                        c_start   = get_column_letter(current_col)
                        c_hatsu   = get_column_letter(current_col+1)
                        c_ty      = get_column_letter(current_col+2)
                        c_5mai    = get_column_letter(current_col+3)
                        c_rate    = get_column_letter(current_col+4)
                        c_sample  = get_column_letter(current_col+5)
                        c_adj_ty  = get_column_letter(current_col+6)
                        
                        # 1列目: 開始G
                        cell = ws_ev.cell(row=row_idx, column=current_col, value=start_g)
                        cell.fill = row_fill
                        cell.border = ev_border
                        
                        # 2列目: 初当たり = (平均初当G - 開始G) / CeilingAnalysis!M6
                        hatsu_formula = f'=IFERROR((AVERAGEIFS({ref_first_g_full}, {ref_r3_thr_full}, "{thr_val}", {ref_judge_full}, "{judge_val}", {ref_first_g_full}, ">="&{c_start}{row_idx}, {ref_max_balance_full}, "<500", {ref_prev_special_full}, "<1000") - {c_start}{row_idx}) / CeilingAnalysis!$M$6, "[-]")'
                        cell = ws_ev.cell(row=row_idx, column=current_col+1, value=hatsu_formula)
                        cell.number_format = '0.0'
                        cell.fill = row_fill
                        cell.border = ev_border
                        
                        # 3列目: TY (純増枚数, <800G限定, 補正値減算, 差枚<500, 前回特殊<1000条件付き)
                        # 平均純増(<800) - 32 * (50/25.3)
                        avg_under800 = f'AVERAGEIFS({ref_total_dedama_full}, {ref_r3_thr_full}, "{thr_val}", {ref_judge_full}, "{judge_val}", {ref_first_g_full}, ">="&{c_start}{row_idx}, {ref_first_g_full}, "<800", {ref_max_balance_full}, "<500", {ref_prev_special_full}, "<1000")'
                        ty_formula = f'=IFERROR({avg_under800} - 32 * (50/25.3), "[-]")'
                        
                        cell = ws_ev.cell(row=row_idx, column=current_col+2, value=ty_formula)
                        cell.number_format = '#,##0'
                        cell.fill = row_fill
                        cell.border = ev_border
                        
                        # 4列目: 5枚等価 = (調整TY - 初当たり * (50/25.3)) * 20  ※調整TY(G列)を参照
                        val_5mai = f'=IF(OR({c_adj_ty}{row_idx}="[-]", {c_hatsu}{row_idx}="[-]"), "[-]", ({c_adj_ty}{row_idx} - {c_hatsu}{row_idx} * (50/25.3)) * 20)'
                        cell = ws_ev.cell(row=row_idx, column=current_col+3, value=val_5mai)
                        cell.number_format = '#,##0'
                        cell.fill = row_fill
                        cell.border = ev_border
                        
                        # 5列目: 出玉率 = (初当たり * 3 + (調整TY / 3.1 * 3) + 5枚等価 / 20) / (初当たり * 3 + (調整TY / 3.1 * 3))
                        val_denom = f'({c_hatsu}{row_idx} * 3 + ({c_adj_ty}{row_idx} / 3.1 * 3))'
                        val_numer = f'({c_hatsu}{row_idx} * 3 + ({c_adj_ty}{row_idx} / 3.1 * 3) + {c_5mai}{row_idx} / 20)'
                        
                        # 分母0やエラー回避
                        val_rate = f'=IF(OR({c_hatsu}{row_idx}="[-]", {c_adj_ty}{row_idx}="[-]", {c_5mai}{row_idx}="[-]", {val_denom}=0), "[-]", {val_numer} / {val_denom})'
                        
                        cell = ws_ev.cell(row=row_idx, column=current_col+4, value=val_rate)
                        cell.number_format = '0.0%'
                        cell.fill = row_fill
                        cell.border = ev_border
                        
                        # 6列目: サンプル数 (差枚<500, 前回特殊<1000条件付き)
                        cnt_sample = f'COUNTIFS({ref_r3_thr_full}, "{thr_val}", {ref_judge_full}, "{judge_val}", {ref_first_g_full}, ">="&{c_start}{row_idx}, {ref_max_balance_full}, "<500", {ref_prev_special_full}, "<1000")'
                        cell = ws_ev.cell(row=row_idx, column=current_col+5, value=f'={cnt_sample}')
                        cell.number_format = '#,##0'
                        cell.fill = row_fill
                        cell.border = ev_border
                        
                        # 7列目: 調整TY = TY × CeilingAnalysis!$M$6
                        adj_ty_formula = f'=IF({c_ty}{row_idx}="[-]", "[-]", {c_ty}{row_idx} * CeilingAnalysis!$M$6)'
                        cell = ws_ev.cell(row=row_idx, column=current_col+6, value=adj_ty_formula)
                        cell.number_format = '#,##0'
                        cell.fill = row_fill
                        cell.border = ev_border
                    
                    # 次のテーブルへシフト (7列データ + 1列空白)
                    current_col += 8

        # ========== 【2連なし】全体期待値表シート ==========
        ws_ev_no2 = writer.book.create_sheet("2連なし期待値表一覧")
        
        # 集計用データを抽出（2連なし + 差枚<500 + 前回特殊<1000条件）
        # chains_dfには 'Loc_全体' 列があり、ここに '2連なし' が入っている
        if 'Loc_全体' in chains_df.columns:
            mask_no2 = (chains_df['Loc_全体'] == '2連なし') & (chains_df['Max_Daily_Balance_Before'] < 500) & (chains_df['Prev_Special_Judge'] < 1000)
            df_target_no2 = chains_df[mask_no2].copy()
        else:
            # 万が一列がない場合（あり得ないが念のため）
            df_target_no2 = pd.DataFrame(columns=chains_df.columns)
        
        # 0スルー目〜9スルー目の10個のテーブルを横並びで生成
        current_col_no2 = 1
        for k in range(10):
            thr_val_no2 = k  # 0スルー目 → k=0
            
            # データ計算用のサブセット抽出
            # 条件1: 進行中 (天国非当選かつスルー回数未到達)
            # (R3_Heaven=False) & (R3_Through <= k) & (R3_End > k)
            mask_prog = (df_target_no2['R3_Heaven'] == False) & (df_target_no2['R3_Through'] <= k) & (df_target_no2['R3_End'] > k)
            
            # 条件2: 確定天国 (天国当選かつスルー回数一致)
            # (R3_Heaven=True) & (R3_Through == k)
            mask_hvn = (df_target_no2['R3_Heaven'] == True) & (df_target_no2['R3_Through'] == k)
            
            sub_prog = df_target_no2[mask_prog]
            sub_hvn = df_target_no2[mask_hvn]
            
            # 列幅調整 (7列分)
            for col_offset in range(7):
                ws_ev_no2.column_dimensions[get_column_letter(current_col_no2 + col_offset)].width = [10, 12, 12, 12, 10, 10, 12][col_offset]
            
            # カテゴリヘッダー
            cell = ws_ev_no2.cell(row=1, column=current_col_no2, value=f"■ {k}スルー目")
            cell.font = Font(bold=True, size=12)
            cell.fill = ev_cat_fill
            cell.border = ev_border
            for c in range(current_col_no2 + 1, current_col_no2 + 7):
                cell = ws_ev_no2.cell(row=1, column=c)
                cell.fill = ev_cat_fill
                cell.border = ev_border
            
            # テーブルヘッダー
            headers_no2 = ["開始G", "初当たり", "TY", "5枚等価", "出玉率", "サンプル数", "調整TY"]
            for i, header in enumerate(headers_no2):
                cell = ws_ev_no2.cell(row=2, column=current_col_no2 + i, value=header)
                cell.fill = ev_header_fill
                cell.font = ev_header_font
                cell.border = ev_border
                cell.alignment = Alignment(horizontal='center')
            
            # データ行
            data_start_row_no2 = 3
            for i, start_g in enumerate(range(35, 801, 5)):
                row_idx = data_start_row_no2 + i
                
                # 交互の背景色
                offset = (start_g - 35) // 5
                row_fill = ev_row_fill_even if offset % 2 == 0 else ev_row_fill_odd
                
                # 列参照用のヘルパー
                c_start   = get_column_letter(current_col_no2)
                c_hatsu   = get_column_letter(current_col_no2 + 1)
                c_ty      = get_column_letter(current_col_no2 + 2)
                c_5mai    = get_column_letter(current_col_no2 + 3)
                c_rate    = get_column_letter(current_col_no2 + 4)
                c_sample  = get_column_letter(current_col_no2 + 5)
                c_adj_ty  = get_column_letter(current_col_no2 + 6)
                
                # --- Python計算 ---
                # First_G >= start_g
                _prog_g = sub_prog[sub_prog['First_G'] >= start_g]
                _hvn_g = sub_hvn[sub_hvn['First_G'] >= start_g]
                
                cnt_p = len(_prog_g)
                cnt_h = len(_hvn_g)
                total_cnt = cnt_p + cnt_h
                
                hatsu_val = "[-]"
                ty_val = "[-]"
                sample_val = "[-]"
                
                if total_cnt > 0:
                    # 初当たり平均 = (ΣStart_prog + ΣStart_hvn) / total - start_g
                    sum_g_p = _prog_g['First_G'].sum()
                    sum_g_h = _hvn_g['First_G'].sum()
                    avg_g = (sum_g_p + sum_g_h) / total_cnt
                    hatsu_val = avg_g - start_g
                    sample_val = total_cnt
                    
                    # TY平均 (<800Gのみ)
                    _prog_ty = _prog_g[_prog_g['First_G'] < 800]
                    _hvn_ty = _hvn_g[_hvn_g['First_G'] < 800]
                    
                    cnt_p_ty = len(_prog_ty)
                    cnt_h_ty = len(_hvn_ty)
                    total_cnt_ty = cnt_p_ty + cnt_h_ty
                    
                    if total_cnt_ty > 0:
                         sum_ty_p = _prog_ty['Total_Dedama'].sum()
                         sum_ty_h = _hvn_ty['Total_Dedama'].sum()
                         avg_ty = (sum_ty_p + sum_ty_h) / total_cnt_ty
                         ty_val = avg_ty - 32 * (50/25.3)
                
                # 1列目: 開始G
                cell = ws_ev_no2.cell(row=row_idx, column=current_col_no2, value=start_g)
                cell.fill = row_fill
                cell.border = ev_border
                
                # 2列目: 初当たり (値 / CeilingAnalysis!M6)
                if hatsu_val != "[-]":
                    cell = ws_ev_no2.cell(row=row_idx, column=current_col_no2 + 1, value=f"={hatsu_val}/CeilingAnalysis!$M$6")
                    cell.number_format = '0.0'
                else:
                    cell = ws_ev_no2.cell(row=row_idx, column=current_col_no2 + 1, value="[-]")
                cell.fill = row_fill
                cell.border = ev_border
                
                # 3列目: TY (値)
                cell = ws_ev_no2.cell(row=row_idx, column=current_col_no2 + 2, value=ty_val)
                if ty_val != "[-]": cell.number_format = '#,##0'
                cell.fill = row_fill
                cell.border = ev_border
                
                # 4列目: 5枚等価 (数式: 調整TYと初当たりを使用)
                val_5mai_no2 = f'=IF(OR({c_adj_ty}{row_idx}="[-]", {c_hatsu}{row_idx}="[-]"), "[-]", ({c_adj_ty}{row_idx} - {c_hatsu}{row_idx} * (50/25.3)) * 20)'
                cell = ws_ev_no2.cell(row=row_idx, column=current_col_no2 + 3, value=val_5mai_no2)
                cell.number_format = '#,##0'
                cell.fill = row_fill
                cell.border = ev_border
                
                # 5列目: 出玉率 (数式)
                val_denom_no2 = f'({c_hatsu}{row_idx} * 3 + ({c_adj_ty}{row_idx} / 3.1 * 3))'
                val_numer_no2 = f'({c_hatsu}{row_idx} * 3 + ({c_adj_ty}{row_idx} / 3.1 * 3) + {c_5mai}{row_idx} / 20)'
                val_rate_no2 = f'=IF(OR({c_hatsu}{row_idx}="[-]", {c_adj_ty}{row_idx}="[-]", {c_5mai}{row_idx}="[-]", {val_denom_no2}=0), "[-]", {val_numer_no2} / {val_denom_no2})'
                cell = ws_ev_no2.cell(row=row_idx, column=current_col_no2 + 4, value=val_rate_no2)
                cell.number_format = '0.0%'
                cell.fill = row_fill
                cell.border = ev_border
                
                # 6列目: サンプル数 (値)
                cell = ws_ev_no2.cell(row=row_idx, column=current_col_no2 + 5, value=sample_val)
                if sample_val != "[-]": cell.number_format = '#,##0'
                cell.fill = row_fill
                cell.border = ev_border
                
                # 7列目: 調整TY (数式: TY * 係数)
                adj_ty_formula_no2 = f'=IF({c_ty}{row_idx}="[-]", "[-]", {c_ty}{row_idx} * CeilingAnalysis!$M$6)'
                cell = ws_ev_no2.cell(row=row_idx, column=current_col_no2 + 6, value=adj_ty_formula_no2)
                cell.number_format = '#,##0'
                cell.fill = row_fill
                cell.border = ev_border
            
            # 次のテーブルへシフト (7列データ + 1列空白)
            current_col_no2 += 8

    print(f"Excel file created: {OUTPUT_FILE}")


if __name__ == "__main__":
    print("=== デュオ天国移行率分析 ===")
    print(f"設定: 天国閾値={HEAVEN_THRESHOLD}G, コイン持ち={COIN_HOLD}G/50枚")
    print()

    # キャッシュ有効性チェック
    use_cache = False
    data_mtime = os.path.getmtime(INPUT_FILE) if os.path.exists(INPUT_FILE) else 0
    
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, 'rb') as f:
                cache = pickle.load(f)
            if cache.get('mtime') == data_mtime:
                print("キャッシュから読み込み中...")
                df = cache['df']
                chains_df = cache['chains_df']
                use_cache = True
                print(f"  読み込み完了: {len(df)}行, チェーン数: {len(chains_df)}")
        except Exception as e:
            print(f"  キャッシュ読み込みエラー: {e}")
            use_cache = False
    
    if not use_cache:
        print("データ読み込み中...")
        df = load_data(INPUT_FILE)
        print(f"  読み込み完了: {len(df)}行")

        print("連チャン解析中...")
        df, chains_df = analyze_chains(df)
        print(f"  チェーン数: {len(chains_df)}")
        
        # キャッシュ保存
        print("キャッシュ保存中...")
        with open(CACHE_FILE, 'wb') as f:
            pickle.dump({'mtime': data_mtime, 'df': df, 'chains_df': chains_df}, f)
        print("  保存完了")

    print("天国移行率計算中...")
    heaven_rate_df = calculate_heaven_rate(chains_df)
    heaven_by_chain_df = calculate_heaven_rate_by_chain(chains_df)
    crosstab_df = calculate_heaven_rate_crosstab(chains_df)
    rule3_df = calculate_3chain_rule_analysis(chains_df)

    print("ドキハナチャンス分析中...")
    chains_df, dokihana_stats = calculate_dokihana_analysis(chains_df)
    print(f"  ドキハナ発生回数: {dokihana_stats['total_count']}")
    print(f"  平均連チャン数: {dokihana_stats['avg_chain_len']:.2f}")
    print(f"  平均獲得枚数: {dokihana_stats['avg_dedama']:.0f}")

    print()
    print("【天国移行率（スルー回数別）】")
    print(heaven_rate_df.to_string(index=False))
    print()
    print("【天国移行率（前回連チャン別）】")
    print(heaven_by_chain_df.to_string(index=False))
    print()

    print("Excel出力中...")
    write_excel(df, heaven_rate_df, heaven_by_chain_df, crosstab_df, rule3_df, chains_df)

    print("完了!")

