import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import pickle
from tqdm import tqdm

# ===== 設定 =====
INPUT_FILE = r'C:\Users\ilove\Desktop\解析\jyogai20251223_duo2_azukun.csv'
OUTPUT_FILE = r'C:\Users\ilove\Desktop\解析\duo解析\keiryou_' + datetime.now().strftime('%Y%m%d_%H%M%S') + '.xlsx'
CACHE_FILE = r'C:\Users\ilove\Desktop\解析\chains_cache.pkl'

# 設定
HEAVEN_THRESHOLD = 35  # 天国連チャンとみなすG数
COIN_HOLD = 25.3  # コイン持ち (G/50枚)
BB_GAMES = 51          # BB獲得G数
RB_GAMES = 21          # RB獲得G数


def load_data(filepath):
    """azukunフォーマットCSVを読み込み"""
    try:
        df = pd.read_csv(filepath, encoding='utf-8-sig')
    except:
        df = pd.read_csv(filepath, encoding='cp932')
    
    # 元の順序を保持
    df['Original_Order'] = range(len(df))
    
    # ID列を保持したまま、内部処理用に分解
    if 'ID' not in df.columns:
        # ID列がない場合は作成を試みる（既存のロジックで対応できない場合はエラーになるが、今回はazukun.csvを使うのでOKなはず）
        pass

    id_parts = df['ID'].str.rsplit('_', n=2, expand=True)
    df['Hall_Name'] = id_parts[0]
    df['Machine_No'] = pd.to_numeric(id_parts[1], errors='coerce').fillna(0).astype(int)
    df['Date'] = id_parts[2].str.replace(r'(\d{4})(\d{2})(\d{2})', r'\1-\2-\3', regex=True)
    
    df['Start'] = pd.to_numeric(df['Start'], errors='coerce').fillna(0).astype(int)
    df['Dedama'] = pd.to_numeric(df['Dedama'], errors='coerce').fillna(0).astype(int)
    
    # Count列作成（時刻順）
    df_sorted = df.sort_values(by=['Hall_Name', 'Date', 'Machine_No', 'Time'])
    df_sorted['Count'] = df_sorted.groupby(['Hall_Name', 'Date', 'Machine_No']).cumcount() + 1
    
    # 元の順序に戻す
    df = df_sorted.sort_values(by='Original_Order').reset_index(drop=True)
    
    return df


def analyze_chains(df):
    """連チャン判定と天国移行率を計算（元のCSV順序で処理）"""
    # 元の順序でソート
    df = df.sort_values(by='Original_Order').reset_index(drop=True)

    all_chains = []
    chain_id = 0
    chain_number_per_id = {}  # {ID: 連チャン回数}

    # DataFrameに Chain_ID, Chain_Position 列を追加
    df['Chain_ID'] = 0
    df['Chain_Position'] = 0

    # 現在のチェーン情報
    current_id = None
    current_chain_start_idx = 0
    current_chain_hits = [] # インデックスのリスト
    prev_chain_len = 0

    for idx, row in tqdm(df.iterrows(), total=len(df), desc="Analyzing Chains"):
        row_id = row['ID']
        start_g = row['Start']

        # 新しい台または天国終了
        if current_id != row_id or start_g > HEAVEN_THRESHOLD:
            # 前のチェーンを保存
            if current_chain_hits:
                chain_id += 1
                chain_len = len(current_chain_hits)
                is_heaven = chain_len >= 2

                # Chain_ID, Positionを設定
                for pos, hit_idx in enumerate(current_chain_hits, start=1):
                    df.at[hit_idx, 'Chain_ID'] = chain_id
                    df.at[hit_idx, 'Chain_Position'] = pos

                # チェーンの差枚計算
                first_start = df.at[current_chain_hits[0], 'Start']
                last_dedama = df.at[current_chain_hits[-1], 'Dedama']
                raw_dedama = df.loc[current_chain_hits, 'Dedama'].sum()  # 出玉合計（生データ）
                total_invest = df.loc[current_chain_hits, 'Start'].sum() / COIN_HOLD * 50
                first_invest = first_start / COIN_HOLD * 50
                # 天国中投資（初当り以外の消費）= 全投資 - 初当り投資
                heaven_invest = total_invest - first_invest
                # 純増枚数 = 出玉合計 - 天国中消費（初当り投資は含まない）
                total_dedama = raw_dedama - heaven_invest
                # 差枚 = 出玉合計 - 全投資（初当り含む）
                net_diff = raw_dedama - total_invest
                # 特殊判定 = 純増 - 最終出玉（最後の出玉を除いた天国中の純増）
                special_judge = total_dedama - last_dedama
                
                # 当選G数のリストを取得
                hit_games = df.loc[current_chain_hits, 'Start'].tolist()
                # Dedamaのリストを取得（ドキハナ計算用）
                hit_dedamas = df.loc[current_chain_hits, 'Dedama'].tolist()
                # Statusのリストを取得（BB/RB判定用）
                hit_statuses = df.loc[current_chain_hits, 'Status'].tolist()

                # 台ごとの連チャン番号をカウント
                machine_id = df.at[current_chain_hits[0], 'ID']
                if machine_id not in chain_number_per_id:
                    chain_number_per_id[machine_id] = 0
                chain_number_per_id[machine_id] += 1

                # 統計用に記録
                all_chains.append({
                    'ID': machine_id,
                    'Chain_ID': chain_id,
                    'Chain_Number': chain_number_per_id[machine_id],
                    'Chain_Length': chain_len,
                    'Hit_Games': hit_games,
                    'Hit_Dedamas': hit_dedamas,
                    'Hit_Statuses': hit_statuses,
                    'Is_Heaven': is_heaven,
                    'First_G': df.at[current_chain_hits[0], 'Start'],
                    'Through_Before': 0,
                    'Prev_Chain_Length': prev_chain_len,
                    'Raw_Dedama': raw_dedama,       # 出玉合計（生データ）
                    'Total_Dedama': total_dedama,   # 純増枚数 = 出玉 - 天国中消費
                    'Total_Invest': total_invest,
                    'Heaven_Invest': heaven_invest,
                    'Net_Diff': net_diff,
                    'Special_Judge': special_judge,
                    'Prev_Special_Judge': 0.0
                })

                if is_heaven:
                    prev_chain_len = chain_len
                else:
                    pass # 連チャンでない場合は前回連チャン数を維持しない（スルー扱いなので、前回の「天国連チャン」数は保持すべきだが、ロジック上は後で計算し直す）
            
            # 新しいチェーン開始
            if current_id != row_id:
                prev_chain_len = 0
            
            current_id = row_id
            current_chain_hits = [idx]
        else:
            # 天国継続
            current_chain_hits.append(idx)
    
    # 最後のチェーンを保存
    if current_chain_hits:
        chain_id += 1
        chain_len = len(current_chain_hits)
        is_heaven = chain_len >= 2

        for pos, hit_idx in enumerate(current_chain_hits, start=1):
            df.at[hit_idx, 'Chain_ID'] = chain_id
            df.at[hit_idx, 'Chain_Position'] = pos

        # チェーンの差枚計算
        first_start = df.at[current_chain_hits[0], 'Start']
        last_dedama = df.at[current_chain_hits[-1], 'Dedama']
        raw_dedama = df.loc[current_chain_hits, 'Dedama'].sum()  # 出玉合計（生データ）
        total_invest = df.loc[current_chain_hits, 'Start'].sum() / COIN_HOLD * 50
        first_invest = first_start / COIN_HOLD * 50
        heaven_invest = total_invest - first_invest
        # 純増枚数 = 出玉合計 - 天国中消費（初当り投資は含まない）
        total_dedama = raw_dedama - heaven_invest
        # 差枚 = 出玉合計 - 全投資（初当り含む）
        net_diff = raw_dedama - total_invest
        # 特殊判定 = 純増 - 最終出玉
        special_judge = total_dedama - last_dedama
        
        # 当選G数のリストを取得
        hit_games = df.loc[current_chain_hits, 'Start'].tolist()
        # Dedamaのリストを取得
        hit_dedamas = df.loc[current_chain_hits, 'Dedama'].tolist()
        # Statusのリストを取得
        hit_statuses = df.loc[current_chain_hits, 'Status'].tolist()

        # 台ごとの連チャン番号をカウント
        machine_id = df.at[current_chain_hits[0], 'ID']
        if machine_id not in chain_number_per_id:
            chain_number_per_id[machine_id] = 0
        chain_number_per_id[machine_id] += 1

        all_chains.append({
            'ID': machine_id,
            'Chain_ID': chain_id,
            'Chain_Number': chain_number_per_id[machine_id],
            'Chain_Length': chain_len,
            'Hit_Games': hit_games,
            'Hit_Dedamas': hit_dedamas,
            'Hit_Statuses': hit_statuses,
            'Is_Heaven': is_heaven,
            'First_G': df.at[current_chain_hits[0], 'Start'],
            'Through_Before': 0,
            'Prev_Chain_Length': prev_chain_len,
            'Raw_Dedama': raw_dedama,       # 出玉合計（生データ）
            'Total_Dedama': total_dedama,   # 純増枚数 = 出玉 - 天国中消費
            'Total_Invest': total_invest,
            'Heaven_Invest': heaven_invest,
            'Net_Diff': net_diff,
            'Special_Judge': special_judge,
            'Prev_Special_Judge': 0.0
        })

    chains_df = pd.DataFrame(all_chains)

    # Through_Before, Prev_Chain_Length, Prev_Special_Judgeを正確に計算
    if not chains_df.empty:
        chains_df['Through_Before'] = 0
        chains_df['Prev_Chain_Length'] = 0
        chains_df['Prev_Special_Judge'] = 0.0  # float型に変更
        chains_df['Daily_Balance_Before'] = 0.0    # 当日差枚（チェーン開始前の累計）
        chains_df['Max_Daily_Balance_Before'] = 0.0  # 当日差枚最大（チェーン開始前までの最大値）
        chains_df['Is_Reset'] = chains_df['Chain_Number'] == 1  # リセット判定（台内初回）

        prev_id = None
        through_count = 0
        prev_heaven_len = 0
        prev_special = 0.0
        cumulative_balance = 0.0   # 累計差枚（前回チェーン終了時点）
        max_balance = 0.0          # 最大差枚（前回チェーン終了時点までの最大）

        for idx in chains_df.index:
            current_id = chains_df.at[idx, 'ID']
            is_heaven = chains_df.at[idx, 'Is_Heaven']
            chain_len = chains_df.at[idx, 'Chain_Length']
            special_judge = chains_df.at[idx, 'Special_Judge']
            net_diff = chains_df.at[idx, 'Net_Diff']  # このチェーンの差枚

            if current_id != prev_id:
                through_count = 0
                prev_heaven_len = 0
                prev_special = 0.0
                cumulative_balance = 0.0
                max_balance = 0.0

            # チェーン開始前の値を記録（このチェーンの差枚を加算する前）
            chains_df.at[idx, 'Through_Before'] = through_count
            chains_df.at[idx, 'Prev_Chain_Length'] = prev_heaven_len
            chains_df.at[idx, 'Prev_Special_Judge'] = prev_special
            chains_df.at[idx, 'Daily_Balance_Before'] = cumulative_balance
            chains_df.at[idx, 'Max_Daily_Balance_Before'] = max_balance

            # このチェーンの差枚を加算（次のチェーンのために）
            cumulative_balance += net_diff
            # 最大値を更新
            if cumulative_balance > max_balance:
                max_balance = cumulative_balance

            if is_heaven:
                through_count = 0
                prev_heaven_len = chain_len
            else:
                through_count += 1

            # 前回の特殊判定を保存（天国チェーンの場合のみ更新、天国以外は直近の天国の値を保持）
            if is_heaven:
                prev_special = special_judge

            prev_id = current_id

    return df, chains_df


def calculate_heaven_rate(chains_df):
    """スルー回数別の天国移行率を計算"""
    results = []
    
    # 0〜10スルー
    for thr in range(11):
        subset = chains_df[chains_df['Through_Before'] == thr]
        total = len(subset)
        heaven = len(subset[subset['Is_Heaven'] == True])
        rate = heaven / total if total > 0 else 0
        
        results.append({
            'スルー回数': f'{thr}スルー',
            '天国移行数': heaven,
            'サンプル数': total,
            '天国移行率': rate
        })
    
    # 11スルー以上
    subset = chains_df[chains_df['Through_Before'] >= 11]
    total = len(subset)
    heaven = len(subset[subset['Is_Heaven'] == True])
    rate = heaven / total if total > 0 else 0
    
    results.append({
        'スルー回数': '11スルー以上',
        '天国移行数': heaven,
        'サンプル数': total,
        '天国移行率': rate
    })
    
    # 合計
    total_all = len(chains_df)
    heaven_all = len(chains_df[chains_df['Is_Heaven'] == True])
    rate_all = heaven_all / total_all if total_all > 0 else 0
    
    results.append({
        'スルー回数': '合計',
        '天国移行数': heaven_all,
        'サンプル数': total_all,
        '天国移行率': rate_all
    })
    
    return pd.DataFrame(results)


def calculate_heaven_rate_by_chain(chains_df):
    """前回連チャン長別の天国移行率を計算"""
    results = []
    
    # 前回連チャン2連〜19連
    for chain_len in range(2, 20):
        subset = chains_df[chains_df['Prev_Chain_Length'] == chain_len]
        total = len(subset)
        heaven = len(subset[subset['Is_Heaven'] == True])
        rate = heaven / total if total > 0 else 0
        
        results.append({
            '前回連チャン': f'{chain_len}連後',
            '天国移行数': heaven,
            'サンプル数': total,
            '天国移行率': rate
        })
    
    # 20連以上
    subset = chains_df[chains_df['Prev_Chain_Length'] >= 20]
    total = len(subset)
    heaven = len(subset[subset['Is_Heaven'] == True])
    rate = heaven / total if total > 0 else 0
    
    results.append({
        '前回連チャン': '20連以上後',
        '天国移行数': heaven,
        'サンプル数': total,
        '天国移行率': rate
    })
    
    # 合計 (前回天国があったもののみ)
    subset_all = chains_df[chains_df['Prev_Chain_Length'] >= 2]
    total_all = len(subset_all)
    heaven_all = len(subset_all[subset_all['Is_Heaven'] == True])
    rate_all = heaven_all / total_all if total_all > 0 else 0
    
    results.append({
        '前回連チャン': '合計',
        '天国移行数': heaven_all,
        'サンプル数': total_all,
        '天国移行率': rate_all
    })
    
    return pd.DataFrame(results)


def calculate_heaven_rate_crosstab(chains_df):
    """前回連チャン長 × スルー回数 のクロス集計（天国移行率とサンプル数を分離）+ 特殊判定分析を統合"""
    results = []

    # 朝イチを除外（Chain_Number > 1のみ）
    chains_with_prev = chains_df[chains_df['Chain_Number'] > 1].copy()

    # ===== 前回連チャン長別 - 天国移行率 =====
    results.append({'': '【前回連チャン長別 - 天国移行率】'})

    for chain_len in range(2, 20):
        rate_row = {'': f'{chain_len}連後'}
        for thr in range(11):
            subset = chains_with_prev[(chains_with_prev['Prev_Chain_Length'] == chain_len) &
                                      (chains_with_prev['Through_Before'] == thr)]
            total = len(subset)
            heaven = len(subset[subset['Is_Heaven'] == True])
            rate = heaven / total if total > 0 else None
            rate_row[f'{thr}スルー'] = rate

        # 合計列には天国移行率を入れる
        total_all = len(chains_with_prev[chains_with_prev['Prev_Chain_Length'] == chain_len])
        heaven_all = len(chains_with_prev[(chains_with_prev['Prev_Chain_Length'] == chain_len) &
                                          (chains_with_prev['Is_Heaven'] == True)])
        rate_all = heaven_all / total_all if total_all > 0 else None
        rate_row['合計'] = rate_all

        results.append(rate_row)

    # 20連以上
    rate_row = {'': '20連以上後'}
    for thr in range(11):
        subset = chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 20) &
                                  (chains_with_prev['Through_Before'] == thr)]
        total = len(subset)
        heaven = len(subset[subset['Is_Heaven'] == True])
        rate = heaven / total if total > 0 else None
        rate_row[f'{thr}スルー'] = rate
    total_all = len(chains_with_prev[chains_with_prev['Prev_Chain_Length'] >= 20])
    heaven_all = len(chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 20) &
                                      (chains_with_prev['Is_Heaven'] == True)])
    rate_all = heaven_all / total_all if total_all > 0 else None
    rate_row['合計'] = rate_all
    results.append(rate_row)

    # 3～9連後（まとめ）
    rate_row = {'': '3～9連後'}
    for thr in range(11):
        subset = chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 3) &
                                  (chains_with_prev['Prev_Chain_Length'] <= 9) &
                                  (chains_with_prev['Through_Before'] == thr)]
        total = len(subset)
        heaven = len(subset[subset['Is_Heaven'] == True])
        rate = heaven / total if total > 0 else None
        rate_row[f'{thr}スルー'] = rate
    total_all = len(chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 3) & (chains_with_prev['Prev_Chain_Length'] <= 9)])
    heaven_all = len(chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 3) & (chains_with_prev['Prev_Chain_Length'] <= 9) &
                                      (chains_with_prev['Is_Heaven'] == True)])
    rate_all = heaven_all / total_all if total_all > 0 else None
    rate_row['合計'] = rate_all
    results.append(rate_row)

    # 10連以上後（まとめ）
    rate_row = {'': '10連以上後'}
    for thr in range(11):
        subset = chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 10) &
                                  (chains_with_prev['Through_Before'] == thr)]
        total = len(subset)
        heaven = len(subset[subset['Is_Heaven'] == True])
        rate = heaven / total if total > 0 else None
        rate_row[f'{thr}スルー'] = rate
    total_all = len(chains_with_prev[chains_with_prev['Prev_Chain_Length'] >= 10])
    heaven_all = len(chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 10) &
                                      (chains_with_prev['Is_Heaven'] == True)])
    rate_all = heaven_all / total_all if total_all > 0 else None
    rate_row['合計'] = rate_all
    results.append(rate_row)

    # 空行
    results.append({})

    # ===== 前回連チャン長別 - サンプル数 =====
    results.append({'': '【前回連チャン長別 - サンプル数】'})

    for chain_len in range(2, 20):
        count_row = {'': f'{chain_len}連後'}
        for thr in range(11):
            subset = chains_with_prev[(chains_with_prev['Prev_Chain_Length'] == chain_len) &
                                      (chains_with_prev['Through_Before'] == thr)]
            count_row[f'{thr}スルー'] = len(subset)

        total_all = len(chains_with_prev[chains_with_prev['Prev_Chain_Length'] == chain_len])
        count_row['合計'] = total_all

        results.append(count_row)

    # 20連以上
    count_row = {'': '20連以上後'}
    for thr in range(11):
        subset = chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 20) &
                                  (chains_with_prev['Through_Before'] == thr)]
        count_row[f'{thr}スルー'] = len(subset)
    total_all = len(chains_with_prev[chains_with_prev['Prev_Chain_Length'] >= 20])
    count_row['合計'] = total_all
    results.append(count_row)

    # 3～9連後（まとめ）
    count_row = {'': '3～9連後'}
    for thr in range(11):
        subset = chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 3) &
                                  (chains_with_prev['Prev_Chain_Length'] <= 9) &
                                  (chains_with_prev['Through_Before'] == thr)]
        count_row[f'{thr}スルー'] = len(subset)
    total_all = len(chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 3) & (chains_with_prev['Prev_Chain_Length'] <= 9)])
    count_row['合計'] = total_all
    results.append(count_row)

    # 10連以上後（まとめ）
    count_row = {'': '10連以上後'}
    for thr in range(11):
        subset = chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 10) &
                                  (chains_with_prev['Through_Before'] == thr)]
        count_row[f'{thr}スルー'] = len(subset)
    total_all = len(chains_with_prev[chains_with_prev['Prev_Chain_Length'] >= 10])
    count_row['合計'] = total_all
    results.append(count_row)

    # 空行
    results.append({})

    # ===== 前回特殊判定別 - 天国移行率 =====
    results.append({'': '【前回特殊判定別 - 天国移行率】'})

    # 100枚刻みのレンジ定義
    spec_ranges = []
    for start in range(0, 2001, 100):
        end = start + 99
        spec_ranges.append((f"{start}～{end}", start, end))
    spec_ranges.append(("2001以上", 2001, None))

    for label, low, high in spec_ranges:
        rate_row = {'': label}

        if high is None:
            range_subset = chains_with_prev[chains_with_prev['Prev_Special_Judge'] >= low]
        else:
            range_subset = chains_with_prev[(chains_with_prev['Prev_Special_Judge'] >= low) &
                                            (chains_with_prev['Prev_Special_Judge'] <= high)]

        for thr in range(11):
            subset = range_subset[range_subset['Through_Before'] == thr]
            total = len(subset)
            heaven = len(subset[subset['Is_Heaven'] == True])
            rate = heaven / total if total > 0 else None
            rate_row[f'{thr}スルー'] = rate

        # 合計列には天国移行率を入れる
        total_all = len(range_subset)
        heaven_all = len(range_subset[range_subset['Is_Heaven'] == True])
        rate_all = heaven_all / total_all if total_all > 0 else None
        rate_row['合計'] = rate_all

        results.append(rate_row)

    # 空行
    results.append({})

    # ===== 前回特殊判定別 - サンプル数 =====
    results.append({'': '【前回特殊判定別 - サンプル数】'})

    for label, low, high in spec_ranges:
        count_row = {'': label}

        if high is None:
            range_subset = chains_with_prev[chains_with_prev['Prev_Special_Judge'] >= low]
        else:
            range_subset = chains_with_prev[(chains_with_prev['Prev_Special_Judge'] >= low) &
                                            (chains_with_prev['Prev_Special_Judge'] <= high)]

        for thr in range(11):
            subset = range_subset[range_subset['Through_Before'] == thr]
            count_row[f'{thr}スルー'] = len(subset)

        count_row['合計'] = len(range_subset)

        
        results.append(rate_row)
        results.append(count_row)
        
    return pd.DataFrame(results)


def calculate_3chain_rule_analysis(chains_df):
    """3連以上を天国、2連を隠れスルーとする分析（率とサンプル数を別列で出力）"""
    # 分析用データを収集
    # 構造: {'Category': [{'2Chain_Loc': ..., 'Through': ..., 'Is_True_Heaven': ...}]}
    analysis_data = {
        '全体': [],
        '2連目1G_RB': [],
        '2連目1G_BB': [],
        '2連目[2-5,32]': [],
        '2連目[6-12]': [],
        '2連目[13-21]': [],
        '2連目[22-31]': [],
        '2連目BB_[2-5,32]': [],
        '2連目BB_[6-12]': [],
        '2連目BB_[13-21]': [],
        '2連目BB_[22-31]': [],
        '2連目RB_[2-5,32]': [],
        '2連目RB_[6-12]': [],
        '2連目RB_[13-21]': [],
        '2連目RB_[22-31]': [],
    }
    
    # 変数初期化
    prev_id = None
    hits_since_true_heaven = 0  # 3連以上基準のスルー回数
    first_2chain_idx = None     # 最初の2連が発生したスルー回数 (None: 未発生)
    prev_true_heaven_valid = False 
    
    # カテゴリ判定用のステート
    current_2chain_feature = None 
    
    trackers = {}
    for cat in analysis_data.keys():
        trackers[cat] = {
            'hits_since': 0,
            'first_2chain_idx': None,
            'active': False, # このカテゴリのデータとして有効か（3連天国後スタート）
            'ignore': False # このカテゴリの条件に合わない「最初の2連」が発生した場合、以降天国まで無視
        }

    # DataFrameに追加するためのリスト初期化
    col_r3_through = []
    col_r3_end = []
    col_r3_heaven = []
    # カテゴリごとのLocation列
    cols_loc = {cat: [] for cat in analysis_data}
    
    for idx, row in chains_df.iterrows():
        row_id = row['ID']
        chain_len = row['Chain_Length']
        hit_games = row.get('Hit_Games', [])
        hit_statuses = row.get('Hit_Statuses', [])
        
        # 台が変わったらリセット
        if row_id != prev_id:
            for cat in trackers:
                trackers[cat]['hits_since'] = 0
                trackers[cat]['first_2chain_idx'] = None
                trackers[cat]['active'] = False
                trackers[cat]['ignore'] = False
            
            if chain_len >= 3:
                for cat in trackers:
                    trackers[cat]['active'] = True
            
            # リセット時は全て空/無効値を入れるが、DataFrameの行数と合わせるため追加が必要
            # ただし、ループは全ての行に対して回るので、ここで追加する値が「その行の値」になる
            # リセット直後の行（=新しい台の1行目）の処理は下流で行うので、ここではcontinueしない方が良いが
            # prev_id logic matches flow.
            # ここでは状態リセットのみ。
            
            # --- ここで continue すると append がスキップされるので注意 ---
            # 元のコードでは continue していた。
            # 今回はリストに必ず値を入れる必要がある（行ズレ防止）。
            # リセット時の行も処理対象となる（e.g. 朝イチHeavenなど）
            prev_id = row_id # ID更新
            
            # 既存ロジックでは continue して次の行へ行っていた（active判定等は次ループで）。
            # しかしリスト追加は必須。
            # ここは「新しい台の1行目」のコンテキストで処理を続行させる。
            # activeチェック等をこの後で行う。
        
        # 前回の天国が3連以上でなければスキップ -> activeフラグで管理
        # 元コード: continueしていた。
        # 今回: continueするとリストが埋まらない。
        # -> 「対象外」として埋める。
        
        if not trackers['全体']['active']:
            # activeでない期間（前任が単発天国等）
            col_r3_through.append(None)
            col_r3_end.append(None)
            col_r3_heaven.append(None)
            for cat in analysis_data:
                cols_loc[cat].append('非アクティブ')
                
            # 次回アクティブ化判定
            if chain_len >= 3:
                for cat in trackers:
                    trackers[cat]['active'] = True
                    trackers[cat]['hits_since'] = 0
                    trackers[cat]['first_2chain_idx'] = None
                    trackers[cat]['ignore'] = False
            prev_id = row_id
            continue
            
        # ここからアクティブな処理
        is_true_heaven = (chain_len >= 3)
        
        # 今回のチェーン情報の解析
        current_2chain_details = {}
        is_valid_2chain = False
        
        if chain_len == 2:
            second_hit_g = hit_games[1] if len(hit_games) > 1 else 0
            if second_hit_g >= 34:
                is_valid_2chain = False
            else:
                is_valid_2chain = True
                status_2 = hit_statuses[1] if len(hit_statuses) > 1 else ''
                is_rb = 'RB' in status_2
                is_bb = 'BB' in status_2
                g = second_hit_g
                rng = None
                if (2 <= g <= 5) or (g == 32): rng = '[2-5,32]'
                elif 6 <= g <= 12: rng = '[6-12]'
                elif 13 <= g <= 21: rng = '[13-21]'
                elif 22 <= g <= 31: rng = '[22-31]'
                current_2chain_details = {
                    'is_1g': (g == 1),
                    'is_rb': is_rb,
                    'is_bb': is_bb,
                    'range': rng
                }

        # 共通値
        current_through = trackers['全体']['hits_since']
        
        # End Through計算 (Excel数式用: Start <= T < End)
        # 天国の場合は Start == T (End不要だが、Startを入れておく)
        if is_true_heaven:
            end_through = current_through
        else:
            # 天国でない場合、Chain Length分だけスルー回数を消費する
            end_through = current_through + chain_len

        col_r3_through.append(current_through)
        col_r3_end.append(end_through)
        col_r3_heaven.append(is_true_heaven)
        
        # 各カテゴリごとに処理
        for cat in analysis_data:
            tr = trackers[cat]
            
            # --- まず状態更新（2連検出を先に行う）---
            if is_true_heaven:
                # 天国到達 -> リセットは後で（ラベル決定後）
                pass
            else:
                # まだ天国でない（スルー中）
                if not tr['ignore']:
                    if is_valid_2chain and tr['first_2chain_idx'] is None:
                        # まだ2連未発生状態で、今回2連が発生した
                        # この2連が「このカテゴリの条件」を満たすか？
                        match = True
                        det = current_2chain_details
                        if cat == '全体': match = True
                        elif cat == '2連目1G_RB': match = (det['is_rb'] and det['is_1g'])
                        elif cat == '2連目1G_BB': match = (det['is_bb'] and det['is_1g'])
                        elif cat == '2連目[2-5,32]': match = (det['range'] == '[2-5,32]')
                        elif cat == '2連目[6-12]': match = (det['range'] == '[6-12]')
                        elif cat == '2連目[13-21]': match = (det['range'] == '[13-21]')
                        elif cat == '2連目[22-31]': match = (det['range'] == '[22-31]')
                        elif cat == '2連目BB_[2-5,32]': match = (det['is_bb'] and det['range'] == '[2-5,32]')
                        elif cat == '2連目BB_[6-12]': match = (det['is_bb'] and det['range'] == '[6-12]')
                        elif cat == '2連目BB_[13-21]': match = (det['is_bb'] and det['range'] == '[13-21]')
                        elif cat == '2連目BB_[22-31]': match = (det['is_bb'] and det['range'] == '[22-31]')
                        elif cat == '2連目RB_[2-5,32]': match = (det['is_rb'] and det['range'] == '[2-5,32]')
                        elif cat == '2連目RB_[6-12]': match = (det['is_rb'] and det['range'] == '[6-12]')
                        elif cat == '2連目RB_[13-21]': match = (det['is_rb'] and det['range'] == '[13-21]')
                        elif cat == '2連目RB_[22-31]': match = (det['is_rb'] and det['range'] == '[22-31]')
                        
                        if match:
                            tr['first_2chain_idx'] = tr['hits_since']
                        else:
                            tr['ignore'] = True
            
            # Locationラベル決定（2連検出後なので、2連発生時はそのスルー目ラベルになる）
            if tr['ignore']:
                loc_label = '対象外'
            elif not tr['active']:
                loc_label = '非アクティブ'
            elif tr['first_2chain_idx'] is None:
                loc_label = '2連なし'
            else:
                loc_label = f"{tr['first_2chain_idx']}スルー目"
            
            cols_loc[cat].append(loc_label)

            # --- データ記録（集計用リストへ） ---
            if not tr['ignore']:
                analysis_data[cat].append({
                    '2Chain_Loc': loc_label,
                    'Through': tr['hits_since'],
                    'Is_True_Heaven': is_true_heaven
                })
            
            # --- 状態更新（スルー回数加算とリセット） ---
            if is_true_heaven:
                tr['hits_since'] = 0
                tr['first_2chain_idx'] = None
                tr['ignore'] = False
            else:
                if not tr['ignore']:
                            
                    # スルー回数加算
                    # ユーザー要望: 2連（失敗）は2スルー分としてカウント（3連以上基準のためボーナス回数でカウント）
                    # 単発なら+1、2連なら+2
                    tr['hits_since'] += chain_len

        prev_id = row_id
    
    # DataFrameに列を追加
    chains_df['R3_Through'] = col_r3_through
    chains_df['R3_End'] = col_r3_end
    chains_df['R3_Heaven'] = col_r3_heaven
    for cat, vals in cols_loc.items():
        chains_df[f'Loc_{cat}'] = vals
        
    # 集計とクロス集計作成
    results = []
    loc_order = ['2連なし'] + [f'{i}スルー目' for i in range(11)] + ['11スルー以上']
    
    # カテゴリ順序
    cat_order = [
         '全体', 
         '2連目1G_RB', '2連目1G_BB',
         '2連目[2-5,32]', '2連目[6-12]', '2連目[13-21]', '2連目[22-31]',
         '2連目BB_[2-5,32]', '2連目BB_[6-12]', '2連目BB_[13-21]', '2連目BB_[22-31]',
         '2連目RB_[2-5,32]', '2連目RB_[6-12]', '2連目RB_[13-21]', '2連目RB_[22-31]'
    ]

    for cat in cat_order:
        data_list = analysis_data[cat]
        if not data_list: continue
        
        df_cat = pd.DataFrame(data_list)
        
        # --- 1. 天国移行率 ---
        results.append({'2連発生位置': f'【{cat} - 天国移行率】'})
        
        for loc in loc_order:
            row_subset = df_cat[df_cat['2Chain_Loc'] == loc] if not df_cat.empty else pd.DataFrame()
            # 空でもスキップしない（行構造を固定するため）
            
            row_data = {'2連発生位置': loc}
            for thr in range(11):
                subset = row_subset[row_subset['Through'] == thr]
                total = len(subset)
                success = len(subset[subset['Is_True_Heaven'] == True])
                rate = success / total if total > 0 else None
                row_data[f'{thr}スルー'] = rate
            
            # 11以降
            subset_over = row_subset[row_subset['Through'] >= 11]
            total = len(subset_over)
            success = len(subset_over[subset_over['Is_True_Heaven'] == True])
            rate = success / total if total > 0 else None
            row_data['11スルー以上'] = rate
            
            # 合計
            total_all = len(row_subset)
            success_all = len(row_subset[row_subset['Is_True_Heaven'] == True])
            rate_all = success_all / total_all if total_all > 0 else None
            row_data['合計'] = rate_all
            
            results.append(row_data)

        # --- 2. 天国移行数（分子） ---
        results.append({'2連発生位置': f'【{cat} - 天国移行数】'})
        
        for loc in loc_order:
            row_subset = df_cat[df_cat['2Chain_Loc'] == loc] if not df_cat.empty else pd.DataFrame()
            # 空でもスキップしない
            
            row_data = {'2連発生位置': loc}
            for thr in range(11):
                subset = row_subset[row_subset['Through'] == thr]
                success = len(subset[subset['Is_True_Heaven'] == True])
                row_data[f'{thr}スルー'] = success
            
            subset_over = row_subset[row_subset['Through'] >= 11]
            success = len(subset_over[subset_over['Is_True_Heaven'] == True])
            row_data['11スルー以上'] = success
            
            # 合計
            success_all = len(row_subset[row_subset['Is_True_Heaven'] == True])
            row_data['合計'] = success_all
            
            results.append(row_data)

        # --- 3. サンプル数（分母） ---
        results.append({'2連発生位置': f'【{cat} - サンプル数】'})
        for loc in loc_order:
            row_subset = df_cat[df_cat['2Chain_Loc'] == loc] if not df_cat.empty else pd.DataFrame()
            # 空でもスキップしない
            
            row_data = {'2連発生位置': loc}
            for thr in range(11):
                subset = row_subset[row_subset['Through'] == thr]
                row_data[f'{thr}スルー'] = len(subset)
            
            subset_over = row_subset[row_subset['Through'] >= 11]
            row_data['11スルー以上'] = len(subset_over)
            
            row_data['合計'] = len(row_subset)
            
            results.append(row_data)

        # 空行
        results.append({})
        
    return pd.DataFrame(results)


def calculate_dokihana_analysis(chains_df):
    """
    ドキハナチャンス分析
    - 天国連チャン中に34Gまたは35Gで当たりが発生（Position>=2）
    - その後も連チャンが続く（合計3連以上）
    - chains_dfにドキハナ判定列を追加
    """
    # ドキハナ列を初期化
    chains_df['Dokihana_Success'] = False
    chains_df['Dokihana_Start_Pos'] = 0
    chains_df['Dokihana_Chain_Len'] = 0
    chains_df['Dokihana_Dedama'] = 0
    
    for idx in chains_df.index:
        chain_len = chains_df.at[idx, 'Chain_Length']
        is_heaven = chains_df.at[idx, 'Is_Heaven']
        hit_games = chains_df.at[idx, 'Hit_Games']
        
        # ドキハナ判定: 天国連チャン中にPosition>=2で34か35Gがあり、その後2連以上続く
        if is_heaven and chain_len >= 3 and len(hit_games) >= 2:
            hit_dedamas = chains_df.at[idx, 'Hit_Dedamas']
            
            # Position>=2で34か35Gを探す
            for pos, g in enumerate(hit_games[1:], start=2):  # Position 2から
                if g in [34, 35]:
                    dokihana_len = chain_len - pos + 1
                    # ドキハナ連チャン数が2以上の場合のみ成功
                    if dokihana_len >= 2:
                        chains_df.at[idx, 'Dokihana_Success'] = True
                        chains_df.at[idx, 'Dokihana_Start_Pos'] = pos
                        chains_df.at[idx, 'Dokihana_Chain_Len'] = dokihana_len
                        
                        # ドキハナ開始位置以降の純増を計算
                        # pos-1はインデックス（0始まり）
                        dokihana_dedamas = sum(hit_dedamas[pos-1:])  # 出玉合計
                        dokihana_games = hit_games[pos-1:]  # 当選G数リスト
                        # ドキハナ区間の天国消費（初回除く = pos以降の当選G数合計）
                        dokihana_invest = sum(dokihana_games[1:]) / COIN_HOLD * 50 if len(dokihana_games) > 1 else 0
                        # 純増 = 出玉 - 天国消費
                        dokihana_net = dokihana_dedamas - dokihana_invest
                        chains_df.at[idx, 'Dokihana_Dedama'] = int(dokihana_net)
                    break
    
    # 統計計算（コンソール表示用）
    dokihana_success_df = chains_df[chains_df['Dokihana_Success'] == True]
    
    stats = {
        'total_count': len(dokihana_success_df),
        'avg_chain_len': dokihana_success_df['Dokihana_Chain_Len'].mean() if len(dokihana_success_df) > 0 else 0,
        'avg_dedama': dokihana_success_df['Dokihana_Dedama'].mean() if len(dokihana_success_df) > 0 else 0,
        'total_dedama': dokihana_success_df['Dokihana_Dedama'].sum() if len(dokihana_success_df) > 0 else 0
    }
    
    return chains_df, stats


def write_excel(df, heaven_rate_df, heaven_by_chain_df, crosstab_df, chains_df):
    """エクセルファイルに出力（軽量化版: 期待値表・3連基準分析なし）"""
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        # Dataシート - 基本データ出力
        data_cols = ['ID', 'Status', 'Start', 'Dedama', 'Time', 'Chain_ID', 'Chain_Position']
        df_out = df[data_cols].copy()
        df_out.columns = ['ID', 'Status', 'Start', 'Dedama', 'Time', '連チャンID', '連チャン位置']
        df_out.to_excel(writer, sheet_name='Data', index=False)
        
        ws_data = writer.sheets['Data']
        max_r = ws_data.max_row
        
        # 追加列ヘッダー (H列から)
        extra_headers = ['朝イチ連', '投資詳細(枚)', '差枚詳細(枚)', '当日差枚', '有利区間G数']
        for i, h in enumerate(extra_headers):
            ws_data.cell(row=1, column=8 + i, value=h)
        
        # Dataシート I-L列の値計算 (Python側で計算して直接書き込み)
        # I: 投資詳細
        df['calc_invest'] = df['Start'] / COIN_HOLD * 50
        
        # J: 差枚詳細
        df['calc_diff'] = df['Dedama'] - df['calc_invest']
        
        # K: 当日差枚 (開始時) = IDごとの累積差枚 (現在行を含まない)
        df['calc_daily_diff'] = df.groupby('ID')['calc_diff'].cumsum() - df['calc_diff']
        
        # L: 有利区間G数 (累積)
        # ボーナス消化G推定: BB=51, RB=21
        def get_bonus_g(k):
            k_str = str(k)
            if 'BB' in k_str: return 51
            if 'RB' in k_str: return 21
            return 0
        
        df['calc_bonus_g'] = df['Status'].apply(get_bonus_g)
        df['calc_total_g'] = df['Start'] + df['calc_bonus_g']
        df['calc_yuuri_g'] = df.groupby('ID')['calc_total_g'].cumsum()

        # 各行に値を書き込み
        for r in range(2, max_r + 1):
            idx = r - 2
            
            # H: 朝イチ連
            is_morning = df.at[idx, 'Is_Morning_Chain']
            ws_data.cell(row=r, column=8, value=is_morning)
            
            # I: 投資詳細
            ws_data.cell(row=r, column=9, value=df.at[idx, 'calc_invest'])
            
            # J: 差枚詳細
            ws_data.cell(row=r, column=10, value=df.at[idx, 'calc_diff'])
            
            # K: 当日差枚
            ws_data.cell(row=r, column=11, value=df.at[idx, 'calc_daily_diff'])
            
            # L: 有利区間G数
            ws_data.cell(row=r, column=12, value=df.at[idx, 'calc_yuuri_g'])
        
        # 一時列削除
        columns_to_drop = ['calc_invest', 'calc_diff', 'calc_daily_diff', 'calc_bonus_g', 'calc_total_g', 'calc_yuuri_g']
        df.drop(columns=columns_to_drop, inplace=True)

        
        # 天国移行率シート（スルー回数別）
        heaven_rate_df.to_excel(writer, sheet_name='天国移行率_スルー', index=False)
        ws = writer.sheets['天国移行率_スルー']
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
            for cell in row:
                cell.number_format = '0.0%'
        ws.column_dimensions['A'].width = 15
        
        # 天国移行率シート（前回連チャン別）
        heaven_by_chain_df.to_excel(writer, sheet_name='天国移行率_前回連', index=False)
        ws2 = writer.sheets['天国移行率_前回連']
        for row in ws2.iter_rows(min_row=2, max_col=ws2.max_column):
            for cell in row:
                cell.number_format = '0.0%'
        ws2.column_dimensions['A'].width = 15
        
        # クロス集計シート（天国移行率とサンプル数交互 + 特殊判定分析統合）
        crosstab_df.to_excel(writer, sheet_name='クロス集計', index=False)
        ws3 = writer.sheets['クロス集計']

        # パーセント表示（天国移行率セクションのみ）
        current_section = None

        for row_idx in range(2, ws3.max_row + 1):
            a_cell_value = ws3.cell(row=row_idx, column=1).value

            # セクションヘッダーを検出
            if a_cell_value and isinstance(a_cell_value, str) and '【' in a_cell_value:
                current_section = a_cell_value
                continue

            # 空行をスキップ
            if not a_cell_value:
                continue

            # 天国移行率セクションのみパーセント表示
            if current_section and '天国移行率' in current_section:
                for col in range(2, 13):  # B-L列
                    cell = ws3.cell(row=row_idx, column=col)
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.0%'


        ws3.column_dimensions['A'].width = 25  # セクション名が長いので幅を広げる



        
        # ChainDataシート (日本語ヘッダー作成)
        rename_dict = {
            'ID': 'ID',
            'Chain_ID': '連チャンID',
            'Chain_Number': '台内連チャン番号',
            'Chain_Length': '連チャン長',
            'Is_Heaven': '天国判定',
            'First_G': '初当G',
            'Through_Before': 'スルー回数',
            'Prev_Chain_Length': '前回連チャン長',
            'Raw_Dedama': '出玉合計',
            'Total_Dedama': '純増枚数',
            'Total_Invest': '投資枚数',
            'Heaven_Invest': '天国中消費',
            'Net_Diff': '差枚',
            'Special_Judge': '特殊判定',
            'Prev_Special_Judge': '前回特殊判定',
            'R3_Through': '3連基準スルー',
            'R3_End': '3連基準End',
            'R3_Heaven': '3連基準天国',
            'Dokihana_Success': 'ドキハナ成功',
            'Dokihana_Start_Pos': 'ドキハナ開始位置',
            'Dokihana_Chain_Len': 'ドキハナ連チャン数',
            'Dokihana_Dedama': 'ドキハナ獲得枚数',
            'Daily_Balance_Before': '当日差枚_開始前',
            'Max_Daily_Balance_Before': '当日差枚最大_開始前',
            'Is_Reset': 'リセット判定',
            'Is_Morning_Chain': '朝イチフラグ'
        }
        # Loc列もリネーム
        for col in chains_df.columns:
            if col.startswith('Loc_'):
                cat_name = col.replace('Loc_', '')
                rename_dict[col] = f'判定_{cat_name}'
                
        chains_jp = chains_df.rename(columns=rename_dict)
        chains_jp.to_excel(writer, sheet_name='ChainData', index=False)
        ws_chain = writer.sheets['ChainData']
        ws_chain.column_dimensions['A'].width = 35
        
        # ChainDataの列マッピング取得
        col_map = {}
        from openpyxl.utils import get_column_letter
        for col in range(1, ws_chain.max_column + 1):
            val = ws_chain.cell(1, col).value
            if val:
                col_map[val] = get_column_letter(col)
        
        # 参照用カラムレター
        col_r3_thr = col_map.get('3連基準スルー')
        col_r3_end = col_map.get('3連基準End')
        col_r3_hvn = col_map.get('3連基準天国')
        col_len = col_map.get('連チャン長')
        

        # ========== ドキハナチャンス分析シート（式計算ベース） ==========
        ws_doki = writer.book.create_sheet("DokihanaAnalysis")
        
        # タイトル
        ws_doki['A1'] = "【ドキハナチャンス分析】"
        ws_doki['A1'].font = Font(bold=True, size=14)
        ws_doki['A2'] = "34Gまたは35Gで発生する特別な天国連チャンの分析"
        
        # ChainDataの行数とドキハナ列参照を取得
        chain_max_row = ws_chain.max_row
        col_doki_success = col_map.get('ドキハナ成功')
        col_doki_len = col_map.get('ドキハナ連チャン数')
        col_doki_dedama = col_map.get('ドキハナ獲得枚数')
        col_doki_start = col_map.get('ドキハナ開始位置')
        col_max_balance = col_map.get('当日差枚最大_開始前')
        
        # ChainDataの範囲参照
        ref_success = f"ChainData!${col_doki_success}$2:${col_doki_success}${chain_max_row}"
        ref_doki_len = f"ChainData!${col_doki_len}$2:${col_doki_len}${chain_max_row}"
        ref_doki_dedama = f"ChainData!${col_doki_dedama}$2:${col_doki_dedama}${chain_max_row}"
        ref_doki_start = f"ChainData!${col_doki_start}$2:${col_doki_start}${chain_max_row}"
        ref_max_balance = f"ChainData!${col_max_balance}$2:${col_max_balance}${chain_max_row}"
        
        # ========== 基本統計 ==========
        ws_doki['A4'] = "■ 基本統計"
        ws_doki['A4'].font = Font(bold=True)
        
        ws_doki['A5'] = "ドキハナ発生回数"
        ws_doki['B5'] = f"=COUNTIF({ref_success}, TRUE)"
        
        ws_doki['A6'] = "平均ドキハナ連チャン数"
        ws_doki['B6'] = f'=IFERROR(AVERAGEIF({ref_success}, TRUE, {ref_doki_len}), "-")'
        ws_doki['B6'].number_format = '0.00'
        
        ws_doki['A7'] = "平均ドキハナ獲得枚数"
        ws_doki['B7'] = f'=IFERROR(AVERAGEIF({ref_success}, TRUE, {ref_doki_dedama})-32/{COIN_HOLD}*50, "-")'
        ws_doki['B7'].number_format = '#,##0'
        
        ws_doki['A8'] = "合計ドキハナ獲得枚数"
        ws_doki['B8'] = f"=SUMIF({ref_success}, TRUE, {ref_doki_dedama})"
        ws_doki['B8'].number_format = '#,##0'
        
        # ========== ドキハナ開始位置別 分析（全体） ==========
        ws_doki['A10'] = "■ ドキハナ開始位置別 分析（全体）"
        ws_doki['A10'].font = Font(bold=True)
        
        ws_doki['A11'] = "開始位置"
        ws_doki['B11'] = "発生回数"
        ws_doki['C11'] = "割合"
        ws_doki['D11'] = "平均ドキハナ連数"
        ws_doki['E11'] = "平均獲得枚数"
        
        row = 12
        for pos in range(2, 21):
            ws_doki.cell(row=row, column=1, value=f"{pos}連目")
            ws_doki.cell(row=row, column=2, value=f"=COUNTIFS({ref_success}, TRUE, {ref_doki_start}, {pos})")
            ws_doki.cell(row=row, column=3, value=f"=IFERROR(B{row}/$B$5, 0)")
            ws_doki.cell(row=row, column=3).number_format = '0.0%'
            ws_doki.cell(row=row, column=4, value=f'=IFERROR(AVERAGEIFS({ref_doki_len}, {ref_success}, TRUE, {ref_doki_start}, {pos}), "-")')
            ws_doki.cell(row=row, column=4).number_format = '0.00'
            ws_doki.cell(row=row, column=5, value=f'=IFERROR(AVERAGEIFS({ref_doki_dedama}, {ref_success}, TRUE, {ref_doki_start}, {pos})-32/{COIN_HOLD}*50, "-")')
            ws_doki.cell(row=row, column=5).number_format = '#,##0'
            row += 1
        
        # 21連目以降
        ws_doki.cell(row=row, column=1, value="21連目以降")
        ws_doki.cell(row=row, column=2, value=f'=COUNTIFS({ref_success}, TRUE, {ref_doki_start}, ">=21")')
        ws_doki.cell(row=row, column=3, value=f"=IFERROR(B{row}/$B$5, 0)")
        ws_doki.cell(row=row, column=3).number_format = '0.0%'
        ws_doki.cell(row=row, column=4, value=f'=IFERROR(AVERAGEIFS({ref_doki_len}, {ref_success}, TRUE, {ref_doki_start}, ">=21"), "-")')
        ws_doki.cell(row=row, column=4).number_format = '0.00'
        ws_doki.cell(row=row, column=5, value=f'=IFERROR(AVERAGEIFS({ref_doki_dedama}, {ref_success}, TRUE, {ref_doki_start}, ">=21")-32/{COIN_HOLD}*50, "-")')
        ws_doki.cell(row=row, column=5).number_format = '#,##0'
        
        # ========== ドキハナ開始位置別 分析（当日差枚最大<1 = 一度もプラスになっていない） ==========
        ws_doki['A35'] = "■ ドキハナ開始位置別 分析（当日差枚が一度も1以上になっていない）"
        ws_doki['A35'].font = Font(bold=True)
        ws_doki['A36'] = "※ その時点までの当日差枚最大が1未満の場合のみカウント"
        
        ws_doki['A37'] = "開始位置"
        ws_doki['B37'] = "発生回数"
        ws_doki['C37'] = "割合"
        ws_doki['D37'] = "平均ドキハナ連数"
        ws_doki['E37'] = "平均獲得枚数"
        
        # 当日差枚条件付きの発生回数合計（割合計算用）
        ws_doki['G35'] = "条件付き合計"
        ws_doki['G36'] = f'=COUNTIFS({ref_success}, TRUE, {ref_max_balance}, "<1")'
        
        row = 38
        for pos in range(2, 21):
            ws_doki.cell(row=row, column=1, value=f"{pos}連目")
            # 条件: ドキハナ成功=TRUE AND 開始位置=pos AND 当日差枚最大<1
            ws_doki.cell(row=row, column=2, value=f'=COUNTIFS({ref_success}, TRUE, {ref_doki_start}, {pos}, {ref_max_balance}, "<1")')
            ws_doki.cell(row=row, column=3, value=f"=IFERROR(B{row}/$G$36, 0)")
            ws_doki.cell(row=row, column=3).number_format = '0.0%'
            ws_doki.cell(row=row, column=4, value=f'=IFERROR(AVERAGEIFS({ref_doki_len}, {ref_success}, TRUE, {ref_doki_start}, {pos}, {ref_max_balance}, "<1"), "-")')
            ws_doki.cell(row=row, column=4).number_format = '0.00'
            ws_doki.cell(row=row, column=5, value=f'=IFERROR(AVERAGEIFS({ref_doki_dedama}, {ref_success}, TRUE, {ref_doki_start}, {pos}, {ref_max_balance}, "<1")-32/{COIN_HOLD}*50, "-")')
            ws_doki.cell(row=row, column=5).number_format = '#,##0'
            row += 1
        
        # 21連目以降
        ws_doki.cell(row=row, column=1, value="21連目以降")
        ws_doki.cell(row=row, column=2, value=f'=COUNTIFS({ref_success}, TRUE, {ref_doki_start}, ">=21", {ref_max_balance}, "<1")')
        ws_doki.cell(row=row, column=3, value=f"=IFERROR(B{row}/$G$36, 0)")
        ws_doki.cell(row=row, column=3).number_format = '0.0%'
        ws_doki.cell(row=row, column=4, value=f'=IFERROR(AVERAGEIFS({ref_doki_len}, {ref_success}, TRUE, {ref_doki_start}, ">=21", {ref_max_balance}, "<1"), "-")')
        ws_doki.cell(row=row, column=4).number_format = '0.00'
        ws_doki.cell(row=row, column=5, value=f'=IFERROR(AVERAGEIFS({ref_doki_dedama}, {ref_success}, TRUE, {ref_doki_start}, ">=21", {ref_max_balance}, "<1")-32/{COIN_HOLD}*50, "-")')
        ws_doki.cell(row=row, column=5).number_format = '#,##0'
        
        # ========== スルー回数別 ドキハナ分析（条件付き） ==========
        # ChainDataのスルー回数列を参照
        col_through = col_map.get('スルー回数')
        ref_through = f"ChainData!${col_through}$2:${col_through}${chain_max_row}"
        
        # スタイル定義
        doki_header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # 青
        doki_header_font = Font(bold=True, color='FFFFFF')  # 白文字
        doki_row_fill_odd = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # 白
        doki_row_fill_even = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # 薄灰色
        doki_total_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')  # 薄青
        doki_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        ws_doki['A62'] = "■ スルー回数別 ドキハナ分析（当日プラス未経験 & 6連目以内開始）"
        ws_doki['A62'].font = Font(bold=True, size=12)
        ws_doki['A63'] = "※ 当日差枚が一度も1以上になっていない & ドキハナ開始位置が6連目以内"
        
        # ヘッダー行（64行目）
        headers = ["スルー回数", "発生回数", "平均ドキハナ連数", "平均獲得枚数"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_doki.cell(row=64, column=col_idx, value=header)
            cell.fill = doki_header_fill
            cell.font = doki_header_font
            cell.border = doki_border
            cell.alignment = Alignment(horizontal='center')
        
        row = 65
        for thr in range(10):  # 0～9スルー
            row_fill = doki_row_fill_even if thr % 2 == 0 else doki_row_fill_odd
            
            # スルー回数ラベル
            cell = ws_doki.cell(row=row, column=1, value=f"{thr}スルー")
            cell.fill = row_fill
            cell.border = doki_border
            
            # 発生回数
            cell = ws_doki.cell(row=row, column=2, value=f'=COUNTIFS({ref_success}, TRUE, {ref_through}, {thr}, {ref_max_balance}, "<1", {ref_doki_start}, "<=6")')
            cell.fill = row_fill
            cell.border = doki_border
            
            # 平均連チャン数
            cell = ws_doki.cell(row=row, column=3, value=f'=IFERROR(AVERAGEIFS({ref_doki_len}, {ref_success}, TRUE, {ref_through}, {thr}, {ref_max_balance}, "<1", {ref_doki_start}, "<=6"), "-")')
            cell.number_format = '0.00'
            cell.fill = row_fill
            cell.border = doki_border
            
            # 平均獲得枚数
            cell = ws_doki.cell(row=row, column=4, value=f'=IFERROR(AVERAGEIFS({ref_doki_dedama}, {ref_success}, TRUE, {ref_through}, {thr}, {ref_max_balance}, "<1", {ref_doki_start}, "<=6")-32/{COIN_HOLD}*50, "-")')
            cell.number_format = '#,##0'
            cell.fill = row_fill
            cell.border = doki_border
            
            row += 1
        
        # 10スルー以上
        row_fill = doki_row_fill_even
        cell = ws_doki.cell(row=row, column=1, value="10スルー以上")
        cell.fill = row_fill
        cell.border = doki_border
        
        cell = ws_doki.cell(row=row, column=2, value=f'=COUNTIFS({ref_success}, TRUE, {ref_through}, ">=10", {ref_max_balance}, "<1", {ref_doki_start}, "<=6")')
        cell.fill = row_fill
        cell.border = doki_border
        
        cell = ws_doki.cell(row=row, column=3, value=f'=IFERROR(AVERAGEIFS({ref_doki_len}, {ref_success}, TRUE, {ref_through}, ">=10", {ref_max_balance}, "<1", {ref_doki_start}, "<=6"), "-")')
        cell.number_format = '0.00'
        cell.fill = row_fill
        cell.border = doki_border
        
        cell = ws_doki.cell(row=row, column=4, value=f'=IFERROR(AVERAGEIFS({ref_doki_dedama}, {ref_success}, TRUE, {ref_through}, ">=10", {ref_max_balance}, "<1", {ref_doki_start}, "<=6")-32/{COIN_HOLD}*50, "-")')
        cell.number_format = '#,##0'
        cell.fill = row_fill
        cell.border = doki_border
        row += 1
        
        # 合計行（薄青背景）
        cell = ws_doki.cell(row=row, column=1, value="合計")
        cell.fill = doki_total_fill
        cell.border = doki_border
        cell.font = Font(bold=True)
        
        cell = ws_doki.cell(row=row, column=2, value=f'=COUNTIFS({ref_success}, TRUE, {ref_max_balance}, "<1", {ref_doki_start}, "<=6")')
        cell.fill = doki_total_fill
        cell.border = doki_border
        cell.font = Font(bold=True)
        
        cell = ws_doki.cell(row=row, column=3, value=f'=IFERROR(AVERAGEIFS({ref_doki_len}, {ref_success}, TRUE, {ref_max_balance}, "<1", {ref_doki_start}, "<=6"), "-")')
        cell.number_format = '0.00'
        cell.fill = doki_total_fill
        cell.border = doki_border
        cell.font = Font(bold=True)
        
        cell = ws_doki.cell(row=row, column=4, value=f'=IFERROR(AVERAGEIFS({ref_doki_dedama}, {ref_success}, TRUE, {ref_max_balance}, "<1", {ref_doki_start}, "<=6")-32/{COIN_HOLD}*50, "-")')
        cell.number_format = '#,##0'
        cell.fill = doki_total_fill
        cell.border = doki_border
        cell.font = Font(bold=True)
        
        # 列幅調整
        ws_doki.column_dimensions['A'].width = 35
        ws_doki.column_dimensions['B'].width = 15
        ws_doki.column_dimensions['C'].width = 18
        ws_doki.column_dimensions['D'].width = 18
        ws_doki.column_dimensions['E'].width = 18

        # ========== 天国当選G数分布シート ==========
        ws_dist = writer.book.create_sheet("HeavenHitDistribution")
        
        ws_dist['A1'] = "【3連以上天国チェーン 当選G数分布】"
        ws_dist['A1'].font = Font(bold=True, size=14)
        ws_dist['A2'] = "※ 3連以上の天国チェーンにおけるPosition 2以降（天国連チャン中）の当選G数分布"
        
        # 3連以上のchainからPosition 2以降のhit_gamesを収集
        all_heaven_hits = []
        for idx in chains_df.index:
            chain_len = chains_df.at[idx, 'Chain_Length']
            is_heaven = chains_df.at[idx, 'Is_Heaven']
            hit_games = chains_df.at[idx, 'Hit_Games']
            
            if is_heaven and chain_len >= 3:
                # Position 2以降（インデックス1以降）の当選G数を収集
                for g in hit_games[1:]:
                    if g <= 35:
                        all_heaven_hits.append(g)
        
        # 分布を集計
        from collections import Counter
        hit_counter = Counter(all_heaven_hits)
        total_hits = len(all_heaven_hits)
        
        # ヘッダー
        ws_dist['A4'] = "当選G数"
        ws_dist['B4'] = "発生回数"
        ws_dist['C4'] = "割合"
        ws_dist['A4'].font = Font(bold=True)
        ws_dist['B4'].font = Font(bold=True)
        ws_dist['C4'].font = Font(bold=True)
        
        row = 5
        for g in range(1, 36):  # 1G〜35G
            ws_dist.cell(row=row, column=1, value=f"{g}G")
            count = hit_counter.get(g, 0)
            ws_dist.cell(row=row, column=2, value=count)
            rate = count / total_hits if total_hits > 0 else 0
            ws_dist.cell(row=row, column=3, value=rate)
            ws_dist.cell(row=row, column=3).number_format = '0.00%'
            row += 1
        
        # 合計行
        ws_dist.cell(row=row, column=1, value="合計")
        ws_dist.cell(row=row, column=1).font = Font(bold=True)
        ws_dist.cell(row=row, column=2, value=total_hits)
        ws_dist.cell(row=row, column=2).font = Font(bold=True)
        ws_dist.cell(row=row, column=3, value=1.0 if total_hits > 0 else 0)
        ws_dist.cell(row=row, column=3).number_format = '0.00%'
        ws_dist.cell(row=row, column=3).font = Font(bold=True)
        
        # 平均G数
        avg_g = sum(all_heaven_hits) / total_hits if total_hits > 0 else 0
        ws_dist.cell(row=row+1, column=1, value="平均G数")
        ws_dist.cell(row=row+1, column=1).font = Font(bold=True)
        ws_dist.cell(row=row+1, column=2, value=avg_g)
        ws_dist.cell(row=row+1, column=2).number_format = '0.0'
        ws_dist.cell(row=row+1, column=2).font = Font(bold=True)
        
        # 列幅調整
        ws_dist.column_dimensions['A'].width = 12
        ws_dist.column_dimensions['B'].width = 12
        ws_dist.column_dimensions['C'].width = 12

        # ========== 天井到達率分析シート ==========
        ws_ceil = writer.book.create_sheet("CeilingAnalysis")
        
        ws_ceil['A1'] = "【天井到達率分析】"
        ws_ceil['A1'].font = Font(bold=True, size=14)
        ws_ceil['A2'] = "※ サンプル数=そのG以上で当選した総数、天井到達=800G以上で当選、到達率=天井到達/サンプル"
        
        # Is_Reset列を参照
        col_reset = col_map.get('リセット判定')
        col_through = col_map.get('スルー回数')
        col_first_g = col_map.get('初当G')
        
        ref_reset = f"ChainData!${col_reset}$2:${col_reset}${chain_max_row}"
        ref_through = f"ChainData!${col_through}$2:${col_through}${chain_max_row}"
        ref_first_g = f"ChainData!${col_first_g}$2:${col_first_g}${chain_max_row}"
        
        # セクション1: リセット0スルー (A-D列)
        ws_ceil['A4'] = "■ リセット時0スルー"
        ws_ceil['A4'].font = Font(bold=True, size=12)
        
        ws_ceil['A5'] = "ゲーム数"
        ws_ceil['B5'] = "サンプル数"
        ws_ceil['C5'] = "天井到達"
        ws_ceil['D5'] = "到達率"
        for col in range(1, 5):
            ws_ceil.cell(row=5, column=col).font = Font(bold=True)
        
        # セクション2: リセット0スルー以外 (F-I列)
        ws_ceil['F4'] = "■ リセット0スルー以外"
        ws_ceil['F4'].font = Font(bold=True, size=12)
        
        ws_ceil['F5'] = "ゲーム数"
        ws_ceil['G5'] = "サンプル数"
        ws_ceil['H5'] = "天井到達"
        ws_ceil['I5'] = "到達率"
        for col in range(6, 10):
            ws_ceil.cell(row=5, column=col).font = Font(bold=True)
        
        # データ行（両セクション同時に出力）
        row = 6
        for g in range(0, 801, 5):
            # セクション1: リセット0スルー (A-D列)
            ws_ceil.cell(row=row, column=1, value=f"{g}G～")
            ws_ceil.cell(row=row, column=2, value=f'=COUNTIFS({ref_reset}, TRUE, {ref_through}, 0, {ref_first_g}, ">={g}")')
            ws_ceil.cell(row=row, column=3, value=f'=COUNTIFS({ref_reset}, TRUE, {ref_through}, 0, {ref_first_g}, ">=800")')
            ws_ceil.cell(row=row, column=4, value=f'=IFERROR(C{row}/B{row}, "-")')
            ws_ceil.cell(row=row, column=4).number_format = '0.00%'
            
            # セクション2: リセット0スルー以外 (F-I列)
            ws_ceil.cell(row=row, column=6, value=f"{g}G～")
            ws_ceil.cell(row=row, column=7, value=f'=COUNTIFS({ref_first_g}, ">={g}")-COUNTIFS({ref_reset}, TRUE, {ref_through}, 0, {ref_first_g}, ">={g}")')
            ws_ceil.cell(row=row, column=8, value=f'=COUNTIFS({ref_first_g}, ">=800")-COUNTIFS({ref_reset}, TRUE, {ref_through}, 0, {ref_first_g}, ">=800")')
            ws_ceil.cell(row=row, column=9, value=f'=IFERROR(H{row}/G{row}, "-")')
            ws_ceil.cell(row=row, column=9).number_format = '0.00%'
            
            row += 1
        
        # 列幅調整
        ws_ceil.column_dimensions['A'].width = 10
        ws_ceil.column_dimensions['B'].width = 12
        ws_ceil.column_dimensions['C'].width = 10
        ws_ceil.column_dimensions['D'].width = 10
        ws_ceil.column_dimensions['E'].width = 2  # 空き列
        ws_ceil.column_dimensions['F'].width = 10
        ws_ceil.column_dimensions['G'].width = 12
        ws_ceil.column_dimensions['H'].width = 10
        ws_ceil.column_dimensions['I'].width = 10

        # ========== 全体出玉率の出力（CeilingAnalysisシート右側） ==========
        # 除外なしで全データから算出
        # IN枚数 = (通常G * 3) + (純増 / 4.1 * 3)  ※純増4.1枚/Gと仮定してボーナス消化G数を逆算
        sum_first_g = chains_df['First_G'].sum()
        sum_ty = chains_df['Total_Dedama'].sum()
        sum_net_diff = chains_df['Net_Diff'].sum()
        
        total_in = sum_first_g * 3 + (sum_ty / 4.1 * 3)
        payout_rate = (total_in + sum_net_diff) / total_in if total_in > 0 else 0
        
        # 出力
        ws_ceil['K2'] = "【全体出玉率】(全データ)"
        ws_ceil['K2'].font = Font(bold=True)
        
        ws_ceil['K3'] = "サンプル数"
        ws_ceil['L3'] = len(chains_df)
        
        ws_ceil['K4'] = "総IN枚数"
        ws_ceil['L4'] = total_in
        ws_ceil['L4'].number_format = '#,##0'
        
        ws_ceil['K5'] = "総差枚"
        ws_ceil['L5'] = sum_net_diff
        ws_ceil['L5'].number_format = '#,##0'
        
        ws_ceil['K6'] = "出玉率"
        ws_ceil['L6'] = payout_rate
        ws_ceil['L6'].number_format = '0.00%'
        
        # 設定1調整係数: M6 = 97.2% / L6 (全体出玉率)
        ws_ceil['M5'] = "設定1調整係数"
        ws_ceil['M5'].font = Font(bold=True)
        ws_ceil['M6'] = f'=0.972/L6'
        ws_ceil['M6'].number_format = '0.000'
        
        ws_ceil.column_dimensions['K'].width = 15
        ws_ceil.column_dimensions['L'].width = 15
        ws_ceil.column_dimensions['M'].width = 15

        # ========== 天国チェーン中 実質純増の計算 ==========
        # 対象: 3連以上の天国チェーン (Is_Heaven=True, Chain_Length>=3)
        heaven_chains = chains_df[(chains_df['Is_Heaven'] == True) & (chains_df['Chain_Length'] >= 3)]
        
        total_heaven_dedama = 0
        total_heaven_wait_g = 0
        total_heaven_bonus_g = 0
        
        for idx in heaven_chains.index:
            hit_games = heaven_chains.at[idx, 'Hit_Games']
            total_dedama = heaven_chains.at[idx, 'Total_Dedama']
            
            # 天国中待機G (2連目以降、35G以下のみ)
            wait_g = sum(g for g in hit_games[1:] if g <= 35)
            
            # 推定ボーナス消化G (純増4枚/Gと仮定)
            bonus_g = total_dedama / 4.0 if total_dedama > 0 else 0
            
            total_heaven_dedama += total_dedama
            total_heaven_wait_g += wait_g
            total_heaven_bonus_g += bonus_g
        
        total_heaven_g = total_heaven_wait_g + total_heaven_bonus_g
        real_dedama_per_g = total_heaven_dedama / total_heaven_g if total_heaven_g > 0 else 0
        
        # 出力
        ws_ceil['K8'] = "【天国中 実質純増】"
        ws_ceil['K8'].font = Font(bold=True)
        
        ws_ceil['K9'] = "天国チェーン数"
        ws_ceil['L9'] = len(heaven_chains)
        
        ws_ceil['K10'] = "天国中待機G計"
        ws_ceil['L10'] = total_heaven_wait_g
        ws_ceil['L10'].number_format = '#,##0'
        
        ws_ceil['K11'] = "ボーナスG計(推定)"
        ws_ceil['L11'] = total_heaven_bonus_g
        ws_ceil['L11'].number_format = '#,##0'
        
        ws_ceil['K12'] = "純増計"
        ws_ceil['L12'] = total_heaven_dedama
        ws_ceil['L12'].number_format = '#,##0'
        
        ws_ceil['K13'] = "実質純増/G"
        ws_ceil['L13'] = real_dedama_per_g
        ws_ceil['L13'].number_format = '0.00'


        # 天国移行率シート（スルー回数別） - 出力しない
        # heaven_rate_df.to_excel(writer, sheet_name='天国移行率_スルー', index=False)
        
        # 天国移行率シート（前回連チャン別） - 出力しない
        # heaven_by_chain_df.to_excel(writer, sheet_name='天国移行率_前回連', index=False)
        
        # クロス集計シート（48行目以降は出力しない -> ヘッダー1行+46行=47行）
        crosstab_df = crosstab_df.iloc[:46]
        crosstab_df.to_excel(writer, sheet_name='クロス集計', index=False)
        ws3 = writer.sheets['クロス集計']

        # パーセント表示（天国移行率セクションのみ）
        current_section = None

        for row_idx in range(2, ws3.max_row + 1):
            a_cell_value = ws3.cell(row=row_idx, column=1).value

            # セクションヘッダーを検出
            if a_cell_value and isinstance(a_cell_value, str) and '【' in a_cell_value:
                current_section = a_cell_value
                continue

            # 空行をスキップ
            if not a_cell_value:
                continue

            # 天国移行率セクションのみパーセント表示
            if current_section and '天国移行率' in current_section:
                for col in range(2, 13):  # B-L列
                    cell = ws3.cell(row=row_idx, column=col)
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.0%'


        ws3.column_dimensions['A'].width = 25  # セクション名が長いので幅を広げる



        
        # ========== ボーナス間期待値表（朝イチ/朝イチ除外） ==========
        
        # 共通生成関数
        def create_bonus_ev_sheet(sheet_title, target_cond_str):
            ws_ev = writer.book.create_sheet(sheet_title)
            
            # 数式用列参照
            col_start = col_map.get('初当G')
            col_dedama = col_map.get('純増枚数')
            col_through = col_map.get('スルー回数')
            col_target = col_map.get('朝イチフラグ')
            col_max_daily = col_map.get('当日差枚最大_開始前')
            col_prev_special = col_map.get('前回特殊判定')
            
            max_r = ws_chain.max_row
            
            ref_start = f"ChainData!${col_start}$2:${col_start}${max_r}"
            ref_dedama = f"ChainData!${col_dedama}$2:${col_dedama}${max_r}"
            ref_through = f"ChainData!${col_through}$2:${col_through}${max_r}"
            ref_target = f"ChainData!${col_target}$2:${col_target}${max_r}"
            ref_max_daily = f"ChainData!${col_max_daily}$2:${col_max_daily}${max_r}"
            ref_prev_special = f"ChainData!${col_prev_special}$2:${col_prev_special}${max_r}"

            # スタイル定義
            ev_header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            ev_row_fill_even = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            ev_row_fill_odd = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            ev_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            # スルー回数ループ (0-9スルー)
            start_row = 3
            current_col = 2
            
            for thr in range(10):
                # ヘッダー作成
                header_cell = ws_ev.cell(row=start_row-1, column=current_col, value=f"{thr}スルー")
                ws_ev.merge_cells(start_row=start_row-1, start_column=current_col, end_row=start_row-1, end_column=current_col+6)
                header_cell.fill = ev_header_fill
                header_cell.font = Font(color="FFFFFF", bold=True)
                header_cell.alignment = Alignment(horizontal='center')
                
                # サブヘッダー
                sub_headers = ["開始G", "初当たり", "TY", "5枚等価", "出玉率", "サンプル", "調整TY"]
                for i, h in enumerate(sub_headers):
                    c = ws_ev.cell(row=start_row, column=current_col + i, value=h)
                    c.fill = ev_header_fill
                    c.font = Font(color="FFFFFF", bold=True)
                    c.alignment = Alignment(horizontal='center')
                    c.border = ev_border
                    ws_ev.column_dimensions[get_column_letter(current_col + i)].width = 12
                    
                col_idx_hatsu = current_col + 1
                col_idx_ty = current_col + 2
                col_idx_5mai = current_col + 3
                col_idx_rate = current_col + 4
                col_idx_sample = current_col + 5
                col_idx_adj_ty = current_col + 6
                
                col_char_hatsu = get_column_letter(col_idx_hatsu)
                col_char_ty = get_column_letter(col_idx_ty)
                col_char_5mai = get_column_letter(col_idx_5mai)
                col_char_rate = get_column_letter(col_idx_rate)
                col_char_sample = get_column_letter(col_idx_sample)
                col_char_adj_ty = get_column_letter(col_idx_adj_ty)

                # 行ループ
                for i, g in enumerate(range(35, 801, 5)):
                    row = start_row + 1 + i
                    row_fill = ev_row_fill_even if i % 2 == 0 else ev_row_fill_odd
                    
                    # 開始G
                    cell = ws_ev.cell(row=row, column=current_col, value=g)
                    cell.fill = row_fill
                    cell.border = ev_border
                    cell.alignment = Alignment(horizontal='center')
                    
                    # スタイル適用
                    for c in range(col_idx_hatsu, col_idx_adj_ty + 1):
                        cell = ws_ev.cell(row=row, column=c)
                        cell.fill = row_fill
                        cell.border = ev_border

                    addr_sample = f"{col_char_sample}{row}"
                    
                    # サンプル数
                    # サンプル数にも条件を入れるべきか？
                    # 初当たり計算の分母になるので入れるべき。
                    f_cnt = f'=COUNTIFS({ref_target}, {target_cond_str}, {ref_through}, {thr}, {ref_start}, ">={g}", {ref_max_daily}, "<=0", {ref_prev_special}, "<1000")'
                    ws_ev.cell(row=row, column=col_idx_sample, value=f_cnt).number_format = '#,##0'

                    # 初当たり
                    # 条件: Is_Target, Through, StartG>=g, MaxDaily<=0, PrevSpec<1000
                    # ユーザー要望: ÷CeilingAnalysis!M6を追加
                    f_hatsu = f'=IF({addr_sample}=0, "[-]", (AVERAGEIFS({ref_start}, {ref_target}, {target_cond_str}, {ref_through}, {thr}, {ref_start}, ">={g}", {ref_max_daily}, "<=0", {ref_prev_special}, "<1000") - {g}) / CeilingAnalysis!$M$6)'
                    ws_ev.cell(row=row, column=col_idx_hatsu, value=f_hatsu).number_format = '0.0'

                    # TY (<800G限定)
                    # 条件: 上記 + StartG<800
                    cost = 32 * (50/COIN_HOLD)
                    f_ty_core = f'AVERAGEIFS({ref_dedama}, {ref_target}, {target_cond_str}, {ref_through}, {thr}, {ref_start}, ">={g}", {ref_start}, "<800", {ref_max_daily}, "<=0", {ref_prev_special}, "<1000")'
                    f_ty = f'=IFERROR({f_ty_core} - {cost}, "[-]")'
                    ws_ev.cell(row=row, column=col_idx_ty, value=f_ty).number_format = '#,##0'
                    
                    # 他の項目
                    addr_ty = f"{col_char_ty}{row}"
                    addr_hatsu = f"{col_char_hatsu}{row}"
                    addr_adj_ty = f"{col_char_adj_ty}{row}"
                    addr_5mai = f"{col_char_5mai}{row}"
                    
                    # 調整TY
                    f_adj_ty = f'=IF({addr_ty}="[-]", "[-]", {addr_ty} * CeilingAnalysis!$M$6)'
                    ws_ev.cell(row=row, column=col_idx_adj_ty, value=f_adj_ty).number_format = '#,##0'
                    
                    # 5枚等価
                    f_5mai = f'=IF(OR({addr_adj_ty}="[-]", {addr_hatsu}="[-]"), "[-]", ({addr_adj_ty} - {addr_hatsu} * (50/{COIN_HOLD})) * 20)'
                    ws_ev.cell(row=row, column=col_idx_5mai, value=f_5mai).number_format = '#,##0'
                    
                    # 出玉率
                    denom = f'({addr_hatsu} * 3 + ({addr_adj_ty} / CeilingAnalysis!$L$13 * 3))'
                    numer = f'({denom} + {addr_5mai} / 20)'
                    f_rate = f'=IF(OR({addr_hatsu}="[-]", {addr_adj_ty}="[-]", {denom}=0), "[-]", {numer} / {denom})'
                    ws_ev.cell(row=row, column=col_idx_rate, value=f_rate).number_format = '0.0%'

                current_col += 8
        
        # シート生成実行
        create_bonus_ev_sheet("朝イチボーナス間期待値表", "TRUE")
        create_bonus_ev_sheet("朝イチ除外ボーナス間期待値表", "FALSE")
    
    print(f"Excel file created: {OUTPUT_FILE}")
        
        # 数式用列参照
if __name__ == "__main__":
    print("=== デュオ天国移行率分析 ===")
    print(f"設定: 天国閾値={HEAVEN_THRESHOLD}G, コイン持ち={COIN_HOLD}G/50枚")
    print()

    # キャッシュ有効性チェック
    use_cache = False
    data_mtime = os.path.getmtime(INPUT_FILE) if os.path.exists(INPUT_FILE) else 0
    
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, 'rb') as f:
                cache = pickle.load(f)
            if cache.get('mtime') == data_mtime:
                print("キャッシュから読み込み中...")
                df = cache['df']
                chains_df = cache['chains_df']
                use_cache = True
                print(f"  読み込み完了: {len(df)}行, チェーン数: {len(chains_df)}")
        except Exception as e:
            print(f"  キャッシュ読み込みエラー: {e}")
            use_cache = False
    
    if not use_cache:
        print("データ読み込み中...")
        df = load_data(INPUT_FILE)
        print(f"  読み込み完了: {len(df)}行")

        print("連チャン解析中...")
        df, chains_df = analyze_chains(df)
        


        # キャッシュ保存
        print("キャッシュ保存中...")
        with open(CACHE_FILE, 'wb') as f:
            pickle.dump({'mtime': data_mtime, 'df': df, 'chains_df': chains_df}, f)
        print("  保存完了")

    # 朝イチ連フラグの計算 (ID毎に最初の天国チェーン終了までをTRUE)
    # キャッシュ有無に関わらず実行
    print("朝イチ連フラグ計算中...")
    if 'Is_Morning_Chain' not in df.columns:
        chains_df['hvn_cumsum_temp'] = chains_df.groupby('ID')['Is_Heaven'].cumsum().astype(int)
        chains_df['Is_Morning_Chain'] = (chains_df['hvn_cumsum_temp'] == 0) | ((chains_df['hvn_cumsum_temp'] == 1) & (chains_df['Is_Heaven']))
        del chains_df['hvn_cumsum_temp']
        
        df = pd.merge(df, chains_df[['Chain_ID', 'Is_Morning_Chain']], on='Chain_ID', how='left')
        df['Is_Morning_Chain'] = df['Is_Morning_Chain'].fillna(False)

    print("天国移行率計算中...")
    heaven_rate_df = calculate_heaven_rate(chains_df)
    heaven_by_chain_df = calculate_heaven_rate_by_chain(chains_df)
    crosstab_df = calculate_heaven_rate_crosstab(chains_df)
    # 3連基準分析は軽量化版ではスキップ
    # rule3_df = calculate_3chain_rule_analysis(chains_df)

    print("ドキハナチャンス分析中...")
    chains_df, dokihana_stats = calculate_dokihana_analysis(chains_df)
    print(f"  ドキハナ発生回数: {dokihana_stats['total_count']}")
    print(f"  平均連チャン数: {dokihana_stats['avg_chain_len']:.2f}")
    print(f"  平均獲得枚数: {dokihana_stats['avg_dedama']:.0f}")

    print()
    print("【天国移行率（スルー回数別）】")
    print(heaven_rate_df.to_string(index=False))
    print()
    print("【天国移行率（前回連チャン別）】")
    print(heaven_by_chain_df.to_string(index=False))
    print()

    print("Excel出力中...")
    write_excel(df, heaven_rate_df, heaven_by_chain_df, crosstab_df, chains_df)

    print("完了!")

