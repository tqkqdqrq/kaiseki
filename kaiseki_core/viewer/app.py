"""kaiseki_core ビューアー (Streamlit, スマホ最適化版)。

起動:
    streamlit run kaiseki_core/viewer/app.py
"""

from __future__ import annotations
import argparse
import glob
import os
import sys
import pandas as pd
import streamlit as st
import plotly.express as px


DEFAULT_OUTPUT_DIRS = [
    r"G:\その他のパソコン\マイ ノートパソコン\解析\ビッグドリーム解析Python\output",
    r"G:\その他のパソコン\マイ ノートパソコン\解析\デュオ解析Python\output",
]

PREVIEW_ROWS = 500
HEAVY_SHEETS = {"Data", "ChainData"}

# 天井狙いシート判定 (シート名に含まれる文字列)
TENJOU_KEYWORDS = ("天井狙い",)


def _parse_args() -> argparse.Namespace:
    args_after_dd = []
    if "--" in sys.argv:
        idx = sys.argv.index("--")
        args_after_dd = sys.argv[idx + 1 :]
    p = argparse.ArgumentParser()
    p.add_argument("--output-dir", action="append", default=None)
    ns, _ = p.parse_known_args(args_after_dd)
    if ns.output_dir is None:
        ns.output_dir = [d for d in DEFAULT_OUTPUT_DIRS if os.path.isdir(d)]
    return ns


@st.cache_data(show_spinner=False, ttl=10)
def _list_xlsx(dirs: list[str]) -> list[str]:
    files: list[str] = []
    for d in dirs:
        if not os.path.isdir(d):
            continue
        files.extend(glob.glob(os.path.join(d, "*.xlsx")))
    files = [f for f in files if os.path.exists(f)]
    files.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return files


@st.cache_data(show_spinner=False)
def _list_sheets(path: str, mtime: float) -> list[str]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


@st.cache_data(show_spinner="シート読込中...")
def _read_sheet(path: str, sheet: str, mtime: float, nrows: int | None = None) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=sheet, nrows=nrows, engine="openpyxl")


def _machine_label(path: str) -> str:
    parts = os.path.normpath(path).split(os.sep)
    for p in reversed(parts):
        if p.endswith("解析Python"):
            return p.replace("解析Python", "")
    return os.path.basename(os.path.dirname(path))


def _is_tenjou_sheet(name: str) -> bool:
    return any(kw in name for kw in TENJOU_KEYWORDS)


def _render_tenjou_sheet(df: pd.DataFrame, name: str) -> None:
    """天井狙いシート専用のスマホ最適化UI。"""
    df = df.dropna(subset=["打ち始めG"]).copy()
    df = df[df["打ち始めG"].astype(str).str.endswith("G")]

    # G値で数値化
    def _g_num(lbl: str) -> int:
        try:
            return int(str(lbl).rstrip("G+(異常)").rstrip("G").split("-")[0])
        except Exception:
            return -1
    df["_G"] = df["打ち始めG"].map(_g_num)
    df = df[df["_G"] >= 0].sort_values("_G").reset_index(drop=True)

    if df.empty:
        st.info("(空)")
        return

    # 1. 主要グラフ (累積EV円 と 機械割推移)
    ev_yen_cols = [c for c in df.columns if "累積EV(円)" in c]
    kikai_cols = [c for c in df.columns if "機械割" in c]
    if ev_yen_cols:
        st.markdown("### 累積EV(円) の推移")
        fig = px.line(df, x="_G", y=ev_yen_cols[0],
                      labels={"_G": "打ち始めG", ev_yen_cols[0]: "期待値(円)"})
        fig.add_hline(y=0, line_dash="dash", line_color="gray")
        fig.update_layout(height=300, margin=dict(l=10, r=10, t=10, b=10))
        st.plotly_chart(fig, use_container_width=True)
    if kikai_cols:
        st.markdown("### 機械割(%) の推移")
        fig = px.line(df, x="_G", y=kikai_cols[0],
                      labels={"_G": "打ち始めG", kikai_cols[0]: "機械割%"})
        fig.add_hline(y=100, line_dash="dash", line_color="red")
        fig.update_layout(height=300, margin=dict(l=10, r=10, t=10, b=10))
        st.plotly_chart(fig, use_container_width=True)

    # 2. フィルタ: 機械割100%超 / プラス期待値
    st.markdown("### フィルタ")
    filter_choice = st.radio("表示行", ["全部", "機械割100%超", "累積EV(円)プラス"], horizontal=True, key=f"f_{name}")
    fdf = df.copy()
    if filter_choice == "機械割100%超" and kikai_cols:
        fdf = fdf[pd.to_numeric(fdf[kikai_cols[0]], errors="coerce") >= 100]
    elif filter_choice == "累積EV(円)プラス" and ev_yen_cols:
        fdf = fdf[pd.to_numeric(fdf[ev_yen_cols[0]], errors="coerce") > 0]

    # 3. 主要列のみ表示 (スマホで見やすく)
    key_cols = ["打ち始めG", "サンプル数", "累積平均初当G", "累積連チャンTY(枚)"]
    if ev_yen_cols:
        key_cols.append(ev_yen_cols[0])
    if kikai_cols:
        key_cols.append(kikai_cols[0])
    show_cols = [c for c in key_cols if c in fdf.columns]
    st.markdown("### 主要列のみ")
    st.dataframe(fdf[show_cols], width="stretch", height=400, hide_index=True)

    # 4. 全列 (折りたたみ)
    with st.expander("全列を表示"):
        all_cols = [c for c in df.columns if c != "_G"]
        st.dataframe(fdf[all_cols], width="stretch", hide_index=True)


def _render_summary_metrics(path: str, sheets: list[str], mtime: float) -> None:
    """トップに主要メトリック (CZ成功率、サンプル数等)。"""
    if "CZ成功率" in sheets:
        cz = _read_sheet(path, "CZ成功率", mtime)
        if "値" in cz.columns and "項目" in cz.columns:
            cols = st.columns(2)
            for i, (_, row) in enumerate(cz.iterrows()):
                if i >= 4:
                    break
                val = row["値"]
                disp = f"{val:.1%}" if isinstance(val, float) and 0 <= val <= 1 else f"{val:,}" if isinstance(val, (int, float)) else str(val)
                cols[i % 2].metric(str(row["項目"]), disp)


def main() -> None:
    st.set_page_config(page_title="パチスロ解析", layout="centered", initial_sidebar_state="collapsed")
    # スマホ向け CSS
    st.markdown("""
        <style>
        .block-container { padding-top: 1rem; padding-bottom: 1rem; }
        h1 { font-size: 1.4rem; }
        h3 { font-size: 1.1rem; margin-top: 1rem; }
        [data-testid="stMetricValue"] { font-size: 1.3rem; }
        </style>
    """, unsafe_allow_html=True)

    args = _parse_args()
    files = _list_xlsx(args.output_dir)
    files = [f for f in files if os.path.exists(f)]

    if not files:
        st.error("xlsxが見つかりません")
        return

    # ファイル選択 (シンプル)
    labels = [f"{_machine_label(p)} {os.path.basename(p)[-19:-5]}" for p in files]
    idx = st.selectbox("ファイル", range(len(files)), format_func=lambda i: labels[i])
    path = files[idx]
    mtime = os.path.getmtime(path)
    all_sheets = _list_sheets(path, mtime)
    value_sheets = {s[:-1] for s in all_sheets if s.endswith("_値")}
    sheet_names = [s for s in all_sheets if s not in value_sheets]

    # 主要メトリック
    _render_summary_metrics(path, sheet_names, mtime)

    # シート選択 (プルダウン、タブよりモバイル親和性高)
    sheet = st.selectbox("シート", sheet_names, key="sheet_sel")

    st.markdown(f"## {sheet}")

    # シート種別で表示切替
    if sheet in HEAVY_SHEETS:
        if not st.checkbox(f"{sheet} を読み込む (プレビュー{PREVIEW_ROWS}行)", key=f"load_{path}_{sheet}"):
            st.info("チェックすると読み込みます")
            return
        df = _read_sheet(path, sheet, mtime, nrows=PREVIEW_ROWS)
        st.caption(f"{len(df):,} 行 (プレビュー)")
        st.dataframe(df, width="stretch")
        return

    df = _read_sheet(path, sheet, mtime)
    if df.empty:
        st.info("(空)")
        return

    if _is_tenjou_sheet(sheet):
        _render_tenjou_sheet(df, sheet)
    else:
        # 一般シート: そのまま + 自動グラフ
        st.dataframe(df, width="stretch", hide_index=True)
        if len(df) <= 500 and len(df.columns) >= 2:
            cnt_col = "件数" if "件数" in df.columns else df.columns[1]
            if pd.api.types.is_numeric_dtype(df[cnt_col]):
                try:
                    fig = px.bar(df, x=df.columns[0], y=cnt_col)
                    fig.update_layout(height=250, margin=dict(l=10, r=10, t=10, b=10))
                    st.plotly_chart(fig, width="stretch")
                except Exception:
                    pass


main()
