"""xlsx → スマホ最適化 HTML ダッシュボード書き出し。

使い方:
    python -m kaiseki_core.viewer.html_export             # dashboard.html 生成
    python -m kaiseki_core.viewer.html_export --serve     # 生成 + http.server 起動 (port 8501)
"""

from __future__ import annotations
import argparse
import glob
import json
import os
import re
import sys
from typing import Any
from openpyxl import load_workbook


# DEFAULT_OUTPUT_DIRS と MACHINE_SPECS は machines.json から導出
import sys as _sys
_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if _REPO_ROOT not in _sys.path:
    _sys.path.insert(0, _REPO_ROOT)
import machines as _m
DEFAULT_OUTPUT_DIRS = _m.output_dirs()

# 単一シート構成の天井狙い (現状 AT間天井狙いのみ)。
# morning/cut の N 回目シートは横並び統合シート (AT間_morning / AT間_cut) から
# extract_xlsx 内で個別キーに分解して out["tenjou"] に格納する。
TENJOU_NAMES = ["AT間天井狙い", "AT間天井狙い_直撃後"]

# 横並び統合シートの 1 ブロック幅 (excel_through._GBIN_BLOCK_WIDTH と整合)
_GBIN_BLOCK_WIDTH = 14
_GBIN_BLOCK_SPACER = 1

# 統合シート内のブロック順序と out["tenjou"] キーのマッピング
# (sheets.py::_build_session_blocks と整合: 1回目 → 2回目以降 → 2..10 回目)
def _session_block_keys(prefix: str) -> list[str]:
    keys = [f"AT間_{prefix}_1", f"AT間_{prefix}_2plus"]
    for n in range(2, 11):
        keys.append(f"AT間_{prefix}_{n}")
    return keys


def _build_session_options(prefix: str) -> list[tuple[str, str]]:
    base = [
        ("1回目", f"AT間_{prefix}_1"),
        ("2回目以降", f"AT間_{prefix}_2plus"),
    ]
    for n in range(2, 11):
        base.append((f"{n}回目", f"AT間_{prefix}_{n}"))
    return base


# セッショングループ (UI集約: 朝/切断後 トグル + 回数プルダウン)
SESSION_GROUP = {
    "label": "回数別AT間",
    "modes": [
        {"key": "morning", "label": "朝", "options": _build_session_options("morning")},
        {"key": "cut", "label": "切断後", "options": _build_session_options("cut")},
    ],
}

SIMPLE_SHEETS = ["状態分布", "連荘初動状態", "連荘長分布"]


def _machine_label(path: str) -> str:
    parts = os.path.normpath(path).split(os.sep)
    for p in reversed(parts):
        if p.endswith("解析Python"):
            return p.replace("解析Python", "")
    return os.path.basename(os.path.dirname(path))


def _g_value(label: Any) -> int | None:
    if not isinstance(label, str):
        return None
    m = re.match(r"^(\d+)G", label)
    return int(m.group(1)) if m else None


def _sheet_rows(ws, max_rows: int | None = None) -> list[dict]:
    headers = [c.value for c in ws[1]]
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if max_rows and i >= max_rows:
            break
        if not any(v is not None for v in row):
            continue
        rows.append(dict(zip(headers, row)))
    return rows


MACHINE_SPECS = _m.machine_specs()


def extract_xlsx(path: str) -> dict:
    wb = load_workbook(path, data_only=True)
    machine = _machine_label(path)
    out: dict[str, Any] = {
        "path": path,
        "name": os.path.basename(path),
        "machine": machine,
        "spec": MACHINE_SPECS.get(machine, {}),
        "mtime": int(os.path.getmtime(path)),
        "kpi": {},
        "tenjou": {},
        "simple": {},
    }

    # 連荘数, データ数 (ChainData / Data シートの行数で代用)
    if "ChainData" in wb.sheetnames:
        out["kpi"].setdefault("連荘数", wb["ChainData"].max_row - 1)
    if "Data" in wb.sheetnames:
        out["kpi"].setdefault("データ行数", wb["Data"].max_row - 1)

    # 総ゲーム数 = 通常時G + AT中G
    at_payout = out["spec"].get("at_payout_per_g", 8.2)
    normal_g = 0
    if "Data" in wb.sheetnames:
        ws = wb["Data"]
        headers = [c.value for c in ws[1]]
        try:
            start_idx = headers.index("Start") + 1
            for row in ws.iter_rows(min_row=2, min_col=start_idx, max_col=start_idx, values_only=True):
                v = row[0]
                if isinstance(v, (int, float)):
                    normal_g += int(v)
        except ValueError:
            pass
    at_g = 0
    if "ChainData" in wb.sheetnames:
        ws = wb["ChainData"]
        headers = [c.value for c in ws[1]]
        try:
            np_idx = headers.index("net_payout") + 1
            ty_sum = 0
            for row in ws.iter_rows(min_row=2, min_col=np_idx, max_col=np_idx, values_only=True):
                v = row[0]
                if isinstance(v, (int, float)):
                    ty_sum += v
            at_g = int(ty_sum / at_payout) if at_payout else 0
        except ValueError:
            pass
    out["kpi"]["総ゲーム数"] = normal_g + at_g
    out["kpi"]["通常G"] = normal_g
    out["kpi"]["AT中G"] = at_g

    # 天井狙いシート (_値 版)
    for name in TENJOU_NAMES:
        val_name = name + "_値"
        if val_name not in wb.sheetnames:
            continue
        ws = wb[val_name]
        headers = [c.value for c in ws[1]]
        # 累積EVラベル特定 (例: "AT到達まで累積EV(円)", "CZ1回目当選まで累積EV(円)")
        # 新旧互換: "累積期待値(円)" (新) or "累積EV(円)" (旧)
        ev_yen_col = next((h for h in headers if isinstance(h, str) and ("累積期待値(円)" in h or "累積EV(円)" in h)), None)
        ev_coin_col = next((h for h in headers if isinstance(h, str) and ("累積期待値(枚)" in h or "累積EV(枚)" in h)), None)
        kikai_col = next((h for h in headers if isinstance(h, str) and "機械割" in h), None)
        # 算出条件 (1行目 column 16 = excel_through.FILTER_DESC_CELL_COL)
        filter_desc = ws.cell(row=1, column=16).value
        filter_description = ""
        if isinstance(filter_desc, str) and filter_desc.startswith("算出条件:"):
            filter_description = filter_desc[len("算出条件:"):].strip()
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            g = _g_value(row[0])
            if g is None:
                continue
            rec = dict(zip(headers, row))
            rec["_g"] = g
            rows.append(rec)
        out["tenjou"][name] = {
            "label": name,
            "rows": rows,
            "ev_yen_col": ev_yen_col,
            "ev_coin_col": ev_coin_col,
            "kikai_col": kikai_col,
            "filter_description": filter_description,
        }

    # 横並び統合シート (AT間_morning_値 / AT間_cut_値) を per-N キーに分解
    for prefix in ("morning", "cut"):
        sheet_name = f"AT間_{prefix}_値"
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        keys = _session_block_keys(prefix)
        block_pitch = _GBIN_BLOCK_WIDTH + _GBIN_BLOCK_SPACER
        for idx, key in enumerate(keys):
            col_start = 1 + idx * block_pitch  # 1-based
            # block 構造: row1=title, row2=filter_desc, row3=header, row4+=data
            title = ws.cell(row=1, column=col_start).value
            fdesc_raw = ws.cell(row=2, column=col_start).value
            filter_description = ""
            if isinstance(fdesc_raw, str) and fdesc_raw.startswith("算出条件:"):
                filter_description = fdesc_raw[len("算出条件:"):].strip()
            headers = [ws.cell(row=3, column=col_start + c).value
                       for c in range(_GBIN_BLOCK_WIDTH)]
            if not headers[0]:
                continue
            ev_yen_col = next((h for h in headers if isinstance(h, str) and ("累積期待値(円)" in h or "累積EV(円)" in h)), None)
            ev_coin_col = next((h for h in headers if isinstance(h, str) and ("累積期待値(枚)" in h or "累積EV(枚)" in h)), None)
            kikai_col = next((h for h in headers if isinstance(h, str) and "機械割" in h), None)
            rows = []
            r = 4
            while r <= ws.max_row:
                lbl = ws.cell(row=r, column=col_start).value
                if lbl is None:
                    r += 1
                    continue
                g = _g_value(lbl)
                if g is None:
                    # max_g+(異常) 行で打ち切り
                    break
                values = [ws.cell(row=r, column=col_start + c).value
                          for c in range(_GBIN_BLOCK_WIDTH)]
                rec = dict(zip(headers, values))
                rec["_g"] = g
                rows.append(rec)
                r += 1
            out["tenjou"][key] = {
                "label": title or key,
                "rows": rows,
                "ev_yen_col": ev_yen_col,
                "ev_coin_col": ev_coin_col,
                "kikai_col": kikai_col,
                "filter_description": filter_description,
            }

    # 直前有利区間G別期待値_値 (8 ブロック横並び、AT間天井タブで切替表示)
    PREV_YURI_KEYS = [
        ("prev_yuri_0_999", "0-999G"),
        ("prev_yuri_1000_1999", "1000-1999G"),
        ("prev_yuri_2000_2999", "2000-2999G"),
        ("prev_yuri_3000_3999", "3000-3999G"),
        ("prev_yuri_4000_4999", "4000-4999G"),
        ("prev_yuri_5000_5999", "5000-5999G"),
        ("prev_yuri_6000_6999", "6000-6999G"),
        ("prev_yuri_7000_plus", "7000G以上"),
    ]
    py_sheet = "直前有利区間G別期待値_値"
    if py_sheet in wb.sheetnames:
        ws = wb[py_sheet]
        block_pitch = _GBIN_BLOCK_WIDTH + _GBIN_BLOCK_SPACER
        for idx, (key, _title) in enumerate(PREV_YURI_KEYS):
            col_start = 1 + idx * block_pitch
            title = ws.cell(row=1, column=col_start).value
            fdesc_raw = ws.cell(row=2, column=col_start).value
            filter_description = ""
            if isinstance(fdesc_raw, str) and fdesc_raw.startswith("算出条件:"):
                filter_description = fdesc_raw[len("算出条件:"):].strip()
            headers = [ws.cell(row=3, column=col_start + c).value
                       for c in range(_GBIN_BLOCK_WIDTH)]
            if not headers[0]:
                continue
            ev_yen_col = next((h for h in headers if isinstance(h, str) and ("累積期待値(円)" in h or "累積EV(円)" in h)), None)
            ev_coin_col = next((h for h in headers if isinstance(h, str) and ("累積期待値(枚)" in h or "累積EV(枚)" in h)), None)
            kikai_col = next((h for h in headers if isinstance(h, str) and "機械割" in h), None)
            rows = []
            r = 4
            while r <= ws.max_row:
                lbl = ws.cell(row=r, column=col_start).value
                if lbl is None:
                    r += 1
                    continue
                g = _g_value(lbl)
                if g is None:
                    break
                values = [ws.cell(row=r, column=col_start + c).value
                          for c in range(_GBIN_BLOCK_WIDTH)]
                rec = dict(zip(headers, values))
                rec["_g"] = g
                rows.append(rec)
                r += 1
            out["tenjou"][key] = {
                "label": title or key,
                "rows": rows,
                "ev_yen_col": ev_yen_col,
                "ev_coin_col": ev_coin_col,
                "kikai_col": kikai_col,
                "filter_description": filter_description,
            }

    # シンプルシート
    for s in SIMPLE_SHEETS:
        if s in wb.sheetnames:
            ws = wb[s]
            out["simple"][s] = {
                "headers": [c.value for c in ws[1]],
                "rows": _sheet_rows(ws),
            }

    # 累積G別AT間 (morning / cut)
    out["cycle_g_at"] = {}
    for key, sheet in [("morning", "累積G別AT間_morning_値"),
                        ("cut", "累積G別AT間_cut_値")]:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        # メインブロックは row 2 (タイトル) + row 3 (header) + row 4-18 (15ビン)
        # 比較ブロックは row 20 (タイトル) + row 21 (header) + row 22-36 (15ビン)
        def _read_block(header_row: int, data_rows: int) -> list[dict]:
            headers = [c.value for c in ws[header_row]]
            block = []
            for r in range(header_row + 1, header_row + 1 + data_rows):
                row = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
                if row[0] is None:
                    continue
                block.append(dict(zip(headers, row)))
            return block
        # excel_through.py の _write_cycle_block:
        # ビン数 = CYCLE_BIN_DEFAULT 長 (現状 8ビン: 1000G刻み + 7000+)
        # 通常G累積: row1=タイトル, row2=ヘッダ, row3-(2+N)=データ (next_row=3+N)
        # 直前有利区間G:  row(next+2)=タイトル, row(next+3)=ヘッダ, row(next+4)-...=データ
        N_BINS = 8  # CYCLE_BIN_DEFAULT 長と一致 (1000G×7 + 7000+)
        normal_rows = _read_block(2, N_BINS)
        # 通常G累積 next_row = 2 + 1 + N_BINS = N_BINS+3
        # 空行+2 → タイトル = N_BINS+5、ヘッダ = N_BINS+6
        total_rows = _read_block(N_BINS + 6, N_BINS)
        # 算出条件メタ (1行目 col 16)
        desc = ws.cell(row=1, column=16).value
        fdesc = ""
        if isinstance(desc, str) and desc.startswith("算出条件:"):
            fdesc = desc[len("算出条件:"):].strip()
        out["cycle_g_at"][key] = {
            "normal_rows": normal_rows,
            "total_rows": total_rows,
            "filter_description": fdesc,
        }

    # 引き戻し分布 (chokugeki と同じ multi-block 構造)
    out["hikimodoshi"] = None
    if "引き戻し分布_値" in wb.sheetnames:
        ws = wb["引き戻し分布_値"]
        blocks = []
        cur_title = None; cur_group_label = None; cur_rows = []
        for r in range(1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            b = ws.cell(row=r, column=2).value
            c = ws.cell(row=r, column=3).value
            d = ws.cell(row=r, column=4).value
            e = ws.cell(row=r, column=5).value
            f = ws.cell(row=r, column=6).value
            if a is None and b is None:
                if cur_title and cur_rows:
                    blocks.append({"title": cur_title, "group_label": cur_group_label, "rows": cur_rows})
                cur_title = None; cur_group_label = None; cur_rows = []
                continue
            if isinstance(a, str) and a.startswith("【"):
                if cur_title and cur_rows:
                    blocks.append({"title": cur_title, "group_label": cur_group_label, "rows": cur_rows})
                cur_title = a; cur_group_label = None; cur_rows = []
                continue
            if isinstance(a, str) and (b == "サンプル数" or b == "サンプル"):
                cur_group_label = a
                continue
            if a is None:
                continue
            cur_rows.append({"label": a, "sample": b,
                             "jyoi": c, "rate_jyoi": d,
                             "n390": e, "rate_n390": f})
        if cur_title and cur_rows:
            blocks.append({"title": cur_title, "group_label": cur_group_label, "rows": cur_rows})
        out["hikimodoshi"] = {"blocks": blocks}

    # 直撃分布 (複数ブロック)
    out["chokugeki"] = None
    if "直撃分布_値" in wb.sheetnames:
        ws = wb["直撃分布_値"]
        blocks = []
        cur_title = None
        cur_group_label = None
        cur_rows = []
        for r in range(1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            b = ws.cell(row=r, column=2).value
            c = ws.cell(row=r, column=3).value
            d = ws.cell(row=r, column=4).value
            e = ws.cell(row=r, column=5).value
            if a is None and b is None:
                if cur_title and cur_rows:
                    blocks.append({"title": cur_title, "group_label": cur_group_label, "rows": cur_rows})
                cur_title = None; cur_group_label = None; cur_rows = []
                continue
            if isinstance(a, str) and a.startswith("【"):
                if cur_title and cur_rows:
                    blocks.append({"title": cur_title, "group_label": cur_group_label, "rows": cur_rows})
                cur_title = a
                cur_group_label = None
                cur_rows = []
                continue
            if isinstance(a, str) and (b == "サンプル数" or b == "サンプル"):
                cur_group_label = a
                continue
            if a is None:
                continue
            cur_rows.append({"label": a, "sample": b, "choku": c, "rate": d, "ty": e})
        if cur_title and cur_rows:
            blocks.append({"title": cur_title, "group_label": cur_group_label, "rows": cur_rows})
        out["chokugeki"] = {"blocks": blocks}

    wb.close()
    return out


def collect_files(dirs: list[str], limit: int = 20) -> list[dict]:
    files = []
    for d in dirs:
        if os.path.isdir(d):
            files.extend(glob.glob(os.path.join(d, "*.xlsx")))
    files = [f for f in files if os.path.exists(f)]
    files.sort(key=os.path.getmtime, reverse=True)
    return [extract_xlsx(f) for f in files[:limit]]


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1,viewport-fit=cover">
<meta name="theme-color" content="#0a0d12">
<title>P/H — Pachislot Hunter</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Bebas+Neue&family=JetBrains+Mono:wght@400;500;700&family=Zen+Kaku+Gothic+Antique:wght@400;500;700&family=Zen+Kaku+Gothic+New:wght@700;900&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<style>
:root{
  /* 親 WP テーマ (suroschool.jp ダーク) と完全融合する gold/dark パレット */
  --bg-void:#06070a;
  --bg-surface:#131419;
  --bg-elev:#1a1c24;
  --bg-row-good:rgba(245,184,0,.05);
  --bg-row-loss:rgba(255,46,126,.05);
  --border:#22242c;
  --border-bright:#2d2f38;
  --text:#ece9df;
  --text-2:#a39c8a;
  --text-dim:#6b6557;
  --lime:#f5b800;            /* 「lime」名前は互換、実体は gold */
  --lime-glow:rgba(245,184,0,.30);
  --amber:#ffd96b;
  --red:#ff2e7e;             /* 親テーマの magenta */
  --cyan:#45e2ff;
  --pink:#ff2e7e;
  --font-display:'Zen Kaku Gothic New','Hiragino Sans',sans-serif;
  --font-mono:ui-monospace,'SF Mono','Menlo','Consolas',monospace;
  --font-body:'Zen Kaku Gothic Antique','Hiragino Sans',system-ui,sans-serif;
}
/* === Light theme (親WP <html data-theme="light"> から ?theme=light で伝搬) === */
html[data-theme="light"]{
  --bg-void:#f4f5f7;
  --bg-surface:#ffffff;
  --bg-elev:#ffffff;
  --bg-row-good:rgba(202,138,4,.06);
  --bg-row-loss:rgba(190,40,90,.06);
  --border:#e6e7eb;
  --border-bright:#d1d3d9;
  --text:#1a1a1a;
  --text-2:#5b5447;
  --text-dim:#8a8474;
  --lime:#c48f00;            /* 明所での gold は彩度落として読みやすく */
  --lime-glow:rgba(196,143,0,.25);
  --amber:#a86f00;
  --red:#c41f63;
  --cyan:#0883a3;
  --pink:#c41f63;
}
html[data-theme="light"] body{
  background-image:
    radial-gradient(ellipse 80% 40% at 50% 0%,rgba(31,157,58,.05) 0%,transparent 70%),
    radial-gradient(ellipse 60% 60% at 80% 100%,rgba(16,131,161,.04) 0%,transparent 70%);
}
html[data-theme="light"] body::before{
  background-image:repeating-linear-gradient(0deg,transparent 0 3px,rgba(0,0,0,.008) 3px 4px);
}
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0;-webkit-tap-highlight-color:transparent;}
/* iframe 内は親 WP テーマと完全融合: 背景透過 */
html{font-size:16px;background:transparent;}
body{
  font-family:var(--font-body);
  color:var(--text);
  background:transparent;
  padding-bottom:env(safe-area-inset-bottom);
  overflow-x:hidden;
}
button{font-family:inherit;color:inherit;background:none;border:none;cursor:pointer;}
select{font-family:inherit;color:inherit;}

/* === Header === iframe 埋め込み前提なのでヘッダ非表示 */
.status-bar{display:none !important;}
.brand,.file-label{display:none;}
/* セッションプルダウン (iOS Safari対応で option も明示色指定) */
.session-pick{
  background:var(--bg-elev);
  border:1px solid var(--lime);
  color:var(--text);
  padding:.45rem 1.7rem .45rem .65rem;
  font-family:var(--font-body);font-size:.85rem;font-weight:600;
  appearance:none;-webkit-appearance:none;
  background-image:url("data:image/svg+xml;utf8,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 12 12' fill='%23f5b800'><polygon points='6,8 2,4 10,4'/></svg>");
  background-repeat:no-repeat;background-position:right .5rem center;background-size:11px;
  border-radius:5px;
  flex:1;min-width:0;
}
.session-pick option{
  background:var(--bg-surface);
  color:var(--text);
}
/* 朝/切断後 toggle と回数プルダウンを横並び */
.session-row{
  display:flex;align-items:center;gap:.5rem;
  margin-bottom:.7rem;
}
.session-row .mode-toggle{margin:0;flex-shrink:0;}
.session-row .gbin-toggle{margin:0;flex-shrink:0;}
.status-tag{
  font-family:var(--font-mono);font-size:.65rem;letter-spacing:.15em;
  color:var(--text-dim);
}

/* === Machine spec === シームレス、外枠なしで親に溶け込む */
.machine-card{
  margin:.4rem .55rem .55rem;
  padding:0;
  background:transparent;
  border:none;
}
.machine-card::before{display:none;}
.machine-name{
  font-family:var(--font-display);font-weight:700;
  font-size:.95rem;line-height:1.2;letter-spacing:.02em;
  color:var(--text);margin-bottom:.45rem;
}
.machine-name .type{
  margin-left:.4rem;
  font-family:var(--font-mono);font-size:.58rem;font-weight:400;
  color:var(--text-dim);letter-spacing:.08em;
}
.cond-grid{
  display:grid;grid-template-columns:1fr 1fr;
  gap:.15rem .8rem;
  font-family:var(--font-mono);font-size:.68rem;
}
.cond-row{
  display:flex;justify-content:space-between;align-items:baseline;gap:.5rem;
  border-bottom:1px solid rgba(255,255,255,.04);
  padding:.18rem 0;
}
.cond-row .k{color:var(--text-dim);letter-spacing:.04em;flex-shrink:0;}
.cond-row .v{color:var(--text);font-weight:600;text-align:right;
  white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.cond-row .v.accent{color:var(--lime);}
html[data-theme="light"] .cond-row{border-bottom-color:rgba(0,0,0,.06);}

/* === Sheet tabs === iframe 上部に sticky、両端 fade マスク */
.sheet-tabs{
  display:flex;gap:0;overflow-x:auto;overflow-y:hidden;
  scrollbar-width:none;
  touch-action:pan-x;
  overscroll-behavior-x:contain;
  border-bottom:1px solid var(--border);
  padding:0 .5rem;
  position:sticky;top:0;z-index:15;
  background:rgba(6,7,10,.94);
  backdrop-filter:blur(10px);-webkit-backdrop-filter:blur(10px);
  -webkit-mask-image:linear-gradient(90deg,transparent 0,#000 12px,#000 calc(100% - 12px),transparent 100%);
          mask-image:linear-gradient(90deg,transparent 0,#000 12px,#000 calc(100% - 12px),transparent 100%);
}
html[data-theme="light"] .sheet-tabs{
  background:rgba(244,245,247,.94);
}
.sheet-tabs::-webkit-scrollbar{display:none;}
.tab{
  flex-shrink:0;
  padding:.7rem .65rem .55rem;
  font-family:var(--font-body);
  font-size:.78rem;font-weight:500;
  color:var(--text-2);
  position:relative;white-space:nowrap;
  transition:color .15s;
}
.tab:hover{color:var(--text);}
.tab.active{color:var(--lime);font-weight:700;}
.tab.active::after{
  content:'';position:absolute;left:.35rem;right:.35rem;bottom:-1px;
  height:2px;background:var(--lime);box-shadow:0 0 6px var(--lime-glow);
  border-radius:1px 1px 0 0;
}
.tab .star{margin-right:.2rem;font-size:.6rem;color:var(--amber);}

/* === Views === シームレス: 横余白最小 */
.view{padding:.7rem .35rem 2rem;}
.section-label{
  font-family:var(--font-mono);font-size:.62rem;
  color:var(--text-dim);letter-spacing:.2em;text-transform:uppercase;
  margin-bottom:.55rem;
}

/* Quick stats */
.quick-stats{
  display:grid;grid-template-columns:1fr 1fr;gap:.45rem;
  margin-bottom:.8rem;
}
.qs-card{
  background:var(--bg-surface);
  border:1px solid var(--border);
  border-radius:4px;
  padding:.55rem .65rem;
  position:relative;
}
.qs-card.hot{
  border-color:var(--lime);
  background:linear-gradient(135deg,rgba(245,184,0,.08) 0%,transparent 60%);
}
.qs-card.hot::before{
  content:'狙い目';
  position:absolute;top:-7px;left:8px;
  background:var(--bg-void);
  color:var(--lime);
  font-family:var(--font-mono);font-size:.55rem;letter-spacing:.14em;
  padding:0 5px;font-weight:700;
}
.qs-card .qs-lbl{
  display:block;font-size:.55rem;color:var(--text-dim);
  letter-spacing:.14em;text-transform:uppercase;margin-bottom:.25rem;font-weight:600;
}
.qs-card .qs-val{
  display:block;font-family:var(--font-mono);
  font-size:1.15rem;font-weight:700;line-height:1;
  font-variant-numeric:tabular-nums;
}
.qs-card.hot .qs-val{color:var(--lime);}
.qs-card .qs-sub{
  display:block;font-family:var(--font-mono);font-size:.65rem;
  color:var(--text-2);margin-top:.25rem;letter-spacing:.04em;
}

/* Chart */
.chart-card{
  background:var(--bg-surface);
  border:1px solid var(--border);
  border-radius:4px;
  padding:.7rem .5rem .4rem;
  margin-bottom:.8rem;
  height:200px;
  position:relative;
}
.chart-card .chart-title{
  font-family:var(--font-mono);font-size:.55rem;font-weight:700;
  color:var(--text-dim);letter-spacing:.14em;text-transform:uppercase;
  position:absolute;top:.5rem;left:.7rem;
}

/* Filter chips */
.filter-bar{
  display:flex;gap:.45rem;
  overflow-x:auto;scrollbar-width:none;
  margin-bottom:.7rem;padding-bottom:.2rem;
}
.filter-bar::-webkit-scrollbar{display:none;}
.chip{
  flex-shrink:0;
  background:var(--bg-surface);
  border:1px solid var(--border-bright);
  color:var(--text-2);
  font-family:var(--font-mono);font-size:.72rem;
  padding:.5rem .85rem;letter-spacing:.06em;
  transition:transform .08s;
}
.chip:active{transform:scale(.96);}
.chip.active{
  background:var(--lime);border-color:var(--lime);color:var(--bg-void);
  font-weight:700;
}

/* Step toggle (5G/50G) */
/* .view 直下のトグルはそれぞれ独立した行で表示 (AT間天井狙いタブ等)
   width:fit-content で内容分の幅だけ確保 (横の黒帯を出さない) */
.view > .mode-toggle{display:flex;width:fit-content;margin:0 0 .5rem;}
.view > .step-toggle{display:flex;width:fit-content;margin:0 0 .7rem;}
.step-toggle{
  display:inline-flex;
  background:var(--bg-surface);
  border:1px solid var(--border-bright);
  margin-bottom:.7rem;
  overflow:hidden;
}
.step-toggle button{
  padding:.5rem .9rem;
  font-family:var(--font-mono);font-size:.72rem;
  color:var(--text-2);
  letter-spacing:.06em;
  border-right:1px solid var(--border-bright);
}
.step-toggle button:last-child{border-right:none;}
.step-toggle button.active{
  background:var(--lime);color:var(--bg-void);font-weight:700;
}
.mode-toggle{
  display:inline-flex;
  background:var(--bg-surface);
  border:1px solid var(--border-bright);
  margin:0 0 .55rem;
  border-radius:6px;
  overflow:hidden;
}
.mode-toggle button{
  padding:.5rem 1.1rem;
  font-family:var(--font-display);font-size:.85rem;
  letter-spacing:.1em;
  color:var(--text-2);
  border-right:1px solid var(--border-bright);
}
.mode-toggle button:last-child{border-right:none;}
.mode-toggle button.active{
  background:var(--amber);color:var(--bg-void);font-weight:700;
}

/* 前回有利区間G サブトグル: 8ボタンをコンパクトに折返し (スマホ最適化) */
.gbin-toggle{
  display:flex;flex-wrap:wrap;gap:.3rem;margin:0 0 .55rem;
}
.gbin-toggle button{
  padding:.32rem .6rem;
  font-family:var(--font-display);font-size:.72rem;letter-spacing:.02em;
  color:var(--text-2);
  background:var(--bg-surface);
  border:1px solid var(--border-bright);
  border-radius:5px;
  white-space:nowrap;
}
.gbin-toggle button.active{
  background:var(--amber);color:var(--bg-void);font-weight:700;border-color:var(--amber);
}

/* Tenjou table — モバイル375pxに最適化、横スクロール可・先頭列sticky
   max-height で内部スクロールにして iframe 高さを予測可能にする
   (5G⇄50G 切替時の iframe 残像対策) */
.table-wrap{
  background:var(--bg-surface);
  border:1px solid var(--border);
  border-radius:4px;
  max-height:560px;
  overflow:auto;
  -webkit-overflow-scrolling:touch;
  position:relative;
}
@media(min-width:600px){
  .table-wrap{max-height:680px;}
}
.filter-banner{
  display:flex;align-items:center;gap:.4rem;
  margin:.4rem .6rem .25rem;padding:.35rem .55rem;
  background:var(--bg-elev);
  border:1px solid var(--border-bright);
  border-left:2px solid var(--lime);
  border-radius:3px;
  font-size:.7rem;
  position:sticky;top:38px;z-index:5;  /* sheet-tabs の下 */
}
#globalFilterBanner{display:none !important;}  /* 算出条件は上の machine-card に集約 */
.filter-banner-label{
  font-family:var(--font-mono);
  font-size:.55rem;font-weight:700;
  letter-spacing:.12em;
  color:var(--lime);
  text-transform:uppercase;
}
.filter-banner-text{
  color:var(--text-2);
  font-family:var(--font-mono);
  font-size:.66rem;
  line-height:1.3;
}
.tenjou-table{
  width:100%;
  border-collapse:collapse;
  font-family:var(--font-mono);
  font-size:.68rem;
  font-variant-numeric:tabular-nums;
  table-layout:auto;
}
.tenjou-table thead{
  position:sticky;top:0;z-index:4;
  background:var(--bg-elev);
}
.tenjou-table thead::after{
  content:'';position:absolute;left:0;right:0;bottom:0;
  height:1px;background:var(--border-bright);
}
.tenjou-table th{
  text-align:right;
  padding:.4rem .2rem;
  font-size:.52rem;font-weight:700;
  color:var(--text-2);
  letter-spacing:.04em;text-transform:uppercase;
  white-space:nowrap;
  background:var(--bg-elev);
}
.tenjou-table th:first-child{
  text-align:left;padding-left:.45rem;
  position:sticky;left:0;z-index:5;
  background:var(--bg-elev);
  border-right:1px solid var(--border);
}
.tenjou-table th:last-child{padding-right:.45rem;}
.tenjou-table td{
  padding:.3rem .2rem;
  text-align:right;
  border-bottom:1px solid var(--border);
  white-space:nowrap;
  line-height:1.15;
}
.tenjou-table td:first-child{
  text-align:left;padding-left:.45rem;
  font-family:var(--font-mono);font-weight:700;
  font-size:.72rem;letter-spacing:.01em;
  color:var(--text);
  position:sticky;left:0;z-index:3;
  background:var(--bg-surface);
  border-right:1px solid var(--border);
}
.tenjou-table td:last-child{padding-right:.45rem;}
.tenjou-table tbody tr{transition:background .08s;}
.tenjou-table tbody tr:nth-child(even) td{background:rgba(255,255,255,.014);}
.tenjou-table tbody tr:nth-child(even) td:first-child{background:var(--bg-surface);
  box-shadow:inset 0 0 0 99px rgba(255,255,255,.014);}
.tenjou-table tbody tr:hover td{background:rgba(245,184,0,.06);}
.tenjou-table tbody tr:hover td:first-child{box-shadow:inset 0 0 0 99px rgba(245,184,0,.06);}
.tenjou-table .num{color:var(--text);}
.tenjou-table .dim{color:var(--text-dim);font-size:.62rem;}
/* ユーザー仕様: プラス&機械割≥100%→青、マイナス|機械割<100%→赤
   .num より後に declare して上書き勝ち */
.tenjou-table .pos{color:#4fb3ff;font-weight:600;}
.tenjou-table .neg{color:var(--red);font-weight:600;}
html[data-theme="light"] .tenjou-table .pos{color:#1e6bcf;}
html[data-theme="light"] .tenjou-table .neg{color:#c41f63;}
/* デスクトップで広めに */
@media(min-width:600px){
  .tenjou-table{font-size:.78rem;}
  .tenjou-table th{font-size:.6rem;padding:.55rem .4rem;}
  .tenjou-table td{padding:.42rem .4rem;}
  .tenjou-table td:first-child{font-size:.85rem;}
}

.list-empty{padding:2rem 1rem;text-align:center;color:var(--text-dim);font-size:.85rem;}
.list-count{padding:.55rem .9rem;font-family:var(--font-mono);font-size:.65rem;
  color:var(--text-dim);letter-spacing:.1em;border-bottom:1px solid var(--border);
  display:flex;justify-content:space-between;}

/* Detail overlay */
.overlay{
  position:fixed;inset:0;z-index:50;
  background:rgba(10,13,18,.84);
  backdrop-filter:blur(10px);-webkit-backdrop-filter:blur(10px);
  display:none;align-items:flex-end;
  animation:fadeIn .2s ease;
}
html[data-theme="light"] .overlay{background:rgba(40,40,50,.55);}
.overlay.open{display:flex;}
@keyframes fadeIn{from{opacity:0;}to{opacity:1;}}
.sheet-panel{
  width:100%;max-height:85vh;overflow-y:auto;
  background:var(--bg-surface);
  border-top:2px solid var(--lime);
  padding:1.25rem 1rem 2rem;padding-bottom:max(2rem,env(safe-area-inset-bottom));
  animation:slideUp .25s cubic-bezier(.2,.8,.3,1);
}
@keyframes slideUp{from{transform:translateY(100%);}to{transform:translateY(0);}}
.sheet-panel .panel-head{
  display:flex;justify-content:space-between;align-items:flex-start;
  margin-bottom:1rem;
}
.sheet-panel .panel-g{
  font-family:var(--font-display);font-size:2.4rem;line-height:1;color:var(--lime);
}
.sheet-panel .panel-close{
  font-family:var(--font-mono);font-size:.75rem;color:var(--text-2);
  border:1px solid var(--border-bright);padding:.45rem .7rem;
}
.detail-grid{display:grid;grid-template-columns:1fr 1fr;gap:.75rem;}
.detail-item{border-left:2px solid var(--border-bright);padding-left:.7rem;}
.detail-item.full{grid-column:span 2;}
.detail-item .di-lbl{
  display:block;font-size:.58rem;color:var(--text-dim);
  letter-spacing:.18em;text-transform:uppercase;margin-bottom:.2rem;
}
.detail-item .di-val{
  display:block;font-family:var(--font-mono);font-size:1rem;font-weight:600;
}
.detail-item.hot .di-val{color:var(--lime);}

/* Simple table */
.simple-card{
  background:var(--bg-surface);
  border:1px solid var(--border);
  border-radius:4px;
  overflow:auto;
}
.simple-tbl{width:100%;border-collapse:collapse;font-family:var(--font-mono);font-size:.7rem;
  font-variant-numeric:tabular-nums;}
.simple-tbl th{
  text-align:left;padding:.45rem .55rem;
  background:var(--bg-elev);
  color:var(--text-2);
  font-size:.55rem;letter-spacing:.1em;text-transform:uppercase;font-weight:700;
  border-bottom:1px solid var(--border-bright);
  white-space:nowrap;
}
.simple-tbl td{
  padding:.45rem .55rem;border-bottom:1px solid var(--border);
  white-space:nowrap;
}
.simple-tbl tr:last-child td{border-bottom:none;}
.simple-tbl tr:nth-child(even) td{background:rgba(255,255,255,.012);}
.simple-tbl .num{text-align:right;}
.simple-tbl .pct{color:var(--cyan);}
@media(min-width:600px){
  .simple-tbl{font-size:.78rem;}
  .simple-tbl th{font-size:.62rem;padding:.55rem .75rem;}
  .simple-tbl td{padding:.55rem .75rem;}
}

@media (min-width:700px){
  .view{max-width:720px;margin:0 auto;}
  .machine-card{max-width:720px;margin:.8rem auto;}
  .quick-stats{grid-template-columns:1.3fr 1fr 1fr 1fr;}
  .detail-grid{grid-template-columns:repeat(3,1fr);}
  .detail-item.full{grid-column:span 3;}
}
/* ライトモードでの暗色 → 明色 inset 影は不要 */
html[data-theme="light"] .tenjou-table tbody tr:nth-child(even) td{
  background:rgba(0,0,0,.025);
}
html[data-theme="light"] .tenjou-table tbody tr:nth-child(even) td:first-child{
  box-shadow:inset 0 0 0 99px rgba(0,0,0,.025);
}
html[data-theme="light"] .tenjou-table tbody tr:hover td{
  background:rgba(196,143,0,.10);
}
html[data-theme="light"] .tenjou-table tbody tr:hover td:first-child{
  box-shadow:inset 0 0 0 99px rgba(196,143,0,.10);
}
html[data-theme="light"] .simple-tbl tr:nth-child(even) td{
  background:rgba(0,0,0,.025);
}
</style>
</head>
<body>

<header class="status-bar">
  <div class="file-label" id="fileLabel">—</div>
</header>

<section class="machine-card">
  <h1 class="machine-name" id="machineName">—<span class="type" id="machineType"></span></h1>
  <div class="cond-grid" id="condGrid"></div>
</section>

<nav class="sheet-tabs" id="sheetTabs"></nav>

<div id="globalFilterBanner" class="filter-banner" style="display:none;"></div>

<main class="view" id="view"></main>

<div class="overlay" id="overlay" onclick="if(event.target===this)closeOverlay()">
  <div class="sheet-panel" id="panel"></div>
</div>

<script>
const FILES = __DATA__;
const SESSION_GROUP = __SESSION_GROUP__;
const STATE = {fi:0, sheet:null, filter:'all', step:50, session:null, sessionMode:'morning', cycleMode:null, atState:'all', atPrevYuri:'none'};
let chartInst = null;

const $ = id => document.getElementById(id);
const fmt = {
  int: n => n==null||isNaN(n) ? '—' : Math.round(n).toLocaleString('ja-JP'),
  signed: n => {
    if(n==null||isNaN(n)) return '—';
    const v = Math.round(n).toLocaleString('ja-JP');
    return n>0 ? '+'+v : v;
  },
  yen: n => {
    if(n==null||isNaN(n)) return '¥—';
    const v = Math.round(n).toLocaleString('ja-JP');
    return n>0 ? '¥+'+v : (n<0?'¥'+v:'¥0');
  },
  pct: n => n==null||isNaN(n) ? '—' : (n*100).toFixed(2)+'%',
  pctRaw: n => n==null||isNaN(n) ? '—' : n.toFixed(2)+'%',
  ty: n => n==null||isNaN(n) ? '—' : Math.round(n)+'枚',
};

function init(){
  // 最新ファイル固定表示
  const f = FILES[0];
  const t = new Date(f.mtime*1000);
  const ts = t.getMonth()+1+'/'+t.getDate()+' '+String(t.getHours()).padStart(2,'0')+':'+String(t.getMinutes()).padStart(2,'0');
  $('fileLabel').textContent = `${f.machine}  ${ts}`;
  render();
}

function currentFile(){return FILES[STATE.fi];}

function render(){
  const f = currentFile();
  renderMachine(f);
  renderTabs(f);
  renderView(f);
}

function renderMachine(f){
  $('machineName').firstChild.textContent = f.machine;
  $('machineType').textContent = f.spec?.type || '';
  const k = f.kpi || {}, sp = f.spec || {};
  const items = [];
  // スペック (連荘数・掛けは省略、解析仕様に直結する項目のみ)
  if(sp.coin_hold) items.push(['コインもち', sp.coin_hold, false]);
  if(sp.invest_per_g) items.push(['純減', sp.invest_per_g.toFixed(3)+'枚/G', false]);
  if(sp.at_payout_per_g) items.push(['AT純増', sp.at_payout_per_g+'枚/G', false]);
  if(sp.yen) items.push(['換金', '等価 '+sp.yen+'円/枚', false]);
  if(sp.ceiling) items.push(['天井', sp.ceiling, false]);
  if(k['総ゲーム数']!=null) items.push(['総G', fmt.int(k['総ゲーム数']), false]);

  // 全シート共通の算出条件 (タブごとに変わる N回目絞り等は除外、共通分のみ)
  // 各シートの filter_description から共通フレーズを抽出
  const tenjou = f.tenjou || {};
  const allFilters = Object.values(tenjou).map(t => t.filter_description||'').filter(Boolean);
  if (allFilters.length) {
    const splitOnAnd = s => s.split(/\s*かつ\s*/).map(x=>x.trim()).filter(Boolean);
    const sets = allFilters.map(splitOnAnd);
    // 全シートに登場する条件のみ採用
    const common = sets[0].filter(c => sets.every(s => s.includes(c)));
    common.forEach(c => items.push(['算出条件', c, 'accent']));
  }

  const grid = $('condGrid');
  grid.innerHTML = items.map(([k_,v,acc])=>
    `<div class="cond-row"><span class="k">${k_}</span><span class="v ${acc?'accent':''}">${v}</span></div>`
  ).join('');
}

function availableSheets(f){
  const list = [];
  const tenjou = f.tenjou || {};
  const sessionKeys = SESSION_GROUP.modes.flatMap(m => m.options.map(o=>o[1]));
  const sessionAvailable = sessionKeys.some(k => k in tenjou);
  // AT間天井狙い タブ内で toggle 表示するので独立タブとしては隠す
  const hiddenInTabs = ['AT間天井狙い_直撃後'];
  // 直前有利区間G別 (prev_yuri_*) も AT間天井狙いタブ内で表示するので独立タブ非表示
  const prevYuriKeys = ['prev_yuri_0_999','prev_yuri_1000_1999','prev_yuri_2000_2999',
                        'prev_yuri_3000_3999','prev_yuri_4000_4999','prev_yuri_5000_5999',
                        'prev_yuri_6000_6999','prev_yuri_7000_plus'];
  Object.keys(tenjou).forEach(n=>{
    if(sessionKeys.includes(n)) return;
    if(hiddenInTabs.includes(n)) return;
    if(prevYuriKeys.includes(n)) return;
    list.push({key:n,label:n,kind:'tenjou'});
    if(n==='AT間天井狙い' && sessionAvailable){
      list.push({key:'__session_group__', label:SESSION_GROUP.label, kind:'session'});
    }
  });
  // フォールバック: AT間天井狙いが無いがsessionあるなら末尾追加
  if(sessionAvailable && !list.some(s=>s.key==='__session_group__')){
    list.push({key:'__session_group__', label:SESSION_GROUP.label, kind:'session'});
  }
  if(f.cycle_g_at && (f.cycle_g_at.morning || f.cycle_g_at.cut)){
    list.push({key:'__cycle_g_at__', label:'累積G別AT間', kind:'cycle_g_at'});
  }
  if(f.chokugeki && f.chokugeki.blocks && f.chokugeki.blocks.length){
    list.push({key:'__chokugeki__', label:'直撃分布', kind:'chokugeki'});
  }
  if(f.hikimodoshi && f.hikimodoshi.blocks && f.hikimodoshi.blocks.length){
    list.push({key:'__hikimodoshi__', label:'引き戻し分布', kind:'hikimodoshi'});
  }
  Object.keys(f.simple||{}).forEach(n=>list.push({key:n,label:n,kind:'simple'}));
  return list;
}

function renderTabs(f){
  const sheets = availableSheets(f);
  if(!STATE.sheet || !sheets.find(s=>s.key===STATE.sheet)){
    STATE.sheet = sheets.length ? sheets[0].key : null;
  }
  const nav = $('sheetTabs');
  nav.innerHTML = '';
  sheets.forEach(s=>{
    const b = document.createElement('button');
    b.className = 'tab'+(s.key===STATE.sheet?' active':'');
    b.innerHTML = (s.star?'<span class="star">★</span>':'')+s.label;
    b.onclick = ()=>{STATE.sheet=s.key;render();};
    nav.appendChild(b);
  });
}

function setGlobalFilterBanner(text){
  const el = document.getElementById('globalFilterBanner');
  if(!el) return;
  if(text){
    el.innerHTML = `<span class="filter-banner-label">算出条件</span><span class="filter-banner-text">${text}</span>`;
    el.style.display = '';
  } else {
    el.style.display = 'none';
    el.innerHTML = '';
  }
}

function renderView(f){
  const v = $('view');
  v.innerHTML = '';
  setGlobalFilterBanner('');  // デフォルトクリア、各シートが必要なら設定
  const s = STATE.sheet;
  if(!s){v.innerHTML='<div class="list-empty">データなし</div>';return;}

  if(s==='__session_group__') return renderSessionGroup(v, f);
  if(s==='__cycle_g_at__') return renderCycleGroup(v, f);
  if(s==='__chokugeki__') return renderChokugeki(v, f);
  if(s==='__hikimodoshi__') return renderHikimodoshi(v, f);
  if(s==='AT間天井狙い') return renderAtTenjouGroup(v, f);
  if(f.tenjou[s]) return renderTenjou(v, f.tenjou[s]);
  if(f.simple[s]) return renderSimple(v, f.simple[s]);
  v.innerHTML='<div class="list-empty">未対応シート</div>';
}

function renderAtTenjouGroup(v, f){
  const tenjou = f.tenjou || {};
  const allKey = 'AT間天井狙い';
  const chokuKey = 'AT間天井狙い_直撃後';
  const hasChoku = chokuKey in tenjou;
  // 直前有利区間G別 (8 ボタン) のキー定義
  const prevYuriBtns = [
    ['0-999G','0_999','prev_yuri_0_999'],
    ['1000-1999G','1000_1999','prev_yuri_1000_1999'],
    ['2000-2999G','2000_2999','prev_yuri_2000_2999'],
    ['3000-3999G','3000_3999','prev_yuri_3000_3999'],
    ['4000-4999G','4000_4999','prev_yuri_4000_4999'],
    ['5000-5999G','5000_5999','prev_yuri_5000_5999'],
    ['6000-6999G','6000_6999','prev_yuri_6000_6999'],
    ['7000G以上','7000_plus','prev_yuri_7000_plus'],
  ];
  const hasPrevYuri = prevYuriBtns.some(([,,sk]) => sk in tenjou);

  // 主トグル: [全状態平均 | 直撃後 | 前回有利区間G数別]
  // STATE.atState in {'all','choku','prev_yuri'}
  const validStates = ['all'];
  if(hasChoku) validStates.push('choku');
  if(hasPrevYuri) validStates.push('prev_yuri');
  if(!validStates.includes(STATE.atState)) STATE.atState = 'all';
  // prev_yuri モードの初期サブ選択
  if(STATE.atState === 'prev_yuri'){
    const cur = prevYuriBtns.find(([,k])=>k===STATE.atPrevYuri);
    if(!cur || !(cur[2] in tenjou)){
      const first = prevYuriBtns.find(([,,sk])=>sk in tenjou);
      STATE.atPrevYuri = first ? first[1] : '0_999';
    }
  }

  const main = document.createElement('div');
  main.className = 'gbin-toggle';
  main.style.margin = '0 0 .4rem';
  const mainBtns = [['全状態平均','all']];
  if(hasChoku) mainBtns.push(['直撃後','choku']);
  if(hasPrevYuri) mainBtns.push(['前回有利区間G数別','prev_yuri']);
  mainBtns.forEach(([lbl,key])=>{
    const b = document.createElement('button');
    b.textContent = lbl;
    b.className = (STATE.atState === key) ? 'active' : '';
    b.onclick = ()=>{
      STATE.atState = key;
      renderView(currentFile());
    };
    main.appendChild(b);
  });
  v.appendChild(main);

  // サブトグル (前回有利区間G数別 選択時のみ) — コンパクトな小ボタン
  if(STATE.atState === 'prev_yuri' && hasPrevYuri){
    const sub = document.createElement('div');
    sub.className = 'gbin-toggle';
    prevYuriBtns.forEach(([lbl,key,sk])=>{
      if(!(sk in tenjou)) return;
      const b = document.createElement('button');
      b.textContent = lbl;
      b.className = (STATE.atPrevYuri === key) ? 'active' : '';
      b.onclick = ()=>{
        STATE.atPrevYuri = key;
        renderView(currentFile());
      };
      sub.appendChild(b);
    });
    v.appendChild(sub);

    // 注記: 前回まで上位/直撃履歴がない台に限定
    const note = document.createElement('div');
    note.style.cssText = 'font-size:.72rem;color:var(--lime);opacity:.75;margin:.1rem 0 .55rem;letter-spacing:.02em;';
    note.textContent = '※ 前回まで上位履歴・直撃履歴がない台';
    v.appendChild(note);
  }

  // 表示対象キー決定
  let targetKey;
  if(STATE.atState === 'choku') targetKey = chokuKey;
  else if(STATE.atState === 'prev_yuri'){
    const sel = prevYuriBtns.find(([,k])=>k===STATE.atPrevYuri);
    targetKey = sel ? sel[2] : allKey;
  } else targetKey = allKey;
  if(tenjou[targetKey]) renderTenjou(v, tenjou[targetKey]);
  else if(tenjou[allKey]) renderTenjou(v, tenjou[allKey]);
}

function renderHikimodoshi(v, f){
  const data = f.hikimodoshi;
  if(!data || !data.blocks || !data.blocks.length){
    v.innerHTML='<div class="list-empty">引き戻し分布データなし</div>';
    return;
  }
  // 上位AT区間ブロックは非表示 (下位AT区間のみ)、ブロックタイトルラベルも省略
  const hblocks = data.blocks.filter(b => !(b.title && b.title.indexOf('上位AT区間') >= 0));
  hblocks.forEach((block, bi)=>{
    const wrap = document.createElement('div');
    wrap.className='table-wrap';
    const tbl = document.createElement('table');
    tbl.className='tenjou-table';
    tbl.innerHTML = `<thead><tr>
        <th>${block.group_label||'区分'}</th><th>サンプル</th><th>上位</th><th>上位率(%)</th><th>3-90内当選</th><th>3-90率(%)</th>
      </tr></thead><tbody></tbody>`;
    const tb = tbl.querySelector('tbody');
    block.rows.forEach(r=>{
      const isTotal = r.label === '合計';
      const tr = document.createElement('tr');
      if (isTotal) tr.style.fontWeight = '700';
      tr.innerHTML = `
        <td>${r.label||''}</td>
        <td class="num">${fmt.int(r.sample)}</td>
        <td class="num">${fmt.int(r.jyoi)}</td>
        <td class="num pos">${r.rate_jyoi!=null?Number(r.rate_jyoi).toFixed(2):'-'}</td>
        <td class="num">${fmt.int(r.n390)}</td>
        <td class="num pos">${r.rate_n390!=null?Number(r.rate_n390).toFixed(2):'-'}</td>`;
      tb.appendChild(tr);
    });
    wrap.appendChild(tbl);
    v.appendChild(wrap);
  });
}

function renderChokugeki(v, f){
  const data = f.chokugeki;
  if(!data || !data.blocks || !data.blocks.length){
    v.innerHTML='<div class="list-empty">直撃分布データなし</div>';
    return;
  }

  data.blocks.forEach((block, bi)=>{
    const title = document.createElement('div');
    title.style.cssText='font-family:var(--font-display);font-weight:700;font-size:.78rem;color:var(--lime);margin:'+(bi===0?'.2rem':'1rem')+' 0 .35rem;letter-spacing:.02em;';
    title.textContent = block.title;
    v.appendChild(title);

    const wrap = document.createElement('div');
    wrap.className='table-wrap';
    const tbl = document.createElement('table');
    tbl.className='tenjou-table';
    tbl.innerHTML = `<thead><tr>
        <th>${block.group_label||'区分'}</th><th>サンプル</th><th>直撃件数</th><th>直撃率(%)</th><th>直撃時TY(枚)</th>
      </tr></thead><tbody></tbody>`;
    const tb = tbl.querySelector('tbody');
    block.rows.forEach(r=>{
      const isTotal = r.label === '合計';
      const tr = document.createElement('tr');
      if (isTotal) tr.style.fontWeight = '700';
      tr.innerHTML = `
        <td>${r.label||''}</td>
        <td class="num">${fmt.int(r.sample)}</td>
        <td class="num">${fmt.int(r.choku)}</td>
        <td class="num pos">${r.rate!=null?Number(r.rate).toFixed(2):'-'}</td>
        <td class="num">${r.ty!=null?fmt.int(r.ty):'-'}</td>`;
      tb.appendChild(tr);
    });
    wrap.appendChild(tbl);
    v.appendChild(wrap);
  });
}

function renderCycleGroup(v, f){
  const cg = f.cycle_g_at || {};
  if(!STATE.cycleMode) STATE.cycleMode = cg.morning ? 'morning' : 'cut';

  // morning/cut トグル
  const toggleWrap = document.createElement('div');
  toggleWrap.className = 'gbin-toggle';
  [['morning','朝'], ['cut','切断後']].forEach(([key, label])=>{
    if(!cg[key]) return;
    const b = document.createElement('button');
    b.textContent = label;
    b.className = (STATE.cycleMode === key) ? 'active' : '';
    b.onclick = ()=>{STATE.cycleMode = key; renderView(currentFile());};
    toggleWrap.appendChild(b);
  });
  v.appendChild(toggleWrap);

  const data = cg[STATE.cycleMode];
  if(!data){v.appendChild(Object.assign(document.createElement('div'), {className:'list-empty', textContent:'データなし'}));return;}

  // 算出条件バナー (タブ群直下の globalFilterBanner に出す)
  setGlobalFilterBanner(data.filter_description || '');

  function _renderBlock(title, rows){
    const t = document.createElement('div');
    t.style.cssText = 'margin:.8rem 1rem .3rem;font-family:var(--font-display);letter-spacing:.06em;color:var(--text-2);font-size:.9rem;';
    t.textContent = title;
    v.appendChild(t);
    const wrap = document.createElement('div');
    wrap.className = 'table-wrap';
    if(!rows || rows.length === 0){
      wrap.appendChild(Object.assign(document.createElement('div'), {className:'list-empty', textContent:'該当なし'}));
      v.appendChild(wrap); return;
    }
    const tbl = document.createElement('table');
    tbl.className = 'tenjou-table';
    tbl.innerHTML = `<thead><tr>
        <th>累積G</th><th>サンプル</th><th>TY(枚)</th><th>AT初当り(G)</th>
        <th>期待値(円)</th><th>機械割(%)</th>
      </tr></thead><tbody></tbody>`;
    const tb = tbl.querySelector('tbody');
    rows.forEach(r=>{
      const ev = r['期待値(円)'] ?? r['EV(円)'];
      const ki = r['機械割(%)'];
      const good = ev>0 && ki>=100;
      const bad  = ev<0 || (ki!=null && ki<100);
      const cls = good ? 'pos' : (bad ? 'neg' : 'dim');
      const evCls = cls, kiCls = cls;
      const tr = document.createElement('tr');
      tr.innerHTML = `
        <td>${r['累積G']||''}</td>
        <td class="num">${fmt.int(r['サンプル数'])}</td>
        <td class="num">${fmt.int(r['平均連チャンTY(枚)'])}</td>
        <td class="num">${fmt.int(r['平均消化G'])}</td>
        <td class="num ${evCls}">${ev!=null?fmt.int(ev):'-'}</td>
        <td class="num ${kiCls}">${ki!=null?ki.toFixed(2):'-'}</td>`;
      tb.appendChild(tr);
    });
    wrap.appendChild(tbl);
    v.appendChild(wrap);
  }
  _renderBlock('【前回通常累積G数別】', data.normal_rows);
  _renderBlock('【直前有利区間G数別】', data.total_rows);
}

function renderSessionGroup(v, f){
  if(!STATE.sessionMode) STATE.sessionMode = SESSION_GROUP.modes[0].key;
  const tenjou = f.tenjou || {};

  // 現モードの options
  const mode = SESSION_GROUP.modes.find(m => m.key === STATE.sessionMode) || SESSION_GROUP.modes[0];
  const opts = mode.options.filter(([_, k]) => k in tenjou);

  // 1行に [朝 | 切断後] トグル + 回数プルダウン を並べる
  const row = document.createElement('div');
  row.className = 'session-row';

  const toggleWrap = document.createElement('div');
  toggleWrap.className = 'gbin-toggle';
  SESSION_GROUP.modes.forEach(m=>{
    const has = m.options.some(([_, k]) => k in tenjou);
    if(!has) return;
    const b = document.createElement('button');
    b.textContent = m.label;
    b.className = (STATE.sessionMode === m.key) ? 'active' : '';
    b.onclick = ()=>{
      STATE.sessionMode = m.key;
      STATE.session = null;  // 再選択させる
      renderView(currentFile());
    };
    toggleWrap.appendChild(b);
  });
  row.appendChild(toggleWrap);

  // プルダウン (回数選択)
  const sel = document.createElement('select');
  sel.className='session-pick';
  opts.forEach(([lbl, key])=>{
    const opt = document.createElement('option');
    opt.value = key;
    opt.textContent = lbl;
    if(key===STATE.session) opt.selected = true;
    sel.appendChild(opt);
  });
  sel.onchange = ()=>{STATE.session = sel.value; renderView(currentFile());};
  row.appendChild(sel);

  v.appendChild(row);

  // 朝モードに「切断/直撃以降除外」ラベルを表記 (一括)
  if(STATE.sessionMode === 'morning'){
    const note = document.createElement('div');
    note.style.cssText = 'font-size:.72rem;color:var(--lime);opacity:.75;margin:.1rem 0 .55rem;letter-spacing:.02em;';
    note.textContent = '※ 切断/直撃以降の連荘は除外';
    v.appendChild(note);
  }

  // 選択された session の天井狙いテーブルを描画
  let cur = STATE.session;
  if(!cur || !(cur in tenjou) || !opts.find(([_, k]) => k === cur)){
    cur = opts.length ? opts[0][1] : null;
    STATE.session = cur;
  }
  if(cur && tenjou[cur]){
    const savedSheet = STATE.sheet;
    STATE.sheet = cur;
    renderTenjou(v, tenjou[cur]);
    STATE.sheet = savedSheet;
  }
}

function renderTenjou(v, t){
  const rows = t.rows;
  const evCol = t.ev_yen_col;
  const evCoinCol = t.ev_coin_col;
  const kikaiCol = t.kikai_col;

  // 算出条件バナー (タブ群直下の globalFilterBanner に統一)
  setGlobalFilterBanner(t.filter_description || '');

  // (グラフは廃止)

  // Step toggle (5G/50G)
  const step = document.createElement('div');
  step.className='step-toggle';
  [5,50].forEach(s=>{
    const b=document.createElement('button');
    b.textContent=s+'G毎';
    b.className=STATE.step===s?'active':'';
    b.onclick=()=>{STATE.step=s;renderView(currentFile());};
    step.appendChild(b);
  });
  v.appendChild(step);

  // ステップ (50G時はG % 50 == 0 の行だけ)
  // AT間天井狙いは 1450G まで (天井近くはサンプル少なく意味薄い)
  // 朝1回目 / 切断後1回目 はサンプル枯渇が早いので 950G まで
  let capG = 1499;
  if (STATE.sheet === 'AT間天井狙い') capG = 1450;
  if (STATE.sheet === 'AT間_morning_1' || STATE.sheet === 'AT間_cut_1') capG = 950;
  const filtered = (STATE.step===50 ? rows.filter(r=>r._g%50===0) : rows)
                     .filter(r => r._g <= capG);

  // Table (slolaboratory風 高密度)
  const wrap = document.createElement('div');
  wrap.className='table-wrap';

  if(filtered.length===0){
    const e = document.createElement('div');
    e.className='list-empty';
    e.textContent='該当行なし';
    wrap.appendChild(e);
    v.appendChild(wrap);
    return;
  }

  // シート種別で「初当り」ラベル切替
  const initLabel = STATE.sheet.includes('AT') ? 'AT初当り' : 'CZ初当り';

  const tbl = document.createElement('table');
  tbl.className='tenjou-table';
  tbl.innerHTML = `<thead><tr>
      <th>打ち始めG</th><th>${initLabel}(G)</th><th>TY(枚)</th>
      <th>期待値(円)</th><th>出玉率(%)</th><th>サンプル数</th>
    </tr></thead><tbody></tbody>`;
  const tb = tbl.querySelector('tbody');
  filtered.forEach(r=>{
    const ev = r[evCol];
    const ki = r[kikaiCol];
    // ユーザー仕様: プラス&機械割≥100→青 (pos), マイナス|機械割<100→赤 (neg)
    const good = ev>0 && ki>=100;
    const bad  = ev<0 || (ki!=null && ki<100);
    const cls = good ? 'pos' : (bad ? 'neg' : 'dim');
    const evCls = cls, kiCls = cls;
    const tr = document.createElement('tr');
    // データ行はヘッダ側に単位があるので素の数値のみ
    tr.innerHTML = `
      <td>${r._g}</td>
      <td class="num">${fmt.int(r['Nから平均消化G'])}</td>
      <td class="num">${fmt.int(r['累積連チャンTY(枚)'])}</td>
      <td class="${evCls}">${ev!=null?fmt.signed(ev):'-'}</td>
      <td class="${kiCls}">${ki!=null?ki.toFixed(2):'-'}</td>
      <td class="num dim">${fmt.int(r['サンプル数'])}</td>`;
    tr.onclick = ()=>openDetail(r, t);
    tb.appendChild(tr);
  });
  wrap.appendChild(tbl);
  v.appendChild(wrap);
}

function drawChart(rows, evCol, kikaiCol){
  if(chartInst){chartInst.destroy();chartInst=null;}
  const ctx = document.getElementById('evChart');
  if(!ctx) return;
  const labels = rows.map(r=>r._g);
  const evData = rows.map(r=>r[evCol]);
  const kiData = rows.map(r=>r[kikaiCol]);
  chartInst = new Chart(ctx,{
    type:'line',
    data:{labels,
      datasets:[
        {label:'期待値(円)',data:evData,borderColor:'#4fb3ff',backgroundColor:'rgba(79,179,255,.08)',
          borderWidth:1.5,pointRadius:0,tension:.2,fill:true,yAxisID:'y'},
        {label:'機械割%',data:kiData,borderColor:'#ffb52e',borderWidth:1.2,pointRadius:0,
          tension:.2,yAxisID:'y2'},
      ]},
    options:{
      responsive:true,maintainAspectRatio:false,
      interaction:{intersect:false,mode:'index'},
      plugins:{
        legend:{labels:{color:'#8a96a8',font:{family:'JetBrains Mono',size:9},boxWidth:12,boxHeight:2},
          position:'bottom'},
        tooltip:{backgroundColor:'#10141b',borderColor:'#2a3340',borderWidth:1,
          titleFont:{family:'JetBrains Mono'},bodyFont:{family:'JetBrains Mono'},titleColor:'#84ff5a'},
      },
      scales:{
        x:{ticks:{color:'#4a5260',font:{family:'JetBrains Mono',size:9},maxTicksLimit:8},
           grid:{color:'rgba(255,255,255,.03)'}},
        y:{ticks:{color:'#84ff5a',font:{family:'JetBrains Mono',size:9},
            callback:v=>(v>0?'+':'')+(v/1000).toFixed(0)+'k'},
           grid:{color:'rgba(255,255,255,.03)'}},
        y2:{position:'right',ticks:{color:'#ffb52e',font:{family:'JetBrains Mono',size:9},
            callback:v=>v.toFixed(0)+'%'},grid:{display:false}},
      },
      animation:{duration:400},
    }
  });
}

function openDetail(r, t){
  const panel = $('panel');
  const items = [];
  const evCol = t.ev_yen_col, kikaiCol = t.kikai_col;
  items.push(['打ち始め', r._g+'G', 'full']);
  items.push(['累積期待値(円)', fmt.yen(r[evCol]), r[evCol]>0?'hot':null]);
  items.push(['機械割', fmt.pctRaw(r[kikaiCol]), r[kikaiCol]>=100?'hot':null]);
  items.push(['サンプル数', fmt.int(r['サンプル数']), null]);
  items.push(['当選数', fmt.int(r['当選数']), null]);
  items.push(['区間当選率', fmt.pct(r['区間当選率']), null]);
  items.push(['Nから平均消化G', fmt.int(r['Nから平均消化G'])+'G', null]);
  items.push(['累積連チャンTY', fmt.ty(r['累積連チャンTY(枚)']), null]);
  items.push(['区間連チャンTY', fmt.ty(r['区間連チャンTY(枚)']), null]);
  items.push(['区間期待値(円)', fmt.yen(r['区間期待値(円)'] ?? r['区間EV(円)']), null]);
  items.push(['区間期待値(枚)', fmt.signed(r['区間期待値(枚)'] ?? r['区間EV(枚)']), null]);
  items.push(['累積期待値(枚)', fmt.signed(r[t.ev_coin_col]), null]);
  panel.innerHTML = `
    <div class="panel-head">
      <div class="panel-g">${r._g}G</div>
      <button class="panel-close" onclick="closeOverlay()">閉じる ✕</button>
    </div>
    <div class="detail-grid">
      ${items.map(([l,v,cls])=>`<div class="detail-item ${cls||''} ${cls==='full'?'full':''}"><span class="di-lbl">${l}</span><span class="di-val">${v}</span></div>`).join('')}
    </div>`;
  $('overlay').classList.add('open');
}
function closeOverlay(){$('overlay').classList.remove('open');}

function renderSimple(v, s){
  const card = document.createElement('div');
  card.className='simple-card';
  card.style.overflowX='auto';
  card.innerHTML = renderTable(s.headers, s.rows);
  v.appendChild(card);
}
function renderTable(headers, rows){
  const ths = headers.map(h=>`<th>${h}</th>`).join('');
  const trs = rows.map(r=>{
    const tds = headers.map(h=>{
      const v = r[h];
      if(v==null) return '<td>—</td>';
      if(typeof v==='number'){
        const isPct = (typeof h==='string') && /率|割/.test(h) && v<=1 && v>=-1;
        if(isPct) return `<td class="num pct">${(v*100).toFixed(2)}%</td>`;
        if(Number.isInteger(v)) return `<td class="num">${v.toLocaleString('ja-JP')}</td>`;
        return `<td class="num">${v.toFixed(2)}</td>`;
      }
      return `<td>${v}</td>`;
    }).join('');
    return `<tr>${tds}</tr>`;
  }).join('');
  return `<table class="simple-tbl"><thead><tr>${ths}</tr></thead><tbody>${trs}</tbody></table>`;
}

init();

// ===== iframe 埋め込み対応: 親WPテーマ追随 + 高さ自動通知 =====
(function(){
  // 1) URL ?theme=light|dark → <html data-theme="..."> 反映
  function applyTheme(t){
    if (t === 'light' || t === 'dark') {
      document.documentElement.setAttribute('data-theme', t);
    }
  }
  try {
    var qs = new URLSearchParams(location.search);
    applyTheme(qs.get('theme'));
  } catch(e) {}

  // 2) 親から theme 切替メッセージ受信
  window.addEventListener('message', function(e){
    var d = e.data;
    if (d && d.type === 'kdashboard-set-theme') applyTheme(d.theme);
  });

  // 3) 自身の高さを親へ送信 (iframe 高さ自動調整用)
  //    揺れ防止: 12px 未満の差は無視 + 250ms デバウンス + 50ms 連続発火を抑制
  var lastH = 0, pending = null, lastSent = 0;
  function measure(){
    // iframe 内: documentElement.scrollHeight は iframe 高さに引きずられて
    // 古い値を返すことがあるため body.scrollHeight のみ採用
    return document.body.scrollHeight;
  }
  function sendHeight(){
    var h = measure();
    if (Math.abs(h - lastH) < 12) return;  // 微差は無視
    var now = Date.now();
    if (now - lastSent < 50) return;       // 連続抑制
    lastH = h; lastSent = now;
    parent.postMessage({type:'kdashboard-height', height: h}, '*');
  }
  function schedule(){
    if (pending) clearTimeout(pending);
    pending = setTimeout(function(){ pending = null; sendHeight(); }, 250);
  }
  if (window.parent !== window) {
    window.addEventListener('load', function(){
      // 初回はデバウンスせず即送信 (1度だけ)
      lastH = measure();
      parent.postMessage({type:'kdashboard-height', height: lastH}, '*');
    });
    if (window.ResizeObserver) {
      new ResizeObserver(schedule).observe(document.body);
    } else {
      setInterval(schedule, 1000);
    }
    // タブ切替直後は確実に高さが変わるので即送信
    document.addEventListener('click', function(){
      setTimeout(function(){ lastH = 0; sendHeight(); }, 100);
    });
  }
})();
</script>
</body>
</html>
"""


def write_dashboard(files: list[dict], out_path: str) -> None:
    data_json = json.dumps(files, ensure_ascii=False, default=str)
    group_json = json.dumps(SESSION_GROUP, ensure_ascii=False)
    html = HTML_TEMPLATE.replace("__DATA__", data_json).replace("__SESSION_GROUP__", group_json)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--output-dir", action="append", default=None,
                   help="xlsx 探索ディレクトリ (複数指定可)")
    p.add_argument("--out", default=None, help="出力 HTML パス (デフォルト: viewer/static/index.html)")
    p.add_argument("--serve", action="store_true", help="生成後 http.server を起動 (port 8501)")
    p.add_argument("--port", type=int, default=8501)
    p.add_argument("--limit", type=int, default=20, help="埋め込む xlsx 最大数 (新しい順)")
    args = p.parse_args(argv)

    dirs = args.output_dir if args.output_dir else DEFAULT_OUTPUT_DIRS
    dirs = [d for d in dirs if os.path.isdir(d)]
    if not dirs:
        print("ERROR: 有効な output ディレクトリなし")
        return 1

    files = collect_files(dirs, limit=args.limit)
    if not files:
        print("ERROR: xlsx が見つからない")
        return 1

    out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static")
    out_path = args.out or os.path.join(out_dir, "index.html")
    write_dashboard(files, out_path)
    print(f"wrote {out_path}  ({len(files)} files, {sum(os.path.getsize(f['path']) for f in files)//1024} KB raw)")

    if args.serve:
        serve_dir = os.path.dirname(out_path)
        print(f"serving {serve_dir} on http://0.0.0.0:{args.port}")
        os.chdir(serve_dir)
        from http.server import HTTPServer, SimpleHTTPRequestHandler
        try:
            HTTPServer(("0.0.0.0", args.port), SimpleHTTPRequestHandler).serve_forever()
        except KeyboardInterrupt:
            print("\nstopped")
    return 0


if __name__ == "__main__":
    sys.exit(main())
