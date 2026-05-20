"""Lビックドリーム 機種固有シート。

スルー集計は core/excel_through.py の汎用テンプレを使う。
機種固有シート (状態分布、連荘初動状態 等) はここで pandas 計算。
"""

from __future__ import annotations
import pandas as pd
from openpyxl import Workbook

from ...core.excel_styles import HEADER_FILL, HEADER_FONT, BORDER_ALL
from ...core.excel_through import (
    build_g_bin_chain_formula_sheet,
    build_g_bin_chain_preview_sheet,
    build_g_bin_chain_multiblock_formula_sheet,
    build_g_bin_chain_multiblock_preview_sheet,
    build_cycle_bin_chain_preview_sheet,
    build_cycle_bin_chain_formula_sheet,
)
from .config import INVEST_PER_G, AT_PAYOUT_PER_G, WARMUP_G


def _write_header(ws, headers: list[str]) -> None:
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_ALL


def _write_row(ws, row_idx: int, values: list) -> None:
    for c_idx, v in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=c_idx, value=v)
        cell.border = BORDER_ALL


# ---- 機種固有: pandas 計算系 ----

def build_state_distribution(data_df, chains_df, wb):
    ws = wb.create_sheet("状態分布")
    _write_header(ws, ["状態", "件数", "平均出玉", "総出玉", "シェア"])
    if data_df.empty:
        return
    total = len(data_df)
    grp = data_df.groupby("State").agg(
        件数=("Dedama", "count"),
        平均出玉=("Dedama", "mean"),
        総出玉=("Dedama", "sum"),
    ).reset_index()
    grp["シェア"] = grp["件数"] / total
    grp = grp.sort_values("件数", ascending=False)
    for r_idx, row in enumerate(grp.itertuples(index=False), start=2):
        _write_row(ws, r_idx, [str(row[0]) or "(空)", int(row[1]), float(row[2]), int(row[3]), float(row[4])])
        ws.cell(row=r_idx, column=5).number_format = "0.00%"


def build_chain_first_state(data_df, chains_df, wb):
    ws = wb.create_sheet("連荘初動状態")
    _write_header(ws, ["初動状態", "件数", "平均連荘長", "平均純増", "平均総出玉"])
    if chains_df.empty:
        return
    grp = chains_df.groupby("first_state").agg(
        件数=("chain_id", "count"),
        平均連荘長=("chain_length", "mean"),
        平均純増=("net_payout", "mean"),
        平均総出玉=("raw_payout", "mean"),
    ).reset_index().sort_values("件数", ascending=False)
    for r_idx, row in enumerate(grp.itertuples(index=False), start=2):
        _write_row(ws, r_idx, [str(row[0]) or "(空)", int(row[1]), float(row[2]), float(row[3]), float(row[4])])


def build_chain_length_distribution(data_df, chains_df, wb):
    ws = wb.create_sheet("連荘長分布")
    _write_header(ws, ["連チャン数", "件数", "平均純増", "平均差枚", "平均総出玉"])
    if chains_df.empty:
        return
    summary = chains_df.groupby("chain_length").agg(
        件数=("chain_id", "count"),
        平均純増=("net_payout", "mean"),
        平均差枚=("net_diff", "mean"),
        平均総出玉=("raw_payout", "mean"),
    ).reset_index().rename(columns={"chain_length": "連チャン数"})
    for r_idx, row in enumerate(summary.itertuples(index=False), start=2):
        _write_row(ws, r_idx, [int(row[0]), int(row[1]), float(row[2]), float(row[3]), float(row[4])])


def build_at_tenjou_summary(data_df, chains_df, wb):
    """AT間天井狙い: Nから打ち始めて AT が当たるまで (AT連荘単位)。
    引き戻し (first_state='引き戻し') は初当たり扱いせず除外。"""
    common = dict(
        invest_per_g=INVEST_PER_G, at_payout_per_g=AT_PAYOUT_PER_G,
        chain_g_col="pre_normal_g",
        chain_payout_col="net_payout",
        chain_exclude_col="first_state",
        chain_exclude_values=["引き戻し"],
        warmup_g=WARMUP_G,
        cumulative_label="AT到達まで累積期待値",
    )
    build_g_bin_chain_formula_sheet(wb, "AT間天井狙い", **common)
    build_g_bin_chain_preview_sheet(wb, "AT間天井狙い", chains_df, **common)


def build_at_tenjou_after_chokugeki(data_df, chains_df, wb):
    """AT間天井狙い (直撃直後): 直前 chain が直撃/朝一直撃 だった chain だけ対象。
    累積G関係なくフィルタ。有利区間差枚<500 条件も追加。"""
    common = dict(
        invest_per_g=INVEST_PER_G, at_payout_per_g=AT_PAYOUT_PER_G,
        chain_g_col="pre_normal_g",
        chain_payout_col="net_payout",
        chain_filter_col="prev_chain_chokugeki",
        chain_filter_min=1,
        chain_filter2_col="yuri_diff_pre",
        chain_filter2_max=499,
        chain_exclude_col="first_state",
        chain_exclude_values=["引き戻し"],
        warmup_g=WARMUP_G,
        cumulative_label="AT到達まで累積期待値",
    )
    build_g_bin_chain_formula_sheet(wb, "AT間天井狙い_直撃後", **common)
    build_g_bin_chain_preview_sheet(wb, "AT間天井狙い_直撃後", chains_df, **common)


# --- 朝/切断後 N回目 横並び統合シート (1枚に 1+2plus+2..10 の 11 ブロック) ---

def _build_session_blocks(prefix: str) -> list[dict]:
    """prefix='morning' または 'cut'。
    ブロック順: 1回目 → 2回目以降(差枚<500) → 2回目 → 3回目 → ... → 10回目
    morning は切断/直撃を経験した chain 以降を集計外にするため
    `chain_number_morning_safe` を使う。"""
    chain_filter_col = ("chain_number_morning_safe" if prefix == "morning"
                        else f"chain_number_{prefix}")
    label_prefix = "朝" if prefix == "morning" else "切断後"
    common_exclude = dict(chain_exclude_col="first_state",
                          chain_exclude_values=["引き戻し"])
    blocks: list[dict] = [
        {"title": f"{label_prefix}1回目",
         "chain_filter_col": chain_filter_col,
         "chain_filter_min": 1, "chain_filter_max": 1,
         **common_exclude},
        {"title": f"{label_prefix}2回目以降 (差枚<500)",
         "chain_filter_col": chain_filter_col,
         "chain_filter_min": 2,
         "chain_filter2_col": "yuri_diff_pre", "chain_filter2_max": 499,
         **common_exclude},
    ]
    for n in range(2, 11):
        blocks.append({
            "title": f"{label_prefix}{n}回目 (差枚<500)",
            "chain_filter_col": chain_filter_col,
            "chain_filter_min": n, "chain_filter_max": n,
            "chain_filter2_col": "yuri_diff_pre", "chain_filter2_max": 499,
            **common_exclude,
        })
    return blocks


def build_morning_all_formula(data_df, chains_df, wb):
    build_g_bin_chain_multiblock_formula_sheet(
        wb, "AT間_morning", _build_session_blocks("morning"),
        invest_per_g=INVEST_PER_G, at_payout_per_g=AT_PAYOUT_PER_G,
        warmup_g=WARMUP_G,
    )


def build_morning_all_preview(data_df, chains_df, wb):
    build_g_bin_chain_multiblock_preview_sheet(
        wb, "AT間_morning", _build_session_blocks("morning"), chains_df,
        invest_per_g=INVEST_PER_G, at_payout_per_g=AT_PAYOUT_PER_G,
        warmup_g=WARMUP_G,
    )


def build_cut_all_formula(data_df, chains_df, wb):
    build_g_bin_chain_multiblock_formula_sheet(
        wb, "AT間_cut", _build_session_blocks("cut"),
        invest_per_g=INVEST_PER_G, at_payout_per_g=AT_PAYOUT_PER_G,
        warmup_g=WARMUP_G,
    )


def build_cut_all_preview(data_df, chains_df, wb):
    build_g_bin_chain_multiblock_preview_sheet(
        wb, "AT間_cut", _build_session_blocks("cut"), chains_df,
        invest_per_g=INVEST_PER_G, at_payout_per_g=AT_PAYOUT_PER_G,
        warmup_g=WARMUP_G,
    )


# === 有利区間内累積G別 AT間天井狙い (500G刻み、15ビン) ===
# preview版のみ (HTML ビューア用)。formula版は未実装

def build_cycle_morning(data_df, chains_df, wb):
    """累積G別AT間 (朝/切断・直撃経験前のみ)
    chain_number_morning_safe >= 1 で 切断/直撃以降の chain を除外、引き戻し除外。"""
    build_cycle_bin_chain_formula_sheet(
        wb, "累積G別AT間_morning",
        chain_filter_col="chain_number_morning_safe",
        chain_filter_min=1,
        chain_exclude_col="first_state",
        chain_exclude_values=["引き戻し"],
        invest_per_g=INVEST_PER_G, at_payout_per_g=AT_PAYOUT_PER_G,
        warmup_g=WARMUP_G,
    )
    build_cycle_bin_chain_preview_sheet(
        wb, "累積G別AT間_morning", chains_df,
        chain_filter_col="chain_number_morning_safe",
        chain_filter_min=1,
        invest_per_g=INVEST_PER_G, at_payout_per_g=AT_PAYOUT_PER_G,
        chain_exclude_col="first_state",
        chain_exclude_values=["引き戻し"],
        warmup_g=WARMUP_G,
    )


PREV_YURI_BINS = [(i * 1000, i * 1000 + 999) for i in range(7)] + [(7000, 10**9)]


def build_prev_yuri_morning_formula(data_df, chains_df, wb):
    """直前有利区間G別 5G刻み 0-300G 期待値表 (8ブロック横並び formula 版)。
    累積G別AT間_morning の【直前有利区間G数別】(cycle_total_g_pre 軸) と同じ軸で、
    1000G ビン毎にブロックを作り、各ブロック内で pre_normal_g 0-300G を 5G 刻みで集計。
    朝N回目集計対象 (chain_number_morning_safe>=1) のみ。"""
    blocks: list[dict] = []
    for low, high in PREV_YURI_BINS:
        if high >= 10**8:
            title = f"{low}G以上"
            filter_max = None
        else:
            title = f"{low}-{high}G"
            filter_max = high
        blocks.append({
            "title": title,
            # 軸は cycle_total_g_pre (累積G別AT間_morning【直前有利区間G数別】と同じ)
            "chain_filter_col": "cycle_total_g_pre",
            "chain_filter_min": low,
            "chain_filter_max": filter_max,
            # 朝N回目集計対象のみ (切断/直撃以降除外)
            "chain_filter2_col": "chain_number_morning_safe",
            "chain_filter2_min": 1,
            "chain_exclude_col": "first_state",
            "chain_exclude_values": ["引き戻し"],
        })
    build_g_bin_chain_multiblock_formula_sheet(
        wb, "直前有利区間G別期待値",
        blocks,
        invest_per_g=INVEST_PER_G,
        at_payout_per_g=AT_PAYOUT_PER_G,
        warmup_g=WARMUP_G,
        bin_size=5,
        max_g=300,
    )


def build_prev_yuri_morning_preview(data_df, chains_df, wb):
    """直前有利区間G別期待値 (preview 値版、HTML ダッシュボード用)。"""
    blocks: list[dict] = []
    for low, high in PREV_YURI_BINS:
        if high >= 10**8:
            title = f"{low}G以上"
            filter_max = None
        else:
            title = f"{low}-{high}G"
            filter_max = high
        blocks.append({
            "title": title,
            "chain_filter_col": "cycle_total_g_pre",
            "chain_filter_min": low,
            "chain_filter_max": filter_max,
            "chain_filter2_col": "chain_number_morning_safe",
            "chain_filter2_min": 1,
            "chain_exclude_col": "first_state",
            "chain_exclude_values": ["引き戻し"],
        })
    build_g_bin_chain_multiblock_preview_sheet(
        wb, "直前有利区間G別期待値", blocks, chains_df,
        invest_per_g=INVEST_PER_G,
        at_payout_per_g=AT_PAYOUT_PER_G,
        warmup_g=WARMUP_G,
        bin_size=5,
        max_g=300,
    )


def build_cycle_cut(data_df, chains_df, wb):
    """累積G別AT間 (切断後) — chain_number_cut >= 1 のみ、引き戻し除外。formula版+preview版"""
    build_cycle_bin_chain_formula_sheet(
        wb, "累積G別AT間_cut",
        invest_per_g=INVEST_PER_G, at_payout_per_g=AT_PAYOUT_PER_G,
        chain_filter_col="chain_number_cut",
        chain_filter_min=1,
        chain_exclude_col="first_state",
        chain_exclude_values=["引き戻し"],
        warmup_g=WARMUP_G,
    )
    build_cycle_bin_chain_preview_sheet(
        wb, "累積G別AT間_cut", chains_df,
        invest_per_g=INVEST_PER_G, at_payout_per_g=AT_PAYOUT_PER_G,
        chain_filter_col="chain_number_cut",
        chain_filter_min=1,
        chain_exclude_col="first_state",
        chain_exclude_values=["引き戻し"],
        warmup_g=WARMUP_G,
    )


# === 累積G 別 直撃当選分布 ===

CHOKUGEKI_FIRST_STATES = ("直撃", "朝一直撃")
CYCLE_BINS = [(i * 1000, i * 1000 + 999) for i in range(7)] + [(7000, 10**9)]


def _write_chokugeki_block(ws, df, start_row, title, bin_col, bins, group_label):
    """直撃分布の 1 ブロックを書く。次の空き行を返す。"""
    ws.cell(row=start_row, column=1, value=title).font = HEADER_FONT
    header_row = start_row + 1
    headers = [group_label, "サンプル数", "直撃件数", "直撃率(%)", "直撃時TY(枚)"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER_ALL

    is_choku_all = df["first_state"].isin(CHOKUGEKI_FIRST_STATES)
    r_idx = header_row + 1
    for low, high in bins:
        if high >= 10**8:
            mask = df[bin_col] >= low
            label = f"{low}G以上"
        else:
            mask = (df[bin_col] >= low) & (df[bin_col] <= high)
            label = f"{low}-{high}G"
        sub = df[mask]
        n = len(sub)
        choku_mask = mask & is_choku_all
        n_choku = int(choku_mask.sum())
        rate = (n_choku / n * 100) if n else 0.0
        avg_ty = float(df.loc[choku_mask, "net_payout"].mean()) if n_choku else None
        _write_row(ws, r_idx, [label, n, n_choku, rate, avg_ty])
        ws.cell(row=r_idx, column=4).number_format = "0.00"
        ws.cell(row=r_idx, column=5).number_format = "0"
        r_idx += 1

    # ブロック合計
    total = len(df)
    total_choku = int(is_choku_all.sum())
    total_rate = (total_choku / total * 100) if total else 0.0
    total_avg_ty = float(df.loc[is_choku_all, "net_payout"].mean()) if total_choku else None
    _write_row(ws, r_idx, ["合計", total, total_choku, total_rate, total_avg_ty])
    ws.cell(row=r_idx, column=4).number_format = "0.00"
    ws.cell(row=r_idx, column=5).number_format = "0"
    for c in range(1, 6):
        ws.cell(row=r_idx, column=c).font = HEADER_FONT
    return r_idx + 2  # 空行 + 次ブロック先頭


# === 引き戻し分布 (一撃枚数別 引き戻し率) ===

HIKIMODOSHI_BINS = [(i * 500, i * 500 + 499) for i in range(14)] + [(7000, 10**9)]


def _write_hikimodoshi_block(ws, df, start_row, title, ty_col, bins, group_label):
    """引き戻し分布 1 ブロック書き込み。
    df: 集計対象 chain (has_jyoi=False or True など事前フィルタ済み)
    ty_col: 一撃枚数として使う列名 (net_payout or jyoi_chuugeki)
    """
    ws.cell(row=start_row, column=1, value=title).font = HEADER_FONT
    header_row = start_row + 1
    headers = [group_label, "サンプル数", "上位あり", "上位率(%)",
               "3-90内当選 件数", "3-90内当選 率(%)"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER_ALL

    r_idx = header_row + 1
    for low, high in bins:
        if high >= 10**8:
            mask = df[ty_col] >= low
            label = f"{low}+枚"
        else:
            mask = (df[ty_col] >= low) & (df[ty_col] <= high)
            label = f"{low}-{high}枚"
        sub = df[mask]
        n = len(sub)
        n_j = int(sub["next_jyoi"].sum()) if n else 0
        n_590 = int(sub.get("next_3_90", pd.Series(0, index=sub.index)).sum()) if n else 0
        rate_j = (n_j / n * 100) if n else 0.0
        rate_590 = (n_590 / n * 100) if n else 0.0
        _write_row(ws, r_idx, [label, n, n_j, rate_j, n_590, rate_590])
        ws.cell(row=r_idx, column=4).number_format = "0.00"
        ws.cell(row=r_idx, column=6).number_format = "0.00"
        r_idx += 1

    total = len(df)
    total_j = int(df["next_jyoi"].sum()) if total else 0
    total_590 = int(df.get("next_3_90", pd.Series(0, index=df.index)).sum()) if total else 0
    total_rate_j = (total_j / total * 100) if total else 0.0
    total_rate_590 = (total_590 / total * 100) if total else 0.0
    _write_row(ws, r_idx, ["合計", total, total_j, total_rate_j, total_590, total_rate_590])
    ws.cell(row=r_idx, column=4).number_format = "0.00"
    ws.cell(row=r_idx, column=6).number_format = "0.00"
    for c in range(1, 7):
        ws.cell(row=r_idx, column=c).font = HEADER_FONT
    return r_idx + 2


def _compute_segments(data_df):
    """preview 版引き戻し分布の入力。
    adapter.detect が data_df の各 segment 起点行に書いた列を直接読む:
      Segment_Type / Segment_Dedama / Is_Shimoi_Start / Is_Jyoi_Start /
      Next_Jyoi / 次イベント_3_90内
    """
    import pandas as pd
    if data_df.empty or "Is_Shimoi_Start" not in data_df.columns:
        return pd.DataFrame(columns=["type", "dedama", "next_jyoi", "next_3_90"])
    mask = (data_df["Is_Shimoi_Start"] == 1) | (data_df["Is_Jyoi_Start"] == 1)
    cols = ["Segment_Type", "Segment_Dedama", "Next_Jyoi", "次イベント_3_90内"]
    seg = data_df.loc[mask, cols].copy()
    seg = seg.rename(columns={
        "Segment_Type": "type",
        "Segment_Dedama": "dedama",
        "Next_Jyoi": "next_jyoi",
        "次イベント_3_90内": "next_3_90",
    })
    return seg.reset_index(drop=True)


def _compute_segments_legacy_unused(data_df):
    """旧 segment 計算 (AT only). 残骸 — 参照なし、後で消す候補。
    segment 区切り:
      - 台日内最初の AT 行
      - Chain_ID 変化 (= 新連荘: 直撃/朝一直撃/CZ成功)
      - State == '引き戻し'
      - State == '上位AT' で前行が上位AT 以外
      - 上位AT 区間終了直後の AT 行
    各 segment: dedama合計, start_state, start_g, 次 segment の起点 state/G を記録。
    """
    import pandas as pd
    segments = []
    at_only = data_df[data_df["Is_AT"]].sort_values("Original_Order")
    for key, grp in at_only.groupby(["Hall", "Date", "Machine_No"], sort=False):
        day_segs = []
        cur_type = None
        cur_dedama = 0
        cur_start_state = None
        cur_start_g = None
        cur_chain_id = None
        prev_state = None
        prev_chain_id = None
        for _, row in grp.iterrows():
            state = row["State"]
            dedama = int(row["Dedama"] or 0)
            chain_id = int(row["Chain_ID"])
            start_g = int(row["Start"] or 0)
            new_seg = False
            seg_type = None
            if cur_type is None:
                new_seg = True
                seg_type = ('上位一撃' if state == '上位AT'
                            else '引き戻し起点' if state == '引き戻し' else 'AT')
            elif chain_id != prev_chain_id:
                # Chain 切替 (直撃/朝一直撃/CZ成功) = 新セグ
                new_seg = True
                seg_type = '上位一撃' if state == '上位AT' else 'AT'
            elif state == '引き戻し':
                new_seg = True; seg_type = '引き戻し起点'
            elif state == '上位AT' and prev_state != '上位AT':
                new_seg = True; seg_type = '上位一撃'
            elif state != '上位AT' and cur_type == '上位一撃':
                new_seg = True; seg_type = 'AT'
            if new_seg:
                if cur_type is not None:
                    day_segs.append({"chain_id": cur_chain_id, "type": cur_type,
                                     "start_state": cur_start_state,
                                     "start_g": cur_start_g,
                                     "dedama": cur_dedama})
                cur_type = seg_type
                cur_start_state = state
                cur_start_g = start_g
                cur_chain_id = chain_id
                cur_dedama = 0
            cur_dedama += dedama
            prev_state = state
            prev_chain_id = chain_id
        if cur_type is not None:
            day_segs.append({"chain_id": cur_chain_id, "type": cur_type,
                             "start_state": cur_start_state,
                             "start_g": cur_start_g,
                             "dedama": cur_dedama})
        # 次 segment 起点情報を各 seg に紐付け (台日内連結)
        for i, seg in enumerate(day_segs):
            next_seg = day_segs[i + 1] if i + 1 < len(day_segs) else None
            ns_state = next_seg["start_state"] if next_seg else None
            ns_g = next_seg["start_g"] if next_seg else None
            seg["next_start_state"] = ns_state
            seg["next_start_g"] = ns_g
            seg["next_jyoi"] = int(ns_state == "上位AT")
            # 3-90 内当選 (上位込み総数): 次 segment 起点 G が 3-90
            in_3_90 = (ns_g is not None and 3 <= int(ns_g) <= 90)
            seg["next_3_90"] = int(in_3_90)
        segments.extend(day_segs)
    return pd.DataFrame(segments)


def _col_letter(df, name):
    from openpyxl.utils import get_column_letter
    return get_column_letter(df.columns.get_loc(name) + 1)


# === Data / ChainData ヘッダー日本語化 ===
DATA_HEADER_JA = {
    "Chain_ID": "連荘ID",
    "Chain_Position": "連荘内位置",
    "Chain_Length": "連荘長",
    "Is_AT": "AT行フラグ",
    "Chain_Pre_Normal_G": "連荘前通常G",
    "Yuri_Diff_Pre": "有利区間差枚_行直前",
    "Segment_Type": "セグメント種別",
    "Is_Shimoi_Start": "下位起点フラグ",
    "Is_Jyoi_Start": "上位起点フラグ",
    "Segment_Dedama": "セグメント合計出玉",
    "Next_Jyoi": "次が上位AT",
}

CHAINDATA_HEADER_JA = {
    "id": "台日ID",
    "chain_id": "連荘ID",
    "chain_number": "連荘番号",
    "chain_length": "連荘長",
    "first_g": "起点G",
    "first_state": "起点状態",
    "pre_normal_g": "連荘前通常G",
    "chain_total_g": "連荘総G",
    "is_streak": "連荘フラグ",
    "raw_payout": "総出玉",
    "net_payout": "純出玉_TY",
    "in_chain_invest": "連荘中投資",
    "net_diff": "純差枚",
    "yuri_diff_pre": "有利区間差枚_直前",
    "has_jyoi": "上位AT発動あり",
    "chain_number_morning": "朝N回目",
    "prev_chain_has_jyoi": "直前連荘が上位",
    "prev_chain_chokugeki": "直前連荘が直撃",
    "next_is_hikimodoshi": "次連荘が引き戻し",
    "jyoi_chuugeki": "上位一撃枚数",
    "chain_number_cut": "切断後N回目",
    "chain_number_morning_safe": "朝N回目_集計用",
    "cycle_g_pre": "有利区間累積通常G",
    "cycle_total_g_pre": "有利区間累積G",
    "prev_yuri_total_g": "前回有利区間総G",
    "is_choku": "直撃フラグ",
    "next_chain_first_state": "次連荘起点状態",
    "next_chain_first_g": "次連荘起点G",
    "chain_via_cz": "CZ経由連荘",
}


def rename_headers_to_ja(data_df, chains_df, wb):
    """Data/ChainData の派生列ヘッダーを日本語に書き換える。
    元データ列 (ID,Hall,Date,Machine_No,Time,Status,Start,Dedama,State,Original_Order) は維持。"""
    for sheet_name, mapping in (("Data", DATA_HEADER_JA), ("ChainData", CHAINDATA_HEADER_JA)):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            if v in mapping:
                ws.cell(row=1, column=c, value=mapping[v])


def _write_hikimodoshi_formula_block(ws, start_row, title, types, bins, group_label,
                                      data_df=None, n_rows=200000):
    """計算式版の引き戻し分布ブロック。data_sheet 参照の COUNTIFS/AVERAGEIFS を入れる。
    types: type 列のフィルタ候補 (例: ['AT','引き戻し起点'] / ['上位一撃'])
    """
    ds = "Data"
    flag_col_name = "Is_Jyoi_Start" if types == ["上位一撃"] else "Is_Shimoi_Start"
    seg_start_l = _col_letter(data_df, flag_col_name)
    type_l = _col_letter(data_df, "Segment_Type")
    dedama_l = _col_letter(data_df, "Segment_Dedama")
    nj_l = _col_letter(data_df, "Next_Jyoi")
    n590_l = _col_letter(data_df, "次イベント_3_90内")
    seg_start_col = f"{ds}!{seg_start_l}2:{seg_start_l}{n_rows+1}"
    type_col = f"{ds}!{type_l}2:{type_l}{n_rows+1}"
    dedama_col = f"{ds}!{dedama_l}2:{dedama_l}{n_rows+1}"
    nj_col = f"{ds}!{nj_l}2:{nj_l}{n_rows+1}"
    n590_col = f"{ds}!{n590_l}2:{n590_l}{n_rows+1}"

    ws.cell(row=start_row, column=1, value=title).font = HEADER_FONT
    header_row = start_row + 1
    headers = [group_label, "サンプル数", "上位あり", "上位率(%)",
               "3-90内当選 件数", "3-90内当選 率(%)"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=i, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.border = BORDER_ALL

    def _type_terms(formula_template):
        if len(types) == 1:
            return formula_template.format(type_filter=f'"{types[0]}"')
        parts = [formula_template.format(type_filter=f'"{t}"') for t in types]
        return "+".join(parts)

    base = f'{seg_start_col},1'
    r_idx = header_row + 1
    for low, high in bins:
        if high >= 10**8:
            label = f"{low}+枚"
            ded_filter = f'{dedama_col},">={low}"'
        else:
            label = f"{low}-{high}枚"
            ded_filter = f'{dedama_col},">={low}",{dedama_col},"<={high}"'

        sample_tpl = f'COUNTIFS({base},{type_col},{{type_filter}},{ded_filter})'
        sample_f = "=" + _type_terms(sample_tpl)
        jyoi_tpl = f'COUNTIFS({base},{type_col},{{type_filter}},{ded_filter},{nj_col},1)'
        jyoi_f = "=" + _type_terms(jyoi_tpl)
        rate_j_f = f'=IFERROR(C{r_idx}/B{r_idx}*100,0)'
        n590_tpl = f'COUNTIFS({base},{type_col},{{type_filter}},{ded_filter},{n590_col},1)'
        n590_count = "=" + _type_terms(n590_tpl)
        rate_n590_f = f'=IFERROR(E{r_idx}/B{r_idx}*100,0)'

        _write_row(ws, r_idx, [label, sample_f, jyoi_f, rate_j_f, n590_count, rate_n590_f])
        ws.cell(row=r_idx, column=4).number_format = "0.00"
        ws.cell(row=r_idx, column=6).number_format = "0.00"
        r_idx += 1

    sample_all_tpl = f'COUNTIFS({base},{type_col},{{type_filter}})'
    sample_all = "=" + _type_terms(sample_all_tpl)
    jyoi_all_tpl = f'COUNTIFS({base},{type_col},{{type_filter}},{nj_col},1)'
    jyoi_all = "=" + _type_terms(jyoi_all_tpl)
    rate_j_all = f'=IFERROR(C{r_idx}/B{r_idx}*100,0)'
    n590_all_tpl = f'COUNTIFS({base},{type_col},{{type_filter}},{n590_col},1)'
    n590_all = "=" + _type_terms(n590_all_tpl)
    rate_n590_all = f'=IFERROR(E{r_idx}/B{r_idx}*100,0)'

    _write_row(ws, r_idx, ["合計", sample_all, jyoi_all, rate_j_all, n590_all, rate_n590_all])
    ws.cell(row=r_idx, column=4).number_format = "0.00"
    ws.cell(row=r_idx, column=6).number_format = "0.00"
    for c in range(1, 7):
        ws.cell(row=r_idx, column=c).font = HEADER_FONT
    return r_idx + 2


def build_hikimodoshi_formula(data_df, chains_df, wb):
    """引き戻し分布 (計算式版)。Data シートの segment 列を直接参照する。"""
    ws = wb.create_sheet("引き戻し分布")
    next_row = _write_hikimodoshi_formula_block(
        ws, 1, "【下位AT区間 戻し判定用一撃枚数 / 500枚刻み】",
        ["AT", "引き戻し起点"], HIKIMODOSHI_BINS, "戻し判定用一撃枚数",
        data_df=data_df,
    )
    _write_hikimodoshi_formula_block(
        ws, next_row, "【上位AT区間 戻し判定用一撃枚数 / 500枚刻み】",
        ["上位一撃"], HIKIMODOSHI_BINS, "戻し判定用一撃枚数",
        data_df=data_df,
    )


def build_hikimodoshi_distribution(data_df, chains_df, wb):
    """引き戻し分布シート (一撃枚数別の引き戻し率)。
    chain を「AT一撃 / 上位一撃 / 引き戻し起点AT一撃」segment に分解して集計。
    Block 1: AT一撃 (= 通常 AT or 引き戻し起点) の dedama を 500枚 bin
    Block 2: 上位一撃 の dedama を 500枚 bin
    引き戻し判定: 次 segment の start_state == '引き戻し'
    """
    ws = wb.create_sheet("引き戻し分布_値")

    segments = _compute_segments(data_df)
    if segments.empty:
        ws.cell(row=1, column=1, value="(segment データなし)")
        return

    df_at = segments[segments["type"].isin(["AT", "引き戻し起点"])]
    next_row = _write_hikimodoshi_block(
        ws, df_at, 1,
        "【下位AT区間 戻し判定用一撃枚数 / 500枚刻み】",
        "dedama", HIKIMODOSHI_BINS, "戻し判定用一撃枚数",
    )

    df_jyoi = segments[segments["type"] == "上位一撃"]
    _write_hikimodoshi_block(
        ws, df_jyoi, next_row,
        "【上位AT区間 戻し判定用一撃枚数 / 500枚刻み】",
        "dedama", HIKIMODOSHI_BINS, "戻し判定用一撃枚数",
    )


def _write_chokugeki_formula_block(ws, start_row, title, bins, group_label,
                                    cut_filter=False,
                                    chains_df=None, n_rows=200000):
    """直撃分布 計算式版 1 ブロック。
    cut_filter: True なら chain_number_cut>=1 (切断後ブロック)
    共通フィルタ: first_state<>引き戻し AND pre_normal_g>=33
    """
    ds = "ChainData"
    fs_l = _col_letter(chains_df, "first_state")
    png_l = _col_letter(chains_df, "pre_normal_g")
    cyc_l = _col_letter(chains_df, "cycle_total_g_pre")
    cut_l = _col_letter(chains_df, "chain_number_cut")
    np_l = _col_letter(chains_df, "net_payout")
    choku_l = _col_letter(chains_df, "is_choku")
    fs_col = f"{ds}!{fs_l}2:{fs_l}{n_rows+1}"
    png_col = f"{ds}!{png_l}2:{png_l}{n_rows+1}"
    cyc_col = f"{ds}!{cyc_l}2:{cyc_l}{n_rows+1}"
    cut_col = f"{ds}!{cut_l}2:{cut_l}{n_rows+1}"
    np_col = f"{ds}!{np_l}2:{np_l}{n_rows+1}"
    choku_col = f"{ds}!{choku_l}2:{choku_l}{n_rows+1}"

    # 共通フィルタ部
    base_filters = (f'{fs_col},"<>引き戻し",{png_col},">={WARMUP_G}"')
    if cut_filter:
        base_filters += f',{cut_col},">=1"'

    ws.cell(row=start_row, column=1, value=title).font = HEADER_FONT
    header_row = start_row + 1
    headers = [group_label, "サンプル数", "直撃件数", "直撃率(%)", "直撃時TY(枚)"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=i, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.border = BORDER_ALL

    r_idx = header_row + 1
    for low, high in bins:
        if high >= 10**8:
            label = f"{low}G以上"
            cyc_range = f',{cyc_col},">={low}"'
        else:
            label = f"{low}-{high}G"
            cyc_range = f',{cyc_col},">={low}",{cyc_col},"<={high}"'
        sample_f = f'=COUNTIFS({base_filters}{cyc_range})'
        choku_f = f'=COUNTIFS({base_filters}{cyc_range},{choku_col},1)'
        rate_f = f'=IFERROR(C{r_idx}/B{r_idx}*100,0)'
        avg_ty_f = f'=IFERROR(AVERAGEIFS({np_col},{base_filters}{cyc_range},{choku_col},1),"")'
        _write_row(ws, r_idx, [label, sample_f, choku_f, rate_f, avg_ty_f])
        ws.cell(row=r_idx, column=4).number_format = "0.00"
        ws.cell(row=r_idx, column=5).number_format = "0"
        r_idx += 1

    # 合計
    sample_all = f'=COUNTIFS({base_filters})'
    choku_all = f'=COUNTIFS({base_filters},{choku_col},1)'
    rate_all = f'=IFERROR(C{r_idx}/B{r_idx}*100,0)'
    avg_ty_all = f'=IFERROR(AVERAGEIFS({np_col},{base_filters},{choku_col},1),"")'
    _write_row(ws, r_idx, ["合計", sample_all, choku_all, rate_all, avg_ty_all])
    ws.cell(row=r_idx, column=4).number_format = "0.00"
    ws.cell(row=r_idx, column=5).number_format = "0"
    for c in range(1, 6):
        ws.cell(row=r_idx, column=c).font = HEADER_FONT
    return r_idx + 2


def build_chokugeki_formula(data_df, chains_df, wb):
    """直撃分布 (計算式版)。ChainData シートを直接参照。"""
    ws = wb.create_sheet("直撃分布")
    next_row = _write_chokugeki_formula_block(
        ws, 1, "【朝】", CYCLE_BINS, "有利区間累積G", cut_filter=False,
        chains_df=chains_df,
    )
    _write_chokugeki_formula_block(
        ws, next_row, "【切断後】", CYCLE_BINS, "有利区間累積G", cut_filter=True,
        chains_df=chains_df,
    )


def build_chokugeki_distribution(data_df, chains_df, wb):
    """直撃分布シート (有利区間累積G 軸 × 今回/前回)。
    Block 1: 今回 有利区間累積G別 / 朝          (cycle_total_g_pre, 全chain)
    Block 2: 今回 有利区間累積G別 / 切断後      (cycle_total_g_pre, chain_number_cut>=1)
    Block 3: 前回 有利区間累積G別                (prev_yuri_total_g, 前回値あり chain のみ)
    フィルタは累積G別AT間と同じ: 引き戻し除外 + 着席G(=pre_normal_g)>=33
    """
    ws = wb.create_sheet("直撃分布_値")

    if "cycle_total_g_pre" not in chains_df.columns:
        ws.cell(row=1, column=1, value="(cycle_total_g_pre 列なし)")
        return

    # 共通フィルタ: 引き戻し除外 + warmup>=33
    base = chains_df[
        (chains_df["first_state"] != "引き戻し") &
        (chains_df["pre_normal_g"] >= WARMUP_G)
    ]
    df_morning = base
    df_cut = base[base["chain_number_cut"] >= 1]

    next_row = _write_chokugeki_block(
        ws, df_morning, 1,
        "【朝】",
        "cycle_total_g_pre", CYCLE_BINS,
        "有利区間累積G",
    )
    _write_chokugeki_block(
        ws, df_cut, next_row,
        "【切断後】",
        "cycle_total_g_pre", CYCLE_BINS,
        "有利区間累積G",
    )
