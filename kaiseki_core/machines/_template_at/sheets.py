"""AT機テンプレ固有シート。骨組みのみ。ゾーン区切りや期待値式は実機ごとに埋める。"""

from __future__ import annotations
import pandas as pd
from openpyxl import Workbook

from ...core.excel_styles import HEADER_FILL, HEADER_FONT, BORDER_ALL


def build_zone_expectation_sheet(
    data_df: pd.DataFrame, chains_df: pd.DataFrame, wb: Workbook
) -> None:
    """ゾーン期待値表 (骨組み)。
    TODO: 機種ごとのゾーン区切り (例: 100-150G, 200-250G) と期待値式を実装。
    """
    ws = wb.create_sheet("ゾーン期待値")
    headers = ["ゾーン", "件数", "平均出玉", "総差枚", "期待値"]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_ALL

    # 代表的ゾーン (実機に合わせて差し替え)
    zones = [
        ("0-99G", 0, 99),
        ("100-149G", 100, 149),
        ("150-199G", 150, 199),
        ("200-299G", 200, 299),
        ("300G+", 300, 9999),
    ]
    for i, (label, _lo, _hi) in enumerate(zones, start=2):
        ws.cell(row=i, column=1, value=label).border = BORDER_ALL
        # TODO: COUNTIFS / AVERAGEIFS で Data シートから集計


def build_at_chain_summary_sheet(
    data_df: pd.DataFrame, chains_df: pd.DataFrame, wb: Workbook
) -> None:
    """AT連荘サマリ (骨組み)。チェーン長別の件数と平均純増。"""
    ws = wb.create_sheet("AT連荘サマリ")
    headers = ["連チャン数", "件数", "平均純増", "平均差枚"]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_ALL

    if chains_df.empty:
        return

    summary = (
        chains_df.groupby("chain_length")
        .agg(
            件数=("chain_id", "count"),
            平均純増=("net_payout", "mean"),
            平均差枚=("net_diff", "mean"),
        )
        .reset_index()
        .rename(columns={"chain_length": "連チャン数"})
    )
    for r_idx, row in enumerate(summary.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=float(value) if c_idx > 1 else int(value)).border = BORDER_ALL
