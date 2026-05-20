"""AT機テンプレ E2E スモーク: CSV → runner → xlsx。"""

import os
import tempfile
from openpyxl import load_workbook

from kaiseki_core.core.runner import main


SAMPLE = os.path.join(
    os.path.dirname(__file__), "..", "samples", "at_dummy.csv"
)


def test_runner_at_template():
    with tempfile.TemporaryDirectory() as tmp:
        out = os.path.join(tmp, "out.xlsx")
        rc = main(["--machine", "_template_at", "--input", SAMPLE, "--output", out])
        assert rc == 0
        assert os.path.exists(out)

        wb = load_workbook(out)
        assert "Data" in wb.sheetnames
        assert "ChainData" in wb.sheetnames
        assert "ゾーン期待値" in wb.sheetnames
        assert "AT連荘サマリ" in wb.sheetnames

        ws_data = wb["Data"]
        # ヘッダ + 5データ行
        assert ws_data.max_row == 6
