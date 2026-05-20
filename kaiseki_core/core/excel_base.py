"""Excel出力基盤: 共通 Data/ChainData シート + 機種固有 SheetSpec 実行。"""

from __future__ import annotations
from dataclasses import dataclass
from typing import Callable
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from .excel_styles import HEADER_FILL, HEADER_FONT, BORDER_ALL


SheetBuilder = Callable[[pd.DataFrame, pd.DataFrame, Workbook], None]


@dataclass
class SheetSpec:
    name: str
    builder: SheetBuilder


def _write_df_to_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(sheet_name)
    rows = list(dataframe_to_rows(df, index=False, header=True))
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            if isinstance(value, (list, dict)):
                value = str(value)
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            cell.border = BORDER_ALL
    ws.freeze_panes = "A2"


def write_workbook(
    output_path: str,
    data_df: pd.DataFrame,
    chains_df: pd.DataFrame,
    extra_sheets: list[SheetSpec] | None = None,
) -> None:
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    _write_df_to_sheet(wb, "Data", data_df)
    _write_df_to_sheet(wb, "ChainData", chains_df)

    for spec in extra_sheets or []:
        spec.builder(data_df, chains_df, wb)

    wb.save(output_path)
