"""スルー別集計の Excel数式埋め込み + pandas プレビューシート生成 (汎用)。

shuuki.md 準拠の累積期待値計算をテンプレ化:
- EV(N) = 成功率(N) × TY(N) + (1 - 成功率(N)) × EV(N+1) - 増分G(N) × 純減

bigdream 以外の機種でも、信号(CZ等)行に Through_Index と Cum_G_At_Hit が付いており、
連荘単位で「起点 Through_Index」と「平均TY」が ChainData にあれば再利用可能。
"""

from __future__ import annotations
from typing import Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .excel_styles import HEADER_FILL, HEADER_FONT, BORDER_ALL

YEN_PER_COIN = 20  # 等価交換
BET_PER_GAME = 3.0  # スマスロ 3枚掛け固定


def _col_map(ws) -> dict[str, str]:
    return {ws.cell(row=1, column=c).value: get_column_letter(c)
            for c in range(1, ws.max_column + 1)}


def _write_header(ws, headers: list[str]) -> None:
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_ALL


def _write_row(ws, row_idx: int, values: list) -> None:
    for c_idx, v in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=c_idx, value=v)
        cell.border = BORDER_ALL


# フィルタ列名の日本語ラベル (HTMLバナー表示用)。機種共通の汎用辞書。
FILTER_LABEL_JA = {
    "chain_number": "N回目",
    "chain_number_morning": "朝N回目",
    "chain_number_cut": "切断後N回目",
    "yuri_diff_pre": "起点直前の有利区間差枚",
    "first_state": "初動状態",
    "starting_cz_through": "起点CZスルー目",
    "starting_cz_cum_g": "起点CZのG",
    "pre_normal_g": "前ハマリG",
    "chain_length": "連荘長",
}


def _build_filter_description(
    chain_through_min_col=None, chain_through_min_value=None,
    chain_through_max_col=None, chain_through_max_value=None,
    chain_filter_col=None, chain_filter_min=None, chain_filter_max=None,
    chain_filter2_col=None, chain_filter2_min=None, chain_filter2_max=None,
    chain_exclude_col=None, chain_exclude_values=None,
    warmup_g: int = 0,
) -> str:
    """フィルタ引数群から表示用条件文字列を組み立てる。
    例: 'chain_number=2 かつ 起点直前の有利区間差枚<500'"""
    parts: list[str] = []

    def _label(col: str) -> str:
        return FILTER_LABEL_JA.get(col, col)

    if chain_through_min_col and chain_through_min_value is not None:
        parts.append(f"{_label(chain_through_min_col)}≥{chain_through_min_value}")
    if chain_through_max_col and chain_through_max_value is not None:
        parts.append(f"{_label(chain_through_max_col)}≤{chain_through_max_value}")
    if chain_filter_col:
        lbl = _label(chain_filter_col)
        if chain_filter_min is not None and chain_filter_max is not None:
            if chain_filter_min == chain_filter_max:
                parts.append(f"{lbl}={chain_filter_min}")
            else:
                parts.append(f"{chain_filter_min}≤{lbl}≤{chain_filter_max}")
        elif chain_filter_min is not None:
            parts.append(f"{lbl}≥{chain_filter_min}")
        elif chain_filter_max is not None:
            parts.append(f"{lbl}≤{chain_filter_max}")
    if chain_filter2_col:
        lbl = _label(chain_filter2_col)
        if chain_filter2_min is not None and chain_filter2_max is not None:
            if chain_filter2_min == chain_filter2_max:
                parts.append(f"{lbl}={chain_filter2_min}")
            else:
                parts.append(f"{chain_filter2_min}≤{lbl}≤{chain_filter2_max}")
        elif chain_filter2_min is not None:
            parts.append(f"{lbl}≥{chain_filter2_min}")
        elif chain_filter2_max is not None:
            # max=499 のとき「<500」と表現したほうが直感的
            v = chain_filter2_max
            if isinstance(v, int) and v >= 0 and (v + 1) % 100 == 0:
                parts.append(f"{lbl}<{v + 1}")
            else:
                parts.append(f"{lbl}≤{v}")
    if chain_exclude_col and chain_exclude_values:
        lbl = FILTER_LABEL_JA.get(chain_exclude_col, chain_exclude_col)
        vals = ", ".join(str(v) for v in chain_exclude_values)
        parts.append(f"{lbl}≠{vals}")
    if warmup_g > 0:
        parts.append(f"着席G≥{warmup_g}")
    return " かつ ".join(parts)


FILTER_DESC_CELL_COL = 16  # 1行目 column 16 に "算出条件:" + 文字列 を書く


def _write_filter_description(ws, description: str) -> None:
    """preview sheet の1行目末尾に算出条件メタを書き込む。"""
    if not description:
        return
    ws.cell(row=1, column=FILTER_DESC_CELL_COL, value=f"算出条件: {description}")


def build_through_summary_formula_sheet(
    wb: Workbook,
    sheet_name: str,
    *,
    invest_per_g: float,
    at_payout_per_g: float = 8.2,  # 未使用 (互換)
    data_sheet: str = "Data",
    chain_sheet: str = "ChainData",
    signal_col: str,
    success_col: str,
    through_col: str,
    cum_g_col: str,
    chain_through_col: str,
    chain_payout_col: str,
    label: str = "スルー目",
    max_through: int = 10,
) -> None:
    """Excel数式でスルー別集計シートを生成。

    必須Data列: signal_col(bool), success_col(bool), through_col(int), cum_g_col(int)
    必須ChainData列: chain_through_col(int), chain_payout_col(float)
    """
    ws = wb.create_sheet(sheet_name)
    _write_header(ws, [label, "件数", "成功件数", "成功率",
                       "平均TY(枚)", "累積消化G", "増分G",
                       "CZ間狙い期待値(枚)", "CZ間狙い期待値(円)",
                       "AT間狙い期待値(枚)", "AT間狙い期待値(円)", "AT間狙い機械割(%)"])

    data_ws = wb[data_sheet]
    chain_ws = wb[chain_sheet]
    data_cols = _col_map(data_ws)
    chain_cols = _col_map(chain_ws)

    needed = {signal_col: data_cols, success_col: data_cols,
              through_col: data_cols, cum_g_col: data_cols,
              chain_through_col: chain_cols, chain_payout_col: chain_cols}
    if any(col not in src for col, src in needed.items()):
        missing = [c for c, s in needed.items() if c not in s]
        ws.cell(row=2, column=1, value=f"(必要列が見つかりません: {missing})")
        return

    sig_l = data_cols[signal_col]
    suc_l = data_cols[success_col]
    thr_l = data_cols[through_col]
    cum_l = data_cols[cum_g_col]
    ct_l = chain_cols[chain_through_col]
    cp_l = chain_cols[chain_payout_col]

    data_ref = f"'{data_sheet}'"
    chain_ref = f"'{chain_sheet}'"

    labels = list(range(1, max_through + 1)) + [f"{max_through + 1}+"]
    n_labels = len(labels)

    for r_idx, lbl in enumerate(labels, start=2):
        if isinstance(lbl, int):
            cond = f'"={lbl}"'
            chain_cond_str = lbl
        else:
            cond = f'">={max_through + 1}"'
            chain_cond_str = f'">={max_through + 1}"'  # 既に引用符込み

        b = (f'=COUNTIFS({data_ref}!{sig_l}:{sig_l},TRUE,'
             f'{data_ref}!{thr_l}:{thr_l},{cond})')
        c = (f'=COUNTIFS({data_ref}!{suc_l}:{suc_l},TRUE,'
             f'{data_ref}!{thr_l}:{thr_l},{cond})')
        d = f"=IFERROR(C{r_idx}/B{r_idx},0)"
        e = (f'=IFERROR(AVERAGEIFS({chain_ref}!{cp_l}:{cp_l},'
             f'{chain_ref}!{ct_l}:{ct_l},{chain_cond_str}),0)')
        f = (f'=IFERROR(AVERAGEIFS({data_ref}!{cum_l}:{cum_l},'
             f'{data_ref}!{sig_l}:{sig_l},TRUE,'
             f'{data_ref}!{thr_l}:{thr_l},{cond}),0)')
        g = f"=F{r_idx}" if r_idx == 2 else f"=F{r_idx}-F{r_idx - 1}"
        h = f"=D{r_idx}*E{r_idx}-G{r_idx}*{invest_per_g}"
        i = f"=H{r_idx}*{YEN_PER_COIN}"
        if r_idx == n_labels + 1:
            j = f"=H{r_idx}"
        else:
            j = f"=D{r_idx}*E{r_idx}+(1-D{r_idx})*J{r_idx + 1}-G{r_idx}*{invest_per_g}"
        k = f"=J{r_idx}*{YEN_PER_COIN}"
        last_row = n_labels + 1
        l_cell = (f"=IFERROR((J{r_idx}+SUM(G{r_idx}:G{last_row})*{invest_per_g})"
                  f"/(SUM(G{r_idx}:G{last_row})*{invest_per_g})*100,0)")

        _write_row(ws, r_idx, [lbl, b, c, d, e, f, g, h, i, j, k, l_cell])
        ws.cell(row=r_idx, column=4).number_format = "0.00%"
        for col_idx in (5, 6, 7, 8, 10):
            ws.cell(row=r_idx, column=col_idx).number_format = "0.0"
        for col_idx in (9, 11):
            ws.cell(row=r_idx, column=col_idx).number_format = "0"
        ws.cell(row=r_idx, column=12).number_format = "0.00"


def build_through_summary_preview_sheet(
    wb: Workbook,
    sheet_name: str,
    data_df: pd.DataFrame,
    chains_df: pd.DataFrame,
    *,
    invest_per_g: float,
    at_payout_per_g: float = 8.2,  # 未使用 (互換)
    signal_col: str,
    success_col: str,
    through_col: str,
    cum_g_col: str,
    chain_through_col: str,
    chain_payout_col: str,
    label: str = "スルー目",
    max_through: int = 10,
    sheet_suffix: str = "_値",
) -> None:
    """pandas 計算済みの値プレビュー版 (Streamlit用)。"""
    ws = wb.create_sheet(sheet_name + sheet_suffix)
    _write_header(ws, [label, "件数", "成功件数", "成功率",
                       "平均TY(枚)", "累積消化G", "増分G",
                       "CZ間狙い期待値(枚)", "CZ間狙い期待値(円)",
                       "AT間狙い期待値(枚)", "AT間狙い期待値(円)", "AT間狙い機械割(%)"])
    labels = list(range(1, max_through + 1)) + [f"{max_through + 1}+"]

    rows = []
    prev_g = 0.0
    for lbl in labels:
        if isinstance(lbl, int):
            data_mask = data_df[signal_col] & (data_df[through_col] == lbl)
            chain_mask = chains_df[chain_through_col] == lbl
        else:
            data_mask = data_df[signal_col] & (data_df[through_col] >= max_through + 1)
            chain_mask = chains_df[chain_through_col] >= max_through + 1
        n_total = int(data_mask.sum())
        n_success = int((data_mask & data_df[success_col]).sum())
        rate = n_success / n_total if n_total else 0.0
        sub = chains_df[chain_mask]
        avg_np = float(sub[chain_payout_col].mean()) if not sub.empty else 0.0
        cum_series = data_df.loc[data_mask, cum_g_col]
        avg_g = float(cum_series.mean()) if not cum_series.empty else 0.0
        step_g = avg_g - prev_g if avg_g else 0.0
        prev_g = avg_g if avg_g else prev_g
        single_ev = rate * avg_np - step_g * invest_per_g
        rows.append({"lbl": lbl, "n_total": n_total, "n_success": n_success,
                     "rate": rate, "avg_np": avg_np, "avg_g": avg_g,
                     "step_g": step_g, "single_ev": single_ev})

    cum = 0.0
    for r in reversed(rows):
        if r is rows[-1]:
            r["cum_ev"] = r["single_ev"]
        else:
            r["cum_ev"] = r["rate"] * r["avg_np"] + (1 - r["rate"]) * cum - r["step_g"] * invest_per_g
        cum = r["cum_ev"]

    sum_step_g = 0.0
    for r in reversed(rows):
        sum_step_g += r["step_g"]
        r["sum_invest"] = sum_step_g * invest_per_g
    for r in rows:
        r["kikai"] = (r["cum_ev"] + r["sum_invest"]) / r["sum_invest"] * 100 if r["sum_invest"] else 0.0

    for r_idx, r in enumerate(rows, start=2):
        _write_row(ws, r_idx, [r["lbl"], r["n_total"], r["n_success"], r["rate"],
                                r["avg_np"], r["avg_g"], r["step_g"],
                                r["single_ev"], r["single_ev"] * YEN_PER_COIN,
                                r["cum_ev"], r["cum_ev"] * YEN_PER_COIN, r["kikai"]])
        ws.cell(row=r_idx, column=4).number_format = "0.00%"
        ws.cell(row=r_idx, column=12).number_format = "0.00"


# ============================================================================
# 5G刻み 期待値表 (CZイベント単位 / AT連荘単位)
# ============================================================================

G_BIN_SIZE = 5
G_BIN_MAX = 1580  # 許容範囲上限 (天井1499G + マージン)


def _g_bin_labels(bin_size: int = G_BIN_SIZE, max_g: int = G_BIN_MAX) -> list[str]:
    """ビンの開始G (打ち始め地点)。例: ['0G', '5G', '10G', ..., '1575G']"""
    return [f"{n}G" for n in range(0, max_g, bin_size)]


def _g_bin_starts(bin_size: int = G_BIN_SIZE, max_g: int = G_BIN_MAX) -> list[int]:
    return list(range(0, max_g, bin_size))


def build_g_bin_signal_formula_sheet(
    wb: Workbook,
    sheet_name: str,
    *,
    invest_per_g: float,
    at_payout_per_g: float = 8.2,  # 未使用 (互換)
    data_sheet: str = "Data",
    chain_sheet: str = "ChainData",
    signal_col: str,                    # Data側: 信号フラグ (Is_CZ)
    success_col: str,                   # Data側: 信号成功フラグ (Is_CZ_Success)
    signal_g_col: str,                  # Data側: 信号自身の累積G (CZ_Cum_G_At_Hit)
    chain_signal_g_col: str,            # ChainData側: 起点信号の累積G (starting_cz_cum_g)
    chain_payout_col: str = "net_payout",
    target_through_col: str | None = None,
    target_through_value: int | None = None,
    chain_through_col: str | None = None,
    warmup_g: int = 0,
    cumulative_label: str = "累積期待値",  # 例: "CZ1回目当選まで累積期待値"
    bin_size: int = G_BIN_SIZE,
    max_g: int = G_BIN_MAX,
) -> None:
    """CZイベント単位 5G刻み期待値表 (Excel数式)。

    各5GビンN:
      件数      = COUNTIFS(Data!signal, TRUE, Data!signal_g, ">=N", "<N+5")
      成功件数  = COUNTIFS(Data!success, TRUE, Data!signal_g, ">=N", "<N+5")
      成功率    = 成功件数/件数
      平均TY    = AVERAGEIFS(ChainData!net_payout, ChainData!chain_signal_g, ">=N", "<N+5")
      CZ間狙いEV = 成功率×TY - 5G×純減
      AT間狙いEV = 再帰 (下行参照)
    """
    ws = wb.create_sheet(sheet_name)
    # 設定セル (純減/G, 円/枚, AT純増/G, 掛枚) を上部に置いて全数式から $絶対参照
    ws.cell(row=1, column=16, value="純減/G").font = HEADER_FONT
    ws.cell(row=1, column=17, value=invest_per_g)
    ws.cell(row=1, column=18, value="円/枚").font = HEADER_FONT
    ws.cell(row=1, column=19, value=YEN_PER_COIN)
    ws.cell(row=1, column=20, value="AT純増/G").font = HEADER_FONT
    ws.cell(row=1, column=21, value=at_payout_per_g)
    ws.cell(row=1, column=22, value="掛枚").font = HEADER_FONT
    ws.cell(row=1, column=23, value=BET_PER_GAME)

    _write_header(ws, ["打ち始めG", "サンプル数", "当選数", "連チャンTYサンプル数", "区間当選率",
                       "区間連チャンTY(枚)", "累積連チャンTY(枚)", "Nから平均消化G",
                       "増分G",
                       "区間期待値(枚)", "区間期待値(円)",
                       f"{cumulative_label}(枚)", f"{cumulative_label}(円)",
                       f"{cumulative_label}_機械割(%)"])

    data_ws = wb[data_sheet]
    chain_ws = wb[chain_sheet]
    data_cols = _col_map(data_ws)
    chain_cols = _col_map(chain_ws)
    needed = {signal_col: data_cols, success_col: data_cols, signal_g_col: data_cols,
              chain_signal_g_col: chain_cols, chain_payout_col: chain_cols}
    missing = [c for c, src in needed.items() if c not in src]
    if missing:
        ws.cell(row=2, column=1, value=f"(必要列が見つかりません: {missing})")
        return

    sig_l = data_cols[signal_col]
    suc_l = data_cols[success_col]
    sg_l = data_cols[signal_g_col]
    csg_l = chain_cols[chain_signal_g_col]
    cp_l = chain_cols[chain_payout_col]
    data_ref = f"'{data_sheet}'"
    chain_ref = f"'{chain_sheet}'"

    # M回目絞り込み (任意)
    through_filter_data = ""
    through_filter_chain = ""
    if target_through_col and target_through_value is not None and target_through_col in data_cols:
        tt_l = data_cols[target_through_col]
        through_filter_data = f',{data_ref}!{tt_l}:{tt_l},"={target_through_value}"'
    if chain_through_col and target_through_value is not None and chain_through_col in chain_cols:
        ct_l = chain_cols[chain_through_col]
        through_filter_chain = f',{chain_ref}!{ct_l}:{ct_l},"={target_through_value}"'

    starts = _g_bin_starts(bin_size, max_g)
    labels = _g_bin_labels(bin_size, max_g)
    n_rows = len(starts)

    invest_ref = "$Q$1"  # 純減/G
    yen_ref = "$S$1"     # 円/枚
    at_payout_ref = "$U$1"  # AT純増/G
    bet_ref = "$W$1"        # 掛枚
    ws.cell(row=1, column=15).value = None  # 旧定数セル位置クリア
    ws.cell(row=1, column=16, value="純減/G").font = HEADER_FONT
    ws.cell(row=1, column=17, value=invest_per_g)
    ws.cell(row=1, column=18, value="円/枚").font = HEADER_FONT
    ws.cell(row=1, column=19, value=YEN_PER_COIN)

    for idx, (n_start, lbl) in enumerate(zip(starts, labels)):
        r_idx = idx + 2
        # 着席ワームアップ: N+warmup_g 以降のサンプルのみ集計
        eff_low = n_start + warmup_g
        n_end = eff_low + bin_size
        lower = f'">="&{eff_low}'
        upper = f'"<"&{n_end}'
        b = (f'=COUNTIFS({data_ref}!{sig_l}:{sig_l},TRUE,'
             f'{data_ref}!{sg_l}:{sg_l},{lower}{through_filter_data})')
        c = (f'=COUNTIFS({data_ref}!{suc_l}:{suc_l},TRUE,'
             f'{data_ref}!{sg_l}:{sg_l},{lower},{data_ref}!{sg_l}:{sg_l},{upper}{through_filter_data})')
        d_ty = (f'=COUNTIFS({chain_ref}!{csg_l}:{csg_l},{lower},'
                f'{chain_ref}!{csg_l}:{csg_l},{upper}{through_filter_chain})')
        e_rate = f"=IFERROR(C{r_idx}/B{r_idx},0)"
        # F: 区間連チャンTY
        f_int_ty = (f'=IFERROR(AVERAGEIFS({chain_ref}!{cp_l}:{cp_l},'
                    f'{chain_ref}!{csg_l}:{csg_l},{lower},{chain_ref}!{csg_l}:{csg_l},{upper}{through_filter_chain}),0)')
        # G: 累積連チャンTY
        g_cum_ty = (f'=IFERROR(AVERAGEIFS({chain_ref}!{cp_l}:{cp_l},'
                    f'{chain_ref}!{csg_l}:{csg_l},{lower}{through_filter_chain}),0)')
        # H: Nから平均消化G (= 絶対G平均 - 打ち始めN)
        h_avg_g = (f'=IFERROR(AVERAGEIFS({chain_ref}!{csg_l}:{csg_l},'
                   f'{chain_ref}!{csg_l}:{csg_l},{lower}{through_filter_chain})-{n_start},0)')
        i_step = f"={bin_size}"
        # J: 区間期待値(枚) = 区間当選率×区間連チャンTY - 増分G×純減
        j_ev = f"=E{r_idx}*F{r_idx}-I{r_idx}*{invest_ref}"
        k_yen = f"=J{r_idx}*{yen_ref}"
        # L: 累積期待値 (ユーザー仕様: シンプル式 = 累積TY - Nから平均消化G × 純減)
        # 「Nから打ち始めて必ずAT当選する」前提、回収-コストの直感モデル
        l_cum = f"=G{r_idx}-H{r_idx}*{invest_ref}"
        m_yen = f"=L{r_idx}*{yen_ref}"
        # N: 機械割 = 100 + EV / 期待総投資 × 100
        # 期待総投資 = (Nから平均消化G + 累積TY/AT純増) × 掛枚
        n_kik = (f"=IFERROR(100+L{r_idx}/"
                 f"((H{r_idx}+G{r_idx}/{at_payout_ref})*{bet_ref})*100,0)")
        _write_row(ws, r_idx, [lbl, b, c, d_ty, e_rate, f_int_ty, g_cum_ty, h_avg_g,
                                i_step, j_ev, k_yen, l_cum, m_yen, n_kik])
        ws.cell(row=r_idx, column=5).number_format = "0.00%"
        for col_idx in (6, 7, 8, 9, 10, 12):
            ws.cell(row=r_idx, column=col_idx).number_format = "0.0"
        for col_idx in (11, 13):
            ws.cell(row=r_idx, column=col_idx).number_format = "0"
        ws.cell(row=r_idx, column=14).number_format = "0.00"

    # 異常データ注記 (1580+)
    note_row = n_rows + 2
    ws.cell(row=note_row, column=1, value=f"{max_g}G+(異常)").font = HEADER_FONT
    ws.cell(row=note_row, column=2,
            value=(f'=COUNTIFS({data_ref}!{sig_l}:{sig_l},TRUE,'
                   f'{data_ref}!{sg_l}:{sg_l},">="&{max_g}{through_filter_data})'))


def build_g_bin_signal_preview_sheet(
    wb: Workbook,
    sheet_name: str,
    data_df: pd.DataFrame,
    chains_df: pd.DataFrame,
    *,
    invest_per_g: float,
    at_payout_per_g: float = 8.2,  # AT中純増 (枚/G)、機械割計算でAT_G逆算に使用
    signal_col: str,
    success_col: str,
    signal_g_col: str,
    chain_signal_g_col: str,
    chain_payout_col: str = "net_payout",
    target_through_col: str | None = None,
    target_through_value: int | None = None,
    chain_through_col: str | None = None,
    warmup_g: int = 0,
    cumulative_label: str = "累積期待値",
    bin_size: int = G_BIN_SIZE,
    max_g: int = G_BIN_MAX,
    sheet_suffix: str = "_値",
) -> None:
    """pandas 計算プレビュー (Streamlit 用)。"""
    ws = wb.create_sheet(sheet_name + sheet_suffix)
    _write_header(ws, ["打ち始めG", "サンプル数", "当選数", "連チャンTYサンプル数", "区間当選率",
                       "区間連チャンTY(枚)", "累積連チャンTY(枚)", "Nから平均消化G",
                       "増分G",
                       "区間期待値(枚)", "区間期待値(円)",
                       f"{cumulative_label}(枚)", f"{cumulative_label}(円)",
                       f"{cumulative_label}_機械割(%)"])
    starts = _g_bin_starts(bin_size, max_g)
    labels = _g_bin_labels(bin_size, max_g)

    sig_mask = data_df[signal_col].astype(bool)
    suc_mask = data_df[success_col].astype(bool)
    if target_through_col and target_through_value is not None and target_through_col in data_df.columns:
        sig_mask = sig_mask & (data_df[target_through_col] == target_through_value)
        suc_mask = suc_mask & (data_df[target_through_col] == target_through_value)
    chain_extra_mask = None
    if chain_through_col and target_through_value is not None and chain_through_col in chains_df.columns:
        chain_extra_mask = (chains_df[chain_through_col] == target_through_value)

    rows = []
    for n_start, lbl in zip(starts, labels):
        eff_low = n_start + warmup_g
        trial_mask = (data_df[signal_g_col] >= eff_low)
        n_trial = int((sig_mask & trial_mask).sum())
        g_mask = (data_df[signal_g_col] >= eff_low) & (data_df[signal_g_col] < eff_low + bin_size)
        n_success = int((suc_mask & g_mask).sum())
        rate = n_success / n_trial if n_trial else 0.0
        cg = chains_df[chain_signal_g_col]
        chain_mask = (cg >= eff_low) & (cg < eff_low + bin_size)
        if chain_extra_mask is not None:
            chain_mask = chain_mask & chain_extra_mask
        n_ty_sample = int(chain_mask.sum())
        avg_ty = float(chains_df.loc[chain_mask, chain_payout_col].mean()) if chain_mask.any() else 0.0
        # 累積平均TY (eff_low以降の全当選TY平均)
        cum_mask = (cg >= eff_low)
        if chain_extra_mask is not None:
            cum_mask = cum_mask & chain_extra_mask
        cum_avg_ty = float(chains_df.loc[cum_mask, chain_payout_col].mean()) if cum_mask.any() else 0.0
        cum_avg_g_abs = float(chains_df.loc[cum_mask, chain_signal_g_col].mean()) if cum_mask.any() else 0.0
        cum_avg_g = max(0.0, cum_avg_g_abs - n_start)  # 打ち始めNを引いた相対G
        step_g = bin_size
        single_ev = rate * avg_ty - step_g * invest_per_g
        rows.append({"lbl": lbl, "n_trial": n_trial, "n_success": n_success,
                     "n_ty_sample": n_ty_sample,
                     "rate": rate, "avg_ty": avg_ty, "cum_avg_ty": cum_avg_ty,
                     "cum_avg_g": cum_avg_g,
                     "step_g": step_g, "single_ev": single_ev})

    cum = 0.0
    for r in reversed(rows):
        if r is rows[-1]:
            r["cum_ev"] = r["single_ev"]
        else:
            r["cum_ev"] = r["rate"] * r["avg_ty"] + (1 - r["rate"]) * cum - r["step_g"] * invest_per_g
        cum = r["cum_ev"]

    # 期待投資累積 (shuuki式再帰、EV計算用)
    inv_cum = [0.0] * len(rows)
    inv_cum[-1] = rows[-1]["step_g"] * invest_per_g
    for i in range(len(rows)-2, -1, -1):
        inv_cum[i] = rows[i]["step_g"] * invest_per_g + (1 - rows[i]["rate"]) * inv_cum[i+1]

    # 機械割 (cum_ev と整合する期待総投資ベース、3枚掛け)
    # EV (cum_ev, 純減ベース) = E[TY獲得] - E[通常G×純減]
    # → E[TY獲得] = cum_ev + inv_cum (=期待通常G×純減)
    # E[通常G]   = inv_cum / 純減
    # E[AT_G]    = E[TY獲得] / 純増
    # 期待総投資 = (E[通常G] + E[AT_G]) × 3
    # 機械割    = 100 + cum_ev / 期待総投資 × 100  ※EV<0 ⇔ 機械割<100% 保証
    for i, r in enumerate(rows):
        r["sum_invest"] = inv_cum[i]
        normal_g_exp = inv_cum[i] / invest_per_g if invest_per_g > 0 else 0
        ty_exp = r["cum_ev"] + inv_cum[i]  # 期待TY獲得 (枚)
        at_g_exp = ty_exp / at_payout_per_g if at_payout_per_g > 0 else 0
        total_invest_exp = (normal_g_exp + at_g_exp) * BET_PER_GAME
        r["kikai"] = 100 + r["cum_ev"] / total_invest_exp * 100 if total_invest_exp > 0 else 0.0

    for r_idx, r in enumerate(rows, start=2):
        _write_row(ws, r_idx, [r["lbl"], r["n_trial"], r["n_success"], r["n_ty_sample"],
                                r["rate"], r["avg_ty"], r["cum_avg_ty"], r["cum_avg_g"],
                                r["step_g"],
                                r["single_ev"], r["single_ev"] * YEN_PER_COIN,
                                r["cum_ev"], r["cum_ev"] * YEN_PER_COIN, r["kikai"]])
        ws.cell(row=r_idx, column=5).number_format = "0.00%"
        ws.cell(row=r_idx, column=14).number_format = "0.00"

    n_over = int((sig_mask & (data_df[signal_g_col] >= max_g)).sum())
    note_row = len(rows) + 3
    ws.cell(row=note_row, column=1, value=f"{max_g}G+(異常)")
    ws.cell(row=note_row, column=2, value=n_over)


def build_g_bin_chain_formula_sheet(
    wb: Workbook,
    sheet_name: str,
    *,
    invest_per_g: float,
    at_payout_per_g: float = 8.2,
    chain_sheet: str = "ChainData",
    chain_g_col: str = "pre_normal_g",
    chain_payout_col: str = "net_payout",
    chain_through_max_col: str | None = None,
    chain_through_max_value: int | None = None,  # 例 2 → <= 2
    chain_through_min_col: str | None = None,
    chain_through_min_value: int | None = None,  # 例 2 → >= 2 (N回目天井狙い用)
    chain_filter_col: str | None = None,  # 任意列 (例: "chain_number")
    chain_filter_min: int | None = None,
    chain_filter_max: int | None = None,
    chain_filter2_col: str | None = None,  # 2軸目フィルタ (例: "yuri_diff_pre")
    chain_filter2_min: int | None = None,
    chain_filter2_max: int | None = None,
    chain_exclude_col: str | None = None,  # 除外フィルタ (例: first_state in [引き戻し])
    chain_exclude_values: list | None = None,
    warmup_g: int = 0,  # 着席から warmup_g G未満は当選なし扱い
    cumulative_label: str = "AT到達まで累積期待値",
    bin_size: int = G_BIN_SIZE,
    max_g: int = G_BIN_MAX,
) -> None:
    """AT連荘単位 5G刻み期待値表 (ハマリG狙い、Excel数式)。

    各5GビンN:
      AT初当り件数 = COUNTIFS(ChainData!pre_normal_g, ">=N", "<N+5")
      残件数(N以降) = COUNTIFS(ChainData!pre_normal_g, ">=N")
      AT当選率(p)  = 件数 / 残件数  (5G離散ハザード)
      平均TY       = AVERAGEIFS(ChainData!net_payout, ChainData!pre_normal_g, ">=N", "<N+5")
      CZ間狙いEV   = p×TY - 5G×純減
      AT間狙いEV   = 再帰
    """
    ws = wb.create_sheet(sheet_name)
    ws.cell(row=1, column=14, value="純減/G").font = HEADER_FONT
    ws.cell(row=1, column=15, value=invest_per_g)
    ws.cell(row=1, column=16, value="円/枚").font = HEADER_FONT
    ws.cell(row=1, column=17, value=YEN_PER_COIN)
    _write_header(ws, ["打ち始めG", "サンプル数", "当選数", "連チャンTYサンプル数", "区間当選率",
                       "区間連チャンTY(枚)", "累積連チャンTY(枚)", "Nから平均消化G",
                       "増分G",
                       "区間期待値(枚)", "区間期待値(円)",
                       f"{cumulative_label}(枚)", f"{cumulative_label}(円)",
                       f"{cumulative_label}_機械割(%)"])
    chain_ws = wb[chain_sheet]
    chain_cols = _col_map(chain_ws)
    if chain_g_col not in chain_cols or chain_payout_col not in chain_cols:
        ws.cell(row=2, column=1, value=f"(必要列が見つかりません)")
        return

    cg_l = chain_cols[chain_g_col]
    cp_l = chain_cols[chain_payout_col]
    chain_ref = f"'{chain_sheet}'"
    ws.cell(row=1, column=15).value = None
    ws.cell(row=1, column=16, value="純減/G").font = HEADER_FONT
    ws.cell(row=1, column=17, value=invest_per_g)
    ws.cell(row=1, column=18, value="円/枚").font = HEADER_FONT
    ws.cell(row=1, column=19, value=YEN_PER_COIN)
    ws.cell(row=1, column=20, value="AT純増/G").font = HEADER_FONT
    ws.cell(row=1, column=21, value=at_payout_per_g)
    ws.cell(row=1, column=22, value="掛枚").font = HEADER_FONT
    ws.cell(row=1, column=23, value=BET_PER_GAME)
    invest_ref = "$Q$1"
    yen_ref = "$S$1"
    at_payout_ref = "$U$1"
    bet_ref = "$W$1"

    # フィルタ用 (≤ または ≥、≥は直撃=0も含む)
    # 注: Excel式は AND のみなので「≥N OR =0」はサンプル件数のみ COUNTIFS で表現困難。
    # 簡易: ≥N のみフィルタ (Excel formula版では直撃は別途加算が必要だが、
    # 現状はpreview版を主に使うため formula版は ≥N の単純フィルタで近似)
    through_filter = ""
    if (chain_through_max_col and chain_through_max_value is not None
            and chain_through_max_col in chain_cols):
        tm_l = chain_cols[chain_through_max_col]
        through_filter = f',{chain_ref}!{tm_l}:{tm_l},"<={chain_through_max_value}"'
    if (chain_through_min_col and chain_through_min_value is not None
            and chain_through_min_col in chain_cols):
        tm_l = chain_cols[chain_through_min_col]
        through_filter += f',{chain_ref}!{tm_l}:{tm_l},">={chain_through_min_value}"'
    if chain_filter_col and chain_filter_col in chain_cols:
        cf_l = chain_cols[chain_filter_col]
        if chain_filter_min is not None:
            through_filter += f',{chain_ref}!{cf_l}:{cf_l},">={chain_filter_min}"'
        if chain_filter_max is not None:
            through_filter += f',{chain_ref}!{cf_l}:{cf_l},"<={chain_filter_max}"'
    if chain_filter2_col and chain_filter2_col in chain_cols:
        cf2_l = chain_cols[chain_filter2_col]
        if chain_filter2_min is not None:
            through_filter += f',{chain_ref}!{cf2_l}:{cf2_l},">={chain_filter2_min}"'
        if chain_filter2_max is not None:
            through_filter += f',{chain_ref}!{cf2_l}:{cf2_l},"<={chain_filter2_max}"'
    # chain_exclude (例: first_state ≠ 引き戻し)
    if chain_exclude_col and chain_exclude_values and chain_exclude_col in chain_cols:
        ex_l = chain_cols[chain_exclude_col]
        for v in chain_exclude_values:
            through_filter += f',{chain_ref}!{ex_l}:{ex_l},"<>{v}"'

    starts = _g_bin_starts(bin_size, max_g)
    labels = _g_bin_labels(bin_size, max_g)
    n_rows = len(starts)

    for idx, (n_start, lbl) in enumerate(zip(starts, labels)):
        r_idx = idx + 2
        # 着席ワームアップ: N+warmup_g 以降のサンプルのみ集計
        eff_low = n_start + warmup_g
        n_end = eff_low + bin_size
        lower = f'">="&{eff_low}'
        upper = f'"<"&{n_end}'
        b = f'=COUNTIFS({chain_ref}!{cg_l}:{cg_l},{lower}{through_filter})'
        c = (f'=COUNTIFS({chain_ref}!{cg_l}:{cg_l},{lower},'
             f'{chain_ref}!{cg_l}:{cg_l},{upper}{through_filter})')
        d_ty = (f'=COUNTIFS({chain_ref}!{cg_l}:{cg_l},{lower},'
                f'{chain_ref}!{cg_l}:{cg_l},{upper}{through_filter})')
        e_rate = f"=IFERROR(C{r_idx}/B{r_idx},0)"
        # F: 区間平均TY (再帰計算用、中間値)。through_filter を AVERAGEIFS にも適用
        f_int_ty = (f'=IFERROR(AVERAGEIFS({chain_ref}!{cp_l}:{cp_l},'
                    f'{chain_ref}!{cg_l}:{cg_l},{lower},'
                    f'{chain_ref}!{cg_l}:{cg_l},{upper}{through_filter}),0)')
        # G: 累積平均TY (N以降、フィルタ後)
        g_cum_ty = (f'=IFERROR(AVERAGEIFS({chain_ref}!{cp_l}:{cp_l},'
                    f'{chain_ref}!{cg_l}:{cg_l},{lower}{through_filter}),0)')
        # H: Nから平均消化G (= 絶対G平均 - 打ち始めN、フィルタ後)
        h_avg_g = (f'=IFERROR(AVERAGEIFS({chain_ref}!{cg_l}:{cg_l},'
                   f'{chain_ref}!{cg_l}:{cg_l},{lower}{through_filter})-{n_start},0)')
        i_step = f"={bin_size}"
        # J: 区間期待値
        j_ev = f"=E{r_idx}*F{r_idx}-I{r_idx}*{invest_ref}"
        k_yen = f"=J{r_idx}*{yen_ref}"
        # L: 累積期待値 (ユーザー仕様: シンプル式 = 累積TY - Nから平均消化G × 純減)
        # 「Nから打ち始めて必ずAT当選する」前提、回収-コストの直感モデル
        l_cum = f"=G{r_idx}-H{r_idx}*{invest_ref}"
        m_yen = f"=L{r_idx}*{yen_ref}"
        # N: 機械割 = 100 + EV / 期待総投資 × 100
        # 期待総投資 = (Nから平均消化G + 累積TY/AT純増) × 掛枚
        n_kik = (f"=IFERROR(100+L{r_idx}/"
                 f"((H{r_idx}+G{r_idx}/{at_payout_ref})*{bet_ref})*100,0)")
        _write_row(ws, r_idx, [lbl, b, c, d_ty, e_rate, f_int_ty, g_cum_ty, h_avg_g,
                                i_step, j_ev, k_yen, l_cum, m_yen, n_kik])
        ws.cell(row=r_idx, column=5).number_format = "0.00%"
        for col_idx in (6, 7, 8, 9, 10, 12):
            ws.cell(row=r_idx, column=col_idx).number_format = "0.0"
        for col_idx in (11, 13):
            ws.cell(row=r_idx, column=col_idx).number_format = "0"
        ws.cell(row=r_idx, column=14).number_format = "0.00"

    note_row = n_rows + 2
    ws.cell(row=note_row, column=1, value=f"{max_g}G+(異常)").font = HEADER_FONT
    ws.cell(row=note_row, column=2,
            value=f'=COUNTIFS({chain_ref}!{cg_l}:{cg_l},">="&{max_g})')


def build_g_bin_chain_preview_sheet(
    wb: Workbook,
    sheet_name: str,
    chains_df: pd.DataFrame,
    *,
    invest_per_g: float,
    at_payout_per_g: float = 8.2,
    chain_g_col: str = "pre_normal_g",
    chain_payout_col: str = "net_payout",
    chain_through_max_col: str | None = None,
    chain_through_max_value: int | None = None,
    chain_through_min_col: str | None = None,
    chain_through_min_value: int | None = None,
    chain_filter_col: str | None = None,
    chain_filter_min: int | None = None,
    chain_filter_max: int | None = None,
    chain_filter2_col: str | None = None,
    chain_filter2_min: int | None = None,
    chain_filter2_max: int | None = None,
    chain_exclude_col: str | None = None,
    chain_exclude_values: list | None = None,
    warmup_g: int = 0,  # 着席から warmup_g G未満は当選なし扱い
    cumulative_label: str = "AT到達まで累積期待値",
    bin_size: int = G_BIN_SIZE,
    max_g: int = G_BIN_MAX,
    sheet_suffix: str = "_値",
) -> None:
    """pandas プレビュー (Streamlit 用)。"""
    ws = wb.create_sheet(sheet_name + sheet_suffix)
    _write_header(ws, ["打ち始めG", "サンプル数", "当選数", "連チャンTYサンプル数", "区間当選率",
                       "区間連チャンTY(枚)", "累積連チャンTY(枚)", "Nから平均消化G",
                       "増分G",
                       "区間期待値(枚)", "区間期待値(円)",
                       f"{cumulative_label}(枚)", f"{cumulative_label}(円)",
                       f"{cumulative_label}_機械割(%)"])
    _write_filter_description(ws, _build_filter_description(
        chain_through_min_col=chain_through_min_col,
        chain_through_min_value=chain_through_min_value,
        chain_through_max_col=chain_through_max_col,
        chain_through_max_value=chain_through_max_value,
        chain_filter_col=chain_filter_col,
        chain_filter_min=chain_filter_min,
        chain_filter_max=chain_filter_max,
        chain_filter2_col=chain_filter2_col,
        chain_filter2_min=chain_filter2_min,
        chain_filter2_max=chain_filter2_max,
        chain_exclude_col=chain_exclude_col,
        chain_exclude_values=chain_exclude_values,
        warmup_g=warmup_g,
    ))
    starts = _g_bin_starts(bin_size, max_g)
    labels = _g_bin_labels(bin_size, max_g)

    # フィルタ (試行/当選/TY 全てフィルタ後、≤ または ≥)
    df_use = chains_df
    if (chain_through_max_col and chain_through_max_value is not None
            and chain_through_max_col in df_use.columns):
        df_use = df_use[df_use[chain_through_max_col] <= chain_through_max_value]
    if (chain_through_min_col and chain_through_min_value is not None
            and chain_through_min_col in df_use.columns):
        # >=N または 直撃 (=0、CZ未経由ATも含む)
        df_use = df_use[(df_use[chain_through_min_col] >= chain_through_min_value)
                        | (df_use[chain_through_min_col] == 0)]
    if chain_filter_col and chain_filter_col in df_use.columns:
        if chain_filter_min is not None:
            df_use = df_use[df_use[chain_filter_col] >= chain_filter_min]
        if chain_filter_max is not None:
            df_use = df_use[df_use[chain_filter_col] <= chain_filter_max]
    if chain_filter2_col and chain_filter2_col in df_use.columns:
        if chain_filter2_min is not None:
            df_use = df_use[df_use[chain_filter2_col] >= chain_filter2_min]
        if chain_filter2_max is not None:
            df_use = df_use[df_use[chain_filter2_col] <= chain_filter2_max]
    if chain_exclude_col and chain_exclude_values and chain_exclude_col in df_use.columns:
        df_use = df_use[~df_use[chain_exclude_col].isin(chain_exclude_values)]

    g = df_use[chain_g_col]
    rows = []
    for n_start, lbl in zip(starts, labels):
        # 着席ワームアップ: N+warmup_g 以降のサンプルのみ集計
        eff_low = n_start + warmup_g
        in_mask = (g >= eff_low) & (g < eff_low + bin_size)
        rest_mask = g >= eff_low
        n_in = int(in_mask.sum())
        n_rest = int(rest_mask.sum())
        p = n_in / n_rest if n_rest else 0.0
        avg_ty = float(df_use.loc[in_mask, chain_payout_col].mean()) if n_in else 0.0
        cum_avg_ty = float(df_use.loc[rest_mask, chain_payout_col].mean()) if n_rest else 0.0
        cum_avg_g_abs = float(df_use.loc[rest_mask, chain_g_col].mean()) if n_rest else 0.0
        cum_avg_g = max(0.0, cum_avg_g_abs - n_start)  # Nから平均消化G (warmup_g以上)
        step_g = bin_size
        # ゾーン狙いEV (= single_ev、その区間だけ打って打ち切り) は区間TY を使う
        # AT到達まで累積期待値 (= cum_ev、打ち続けて当選まで) は累積TY を使う (ユーザー仕様)
        single_ev = p * avg_ty - step_g * invest_per_g
        rows.append({"lbl": lbl, "n_trial": n_rest, "n_win": n_in, "p": p,
                     "avg_ty": avg_ty, "cum_avg_ty": cum_avg_ty, "cum_avg_g": cum_avg_g,
                     "step_g": step_g, "single_ev": single_ev})

    # 累積期待値 (AT到達まで打ち続けるモデル): シンプル式
    # EV = 累積TY - Nから平均消化G × 純減 (= 回収 - コスト、ユーザー仕様)
    # 「Nから打ち始めて必ず当選する」前提
    for r in rows:
        r["cum_ev"] = r["cum_avg_ty"] - r["cum_avg_g"] * invest_per_g

    # 機械割 (シンプル式 cum_ev 整合): 通常G + AT中G を3枚掛けで投資合算
    # 通常G = cum_avg_g (Nから平均消化G)、AT中G = cum_avg_ty / 純増レート
    for r in rows:
        normal_g_exp = r["cum_avg_g"]
        at_g_exp = r["cum_avg_ty"] / at_payout_per_g if at_payout_per_g > 0 else 0
        total_invest_exp = (normal_g_exp + at_g_exp) * BET_PER_GAME
        r["kikai"] = 100 + r["cum_ev"] / total_invest_exp * 100 if total_invest_exp > 0 else 0.0

    for r_idx, r in enumerate(rows, start=2):
        _write_row(ws, r_idx, [r["lbl"], r["n_trial"], r["n_win"], r["n_win"],
                                r["p"], r["avg_ty"], r["cum_avg_ty"], r["cum_avg_g"],
                                r["step_g"],
                                r["single_ev"], r["single_ev"] * YEN_PER_COIN,
                                r["cum_ev"], r["cum_ev"] * YEN_PER_COIN, r["kikai"]])
        ws.cell(row=r_idx, column=5).number_format = "0.00%"
        ws.cell(row=r_idx, column=14).number_format = "0.00"

    n_over = int((g >= max_g).sum())
    note_row = len(rows) + 3
    ws.cell(row=note_row, column=1, value=f"{max_g}G+(異常)")
    ws.cell(row=note_row, column=2, value=n_over)


# === 累積G別 (500G刻み) AT間天井狙い (preview版のみ) ===

CYCLE_BIN_DEFAULT = [(i * 1000, i * 1000 + 999) for i in range(7)] + [(7000, 10**9)]


def _write_cycle_block(ws, df_use, chain_cycle_col, chain_g_col, chain_payout_col,
                       invest_per_g, at_payout_per_g, bins, start_row, title):
    """1ブロック (タイトル行+ヘッダ+データ) を書く。次の使用可能行を返す。"""
    ws.cell(row=start_row, column=1, value=title).font = HEADER_FONT
    header_row = start_row + 1
    for c_idx, h in enumerate(["累積G", "サンプル数", "平均連チャンTY(枚)", "平均消化G",
                                "期待値(枚)", "期待値(円)", "機械割(%)"], start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_ALL
    if chain_cycle_col not in df_use.columns:
        ws.cell(row=header_row + 1, column=1, value=f"(列なし: {chain_cycle_col})")
        return header_row + 2
    for i, (low, high) in enumerate(bins):
        r_idx = header_row + 1 + i
        if high >= 10**8:
            mask = df_use[chain_cycle_col] >= low
            label = f"{low}G以上"
        else:
            mask = (df_use[chain_cycle_col] >= low) & (df_use[chain_cycle_col] <= high)
            label = f"{low}-{high}G"
        sub = df_use[mask]
        n = len(sub)
        avg_ty = float(sub[chain_payout_col].mean()) if n else 0.0
        avg_g = float(sub[chain_g_col].mean()) if n else 0.0
        ev = avg_ty - avg_g * invest_per_g
        at_g_exp = avg_ty / at_payout_per_g if at_payout_per_g else 0
        total_invest = (avg_g + at_g_exp) * BET_PER_GAME
        kikai = 100 + ev / total_invest * 100 if total_invest > 0 else 0.0
        _write_row(ws, r_idx, [label, n, avg_ty, avg_g,
                                ev, ev * YEN_PER_COIN, kikai])
        for col_idx in (3, 4, 5):
            ws.cell(row=r_idx, column=col_idx).number_format = "0.0"
        ws.cell(row=r_idx, column=6).number_format = "0"
        ws.cell(row=r_idx, column=7).number_format = "0.00"
    return header_row + 1 + len(bins)


def build_cycle_bin_chain_preview_sheet(
    wb: Workbook,
    sheet_name: str,
    chains_df: pd.DataFrame,
    *,
    invest_per_g: float,
    at_payout_per_g: float = 8.2,
    chain_cycle_col: str = "cycle_g_pre",
    chain_cycle_total_col: str = "cycle_total_g_pre",
    chain_g_col: str = "pre_normal_g",
    chain_payout_col: str = "net_payout",
    chain_filter_col: str | None = None,
    chain_filter_min: int | None = None,
    chain_filter_max: int | None = None,
    chain_exclude_col: str | None = None,
    chain_exclude_values: list | None = None,
    warmup_g: int = 0,
    bins: list[tuple[int, int]] | None = None,
    sheet_suffix: str = "_値",
) -> None:
    """有利区間内累積G別 AT間天井狙い (500G刻み、各ビン独立EV)。

    上ブロック: cycle_g_pre (通常G累積) ベース
    下ブロック: cycle_total_g_pre (全G累積、通常G+AT中G) ベース 比較表
    """
    if bins is None:
        bins = CYCLE_BIN_DEFAULT
    ws = wb.create_sheet(sheet_name + sheet_suffix)

    df_use = chains_df
    if chain_filter_col and chain_filter_col in df_use.columns:
        if chain_filter_min is not None:
            df_use = df_use[df_use[chain_filter_col] >= chain_filter_min]
        if chain_filter_max is not None:
            df_use = df_use[df_use[chain_filter_col] <= chain_filter_max]
    if chain_exclude_col and chain_exclude_values and chain_exclude_col in df_use.columns:
        df_use = df_use[~df_use[chain_exclude_col].isin(chain_exclude_values)]
    # 着席ワームアップ: 連荘起点から warmup_g 未満で当選した連荘を除外
    if warmup_g > 0:
        df_use = df_use[df_use[chain_g_col] >= warmup_g]

    _write_filter_description(ws, _build_filter_description(
        chain_filter_col=chain_filter_col,
        chain_filter_min=chain_filter_min,
        chain_filter_max=chain_filter_max,
        chain_exclude_col=chain_exclude_col,
        chain_exclude_values=chain_exclude_values,
        warmup_g=warmup_g,
    ))

    # 上ブロック: 前回通常累積G数別
    next_row = _write_cycle_block(
        ws, df_use, chain_cycle_col, chain_g_col, chain_payout_col,
        invest_per_g, at_payout_per_g, bins,
        start_row=1, title="【前回通常累積G数別】"
    )
    # 空行 + 下ブロック: 直前有利区間G数別
    _write_cycle_block(
        ws, df_use, chain_cycle_total_col, chain_g_col, chain_payout_col,
        invest_per_g, at_payout_per_g, bins,
        start_row=next_row + 2,
        title="【直前有利区間G数別】"
    )


def build_cycle_bin_chain_formula_sheet(
    wb: Workbook,
    sheet_name: str,
    *,
    invest_per_g: float,
    at_payout_per_g: float = 8.2,
    chain_sheet: str = "ChainData",
    chain_cycle_col: str = "cycle_g_pre",
    chain_cycle_total_col: str = "cycle_total_g_pre",
    chain_g_col: str = "pre_normal_g",
    chain_payout_col: str = "net_payout",
    chain_filter_col: str | None = None,
    chain_filter_min: int | None = None,
    chain_filter_max: int | None = None,
    chain_exclude_col: str | None = None,
    chain_exclude_values: list | None = None,
    warmup_g: int = 0,
    bins: list[tuple[int, int]] | None = None,
) -> None:
    """有利区間内累積G別 AT間天井狙い (Excel数式版、上下2ブロック)。"""
    if bins is None:
        bins = CYCLE_BIN_DEFAULT
    ws = wb.create_sheet(sheet_name)
    # 設定セル (1行目右側)
    ws.cell(row=1, column=10, value="純減/G").font = HEADER_FONT
    ws.cell(row=1, column=11, value=invest_per_g)
    ws.cell(row=1, column=12, value="円/枚").font = HEADER_FONT
    ws.cell(row=1, column=13, value=YEN_PER_COIN)
    ws.cell(row=1, column=14, value="AT純増/G").font = HEADER_FONT
    ws.cell(row=1, column=15, value=at_payout_per_g)
    ws.cell(row=1, column=16, value="掛枚").font = HEADER_FONT
    ws.cell(row=1, column=17, value=BET_PER_GAME)
    invest_ref = "$K$1"
    yen_ref = "$M$1"
    at_payout_ref = "$O$1"
    bet_ref = "$Q$1"

    chain_ws = wb[chain_sheet]
    chain_cols = _col_map(chain_ws)
    if (chain_cycle_col not in chain_cols or chain_g_col not in chain_cols
            or chain_payout_col not in chain_cols):
        ws.cell(row=2, column=1, value="(必要列なし)")
        return
    chain_ref = f"'{chain_sheet}'"
    g_l = chain_cols[chain_g_col]
    p_l = chain_cols[chain_payout_col]

    # 追加フィルタ (chain_filter_col の min/max + chain_exclude + warmup_g)
    extra_filter = ""
    if chain_filter_col and chain_filter_col in chain_cols:
        cf_l = chain_cols[chain_filter_col]
        if chain_filter_min is not None:
            extra_filter += f',{chain_ref}!{cf_l}:{cf_l},">={chain_filter_min}"'
        if chain_filter_max is not None:
            extra_filter += f',{chain_ref}!{cf_l}:{cf_l},"<={chain_filter_max}"'
    if chain_exclude_col and chain_exclude_values and chain_exclude_col in chain_cols:
        ex_l = chain_cols[chain_exclude_col]
        for v in chain_exclude_values:
            extra_filter += f',{chain_ref}!{ex_l}:{ex_l},"<>{v}"'
    if warmup_g > 0:
        extra_filter += f',{chain_ref}!{g_l}:{g_l},">={warmup_g}"'

    def _write_block(start_row: int, title: str, cycle_col: str):
        ws.cell(row=start_row, column=1, value=title).font = HEADER_FONT
        header_row = start_row + 1
        for c_idx, h in enumerate(["累積G", "サンプル数", "平均連チャンTY(枚)", "平均消化G",
                                    "期待値(枚)", "期待値(円)", "機械割(%)"], start=1):
            cell = ws.cell(row=header_row, column=c_idx, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER_ALL
        if cycle_col not in chain_cols:
            ws.cell(row=header_row + 1, column=1, value=f"(列なし: {cycle_col})")
            return header_row + 2
        c_l = chain_cols[cycle_col]
        for i, (low, high) in enumerate(bins):
            r_idx = header_row + 1 + i
            if high >= 10**8:
                rng = f'{chain_ref}!{c_l}:{c_l},">="&{low}'
                label = f"{low}G以上"
            else:
                rng = f'{chain_ref}!{c_l}:{c_l},">="&{low},{chain_ref}!{c_l}:{c_l},"<="&{high}'
                label = f"{low}-{high}G"
            rng += extra_filter  # chain_filter (例: chain_number_cut>=1) 追加
            b = f'=COUNTIFS({rng})'
            c_ty = f'=IFERROR(AVERAGEIFS({chain_ref}!{p_l}:{p_l},{rng}),0)'
            d_g = f'=IFERROR(AVERAGEIFS({chain_ref}!{g_l}:{g_l},{rng}),0)'
            e_ev = f"=C{r_idx}-D{r_idx}*{invest_ref}"
            f_yen = f"=E{r_idx}*{yen_ref}"
            g_kik = (f"=IFERROR(100+E{r_idx}/"
                     f"((D{r_idx}+C{r_idx}/{at_payout_ref})*{bet_ref})*100,0)")
            _write_row(ws, r_idx, [label, b, c_ty, d_g, e_ev, f_yen, g_kik])
            for col_idx in (3, 4, 5):
                ws.cell(row=r_idx, column=col_idx).number_format = "0.0"
            ws.cell(row=r_idx, column=6).number_format = "0"
            ws.cell(row=r_idx, column=7).number_format = "0.00"
        return header_row + 1 + len(bins)

    next_row = _write_block(1, "【前回通常累積G数別】", chain_cycle_col)
    _write_block(next_row + 2, "【直前有利区間G数別】", chain_cycle_total_col)


# ============================================================================
# 横並び複数ブロック版 (G_bin chain): morning/cut の N回目シートを1枚に集約
# ============================================================================

_GBIN_BLOCK_HEADERS = ["打ち始めG", "サンプル数", "当選数", "連チャンTYサンプル数", "区間当選率",
                       "区間連チャンTY(枚)", "累積連チャンTY(枚)", "Nから平均消化G",
                       "増分G",
                       "区間期待値(枚)", "区間期待値(円)"]
_GBIN_BLOCK_WIDTH = 14   # ヘッダ列数 (累積期待値枚/円/機械割 含む)
_GBIN_BLOCK_SPACER = 1   # ブロック間スペーサー列数


def _build_g_bin_chain_through_filter(
    chain_ref: str, chain_cols: dict[str, str],
    chain_through_max_col=None, chain_through_max_value=None,
    chain_through_min_col=None, chain_through_min_value=None,
    chain_filter_col=None, chain_filter_min=None, chain_filter_max=None,
    chain_filter2_col=None, chain_filter2_min=None, chain_filter2_max=None,
    chain_exclude_col=None, chain_exclude_values=None,
) -> str:
    """COUNTIFS/AVERAGEIFS 末尾に付けるフィルタ文字列を組み立てる。"""
    tf = ""
    if (chain_through_max_col and chain_through_max_value is not None
            and chain_through_max_col in chain_cols):
        tm_l = chain_cols[chain_through_max_col]
        tf += f',{chain_ref}!{tm_l}:{tm_l},"<={chain_through_max_value}"'
    if (chain_through_min_col and chain_through_min_value is not None
            and chain_through_min_col in chain_cols):
        tm_l = chain_cols[chain_through_min_col]
        tf += f',{chain_ref}!{tm_l}:{tm_l},">={chain_through_min_value}"'
    if chain_filter_col and chain_filter_col in chain_cols:
        cf_l = chain_cols[chain_filter_col]
        if chain_filter_min is not None:
            tf += f',{chain_ref}!{cf_l}:{cf_l},">={chain_filter_min}"'
        if chain_filter_max is not None:
            tf += f',{chain_ref}!{cf_l}:{cf_l},"<={chain_filter_max}"'
    if chain_filter2_col and chain_filter2_col in chain_cols:
        cf2_l = chain_cols[chain_filter2_col]
        if chain_filter2_min is not None:
            tf += f',{chain_ref}!{cf2_l}:{cf2_l},">={chain_filter2_min}"'
        if chain_filter2_max is not None:
            tf += f',{chain_ref}!{cf2_l}:{cf2_l},"<={chain_filter2_max}"'
    # chain_exclude (例: first_state ≠ 引き戻し)
    if chain_exclude_col and chain_exclude_values and chain_exclude_col in chain_cols:
        ex_l = chain_cols[chain_exclude_col]
        for v in chain_exclude_values:
            tf += f',{chain_ref}!{ex_l}:{ex_l},"<>{v}"'
    return tf


def _apply_g_bin_chain_filters(df, *,
    chain_through_max_col=None, chain_through_max_value=None,
    chain_through_min_col=None, chain_through_min_value=None,
    chain_filter_col=None, chain_filter_min=None, chain_filter_max=None,
    chain_filter2_col=None, chain_filter2_min=None, chain_filter2_max=None,
    chain_exclude_col=None, chain_exclude_values=None,
    chain_g_col: str = "pre_normal_g", warmup_g: int = 0,
):
    """preview 用の df フィルタリング (既存 build_g_bin_chain_preview_sheet と同等)。"""
    df_use = df
    if (chain_through_max_col and chain_through_max_value is not None
            and chain_through_max_col in df_use.columns):
        df_use = df_use[df_use[chain_through_max_col] <= chain_through_max_value]
    if (chain_through_min_col and chain_through_min_value is not None
            and chain_through_min_col in df_use.columns):
        df_use = df_use[(df_use[chain_through_min_col] >= chain_through_min_value)
                        | (df_use[chain_through_min_col] == 0)]
    if chain_filter_col and chain_filter_col in df_use.columns:
        if chain_filter_min is not None:
            df_use = df_use[df_use[chain_filter_col] >= chain_filter_min]
        if chain_filter_max is not None:
            df_use = df_use[df_use[chain_filter_col] <= chain_filter_max]
    if chain_filter2_col and chain_filter2_col in df_use.columns:
        if chain_filter2_min is not None:
            df_use = df_use[df_use[chain_filter2_col] >= chain_filter2_min]
        if chain_filter2_max is not None:
            df_use = df_use[df_use[chain_filter2_col] <= chain_filter2_max]
    if chain_exclude_col and chain_exclude_values and chain_exclude_col in df_use.columns:
        df_use = df_use[~df_use[chain_exclude_col].isin(chain_exclude_values)]
    if warmup_g > 0 and chain_g_col in df_use.columns:
        df_use = df_use[df_use[chain_g_col] >= warmup_g]
    return df_use


def _write_g_bin_chain_formula_block(
    ws, *, start_col: int, header_row: int,
    chain_ref: str, cg_l: str, cp_l: str, through_filter: str,
    invest_ref: str, yen_ref: str, at_payout_ref: str, bet_ref: str,
    warmup_g: int, bin_size: int, max_g: int, cumulative_label: str,
    title: str | None = None,
    filter_description: str | None = None,
) -> None:
    """(start_col, header_row) を起点に 14列幅のブロックを書く。
    title 指定時は header_row-2 にタイトル、header_row-1 に算出条件を書く。"""
    if title:
        c = ws.cell(row=header_row - 2, column=start_col, value=title)
        c.font = HEADER_FONT
    if filter_description:
        ws.cell(row=header_row - 1, column=start_col,
                value=f"算出条件: {filter_description}")

    headers = _GBIN_BLOCK_HEADERS + [
        f"{cumulative_label}(枚)", f"{cumulative_label}(円)",
        f"{cumulative_label}_機械割(%)",
    ]
    for c_offset, h in enumerate(headers):
        cell = ws.cell(row=header_row, column=start_col + c_offset, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_ALL

    cols = [get_column_letter(start_col + i) for i in range(_GBIN_BLOCK_WIDTH)]
    A_, B_, C_, D_, E_, F_, G_, H_, I_, J_, K_, L_, M_, N_ = cols

    starts = _g_bin_starts(bin_size, max_g)
    labels = _g_bin_labels(bin_size, max_g)
    n_rows = len(starts)

    for idx, (n_start, lbl) in enumerate(zip(starts, labels)):
        r = header_row + 1 + idx
        eff_low = n_start + warmup_g
        n_end = eff_low + bin_size
        lower = f'">="&{eff_low}'
        upper = f'"<"&{n_end}'
        b = f'=COUNTIFS({chain_ref}!{cg_l}:{cg_l},{lower}{through_filter})'
        c_val = (f'=COUNTIFS({chain_ref}!{cg_l}:{cg_l},{lower},'
                 f'{chain_ref}!{cg_l}:{cg_l},{upper}{through_filter})')
        d_ty = (f'=COUNTIFS({chain_ref}!{cg_l}:{cg_l},{lower},'
                f'{chain_ref}!{cg_l}:{cg_l},{upper}{through_filter})')
        e_rate = f"=IFERROR({C_}{r}/{B_}{r},0)"
        f_int_ty = (f'=IFERROR(AVERAGEIFS({chain_ref}!{cp_l}:{cp_l},'
                    f'{chain_ref}!{cg_l}:{cg_l},{lower},'
                    f'{chain_ref}!{cg_l}:{cg_l},{upper}{through_filter}),0)')
        g_cum_ty = (f'=IFERROR(AVERAGEIFS({chain_ref}!{cp_l}:{cp_l},'
                    f'{chain_ref}!{cg_l}:{cg_l},{lower}{through_filter}),0)')
        h_avg_g = (f'=IFERROR(AVERAGEIFS({chain_ref}!{cg_l}:{cg_l},'
                   f'{chain_ref}!{cg_l}:{cg_l},{lower}{through_filter})-{n_start},0)')
        i_step = f"={bin_size}"
        j_ev = f"={E_}{r}*{F_}{r}-{I_}{r}*{invest_ref}"
        k_yen = f"={J_}{r}*{yen_ref}"
        l_cum = f"={G_}{r}-{H_}{r}*{invest_ref}"
        m_yen = f"={L_}{r}*{yen_ref}"
        n_kik = (f"=IFERROR(100+{L_}{r}/"
                 f"(({H_}{r}+{G_}{r}/{at_payout_ref})*{bet_ref})*100,0)")
        values = [lbl, b, c_val, d_ty, e_rate, f_int_ty, g_cum_ty, h_avg_g,
                  i_step, j_ev, k_yen, l_cum, m_yen, n_kik]
        for c_offset, v in enumerate(values):
            cell = ws.cell(row=r, column=start_col + c_offset, value=v)
            cell.border = BORDER_ALL
        ws.cell(row=r, column=start_col + 4).number_format = "0.00%"
        for c_off in (5, 6, 7, 8, 9, 11):
            ws.cell(row=r, column=start_col + c_off).number_format = "0.0"
        for c_off in (10, 12):
            ws.cell(row=r, column=start_col + c_off).number_format = "0"
        ws.cell(row=r, column=start_col + 13).number_format = "0.00"

    note_row = header_row + 1 + n_rows
    ws.cell(row=note_row, column=start_col, value=f"{max_g}G+(異常)").font = HEADER_FONT
    ws.cell(row=note_row, column=start_col + 1,
            value=f'=COUNTIFS({chain_ref}!{cg_l}:{cg_l},">="&{max_g}{through_filter})')


def _write_g_bin_chain_preview_block(
    ws, df_use, *, start_col: int, header_row: int,
    chain_g_col: str, chain_payout_col: str,
    invest_per_g: float, at_payout_per_g: float,
    warmup_g: int, bin_size: int, max_g: int, cumulative_label: str,
    title: str | None = None,
    filter_description: str | None = None,
) -> None:
    """preview版ブロック。df_use は既にフィルタ適用済みの DataFrame。"""
    if title:
        c = ws.cell(row=header_row - 2, column=start_col, value=title)
        c.font = HEADER_FONT
    if filter_description:
        ws.cell(row=header_row - 1, column=start_col,
                value=f"算出条件: {filter_description}")

    headers = _GBIN_BLOCK_HEADERS + [
        f"{cumulative_label}(枚)", f"{cumulative_label}(円)",
        f"{cumulative_label}_機械割(%)",
    ]
    for c_offset, h in enumerate(headers):
        cell = ws.cell(row=header_row, column=start_col + c_offset, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_ALL

    starts = _g_bin_starts(bin_size, max_g)
    labels = _g_bin_labels(bin_size, max_g)
    g = df_use[chain_g_col]

    rows = []
    for n_start, lbl in zip(starts, labels):
        eff_low = n_start + warmup_g
        in_mask = (g >= eff_low) & (g < eff_low + bin_size)
        rest_mask = g >= eff_low
        n_in = int(in_mask.sum())
        n_rest = int(rest_mask.sum())
        p = n_in / n_rest if n_rest else 0.0
        avg_ty = float(df_use.loc[in_mask, chain_payout_col].mean()) if n_in else 0.0
        cum_avg_ty = float(df_use.loc[rest_mask, chain_payout_col].mean()) if n_rest else 0.0
        cum_avg_g_abs = float(df_use.loc[rest_mask, chain_g_col].mean()) if n_rest else 0.0
        cum_avg_g = max(0.0, cum_avg_g_abs - n_start)
        step_g = bin_size
        single_ev = p * avg_ty - step_g * invest_per_g
        rows.append({"lbl": lbl, "n_trial": n_rest, "n_win": n_in, "p": p,
                     "avg_ty": avg_ty, "cum_avg_ty": cum_avg_ty, "cum_avg_g": cum_avg_g,
                     "step_g": step_g, "single_ev": single_ev})

    for r in rows:
        r["cum_ev"] = r["cum_avg_ty"] - r["cum_avg_g"] * invest_per_g
        normal_g_exp = r["cum_avg_g"]
        at_g_exp = r["cum_avg_ty"] / at_payout_per_g if at_payout_per_g > 0 else 0
        total_invest_exp = (normal_g_exp + at_g_exp) * BET_PER_GAME
        r["kikai"] = 100 + r["cum_ev"] / total_invest_exp * 100 if total_invest_exp > 0 else 0.0

    for r_idx, r in enumerate(rows):
        row_num = header_row + 1 + r_idx
        values = [r["lbl"], r["n_trial"], r["n_win"], r["n_win"],
                  r["p"], r["avg_ty"], r["cum_avg_ty"], r["cum_avg_g"],
                  r["step_g"],
                  r["single_ev"], r["single_ev"] * YEN_PER_COIN,
                  r["cum_ev"], r["cum_ev"] * YEN_PER_COIN, r["kikai"]]
        for c_offset, v in enumerate(values):
            cell = ws.cell(row=row_num, column=start_col + c_offset, value=v)
            cell.border = BORDER_ALL
        ws.cell(row=row_num, column=start_col + 4).number_format = "0.00%"
        ws.cell(row=row_num, column=start_col + 13).number_format = "0.00"

    n_over = int((g >= max_g).sum())
    note_row = header_row + 1 + len(rows) + 1
    ws.cell(row=note_row, column=start_col, value=f"{max_g}G+(異常)")
    ws.cell(row=note_row, column=start_col + 1, value=n_over)


def build_g_bin_chain_multiblock_formula_sheet(
    wb: Workbook,
    sheet_name: str,
    blocks: list[dict],
    *,
    invest_per_g: float,
    at_payout_per_g: float = 8.2,
    chain_sheet: str = "ChainData",
    chain_g_col: str = "pre_normal_g",
    chain_payout_col: str = "net_payout",
    warmup_g: int = 0,
    bin_size: int = G_BIN_SIZE,
    max_g: int = G_BIN_MAX,
    cumulative_label: str = "AT到達まで累積期待値",
) -> None:
    """各 block dict は build_g_bin_chain_formula_sheet のフィルタ引数群 +
    "title" キーを持つ。複数ブロックを横並びで配置。"""
    ws = wb.create_sheet(sheet_name)
    # 設定セル (シート共通) を 1行目左端に配置
    ws.cell(row=1, column=1, value="純減/G").font = HEADER_FONT
    ws.cell(row=1, column=2, value=invest_per_g)
    ws.cell(row=1, column=3, value="円/枚").font = HEADER_FONT
    ws.cell(row=1, column=4, value=YEN_PER_COIN)
    ws.cell(row=1, column=5, value="AT純増/G").font = HEADER_FONT
    ws.cell(row=1, column=6, value=at_payout_per_g)
    ws.cell(row=1, column=7, value="掛枚").font = HEADER_FONT
    ws.cell(row=1, column=8, value=BET_PER_GAME)
    invest_ref = "$B$1"
    yen_ref = "$D$1"
    at_payout_ref = "$F$1"
    bet_ref = "$H$1"

    chain_ws = wb[chain_sheet]
    chain_cols = _col_map(chain_ws)
    if chain_g_col not in chain_cols or chain_payout_col not in chain_cols:
        ws.cell(row=2, column=1, value="(必要列が見つかりません)")
        return
    cg_l = chain_cols[chain_g_col]
    cp_l = chain_cols[chain_payout_col]
    chain_ref = f"'{chain_sheet}'"

    # ブロック配置: row 3 = title, row 4 = filter description, row 5 = header, row 6+ = data
    title_row = 3
    fdesc_row = 4
    header_row = 5
    cur_col = 1
    for block in blocks:
        tf = _build_g_bin_chain_through_filter(
            chain_ref, chain_cols,
            chain_through_max_col=block.get("chain_through_max_col"),
            chain_through_max_value=block.get("chain_through_max_value"),
            chain_through_min_col=block.get("chain_through_min_col"),
            chain_through_min_value=block.get("chain_through_min_value"),
            chain_filter_col=block.get("chain_filter_col"),
            chain_filter_min=block.get("chain_filter_min"),
            chain_filter_max=block.get("chain_filter_max"),
            chain_filter2_col=block.get("chain_filter2_col"),
            chain_filter2_min=block.get("chain_filter2_min"),
            chain_filter2_max=block.get("chain_filter2_max"),
            chain_exclude_col=block.get("chain_exclude_col"),
            chain_exclude_values=block.get("chain_exclude_values"),
        )
        fdesc = _build_filter_description(
            chain_through_min_col=block.get("chain_through_min_col"),
            chain_through_min_value=block.get("chain_through_min_value"),
            chain_through_max_col=block.get("chain_through_max_col"),
            chain_through_max_value=block.get("chain_through_max_value"),
            chain_filter_col=block.get("chain_filter_col"),
            chain_filter_min=block.get("chain_filter_min"),
            chain_filter_max=block.get("chain_filter_max"),
            chain_filter2_col=block.get("chain_filter2_col"),
            chain_filter2_min=block.get("chain_filter2_min"),
            chain_filter2_max=block.get("chain_filter2_max"),
            chain_exclude_col=block.get("chain_exclude_col"),
            chain_exclude_values=block.get("chain_exclude_values"),
            warmup_g=warmup_g,
        )
        _write_g_bin_chain_formula_block(
            ws, start_col=cur_col, header_row=header_row,
            chain_ref=chain_ref, cg_l=cg_l, cp_l=cp_l, through_filter=tf,
            invest_ref=invest_ref, yen_ref=yen_ref,
            at_payout_ref=at_payout_ref, bet_ref=bet_ref,
            warmup_g=warmup_g, bin_size=bin_size, max_g=max_g,
            cumulative_label=cumulative_label,
            title=block.get("title"),
            filter_description=fdesc,
        )
        cur_col += _GBIN_BLOCK_WIDTH + _GBIN_BLOCK_SPACER


def build_g_bin_chain_multiblock_preview_sheet(
    wb: Workbook,
    sheet_name: str,
    blocks: list[dict],
    chains_df,
    *,
    invest_per_g: float,
    at_payout_per_g: float = 8.2,
    chain_g_col: str = "pre_normal_g",
    chain_payout_col: str = "net_payout",
    warmup_g: int = 0,
    bin_size: int = G_BIN_SIZE,
    max_g: int = G_BIN_MAX,
    cumulative_label: str = "AT到達まで累積期待値",
    sheet_suffix: str = "_値",
) -> None:
    """preview版 横並びマルチブロックシート。"""
    ws = wb.create_sheet(sheet_name + sheet_suffix)
    # row 1 = title, row 2 = filter description, row 3 = header, row 4+ = data
    title_row = 1
    fdesc_row = 2
    header_row = 3
    cur_col = 1
    for block in blocks:
        df_use = _apply_g_bin_chain_filters(
            chains_df,
            chain_through_max_col=block.get("chain_through_max_col"),
            chain_through_max_value=block.get("chain_through_max_value"),
            chain_through_min_col=block.get("chain_through_min_col"),
            chain_through_min_value=block.get("chain_through_min_value"),
            chain_filter_col=block.get("chain_filter_col"),
            chain_filter_min=block.get("chain_filter_min"),
            chain_filter_max=block.get("chain_filter_max"),
            chain_filter2_col=block.get("chain_filter2_col"),
            chain_filter2_min=block.get("chain_filter2_min"),
            chain_filter2_max=block.get("chain_filter2_max"),
            chain_exclude_col=block.get("chain_exclude_col"),
            chain_exclude_values=block.get("chain_exclude_values"),
            chain_g_col=chain_g_col, warmup_g=warmup_g,
        )
        fdesc = _build_filter_description(
            chain_through_min_col=block.get("chain_through_min_col"),
            chain_through_min_value=block.get("chain_through_min_value"),
            chain_through_max_col=block.get("chain_through_max_col"),
            chain_through_max_value=block.get("chain_through_max_value"),
            chain_filter_col=block.get("chain_filter_col"),
            chain_filter_min=block.get("chain_filter_min"),
            chain_filter_max=block.get("chain_filter_max"),
            chain_filter2_col=block.get("chain_filter2_col"),
            chain_filter2_min=block.get("chain_filter2_min"),
            chain_filter2_max=block.get("chain_filter2_max"),
            chain_exclude_col=block.get("chain_exclude_col"),
            chain_exclude_values=block.get("chain_exclude_values"),
            warmup_g=warmup_g,
        )
        _write_g_bin_chain_preview_block(
            ws, df_use, start_col=cur_col, header_row=header_row,
            chain_g_col=chain_g_col, chain_payout_col=chain_payout_col,
            invest_per_g=invest_per_g, at_payout_per_g=at_payout_per_g,
            warmup_g=warmup_g, bin_size=bin_size, max_g=max_g,
            cumulative_label=cumulative_label,
            title=block.get("title"),
            filter_description=fdesc,
        )
        cur_col += _GBIN_BLOCK_WIDTH + _GBIN_BLOCK_SPACER
