"""Excel スタイル定数。"""

from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
CATEGORY_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
ZEBRA_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ACCENT_FILL = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

HEADER_FONT = Font(bold=True, color="FFFFFF")
BOLD_FONT = Font(bold=True)

THIN = Side(border_style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center")
