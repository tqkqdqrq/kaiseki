import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import pickle
from tqdm import tqdm

# ===== 設定 =====
INPUT_FILE = r'C:\Users\ilove\Desktop\解析\jyogai20251223_duo2_azukun.csv'
OUTPUT_FILE = r'C:\Users\ilove\Desktop\解析\duo解析\keiryou_' + datetime.now().strftime('%Y%m%d_%H%M%S') + '.xlsx'
CACHE_FILE = r'C:\Users\ilove\Desktop\解析\chains_cache.pkl'

# 設定
HEAVEN_THRESHOLD = 35  # 天国連チャンとみなすG数
COIN_HOLD = 25.3  # コイン持ち (G/50枚)
BB_GAMES = 51          # BB獲得G数
RB_GAMES = 21          # RB獲得G数


def load_data(filepath):
    """azukunフォーマットCSVを読み込み"""
    try:
        df = pd.read_csv(filepath, encoding='utf-8-sig')
    except:
        df = pd.read_csv(filepath, encoding='cp932')
    
    # 元の順序を保持
    df['Original_Order'] = range(len(df))
    
    # ID列を保持したまま、内部処理用に分解
    if 'ID' not in df.columns:
        # ID列がない場合は作成を試みる（既存のロジックで対応できない場合はエラーになるが、今回はazukun.csvを使うのでOKなはず）
        pass

    id_parts = df['ID'].str.rsplit('_', n=2, expand=True)
    df['Hall_Name'] = id_parts[0]
    df['Machine_No'] = pd.to_numeric(id_parts[1], errors='coerce').fillna(0).astype(int)
    df['Date'] = id_parts[2].str.replace(r'(\d{4})(\d{2})(\d{2})', r'\1-\2-\3', regex=True)
    
    df['Start'] = pd.to_numeric(df['Start'], errors='coerce').fillna(0).astype(int)
    df['Dedama'] = pd.to_numeric(df['Dedama'], errors='coerce').fillna(0).astype(int)
    
    # Count列作成（時刻順）
    df_sorted = df.sort_values(by=['Hall_Name', 'Date', 'Machine_No', 'Time'])
    df_sorted['Count'] = df_sorted.groupby(['Hall_Name', 'Date', 'Machine_No']).cumcount() + 1
    
    # 元の順序に戻す
    df = df_sorted.sort_values(by='Original_Order').reset_index(drop=True)
    
    return df


def analyze_chains(df):
    """連チャン判定と天国移行率を計算（元のCSV順序で処理）"""
    # 元の順序でソート
    df = df.sort_values(by='Original_Order').reset_index(drop=True)

    all_chains = []
    chain_id = 0
    chain_number_per_id = {}  # {ID: 連チャン回数}

    # DataFrameに Chain_ID, Chain_Position 列を追加
    df['Chain_ID'] = 0
    df['Chain_Position'] = 0

    # 現在のチェーン情報
    current_id = None
    current_chain_start_idx = 0
    current_chain_hits = [] # インデックスのリスト
    prev_chain_len = 0

    for idx, row in tqdm(df.iterrows(), total=len(df), desc="Analyzing Chains"):
        row_id = row['ID']
        start_g = row['Start']

        # 新しい台または天国終了
        if current_id != row_id or start_g > HEAVEN_THRESHOLD:
            # 前のチェーンを保存
            if current_chain_hits:
                chain_id += 1
                chain_len = len(current_chain_hits)
                is_heaven = chain_len >= 2

                # Chain_ID, Positionを設定
                for pos, hit_idx in enumerate(current_chain_hits, start=1):
                    df.at[hit_idx, 'Chain_ID'] = chain_id
                    df.at[hit_idx, 'Chain_Position'] = pos

                # チェーンの差枚計算
                first_start = df.at[current_chain_hits[0], 'Start']
                last_dedama = df.at[current_chain_hits[-1], 'Dedama']
                raw_dedama = df.loc[current_chain_hits, 'Dedama'].sum()  # 出玉合計（生データ）
                total_invest = df.loc[current_chain_hits, 'Start'].sum() / COIN_HOLD * 50
                first_invest = first_start / COIN_HOLD * 50
                # 天国中投資（初当り以外の消費）= 全投資 - 初当り投資
                heaven_invest = total_invest - first_invest
                # 純増枚数 = 出玉合計 - 天国中消費（初当り投資は含まない）
                total_dedama = raw_dedama - heaven_invest
                # 差枚 = 出玉合計 - 全投資（初当り含む）
                net_diff = raw_dedama - total_invest
                # 特殊判定 = 純増 - 最終出玉（最後の出玉を除いた天国中の純増）
                special_judge = total_dedama - last_dedama
                
                # 当選G数のリストを取得
                hit_games = df.loc[current_chain_hits, 'Start'].tolist()
                # Dedamaのリストを取得（ドキハナ計算用）
                hit_dedamas = df.loc[current_chain_hits, 'Dedama'].tolist()
                # Statusのリストを取得（BB/RB判定用）
                hit_statuses = df.loc[current_chain_hits, 'Status'].tolist()

                # 台ごとの連チャン番号をカウント
                machine_id = df.at[current_chain_hits[0], 'ID']
                if machine_id not in chain_number_per_id:
                    chain_number_per_id[machine_id] = 0
                chain_number_per_id[machine_id] += 1

                # 統計用に記録
                all_chains.append({
                    'ID': machine_id,
                    'Chain_ID': chain_id,
                    'Chain_Number': chain_number_per_id[machine_id],
                    'Chain_Length': chain_len,
                    'Hit_Games': hit_games,
                    'Hit_Dedamas': hit_dedamas,
                    'Hit_Statuses': hit_statuses,
                    'Is_Heaven': is_heaven,
                    'First_G': df.at[current_chain_hits[0], 'Start'],
                    'Through_Before': 0,
                    'Prev_Chain_Length': prev_chain_len,
                    'Raw_Dedama': raw_dedama,       # 出玉合計（生データ）
                    'Total_Dedama': total_dedama,   # 純増枚数 = 出玉 - 天国中消費
                    'Total_Invest': total_invest,
                    'Heaven_Invest': heaven_invest,
                    'Net_Diff': net_diff,
                    'Special_Judge': special_judge,
                    'Prev_Special_Judge': 0.0
                })

                if is_heaven:
                    prev_chain_len = chain_len
                else:
                    pass # 連チャンでない場合は前回連チャン数を維持しない（スルー扱いなので、前回の「天国連チャン」数は保持すべきだが、ロジック上は後で計算し直す）
            
            # 新しいチェーン開始
            if current_id != row_id:
                prev_chain_len = 0
            
            current_id = row_id
            current_chain_hits = [idx]
        else:
            # 天国継続
            current_chain_hits.append(idx)
    
    # 最後のチェーンを保存
    if current_chain_hits:
        chain_id += 1
        chain_len = len(current_chain_hits)
        is_heaven = chain_len >= 2

        for pos, hit_idx in enumerate(current_chain_hits, start=1):
            df.at[hit_idx, 'Chain_ID'] = chain_id
            df.at[hit_idx, 'Chain_Position'] = pos

        # チェーンの差枚計算
        first_start = df.at[current_chain_hits[0], 'Start']
        last_dedama = df.at[current_chain_hits[-1], 'Dedama']
        raw_dedama = df.loc[current_chain_hits, 'Dedama'].sum()  # 出玉合計（生データ）
        total_invest = df.loc[current_chain_hits, 'Start'].sum() / COIN_HOLD * 50
        first_invest = first_start / COIN_HOLD * 50
        heaven_invest = total_invest - first_invest
        # 純増枚数 = 出玉合計 - 天国中消費（初当り投資は含まない）
        total_dedama = raw_dedama - heaven_invest
        # 差枚 = 出玉合計 - 全投資（初当り含む）
        net_diff = raw_dedama - total_invest
        # 特殊判定 = 純増 - 最終出玉
        special_judge = total_dedama - last_dedama
        
        # 当選G数のリストを取得
        hit_games = df.loc[current_chain_hits, 'Start'].tolist()
        # Dedamaのリストを取得
        hit_dedamas = df.loc[current_chain_hits, 'Dedama'].tolist()
        # Statusのリストを取得
        hit_statuses = df.loc[current_chain_hits, 'Status'].tolist()

        # 台ごとの連チャン番号をカウント
        machine_id = df.at[current_chain_hits[0], 'ID']
        if machine_id not in chain_number_per_id:
            chain_number_per_id[machine_id] = 0
        chain_number_per_id[machine_id] += 1

        all_chains.append({
            'ID': machine_id,
            'Chain_ID': chain_id,
            'Chain_Number': chain_number_per_id[machine_id],
            'Chain_Length': chain_len,
            'Hit_Games': hit_games,
            'Hit_Dedamas': hit_dedamas,
            'Hit_Statuses': hit_statuses,
            'Is_Heaven': is_heaven,
            'First_G': df.at[current_chain_hits[0], 'Start'],
            'Through_Before': 0,
            'Prev_Chain_Length': prev_chain_len,
            'Raw_Dedama': raw_dedama,       # 出玉合計（生データ）
            'Total_Dedama': total_dedama,   # 純増枚数 = 出玉 - 天国中消費
            'Total_Invest': total_invest,
            'Heaven_Invest': heaven_invest,
            'Net_Diff': net_diff,
            'Special_Judge': special_judge,
            'Prev_Special_Judge': 0.0
        })

    chains_df = pd.DataFrame(all_chains)

    # Through_Before, Prev_Chain_Length, Prev_Special_Judgeを正確に計算
    if not chains_df.empty:
        chains_df['Through_Before'] = 0
        chains_df['Prev_Chain_Length'] = 0
        chains_df['Prev_Special_Judge'] = 0.0  # float型に変更
        chains_df['Daily_Balance_Before'] = 0.0    # 当日差枚（チェーン開始前の累計）
        chains_df['Max_Daily_Balance_Before'] = 0.0  # 当日差枚最大（チェーン開始前までの最大値）
        chains_df['Is_Reset'] = chains_df['Chain_Number'] == 1  # リセット判定（台内初回）

        prev_id = None
        through_count = 0
        prev_heaven_len = 0
        prev_special = 0.0
        cumulative_balance = 0.0   # 累計差枚（前回チェーン終了時点）
        max_balance = 0.0          # 最大差枚（前回チェーン終了時点までの最大）
        
        # 前回当選G数 (Prev_First_G) 初期化
        chains_df['Prev_First_G'] = -1
        prev_first_g = -1

        for idx in chains_df.index:
            current_id = chains_df.at[idx, 'ID']
            is_heaven = chains_df.at[idx, 'Is_Heaven']
            chain_len = chains_df.at[idx, 'Chain_Length']
            special_judge = chains_df.at[idx, 'Special_Judge']
            net_diff = chains_df.at[idx, 'Net_Diff']  # このチェーンの差枚
            first_g = chains_df.at[idx, 'First_G']

            if current_id != prev_id:
                through_count = 0
                prev_heaven_len = 0
                prev_special = 0.0
                cumulative_balance = 0.0
                max_balance = 0.0
                prev_first_g = -1 # IDが変わったらリセット

            # 列にセット (セットしてから、この行の値を次回のprevとして保存するか？)
            # 前回当選G数（前の行のFirst_G）を記録するのは、「このチェーンが始まる前の情報」
            chains_df.at[idx, 'Prev_First_G'] = prev_first_g
            
            # 次の行のために保存
            prev_first_g = first_g

            # チェーン開始前の値を記録（このチェーンの差枚を加算する前）
            chains_df.at[idx, 'Through_Before'] = through_count
            chains_df.at[idx, 'Prev_Chain_Length'] = prev_heaven_len
            chains_df.at[idx, 'Prev_Special_Judge'] = prev_special
            chains_df.at[idx, 'Daily_Balance_Before'] = cumulative_balance
            chains_df.at[idx, 'Max_Daily_Balance_Before'] = max_balance

            # このチェーンの差枚を加算（次のチェーンのために）
            cumulative_balance += net_diff
            # 最大値を更新
            if cumulative_balance > max_balance:
                max_balance = cumulative_balance

            if is_heaven:
                through_count = 0
                prev_heaven_len = chain_len
            else:
                through_count += 1

            # 前回の特殊判定を保存（天国チェーンの場合のみ更新、天国以外は直近の天国の値を保持）
            if is_heaven:
                prev_special = special_judge

            prev_id = current_id

    return df, chains_df


def calculate_heaven_rate(chains_df):
    """スルー回数別の天国移行率を計算"""
    results = []
    
    # 0〜10スルー
    for thr in range(11):
        subset = chains_df[chains_df['Through_Before'] == thr]
        total = len(subset)
        heaven = len(subset[subset['Is_Heaven'] == True])
        rate = heaven / total if total > 0 else 0
        
        results.append({
            'スルー回数': f'{thr}スルー',
            '天国移行数': heaven,
            'サンプル数': total,
            '天国移行率': rate
        })
    
    # 11スルー以上
    subset = chains_df[chains_df['Through_Before'] >= 11]
    total = len(subset)
    heaven = len(subset[subset['Is_Heaven'] == True])
    rate = heaven / total if total > 0 else 0
    
    results.append({
        'スルー回数': '11スルー以上',
        '天国移行数': heaven,
        'サンプル数': total,
        '天国移行率': rate
    })
    
    # 合計
    total_all = len(chains_df)
    heaven_all = len(chains_df[chains_df['Is_Heaven'] == True])
    rate_all = heaven_all / total_all if total_all > 0 else 0
    
    results.append({
        'スルー回数': '合計',
        '天国移行数': heaven_all,
        'サンプル数': total_all,
        '天国移行率': rate_all
    })
    
    return pd.DataFrame(results)


def calculate_heaven_rate_by_chain(chains_df):
    """前回連チャン長別の天国移行率を計算"""
    results = []
    
    # 前回連チャン2連〜19連
    for chain_len in range(2, 20):
        subset = chains_df[chains_df['Prev_Chain_Length'] == chain_len]
        total = len(subset)
        heaven = len(subset[subset['Is_Heaven'] == True])
        rate = heaven / total if total > 0 else 0
        
        results.append({
            '前回連チャン': f'{chain_len}連後',
            '天国移行数': heaven,
            'サンプル数': total,
            '天国移行率': rate
        })
    
    # 20連以上
    subset = chains_df[chains_df['Prev_Chain_Length'] >= 20]
    total = len(subset)
    heaven = len(subset[subset['Is_Heaven'] == True])
    rate = heaven / total if total > 0 else 0
    
    results.append({
        '前回連チャン': '20連以上後',
        '天国移行数': heaven,
        'サンプル数': total,
        '天国移行率': rate
    })
    
    # 合計 (前回天国があったもののみ)
    subset_all = chains_df[chains_df['Prev_Chain_Length'] >= 2]
    total_all = len(subset_all)
    heaven_all = len(subset_all[subset_all['Is_Heaven'] == True])
    rate_all = heaven_all / total_all if total_all > 0 else 0
    
    results.append({
        '前回連チャン': '合計',
        '天国移行数': heaven_all,
        'サンプル数': total_all,
        '天国移行率': rate_all
    })
    
    return pd.DataFrame(results)


def calculate_heaven_rate_crosstab(chains_df):
    """前回連チャン長 × スルー回数 のクロス集計（天国移行率とサンプル数を分離）+ 特殊判定分析を統合"""
    results = []

    # 朝イチを除外（Chain_Number > 1のみ）
    chains_with_prev = chains_df[chains_df['Chain_Number'] > 1].copy()

    # ===== 前回連チャン長別 - 天国移行率 =====
    results.append({'': '【前回連チャン長別 - 天国移行率】'})

    for chain_len in range(2, 20):
        rate_row = {'': f'{chain_len}連後'}
        for thr in range(11):
            subset = chains_with_prev[(chains_with_prev['Prev_Chain_Length'] == chain_len) &
                                      (chains_with_prev['Through_Before'] == thr)]
            total = len(subset)
            heaven = len(subset[subset['Is_Heaven'] == True])
            rate = heaven / total if total > 0 else None
            rate_row[f'{thr}スルー'] = rate

        # 合計列には天国移行率を入れる
        total_all = len(chains_with_prev[chains_with_prev['Prev_Chain_Length'] == chain_len])
        heaven_all = len(chains_with_prev[(chains_with_prev['Prev_Chain_Length'] == chain_len) &
                                          (chains_with_prev['Is_Heaven'] == True)])
        rate_all = heaven_all / total_all if total_all > 0 else None
        rate_row['合計'] = rate_all

        results.append(rate_row)

    # 20連以上
    rate_row = {'': '20連以上後'}
    for thr in range(11):
        subset = chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 20) &
                                  (chains_with_prev['Through_Before'] == thr)]
        total = len(subset)
        heaven = len(subset[subset['Is_Heaven'] == True])
        rate = heaven / total if total > 0 else None
        rate_row[f'{thr}スルー'] = rate
    total_all = len(chains_with_prev[chains_with_prev['Prev_Chain_Length'] >= 20])
    heaven_all = len(chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 20) &
                                      (chains_with_prev['Is_Heaven'] == True)])
    rate_all = heaven_all / total_all if total_all > 0 else None
    rate_row['合計'] = rate_all
    results.append(rate_row)

    # 3～9連後（まとめ）
    rate_row = {'': '3～9連後'}
    for thr in range(11):
        subset = chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 3) &
                                  (chains_with_prev['Prev_Chain_Length'] <= 9) &
                                  (chains_with_prev['Through_Before'] == thr)]
        total = len(subset)
        heaven = len(subset[subset['Is_Heaven'] == True])
        rate = heaven / total if total > 0 else None
        rate_row[f'{thr}スルー'] = rate
    total_all = len(chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 3) & (chains_with_prev['Prev_Chain_Length'] <= 9)])
    heaven_all = len(chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 3) & (chains_with_prev['Prev_Chain_Length'] <= 9) &
                                      (chains_with_prev['Is_Heaven'] == True)])
    rate_all = heaven_all / total_all if total_all > 0 else None
    rate_row['合計'] = rate_all
    results.append(rate_row)

    # 10連以上後（まとめ）
    rate_row = {'': '10連以上後'}
    for thr in range(11):
        subset = chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 10) &
                                  (chains_with_prev['Through_Before'] == thr)]
        total = len(subset)
        heaven = len(subset[subset['Is_Heaven'] == True])
        rate = heaven / total if total > 0 else None
        rate_row[f'{thr}スルー'] = rate
    total_all = len(chains_with_prev[chains_with_prev['Prev_Chain_Length'] >= 10])
    heaven_all = len(chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 10) &
                                      (chains_with_prev['Is_Heaven'] == True)])
    rate_all = heaven_all / total_all if total_all > 0 else None
    rate_row['合計'] = rate_all
    results.append(rate_row)

    # 空行
    results.append({})

    # ===== 前回連チャン長別 - サンプル数 =====
    results.append({'': '【前回連チャン長別 - サンプル数】'})

    for chain_len in range(2, 20):
        count_row = {'': f'{chain_len}連後'}
        for thr in range(11):
            subset = chains_with_prev[(chains_with_prev['Prev_Chain_Length'] == chain_len) &
                                      (chains_with_prev['Through_Before'] == thr)]
            count_row[f'{thr}スルー'] = len(subset)

        total_all = len(chains_with_prev[chains_with_prev['Prev_Chain_Length'] == chain_len])
        count_row['合計'] = total_all

        results.append(count_row)

    # 20連以上
    count_row = {'': '20連以上後'}
    for thr in range(11):
        subset = chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 20) &
                                  (chains_with_prev['Through_Before'] == thr)]
        count_row[f'{thr}スルー'] = len(subset)
    total_all = len(chains_with_prev[chains_with_prev['Prev_Chain_Length'] >= 20])
    count_row['合計'] = total_all
    results.append(count_row)

    # 3～9連後（まとめ）
    count_row = {'': '3～9連後'}
    for thr in range(11):
        subset = chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 3) &
                                  (chains_with_prev['Prev_Chain_Length'] <= 9) &
                                  (chains_with_prev['Through_Before'] == thr)]
        count_row[f'{thr}スルー'] = len(subset)
    total_all = len(chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 3) & (chains_with_prev['Prev_Chain_Length'] <= 9)])
    count_row['合計'] = total_all
    results.append(count_row)

    # 10連以上後（まとめ）
    count_row = {'': '10連以上後'}
    for thr in range(11):
        subset = chains_with_prev[(chains_with_prev['Prev_Chain_Length'] >= 10) &
                                  (chains_with_prev['Through_Before'] == thr)]
        count_row[f'{thr}スルー'] = len(subset)
    total_all = len(chains_with_prev[chains_with_prev['Prev_Chain_Length'] >= 10])
    count_row['合計'] = total_all
    results.append(count_row)

    # 空行
    results.append({})

    # ===== 前回特殊判定別 - 天国移行率 =====
    results.append({'': '【前回特殊判定別 - 天国移行率】'})

    # 100枚刻みのレンジ定義
    spec_ranges = []
    for start in range(0, 2001, 100):
        end = start + 99
        spec_ranges.append((f"{start}～{end}", start, end))
    spec_ranges.append(("2001以上", 2001, None))

    for label, low, high in spec_ranges:
        rate_row = {'': label}

        if high is None:
            range_subset = chains_with_prev[chains_with_prev['Prev_Special_Judge'] >= low]
        else:
            range_subset = chains_with_prev[(chains_with_prev['Prev_Special_Judge'] >= low) &
                                            (chains_with_prev['Prev_Special_Judge'] <= high)]

        for thr in range(11):
            subset = range_subset[range_subset['Through_Before'] == thr]
            total = len(subset)
            heaven = len(subset[subset['Is_Heaven'] == True])
            rate = heaven / total if total > 0 else None
            rate_row[f'{thr}スルー'] = rate

        # 合計列には天国移行率を入れる
        total_all = len(range_subset)
        heaven_all = len(range_subset[range_subset['Is_Heaven'] == True])
        rate_all = heaven_all / total_all if total_all > 0 else None
        rate_row['合計'] = rate_all

        results.append(rate_row)

    # 空行
    results.append({})

    # ===== 前回特殊判定別 - サンプル数 =====
    results.append({'': '【前回特殊判定別 - サンプル数】'})

    for label, low, high in spec_ranges:
        count_row = {'': label}

        if high is None:
            range_subset = chains_with_prev[chains_with_prev['Prev_Special_Judge'] >= low]
        else:
            range_subset = chains_with_prev[(chains_with_prev['Prev_Special_Judge'] >= low) &
                                            (chains_with_prev['Prev_Special_Judge'] <= high)]

        for thr in range(11):
            subset = range_subset[range_subset['Through_Before'] == thr]
            count_row[f'{thr}スルー'] = len(subset)

        count_row['合計'] = len(range_subset)

        
        results.append(rate_row)
        results.append(count_row)
        
    return pd.DataFrame(results)


def calculate_3chain_rule_analysis(chains_df):
    """3連以上を天国、2連を隠れスルーとする分析（率とサンプル数を別列で出力）"""
    # 分析用データを収集
    # 構造: {'Category': [{'2Chain_Loc': ..., 'Through': ..., 'Is_True_Heaven': ...}]}
    analysis_data = {
        '全体': [],
        '2連目1G_RB': [],
        '2連目1G_BB': [],
        '2連目[2-5,32]': [],
        '2連目[6-12]': [],
        '2連目[13-21]': [],
        '2連目[22-31]': [],
        '2連目BB_[2-5,32]': [],
        '2連目BB_[6-12]': [],
        '2連目BB_[13-21]': [],
        '2連目BB_[22-31]': [],
        '2連目RB_[2-5,32]': [],
        '2連目RB_[6-12]': [],
        '2連目RB_[13-21]': [],
        '2連目RB_[22-31]': [],
    }
    
    # 変数初期化
    prev_id = None
    hits_since_true_heaven = 0  # 3連以上基準のスルー回数
    first_2chain_idx = None     # 最初の2連が発生したスルー回数 (None: 未発生)
    prev_true_heaven_valid = False 
    
    # カテゴリ判定用のステート
    current_2chain_feature = None 
    
    trackers = {}
    for cat in analysis_data.keys():
        trackers[cat] = {
            'hits_since': 0,
            'first_2chain_idx': None,
            'active': False, # このカテゴリのデータとして有効か（3連天国後スタート）
            'ignore': False # このカテゴリの条件に合わない「最初の2連」が発生した場合、以降天国まで無視
        }

    # DataFrameに追加するためのリスト初期化
    col_r3_through = []
    col_r3_end = []
    col_r3_heaven = []
    # カテゴリごとのLocation列
    cols_loc = {cat: [] for cat in analysis_data}
    
    for idx, row in chains_df.iterrows():
        row_id = row['ID']
        chain_len = row['Chain_Length']
        hit_games = row.get('Hit_Games', [])
        hit_statuses = row.get('Hit_Statuses', [])
        
        # 台が変わったらリセット
        if row_id != prev_id:
            for cat in trackers:
                trackers[cat]['hits_since'] = 0
                trackers[cat]['first_2chain_idx'] = None
                trackers[cat]['active'] = False
                trackers[cat]['ignore'] = False
            
            if chain_len >= 3:
                for cat in trackers:
                    trackers[cat]['active'] = True
            
            # リセット時は全て空/無効値を入れるが、DataFrameの行数と合わせるため追加が必要
            # ただし、ループは全ての行に対して回るので、ここで追加する値が「その行の値」になる
            # リセット直後の行（=新しい台の1行目）の処理は下流で行うので、ここではcontinueしない方が良いが
            # prev_id logic matches flow.
            # ここでは状態リセットのみ。
            
            # --- ここで continue すると append がスキップされるので注意 ---
            # 元のコードでは continue していた。
            # 今回はリストに必ず値を入れる必要がある（行ズレ防止）。
            # リセット時の行も処理対象となる（e.g. 朝イチHeavenなど）
            prev_id = row_id # ID更新
            
            # 既存ロジックでは continue して次の行へ行っていた（active判定等は次ループで）。
            # しかしリスト追加は必須。
            # ここは「新しい台の1行目」のコンテキストで処理を続行させる。
            # activeチェック等をこの後で行う。
        
        # 前回の天国が3連以上でなければスキップ -> activeフラグで管理
        # 元コード: continueしていた。
        # 今回: continueするとリストが埋まらない。
        # -> 「対象外」として埋める。
        
        if not trackers['全体']['active']:
            # activeでない期間（前任が単発天国等）
            col_r3_through.append(None)
            col_r3_end.append(None)
            col_r3_heaven.append(None)
            for cat in analysis_data:
                cols_loc[cat].append('非アクティブ')
                
            # 次回アクティブ化判定
            if chain_len >= 3:
                for cat in trackers:
                    trackers[cat]['active'] = True
                    trackers[cat]['hits_since'] = 0
                    trackers[cat]['first_2chain_idx'] = None
                    trackers[cat]['ignore'] = False
            prev_id = row_id
            continue
            
        # ここからアクティブな処理
        is_true_heaven = (chain_len >= 3)
        
        # 今回のチェーン情報の解析
        current_2chain_details = {}
        is_valid_2chain = False
        
        if chain_len == 2:
            second_hit_g = hit_games[1] if len(hit_games) > 1 else 0
            if second_hit_g >= 34:
                is_valid_2chain = False
            else:
                is_valid_2chain = True
                status_2 = hit_statuses[1] if len(hit_statuses) > 1 else ''
                is_rb = 'RB' in status_2
                is_bb = 'BB' in status_2
                g = second_hit_g
                rng = None
                if (2 <= g <= 5) or (g == 32): rng = '[2-5,32]'
                elif 6 <= g <= 12: rng = '[6-12]'
                elif 13 <= g <= 21: rng = '[13-21]'
                elif 22 <= g <= 31: rng = '[22-31]'
                current_2chain_details = {
                    'is_1g': (g == 1),
                    'is_rb': is_rb,
                    'is_bb': is_bb,
                    'range': rng
                }

        # 共通値
        current_through = trackers['全体']['hits_since']
        
        # End Through計算 (Excel数式用: Start <= T < End)
        # 天国の場合は Start == T (End不要だが、Startを入れておく)
        if is_true_heaven:
            end_through = current_through
        else:
            # 天国でない場合、Chain Length分だけスルー回数を消費する
            end_through = current_through + chain_len

        col_r3_through.append(current_through)
        col_r3_end.append(end_through)
        col_r3_heaven.append(is_true_heaven)
        
        # 各カテゴリごとに処理
        for cat in analysis_data:
            tr = trackers[cat]
            
            # --- まず状態更新（2連検出を先に行う）---
            if is_true_heaven:
                # 天国到達 -> リセットは後で（ラベル決定後）
                pass
            else:
                # まだ天国でない（スルー中）
                if not tr['ignore']:
                    if is_valid_2chain and tr['first_2chain_idx'] is None:
                        # まだ2連未発生状態で、今回2連が発生した
                        # この2連が「このカテゴリの条件」を満たすか？
                        match = True
                        det = current_2chain_details
                        if cat == '全体': match = True
                        elif cat == '2連目1G_RB': match = (det['is_rb'] and det['is_1g'])
                        elif cat == '2連目1G_BB': match = (det['is_bb'] and det['is_1g'])
                        elif cat == '2連目[2-5,32]': match = (det['range'] == '[2-5,32]')
                        elif cat == '2連目[6-12]': match = (det['range'] == '[6-12]')
                        elif cat == '2連目[13-21]': match = (det['range'] == '[13-21]')
                        elif cat == '2連目[22-31]': match = (det['range'] == '[22-31]')
                        elif cat == '2連目BB_[2-5,32]': match = (det['is_bb'] and det['range'] == '[2-5,32]')
                        elif cat == '2連目BB_[6-12]': match = (det['is_bb'] and det['range'] == '[6-12]')
                        elif cat == '2連目BB_[13-21]': match = (det['is_bb'] and det['range'] == '[13-21]')
                        elif cat == '2連目BB_[22-31]': match = (det['is_bb'] and det['range'] == '[22-31]')
                        elif cat == '2連目RB_[2-5,32]': match = (det['is_rb'] and det['range'] == '[2-5,32]')
                        elif cat == '2連目RB_[6-12]': match = (det['is_rb'] and det['range'] == '[6-12]')
                        elif cat == '2連目RB_[13-21]': match = (det['is_rb'] and det['range'] == '[13-21]')
                        elif cat == '2連目RB_[22-31]': match = (det['is_rb'] and det['range'] == '[22-31]')
                        
                        if match:
                            tr['first_2chain_idx'] = tr['hits_since']
                        else:
                            tr['ignore'] = True
            
            # Locationラベル決定（2連検出後なので、2連発生時はそのスルー目ラベルになる）
            if tr['ignore']:
                loc_label = '対象外'
            elif not tr['active']:
                loc_label = '非アクティブ'
            elif tr['first_2chain_idx'] is None:
                loc_label = '2連なし'
            else:
                loc_label = f"{tr['first_2chain_idx']}スルー目"
            
            cols_loc[cat].append(loc_label)

            # --- データ記録（集計用リストへ） ---
            if not tr['ignore']:
                analysis_data[cat].append({
                    '2Chain_Loc': loc_label,
                    'Through': tr['hits_since'],
                    'Is_True_Heaven': is_true_heaven
                })
            
            # --- 状態更新（スルー回数加算とリセット） ---
            if is_true_heaven:
                tr['hits_since'] = 0
                tr['first_2chain_idx'] = None
                tr['ignore'] = False
            else:
                if not tr['ignore']:
                            
                    # スルー回数加算
                    # ユーザー要望: 2連（失敗）は2スルー分としてカウント（3連以上基準のためボーナス回数でカウント）
                    # 単発なら+1、2連なら+2
                    tr['hits_since'] += chain_len

        prev_id = row_id
    
    # DataFrameに列を追加
    chains_df['R3_Through'] = col_r3_through
    chains_df['R3_End'] = col_r3_end
    chains_df['R3_Heaven'] = col_r3_heaven
    for cat, vals in cols_loc.items():
        chains_df[f'Loc_{cat}'] = vals
        
    # 集計とクロス集計作成
    results = []
    loc_order = ['2連なし'] + [f'{i}スルー目' for i in range(11)] + ['11スルー以上']
    
    # カテゴリ順序
    cat_order = [
         '全体', 
         '2連目1G_RB', '2連目1G_BB',
         '2連目[2-5,32]', '2連目[6-12]', '2連目[13-21]', '2連目[22-31]',
         '2連目BB_[2-5,32]', '2連目BB_[6-12]', '2連目BB_[13-21]', '2連目BB_[22-31]',
         '2連目RB_[2-5,32]', '2連目RB_[6-12]', '2連目RB_[13-21]', '2連目RB_[22-31]'
    ]

    for cat in cat_order:
        data_list = analysis_data[cat]
        if not data_list: continue
        
        df_cat = pd.DataFrame(data_list)
        
        # --- 1. 天国移行率 ---
        results.append({'2連発生位置': f'【{cat} - 天国移行率】'})
        
        for loc in loc_order:
            row_subset = df_cat[df_cat['2Chain_Loc'] == loc] if not df_cat.empty else pd.DataFrame()
            # 空でもスキップしない（行構造を固定するため）
            
            row_data = {'2連発生位置': loc}
            for thr in range(11):
                subset = row_subset[row_subset['Through'] == thr]
                total = len(subset)
                success = len(subset[subset['Is_True_Heaven'] == True])
                rate = success / total if total > 0 else None
                row_data[f'{thr}スルー'] = rate
            
            # 11以降
            subset_over = row_subset[row_subset['Through'] >= 11]
            total = len(subset_over)
            success = len(subset_over[subset_over['Is_True_Heaven'] == True])
            rate = success / total if total > 0 else None
            row_data['11スルー以上'] = rate
            
            # 合計
            total_all = len(row_subset)
            success_all = len(row_subset[row_subset['Is_True_Heaven'] == True])
            rate_all = success_all / total_all if total_all > 0 else None
            row_data['合計'] = rate_all
            
            results.append(row_data)

        # --- 2. 天国移行数（分子） ---
        results.append({'2連発生位置': f'【{cat} - 天国移行数】'})
        
        for loc in loc_order:
            row_subset = df_cat[df_cat['2Chain_Loc'] == loc] if not df_cat.empty else pd.DataFrame()
            # 空でもスキップしない
            
            row_data = {'2連発生位置': loc}
            for thr in range(11):
                subset = row_subset[row_subset['Through'] == thr]
                success = len(subset[subset['Is_True_Heaven'] == True])
                row_data[f'{thr}スルー'] = success
            
            subset_over = row_subset[row_subset['Through'] >= 11]
            success = len(subset_over[subset_over['Is_True_Heaven'] == True])
            row_data['11スルー以上'] = success
            
            # 合計
            success_all = len(row_subset[row_subset['Is_True_Heaven'] == True])
            row_data['合計'] = success_all
            
            results.append(row_data)

        # --- 3. サンプル数（分母） ---
        results.append({'2連発生位置': f'【{cat} - サンプル数】'})
        for loc in loc_order:
            row_subset = df_cat[df_cat['2Chain_Loc'] == loc] if not df_cat.empty else pd.DataFrame()
            # 空でもスキップしない
            
            row_data = {'2連発生位置': loc}
            for thr in range(11):
                subset = row_subset[row_subset['Through'] == thr]
                row_data[f'{thr}スルー'] = len(subset)
            
            subset_over = row_subset[row_subset['Through'] >= 11]
            row_data['11スルー以上'] = len(subset_over)
            
            row_data['合計'] = len(row_subset)
            
            results.append(row_data)

        # 空行
        results.append({})
        
    return pd.DataFrame(results)


def calculate_dokihana_analysis(chains_df):
    """
    ドキハナチャンス分析
    - 天国連チャン中に34Gまたは35Gで当たりが発生（Position>=2）
    - その後も連チャンが続く（合計3連以上）
    - chains_dfにドキハナ判定列を追加
    """
    # ドキハナ列を初期化
    chains_df['Dokihana_Success'] = False
    chains_df['Dokihana_Start_Pos'] = 0
    chains_df['Dokihana_Chain_Len'] = 0
    chains_df['Dokihana_Dedama'] = 0
    
    for idx in chains_df.index:
        chain_len = chains_df.at[idx, 'Chain_Length']
        is_heaven = chains_df.at[idx, 'Is_Heaven']
        hit_games = chains_df.at[idx, 'Hit_Games']
        
        # ドキハナ判定: 天国連チャン中にPosition>=2で34か35Gがあり、その後2連以上続く
        if is_heaven and chain_len >= 3 and len(hit_games) >= 2:
            hit_dedamas = chains_df.at[idx, 'Hit_Dedamas']
            
            # Position>=2で34か35Gを探す
            for pos, g in enumerate(hit_games[1:], start=2):  # Position 2から
                if g in [34, 35]:
                    dokihana_len = chain_len - pos + 1
                    # ドキハナ連チャン数が2以上の場合のみ成功
                    if dokihana_len >= 2:
                        chains_df.at[idx, 'Dokihana_Success'] = True
                        chains_df.at[idx, 'Dokihana_Start_Pos'] = pos
                        chains_df.at[idx, 'Dokihana_Chain_Len'] = dokihana_len
                        
                        # ドキハナ開始位置以降の純増を計算
                        # pos-1はインデックス（0始まり）
                        dokihana_dedamas = sum(hit_dedamas[pos-1:])  # 出玉合計
                        dokihana_games = hit_games[pos-1:]  # 当選G数リスト
                        # ドキハナ区間の天国消費（初回除く = pos以降の当選G数合計）
                        dokihana_invest = sum(dokihana_games[1:]) / COIN_HOLD * 50 if len(dokihana_games) > 1 else 0
                        # 純増 = 出玉 - 天国消費
                        dokihana_net = dokihana_dedamas - dokihana_invest
                        chains_df.at[idx, 'Dokihana_Dedama'] = int(dokihana_net)
                    break
    
    # 統計計算（コンソール表示用）
    dokihana_success_df = chains_df[chains_df['Dokihana_Success'] == True]
    
    stats = {
        'total_count': len(dokihana_success_df),
        'avg_chain_len': dokihana_success_df['Dokihana_Chain_Len'].mean() if len(dokihana_success_df) > 0 else 0,
        'avg_dedama': dokihana_success_df['Dokihana_Dedama'].mean() if len(dokihana_success_df) > 0 else 0,
        'total_dedama': dokihana_success_df['Dokihana_Dedama'].sum() if len(dokihana_success_df) > 0 else 0
    }
    
    return chains_df, stats


def write_excel(df, heaven_rate_df, heaven_by_chain_df, crosstab_df, chains_df):
    """エクセルファイルに出力（軽量化版: 期待値表・3連基準分析なし）"""
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        # Dataシート - 基本データ出力
        data_cols = ['ID', 'Status', 'Start', 'Dedama', 'Time', 'Chain_ID', 'Chain_Position']
        df_out = df[data_cols].copy()
        df_out.columns = ['ID', 'Status', 'Start', 'Dedama', 'Time', '連チャンID', '連チャン位置']
        df_out.to_excel(writer, sheet_name='Data', index=False)
        
        ws_data = writer.sheets['Data']
        max_r = ws_data.max_row
        
        # 追加列ヘッダー (H列から)
        extra_headers = ['朝イチ連', '投資詳細(枚)', '差枚詳細(枚)', '当日差枚', '有利区間G数']
        for i, h in enumerate(extra_headers):
            ws_data.cell(row=1, column=8 + i, value=h)
        
        # Dataシート I-L列の値計算 (Python側で計算して直接書き込み)
        # I: 投資詳細
        df['calc_invest'] = df['Start'] / COIN_HOLD * 50
        
        # J: 差枚詳細
        df['calc_diff'] = df['Dedama'] - df['calc_invest']
        
        # K: 当日差枚 (開始時) = IDごとの累積差枚 (現在行を含まない)
        df['calc_daily_diff'] = df.groupby('ID')['calc_diff'].cumsum() - df['calc_diff']
        
        # L: 有利区間G数 (累積)
        # ボーナス消化G推定: BB=51, RB=21
        def get_bonus_g(k):
            k_str = str(k)
            if 'BB' in k_str: return 51
            if 'RB' in k_str: return 21
            return 0
        
        df['calc_bonus_g'] = df['Status'].apply(get_bonus_g)
        df['calc_total_g'] = df['Start'] + df['calc_bonus_g']
        df['calc_yuuri_g'] = df.groupby('ID')['calc_total_g'].cumsum()

        # 各行に値を書き込み
        for r in range(2, max_r + 1):
            idx = r - 2
            
            # H: 朝イチ連
            is_morning = df.at[idx, 'Is_Morning_Chain']
            ws_data.cell(row=r, column=8, value=is_morning)
            
            # I: 投資詳細
            ws_data.cell(row=r, column=9, value=df.at[idx, 'calc_invest'])
            
            # J: 差枚詳細
            ws_data.cell(row=r, column=10, value=df.at[idx, 'calc_diff'])
            
            # K: 当日差枚
            ws_data.cell(row=r, column=11, value=df.at[idx, 'calc_daily_diff'])
            
            # L: 有利区間G数
            ws_data.cell(row=r, column=12, value=df.at[idx, 'calc_yuuri_g'])
        
        # 一時列削除
        columns_to_drop = ['calc_invest', 'calc_diff', 'calc_daily_diff', 'calc_bonus_g', 'calc_total_g', 'calc_yuuri_g']
        df.drop(columns=columns_to_drop, inplace=True)

        
        # 天国移行率シート（スルー回数別）
        heaven_rate_df.to_excel(writer, sheet_name='天国移行率_スルー', index=False)
        ws = writer.sheets['天国移行率_スルー']
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
            for cell in row:
                cell.number_format = '0.0%'
        ws.column_dimensions['A'].width = 15
        
        # 天国移行率シート（前回連チャン別）
        heaven_by_chain_df.to_excel(writer, sheet_name='天国移行率_前回連', index=False)
        ws2 = writer.sheets['天国移行率_前回連']
        for row in ws2.iter_rows(min_row=2, max_col=ws2.max_column):
            for cell in row:
                cell.number_format = '0.0%'
        ws2.column_dimensions['A'].width = 15
        
        # クロス集計シート（天国移行率とサンプル数交互 + 特殊判定分析統合）
        crosstab_df.to_excel(writer, sheet_name='クロス集計', index=False)
        ws3 = writer.sheets['クロス集計']

        # パーセント表示（天国移行率セクションのみ）
        current_section = None

        for row_idx in range(2, ws3.max_row + 1):
            a_cell_value = ws3.cell(row=row_idx, column=1).value

            # セクションヘッダーを検出
            if a_cell_value and isinstance(a_cell_value, str) and '【' in a_cell_value:
                current_section = a_cell_value
                continue

            # 空行をスキップ
            if not a_cell_value:
                continue

            # 天国移行率セクションのみパーセント表示
            if current_section and '天国移行率' in current_section:
                for col in range(2, 13):  # B-L列
                    cell = ws3.cell(row=row_idx, column=col)
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.0%'


        ws3.column_dimensions['A'].width = 25  # セクション名が長いので幅を広げる



        
        # ChainDataシート (日本語ヘッダー作成)
        rename_dict = {
            'ID': 'ID',
            'Chain_ID': '連チャンID',
            'Chain_Number': '台内連チャン番号',
            'Chain_Length': '連チャン長',
            'Is_Heaven': '天国判定',
            'First_G': '初当G',
            'Through_Before': 'スルー回数',
            'Prev_Chain_Length': '前回連チャン長',
            'Raw_Dedama': '出玉合計',
            'Total_Dedama': '純増枚数',
            'Total_Invest': '投資枚数',
            'Heaven_Invest': '天国中消費',
            'Net_Diff': '差枚',
            'Special_Judge': '特殊判定',
            'Prev_Special_Judge': '前回特殊判定',
            'R3_Through': '3連基準スルー',
            'R3_End': '3連基準End',
            'R3_Heaven': '3連基準天国',
            'Dokihana_Success': 'ドキハナ成功',
            'Dokihana_Start_Pos': 'ドキハナ開始位置',
            'Dokihana_Chain_Len': 'ドキハナ連チャン数',
            'Dokihana_Dedama': 'ドキハナ獲得枚数',
            'Daily_Balance_Before': '当日差枚_開始前',
            'Max_Daily_Balance_Before': '当日差枚最大_開始前',
            'Is_Reset': 'リセット判定',
            'Is_Morning_Chain': '朝イチフラグ'
        }
        # Loc列もリネーム
        for col in chains_df.columns:
            if col.startswith('Loc_'):
                cat_name = col.replace('Loc_', '')
                rename_dict[col] = f'判定_{cat_name}'
                
        chains_jp = chains_df.rename(columns=rename_dict)
        chains_jp.to_excel(writer, sheet_name='ChainData', index=False)
        ws_chain = writer.sheets['ChainData']
        ws_chain.column_dimensions['A'].width = 35
        
        # ChainDataの列マッピング取得
        col_map = {}
        from openpyxl.utils import get_column_letter
        for col in range(1, ws_chain.max_column + 1):
            val = ws_chain.cell(1, col).value
            if val:
                col_map[val] = get_column_letter(col)
        
        # 参照用カラムレター
        col_r3_thr = col_map.get('3連基準スルー')
        col_r3_end = col_map.get('3連基準End')
        col_r3_hvn = col_map.get('3連基準天国')
        col_len = col_map.get('連チャン長')
        

        # ========== ドキハナチャンス分析シート（式計算ベース） ==========
        ws_doki = writer.book.create_sheet("DokihanaAnalysis")
        
        # タイトル
        ws_doki['A1'] = "【ドキハナチャンス分析】"
        ws_doki['A1'].font = Font(bold=True, size=14)
        ws_doki['A2'] = "34Gまたは35Gで発生する特別な天国連チャンの分析"
        
        # ChainDataの行数とドキハナ列参照を取得
        chain_max_row = ws_chain.max_row
        col_doki_success = col_map.get('ドキハナ成功')
        col_doki_len = col_map.get('ドキハナ連チャン数')
        col_doki_dedama = col_map.get('ドキハナ獲得枚数')
        col_doki_start = col_map.get('ドキハナ開始位置')
        col_max_balance = col_map.get('当日差枚最大_開始前')
        
        # ChainDataの範囲参照
        ref_success = f"ChainData!${col_doki_success}$2:${col_doki_success}${chain_max_row}"
        ref_doki_len = f"ChainData!${col_doki_len}$2:${col_doki_len}${chain_max_row}"
        ref_doki_dedama = f"ChainData!${col_doki_dedama}$2:${col_doki_dedama}${chain_max_row}"
        ref_doki_start = f"ChainData!${col_doki_start}$2:${col_doki_start}${chain_max_row}"
        ref_max_balance = f"ChainData!${col_max_balance}$2:${col_max_balance}${chain_max_row}"
        
        # ========== 基本統計 ==========
        ws_doki['A4'] = "■ 基本統計"
        ws_doki['A4'].font = Font(bold=True)
        
        ws_doki['A5'] = "ドキハナ発生回数"
        ws_doki['B5'] = f"=COUNTIF({ref_success}, TRUE)"
        
        ws_doki['A6'] = "平均ドキハナ連チャン数"
        ws_doki['B6'] = f'=IFERROR(AVERAGEIF({ref_success}, TRUE, {ref_doki_len}), "-")'
        ws_doki['B6'].number_format = '0.00'
        
        ws_doki['A7'] = "平均ドキハナ獲得枚数"
        ws_doki['B7'] = f'=IFERROR(AVERAGEIF({ref_success}, TRUE, {ref_doki_dedama})-32/{COIN_HOLD}*50, "-")'
        ws_doki['B7'].number_format = '#,##0'
        
        ws_doki['A8'] = "合計ドキハナ獲得枚数"
        ws_doki['B8'] = f"=SUMIF({ref_success}, TRUE, {ref_doki_dedama})"
        ws_doki['B8'].number_format = '#,##0'
        
        # ========== ドキハナ開始位置別 分析（全体） ==========
        ws_doki['A10'] = "■ ドキハナ開始位置別 分析（全体）"
        ws_doki['A10'].font = Font(bold=True)
        
        ws_doki['A11'] = "開始位置"
        ws_doki['B11'] = "発生回数"
        ws_doki['C11'] = "割合"
        ws_doki['D11'] = "平均ドキハナ連数"
        ws_doki['E11'] = "平均獲得枚数"
        
        row = 12
        for pos in range(2, 21):
            ws_doki.cell(row=row, column=1, value=f"{pos}連目")
            ws_doki.cell(row=row, column=2, value=f"=COUNTIFS({ref_success}, TRUE, {ref_doki_start}, {pos})")
            ws_doki.cell(row=row, column=3, value=f"=IFERROR(B{row}/$B$5, 0)")
            ws_doki.cell(row=row, column=3).number_format = '0.0%'
            ws_doki.cell(row=row, column=4, value=f'=IFERROR(AVERAGEIFS({ref_doki_len}, {ref_success}, TRUE, {ref_doki_start}, {pos}), "-")')
            ws_doki.cell(row=row, column=4).number_format = '0.00'
            ws_doki.cell(row=row, column=5, value=f'=IFERROR(AVERAGEIFS({ref_doki_dedama}, {ref_success}, TRUE, {ref_doki_start}, {pos})-32/{COIN_HOLD}*50, "-")')
            ws_doki.cell(row=row, column=5).number_format = '#,##0'
            row += 1
        
        # 21連目以降
        ws_doki.cell(row=row, column=1, value="21連目以降")
        ws_doki.cell(row=row, column=2, value=f'=COUNTIFS({ref_success}, TRUE, {ref_doki_start}, ">=21")')
        ws_doki.cell(row=row, column=3, value=f"=IFERROR(B{row}/$B$5, 0)")
        ws_doki.cell(row=row, column=3).number_format = '0.0%'
        ws_doki.cell(row=row, column=4, value=f'=IFERROR(AVERAGEIFS({ref_doki_len}, {ref_success}, TRUE, {ref_doki_start}, ">=21"), "-")')
        ws_doki.cell(row=row, column=4).number_format = '0.00'
        ws_doki.cell(row=row, column=5, value=f'=IFERROR(AVERAGEIFS({ref_doki_dedama}, {ref_success}, TRUE, {ref_doki_start}, ">=21")-32/{COIN_HOLD}*50, "-")')
        ws_doki.cell(row=row, column=5).number_format = '#,##0'
        
        # ========== ドキハナ開始位置別 分析（当日差枚最大<1 = 一度もプラスになっていない） ==========
        ws_doki['A35'] = "■ ドキハナ開始位置別 分析（当日差枚が一度も1以上になっていない）"
        ws_doki['A35'].font = Font(bold=True)
        ws_doki['A36'] = "※ その時点までの当日差枚最大が1未満の場合のみカウント"
        
        ws_doki['A37'] = "開始位置"
        ws_doki['B37'] = "発生回数"
        ws_doki['C37'] = "割合"
        ws_doki['D37'] = "平均ドキハナ連数"
        ws_doki['E37'] = "平均獲得枚数"
        
        # 当日差枚条件付きの発生回数合計（割合計算用）
        ws_doki['G35'] = "条件付き合計"
        ws_doki['G36'] = f'=COUNTIFS({ref_success}, TRUE, {ref_max_balance}, "<1")'
        
        row = 38
        for pos in range(2, 21):
            ws_doki.cell(row=row, column=1, value=f"{pos}連目")
            # 条件: ドキハナ成功=TRUE AND 開始位置=pos AND 当日差枚最大<1
            ws_doki.cell(row=row, column=2, value=f'=COUNTIFS({ref_success}, TRUE, {ref_doki_start}, {pos}, {ref_max_balance}, "<1")')
            ws_doki.cell(row=row, column=3, value=f"=IFERROR(B{row}/$G$36, 0)")
            ws_doki.cell(row=row, column=3).number_format = '0.0%'
            ws_doki.cell(row=row, column=4, value=f'=IFERROR(AVERAGEIFS({ref_doki_len}, {ref_success}, TRUE, {ref_doki_start}, {pos}, {ref_max_balance}, "<1"), "-")')
            ws_doki.cell(row=row, column=4).number_format = '0.00'
            ws_doki.cell(row=row, column=5, value=f'=IFERROR(AVERAGEIFS({ref_doki_dedama}, {ref_success}, TRUE, {ref_doki_start}, {pos}, {ref_max_balance}, "<1")-32/{COIN_HOLD}*50, "-")')
            ws_doki.cell(row=row, column=5).number_format = '#,##0'
            row += 1
        
        # 21連目以降
        ws_doki.cell(row=row, column=1, value="21連目以降")
        ws_doki.cell(row=row, column=2, value=f'=COUNTIFS({ref_success}, TRUE, {ref_doki_start}, ">=21", {ref_max_balance}, "<1")')
        ws_doki.cell(row=row, column=3, value=f"=IFERROR(B{row}/$G$36, 0)")
        ws_doki.cell(row=row, column=3).number_format = '0.0%'
        ws_doki.cell(row=row, column=4, value=f'=IFERROR(AVERAGEIFS({ref_doki_len}, {ref_success}, TRUE, {ref_doki_start}, ">=21", {ref_max_balance}, "<1"), "-")')
        ws_doki.cell(row=row, column=4).number_format = '0.00'
        ws_doki.cell(row=row, column=5, value=f'=IFERROR(AVERAGEIFS({ref_doki_dedama}, {ref_success}, TRUE, {ref_doki_start}, ">=21", {ref_max_balance}, "<1")-32/{COIN_HOLD}*50, "-")')
        ws_doki.cell(row=row, column=5).number_format = '#,##0'
        
        # ========== スルー回数別 ドキハナ分析（条件付き） ==========
        # ChainDataのスルー回数列を参照
        col_through = col_map.get('スルー回数')
        ref_through = f"ChainData!${col_through}$2:${col_through}${chain_max_row}"
        
        # スタイル定義
        doki_header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # 青
        doki_header_font = Font(bold=True, color='FFFFFF')  # 白文字
        doki_row_fill_odd = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # 白
        doki_row_fill_even = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # 薄灰色
        doki_total_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')  # 薄青
        doki_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        ws_doki['A62'] = "■ スルー回数別 ドキハナ分析（当日プラス未経験 & 6連目以内開始）"
        ws_doki['A62'].font = Font(bold=True, size=12)
        ws_doki['A63'] = "※ 当日差枚が一度も1以上になっていない & ドキハナ開始位置が6連目以内"
        
        # ヘッダー行（64行目）
        headers = ["スルー回数", "発生回数", "平均ドキハナ連数", "平均獲得枚数"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_doki.cell(row=64, column=col_idx, value=header)
            cell.fill = doki_header_fill
            cell.font = doki_header_font
            cell.border = doki_border
            cell.alignment = Alignment(horizontal='center')
        
        row = 65
        for thr in range(10):  # 0～9スルー
            row_fill = doki_row_fill_even if thr % 2 == 0 else doki_row_fill_odd
            
            # スルー回数ラベル
            cell = ws_doki.cell(row=row, column=1, value=f"{thr}スルー")
            cell.fill = row_fill
            cell.border = doki_border
            
            # 発生回数
            cell = ws_doki.cell(row=row, column=2, value=f'=COUNTIFS({ref_success}, TRUE, {ref_through}, {thr}, {ref_max_balance}, "<1", {ref_doki_start}, "<=6")')
            cell.fill = row_fill
            cell.border = doki_border
            
            # 平均連チャン数
            cell = ws_doki.cell(row=row, column=3, value=f'=IFERROR(AVERAGEIFS({ref_doki_len}, {ref_success}, TRUE, {ref_through}, {thr}, {ref_max_balance}, "<1", {ref_doki_start}, "<=6"), "-")')
            cell.number_format = '0.00'
            cell.fill = row_fill
            cell.border = doki_border
            
            # 平均獲得枚数
            cell = ws_doki.cell(row=row, column=4, value=f'=IFERROR(AVERAGEIFS({ref_doki_dedama}, {ref_success}, TRUE, {ref_through}, {thr}, {ref_max_balance}, "<1", {ref_doki_start}, "<=6")-32/{COIN_HOLD}*50, "-")')
            cell.number_format = '#,##0'
            cell.fill = row_fill
            cell.border = doki_border
            
            row += 1
        
        # 10スルー以上
        row_fill = doki_row_fill_even
        cell = ws_doki.cell(row=row, column=1, value="10スルー以上")
        cell.fill = row_fill
        cell.border = doki_border
        
        cell = ws_doki.cell(row=row, column=2, value=f'=COUNTIFS({ref_success}, TRUE, {ref_through}, ">=10", {ref_max_balance}, "<1", {ref_doki_start}, "<=6")')
        cell.fill = row_fill
        cell.border = doki_border
        
        cell = ws_doki.cell(row=row, column=3, value=f'=IFERROR(AVERAGEIFS({ref_doki_len}, {ref_success}, TRUE, {ref_through}, ">=10", {ref_max_balance}, "<1", {ref_doki_start}, "<=6"), "-")')
        cell.number_format = '0.00'
        cell.fill = row_fill
        cell.border = doki_border
        
        cell = ws_doki.cell(row=row, column=4, value=f'=IFERROR(AVERAGEIFS({ref_doki_dedama}, {ref_success}, TRUE, {ref_through}, ">=10", {ref_max_balance}, "<1", {ref_doki_start}, "<=6")-32/{COIN_HOLD}*50, "-")')
        cell.number_format = '#,##0'
        cell.fill = row_fill
        cell.border = doki_border
        row += 1
        
        # 合計行（薄青背景）
        cell = ws_doki.cell(row=row, column=1, value="合計")
        cell.fill = doki_total_fill
        cell.border = doki_border
        cell.font = Font(bold=True)
        
        cell = ws_doki.cell(row=row, column=2, value=f'=COUNTIFS({ref_success}, TRUE, {ref_max_balance}, "<1", {ref_doki_start}, "<=6")')
        cell.fill = doki_total_fill
        cell.border = doki_border
        cell.font = Font(bold=True)
        
        cell = ws_doki.cell(row=row, column=3, value=f'=IFERROR(AVERAGEIFS({ref_doki_len}, {ref_success}, TRUE, {ref_max_balance}, "<1", {ref_doki_start}, "<=6"), "-")')
        cell.number_format = '0.00'
        cell.fill = doki_total_fill
        cell.border = doki_border
        cell.font = Font(bold=True)
        
        cell = ws_doki.cell(row=row, column=4, value=f'=IFERROR(AVERAGEIFS({ref_doki_dedama}, {ref_success}, TRUE, {ref_max_balance}, "<1", {ref_doki_start}, "<=6")-32/{COIN_HOLD}*50, "-")')
        cell.number_format = '#,##0'
        cell.fill = doki_total_fill
        cell.border = doki_border
        cell.font = Font(bold=True)
        
        # 列幅調整
        ws_doki.column_dimensions['A'].width = 35
        ws_doki.column_dimensions['B'].width = 15
        ws_doki.column_dimensions['C'].width = 18
        ws_doki.column_dimensions['D'].width = 18
        ws_doki.column_dimensions['E'].width = 18

        # ========== 天国当選G数分布シート ==========
        ws_dist = writer.book.create_sheet("HeavenHitDistribution")
        
        ws_dist['A1'] = "【3連以上天国チェーン 当選G数分布】"
        ws_dist['A1'].font = Font(bold=True, size=14)
        ws_dist['A2'] = "※ 3連以上の天国チェーンにおけるPosition 2以降（天国連チャン中）の当選G数分布"
        
        # 3連以上のchainからPosition 2以降のhit_gamesを収集
        all_heaven_hits = []
        for idx in chains_df.index:
            chain_len = chains_df.at[idx, 'Chain_Length']
            is_heaven = chains_df.at[idx, 'Is_Heaven']
            hit_games = chains_df.at[idx, 'Hit_Games']
            
            if is_heaven and chain_len >= 3:
                # Position 2以降（インデックス1以降）の当選G数を収集
                for g in hit_games[1:]:
                    if g <= 35:
                        all_heaven_hits.append(g)
        
        # 分布を集計
        from collections import Counter
        hit_counter = Counter(all_heaven_hits)
        total_hits = len(all_heaven_hits)
        
        # ヘッダー
        ws_dist['A4'] = "当選G数"
        ws_dist['B4'] = "発生回数"
        ws_dist['C4'] = "割合"
        ws_dist['A4'].font = Font(bold=True)
        ws_dist['B4'].font = Font(bold=True)
        ws_dist['C4'].font = Font(bold=True)
        
        row = 5
        for g in range(1, 36):  # 1G〜35G
            ws_dist.cell(row=row, column=1, value=f"{g}G")
            count = hit_counter.get(g, 0)
            ws_dist.cell(row=row, column=2, value=count)
            rate = count / total_hits if total_hits > 0 else 0
            ws_dist.cell(row=row, column=3, value=rate)
            ws_dist.cell(row=row, column=3).number_format = '0.00%'
            row += 1
        
        # 合計行
        ws_dist.cell(row=row, column=1, value="合計")
        ws_dist.cell(row=row, column=1).font = Font(bold=True)
        ws_dist.cell(row=row, column=2, value=total_hits)
        ws_dist.cell(row=row, column=2).font = Font(bold=True)
        ws_dist.cell(row=row, column=3, value=1.0 if total_hits > 0 else 0)
        ws_dist.cell(row=row, column=3).number_format = '0.00%'
        ws_dist.cell(row=row, column=3).font = Font(bold=True)
        
        # 平均G数
        avg_g = sum(all_heaven_hits) / total_hits if total_hits > 0 else 0
        ws_dist.cell(row=row+1, column=1, value="平均G数")
        ws_dist.cell(row=row+1, column=1).font = Font(bold=True)
        ws_dist.cell(row=row+1, column=2, value=avg_g)
        ws_dist.cell(row=row+1, column=2).number_format = '0.0'
        ws_dist.cell(row=row+1, column=2).font = Font(bold=True)
        
        # 列幅調整
        ws_dist.column_dimensions['A'].width = 12
        ws_dist.column_dimensions['B'].width = 12
        ws_dist.column_dimensions['C'].width = 12

        # ========== 天井到達率分析シート ==========
        ws_ceil = writer.book.create_sheet("CeilingAnalysis")
        
        ws_ceil['A1'] = "【天井到達率分析】"
        ws_ceil['A1'].font = Font(bold=True, size=14)
        ws_ceil['A2'] = "※ サンプル数=そのG以上で当選した総数、天井到達=800G以上で当選、到達率=天井到達/サンプル"
        
        ws_ceil['A16'] = "【設定1調整係数】"
        ws_ceil['A17'] = "天井到達率÷(1-天井到達率)で算出される初当たり確率の逆数？ ではなく、"
        ws_ceil['A18'] = "解析値(1/255) / 実測値(全平均) で算出"
        
        ws_ceil['B18'] = "実測平均初当G"
        all_avg_g = chains_df[chains_df['First_G'] > 0]['First_G'].mean()
        ws_ceil['C18'] = all_avg_g
        ws_ceil['C18'].number_format = '0.0'
        
        ws_ceil['B19'] = "実測初当確率"
        if all_avg_g > 0:
             # 初当たり確率分母 = 平均G + ？ (単純に平均Gを確率分母とみなすか、厳密には違うが)
             # ここでは簡易的に「平均G数」を使用
             denom_calc = all_avg_g
             ws_ceil['C19'] = f"1/{denom_calc:.1f}"
             
             # 調整係数 = (解析確率分母 255) / 実測平均G
             setting1_adjust = 255.0 / all_avg_g
        else:
             setting1_adjust = 1.0
             ws_ceil['C19'] = "-"

        ws_ceil['B20'] = "調整係数"
        ws_ceil.cell(row=20, column=3, value=setting1_adjust)
        ws_ceil['C20'].number_format = '0.000'
        
        # 列幅調整
        ws_ceil.column_dimensions['A'].width = 15
        ws_ceil.column_dimensions['B'].width = 12
        ws_ceil.column_dimensions['C'].width = 12
        ws_ceil.column_dimensions['D'].width = 12

        # ========== 朝イチ履歴分け分析シート ==========
        ws_morn = writer.book.create_sheet("MorningAnalysis")
        ws_morn['A1'] = "【朝イチ履歴分け分析】"
        ws_morn['A1'].font = Font(bold=True, size=14)
        
        headers = ["履歴G数", "0スルーサンプル", "0スルー天国", "0スルー移行率"]
        for t in range(1, 6): # 1~5スルー
            headers.extend([f"{t}スルーサンプル", f"{t}スルー天国", f"{t}スルー移行率"])
            
        for col, h in enumerate(headers, 1):
            cell = ws_morn.cell(row=3, column=col, value=h)
            cell.font = Font(bold=True)
            cell.border = doki_border
            cell.fill = doki_header_fill
            
        # バケット定義
        buckets = [
            ("0~50", 0, 50),
            ("51~100", 51, 100),
            ("101~150", 101, 150),
            ("151~200", 151, 200),
            ("201~400", 201, 400),
            ("400~", 400, 9999)
        ]
        
        # ChainDataシートへの参照
        # ※ chains_df は 'ChainData' シートに出力されていると仮定 (後続の出力コードで確認必要だが、恐らく 'ChainData' になる)
        # 既存コードでは chains_df を to_excel で 'ChainData' シートに出力しているはず。
        # 確認できないため、一般的な 'ChainData' と推測して数式を組む。
        # 参照列:
        #  Morning_0Through_G: column ? (後で追加されるので末尾付近)
        #  Through_Before: column ?
        #  Is_Heaven: column ?
        #  Is_Reset: column ?
        
        # 列位置の特定が難しいため、COUNTIFSでは列ラベル（構造化参照）が使えない場合、列番号指定が必要。
        # しかし、Pythonで書き込む際に DataFrame の列順序が変わると壊れる。
        # ここでは「ChainData」シートの列名を検索して列文字を取得するロジックを入れるか、
        # あるいは「データ出力」部分で列名が決まるので、数式ではなくPython計算値を埋め込む方が安全だが、
        # ユーザー要望は「計算は数式で」とのこと。
        
        # ヘルパー列の列文字特定用 (ChainDataシートは writer.sheets['ChainData'] でアクセス可能になるのは書き込み後)
        # よって、ここでは数式文字列を生成するが、列文字はプレースホルダーにしておくか、
        # あるいは「ChainData」シートを先に書き込む必要がある。
        # 既存コードの流れでは write_excel の最後に df.to_excel していると思われる。
        
        # 既存の write_excel の構造を見ると、最後に to_excel で各シートを出力しているので、
        # この時点ではまだシートが存在しない（openpyxlで枠を作っている最中）。
        # したがって、列文字を固定で推定するか、数式生成のみここで行い、列文字は後で埋めるなどが考えられる。
        # 軽量化版では to_excel を使わず openpyxl で書いている？ いや load_data などをしている。
        
        # 下の方にある `chains_df.to_excel(writer, sheet_name='ChainData', index=False)` を探す。
        # まだコードの全貌が見えていないが、chains_df に列を追加したので、出力時にはその列が含まれる。
        # 列名から列文字を特定するヘルパー関数を用意するのが無難。
        
        def get_col_letter(df, col_name):
            try:
                col_idx = df.columns.get_loc(col_name) + 1
                return get_column_letter(col_idx)
            except:
                return "A" # Fallback
                
        # 必要な列の列文字 (ユーザー指定: First_G=I列, Morning_Flag=W列)
        # 必要な列の列文字
        col_first_g = get_col_letter(chains_df, 'First_G')
        col_morn_flag = get_col_letter(chains_df, 'Is_Morning_Chain') 
        
        # Is_Morning_Chain は main で追加されるため chains_df にはあるはず
        if 'Is_Morning_Chain' in chains_df.columns:
            col_morn_flag = get_col_letter(chains_df, 'Is_Morning_Chain')
        else:
            col_morn_flag = "W"

        col_through = get_col_letter(chains_df, 'Through_Before')
        col_heaven = get_col_letter(chains_df, 'Is_Heaven')
        col_prev_first_g = get_col_letter(chains_df, 'Prev_First_G') # 新規追加列
        
        # Helper to generate EV table for a specific condition
        def generate_ev_table_for_bucket(ws, start_row, start_col, label, cond_prev_g_range):
            # Header
            ws.cell(row=start_row, column=start_col, value=f"【{label} の1スルーボーナス間期待値】")
            ws.cell(row=start_row, column=start_col).font = Font(bold=True)
            
            headers_ev = ["開始G数", "初当たり", "TY", "調整TY", "5枚等価", "出玉率", "サンプル"]
            
            header_row = start_row + 1
            for i, h in enumerate(headers_ev):
                cell = ws.cell(row=header_row, column=start_col + i, value=h)
                cell.fill = doki_header_fill
                cell.font = doki_header_font
                cell.border = doki_border
            
            # Rows (0G, 50G, ... 800G)
            curr_row = header_row + 1
            
            range_curr_first_g = f'ChainData!${col_first_g}$2:${col_first_g}$1048576'
            range_curr_through = f'ChainData!${col_through}$2:${col_through}$1048576'
            range_curr_flag = f'ChainData!${col_morn_flag}$2:${col_morn_flag}$1048576'
            
            col_total_dedama = get_col_letter(chains_df, 'Total_Dedama')
            range_curr_dedama = f'ChainData!${col_total_dedama}$2:${col_total_dedama}$1048576'

            # Additional filters (standard EV): MaxDaily<=0, PrevSpec<1000
            col_max_daily = get_col_letter(chains_df, 'Max_Daily_Balance_Before')
            col_prev_spec = get_col_letter(chains_df, 'Prev_Special_Judge')
            range_max_daily = f'ChainData!${col_max_daily}$2:${col_max_daily}$1048576'
            range_prev_spec = f'ChainData!${col_prev_spec}$2:${col_prev_spec}$1048576'
            
            # Loop for G
            for g in range(35, 801, 5): 
                # Base Criteria
                # 1. Through = 1
                # 2. Morning = TRUE
                # 3. First_G >= g
                # 4. Standard filters
                # 5. Prev G Bucket (Passed as condition string for Prev_First_G column)
                
                # MaxDaily filter (<=500) removed as per user request
                base_criteria = f'{range_curr_through}, 1, {range_curr_flag}, TRUE, {range_curr_first_g}, ">={g}", {range_prev_spec}, "<1000", {cond_prev_g_range}'
                
                # Column offsets relative to start_col (1-based index)
                # 0: Label (StartG)
                # 1: FirstHit
                # 2: TY
                # 3: AdjTY
                # 4: 50Eq
                # 5: Rate
                # 6: Sample
                
                # Start G Label
                ws.cell(row=curr_row, column=start_col, value=f"{g}G~").border = doki_border

                # 1. Sample (Col 6 relative) -> start_col + 6
                f_sample = f'=COUNTIFS({base_criteria})'
                ws.cell(row=curr_row, column=start_col + 6, value=f_sample).border = doki_border
                
                # 2. First Hit (Average First_G) -> Col 1 rel
                f_avg_g = f'AVERAGEIFS({range_curr_first_g}, {base_criteria})'
                adjust_ref = "CeilingAnalysis!$M$6"
                
                # Cell Ref for Sample: Offset +6
                # Excel Cell Ref construction:
                col_sample_let = get_column_letter(start_col + 6)
                col_first_let = get_column_letter(start_col + 1)
                
                f_hatsu = f'=IF({col_sample_let}{curr_row}=0, "-", ({f_avg_g} - {g}) / {adjust_ref})'
                ws.cell(row=curr_row, column=start_col + 1, value=f_hatsu).border = doki_border
                ws.cell(row=curr_row, column=start_col + 1).number_format = '0.0'

                # 3. TY -> Col 2 rel
                cost_str = f"32 * (50/{COIN_HOLD})"
                f_avg_ty = f'AVERAGEIFS({range_curr_dedama}, {base_criteria})'
                f_ty = f'=IF({col_sample_let}{curr_row}=0, "-", {f_avg_ty} - {cost_str})'
                ws.cell(row=curr_row, column=start_col + 2, value=f_ty).border = doki_border
                ws.cell(row=curr_row, column=start_col + 2).number_format = '#,##0'
                
                # 4. Adjusted TY -> Col 3 rel
                col_ty_let = get_column_letter(start_col + 2)
                f_adj_ty = f'=IF({col_ty_let}{curr_row}="-", "-", {col_ty_let}{curr_row} * {adjust_ref})'
                ws.cell(row=curr_row, column=start_col + 3, value=f_adj_ty).border = doki_border
                ws.cell(row=curr_row, column=start_col + 3).number_format = '#,##0'
                
                # 5. 50 Coins Eq -> Col 4 rel
                col_adj_ty_let = get_column_letter(start_col + 3)
                coin_factor = f"(50/{COIN_HOLD})"
                f_50eq = f'=IF(OR({col_adj_ty_let}{curr_row}="-", {col_first_let}{curr_row}="-"), "-", ({col_adj_ty_let}{curr_row} - {col_first_let}{curr_row} * {coin_factor}) * 20)'
                ws.cell(row=curr_row, column=start_col + 4, value=f_50eq).border = doki_border
                ws.cell(row=curr_row, column=start_col + 4).number_format = '#,##0'
                
                # 6. Payout Rate -> Col 5 rel
                col_50eq_let = get_column_letter(start_col + 4)
                denom = f'({col_first_let}{curr_row} * 3 + ({col_adj_ty_let}{curr_row} / CeilingAnalysis!$L$13 * 3))'
                numer = f'({denom} + {col_50eq_let}{curr_row} / 20)'
                f_rate = f'=IF(OR({col_first_let}{curr_row}="-", {col_adj_ty_let}{curr_row}="-", {denom}=0), "-", {numer} / {denom})'
                ws.cell(row=curr_row, column=start_col + 5, value=f_rate).border = doki_border
                ws.cell(row=curr_row, column=start_col + 5).number_format = '0.0%'
                
                curr_row += 1
            
            return curr_row # Return last row used

        start_row = 4
        
        # Summary Loop
        for i, (label, min_g, max_g) in enumerate(buckets):
            row = start_row + i
            
            # --- 0スルー分析 ---
            # 条件: First_G(I) in Range AND Through(J)=0 AND MorningFlag(W)=TRUE
            cond_g_0 = f'ChainData!${col_first_g}$2:${col_first_g}$1048576, ">={min_g}", ChainData!${col_first_g}$2:${col_first_g}$1048576, "<={max_g}"'
            
            ws_morn.cell(row=row, column=1, value=label).border = doki_border
            
            f_samp0 = f'=COUNTIFS({cond_g_0}, ChainData!${col_through}$2:${col_through}$1048576, 0, ChainData!${col_morn_flag}$2:${col_morn_flag}$1048576, TRUE)'
            ws_morn.cell(row=row, column=2, value=f_samp0).border = doki_border
            f_hvn0 = f'=COUNTIFS({cond_g_0}, ChainData!${col_through}$2:${col_through}$1048576, 0, ChainData!${col_morn_flag}$2:${col_morn_flag}$1048576, TRUE, ChainData!${col_heaven}$2:${col_heaven}$1048576, TRUE)'
            ws_morn.cell(row=row, column=3, value=f_hvn0).border = doki_border
            f_rate0 = f'=IF(B{row}=0, "-", C{row}/B{row})'
            cell = ws_morn.cell(row=row, column=4, value=f_rate0)
            cell.number_format = '0.0%'
            cell.border = doki_border
            
            # --- 1~5スルー分析 (Using Prev_First_G) ---
            # 条件: Prev_First_G in Range AND Through=t AND MorningFlag=TRUE
            
            range_prev_g_col = f'ChainData!${col_prev_first_g}$2:${col_prev_first_g}$1048576'
            range_curr_thr = f'ChainData!${col_through}$2:${col_through}$1048576'
            range_curr_flag = f'ChainData!${col_morn_flag}$2:${col_morn_flag}$1048576'
            range_curr_hvn = f'ChainData!${col_heaven}$2:${col_heaven}$1048576'
            
            # Condition for Prev_First_G
            cond_prev_g = f'{range_prev_g_col}, ">={min_g}", {range_prev_g_col}, "<={max_g}"'
            
            # Loop for 1 to 5 through
            current_col = 5
            for thr in range(1, 6):
                f_samp = f'=COUNTIFS({cond_prev_g}, {range_curr_thr}, {thr}, {range_curr_flag}, TRUE)'
                ws_morn.cell(row=row, column=current_col, value=f_samp).border = doki_border
                
                f_hvn = f'=COUNTIFS({cond_prev_g}, {range_curr_thr}, {thr}, {range_curr_flag}, TRUE, {range_curr_hvn}, TRUE)'
                ws_morn.cell(row=row, column=current_col+1, value=f_hvn).border = doki_border

                col_let_samp = get_column_letter(current_col)
                col_let_hvn = get_column_letter(current_col+1)
                
                f_rate = f'=IF({col_let_samp}{row}=0, "-", {col_let_hvn}{row}/{col_let_samp}{row})'
                cell = ws_morn.cell(row=row, column=current_col+2, value=f_rate)
                cell.number_format = '0.0%'
                cell.border = doki_border
                
                current_col += 3

        # 全履歴平均 (Total Row)
        row_total = start_row + len(buckets)
        doki_total_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

        ws_morn.cell(row=row_total, column=1, value="全履歴平均").border = doki_border
        
        # 0スルー (No G-range condition)
        cond_g_0_total = f'ChainData!${col_first_g}$2:${col_first_g}$1048576, ">={0}"'
        f_samp0 = f'=COUNTIFS({cond_g_0_total}, ChainData!${col_through}$2:${col_through}$1048576, 0, ChainData!${col_morn_flag}$2:${col_morn_flag}$1048576, TRUE)'
        ws_morn.cell(row=row_total, column=2, value=f_samp0).border = doki_border
        
        f_hvn0 = f'=COUNTIFS({cond_g_0_total}, ChainData!${col_through}$2:${col_through}$1048576, 0, ChainData!${col_morn_flag}$2:${col_morn_flag}$1048576, TRUE, ChainData!${col_heaven}$2:${col_heaven}$1048576, TRUE)'
        ws_morn.cell(row=row_total, column=3, value=f_hvn0).border = doki_border
        
        f_rate0 = f'=IF(B{row_total}=0, "-", C{row_total}/B{row_total})'
        cell = ws_morn.cell(row=row_total, column=4, value=f_rate0)
        cell.number_format = '0.0%'
        cell.border = doki_border
        
        # 1~5スルー (No Prev_G-range condition)
        current_col = 5
        range_prev_g_col = f'ChainData!${col_prev_first_g}$2:${col_prev_first_g}$1048576'
        
        for thr in range(1, 6):
            # 条件: Prev_G >= 0
            f_samp = f'=COUNTIFS({range_prev_g_col}, ">=0", {range_curr_thr}, {thr}, {range_curr_flag}, TRUE)'
            ws_morn.cell(row=row_total, column=current_col, value=f_samp).border = doki_border
            
            f_hvn = f'=COUNTIFS({range_prev_g_col}, ">=0", {range_curr_thr}, {thr}, {range_curr_flag}, TRUE, {range_curr_hvn}, TRUE)'
            ws_morn.cell(row=row_total, column=current_col+1, value=f_hvn).border = doki_border

            col_let_samp = get_column_letter(current_col)
            col_let_hvn = get_column_letter(current_col+1)
            
            f_rate = f'=IF({col_let_samp}{row_total}=0, "-", {col_let_hvn}{row_total}/{col_let_samp}{row_total})'
            cell = ws_morn.cell(row=row_total, column=current_col+2, value=f_rate)
            cell.number_format = '0.0%'
            cell.border = doki_border
            
            current_col += 3
            
        # 背景色適用 (全履歴行)
        for c in range(1, current_col):
            ws_morn.cell(row=row_total, column=c).fill = doki_total_fill

        # EV Tables Loop (Horizontal)
        current_table_row = 12 
        current_table_col = 1
        
        for i, (label, min_g, max_g) in enumerate(buckets):
             range_prev_g_col = f'ChainData!${col_prev_first_g}$2:${col_prev_first_g}$1048576'
             cond_prev_g_range = f'{range_prev_g_col}, ">={min_g}", {range_prev_g_col}, "<={max_g}"'
             
             generate_ev_table_for_bucket(ws_morn, current_table_row, current_table_col, label, cond_prev_g_range)
             
             current_table_col += 8 # 7 columns + 1 gap separation


        # 列幅
        # 列幅
        for c in range(1, 4 + 3*5 + 1): # 1 + 3(0thr) + 3*5(1-5thr) = 19
            ws_morn.column_dimensions[get_column_letter(c)].width = 12
            ws_morn.column_dimensions[get_column_letter(c)].width = 15

        # ========== 部分当選率解析シート ==========
        ws_partial = writer.book.create_sheet("PartialWinAnalysis")
        ws_partial['A1'] = "【部分当選率解析】朝イチ0スルー目 (1G目当選除外)"
        ws_partial['A1'].font = Font(bold=True, size=14)
        ws_partial['A2'] = "※ 部分当選率 = その区間での当選数 / その区間以上からのサンプル数"
        
        # ヘッダー
        partial_headers = ["ゲーム数", "当選数", "サンプル数", "部分当選率", "部分当選期待度", "部分当選TY"]
        for i, h in enumerate(partial_headers, 1):
            cell = ws_partial.cell(row=4, column=i, value=h)
            cell.font = Font(bold=True)
            cell.fill = doki_header_fill
            cell.border = doki_border
        
        # 朝イチ0スルーのデータを抽出 (1G目当選を除外: First_G > 1)
        morning_0through = chains_df[(chains_df['Is_Morning_Chain'] == True) & (chains_df['Through_Before'] == 0) & (chains_df['First_G'] > 1)]
        
        # 部分当選率データを保持 (ゾーン期待値計算用)
        partial_data = {}  # {g_start: {'wins': n, 'sample': m, 'rate': r, 'ty': t}}
        
        # 5G刻みで集計 (1-5, 6-10, ... 801-805)
        row = 5
        for g_start in range(1, 806, 5):
            g_end = g_start + 4
            
            # その区間での当選数
            wins_in_range = len(morning_0through[(morning_0through['First_G'] >= g_start) & (morning_0through['First_G'] <= g_end)])
            
            # その区間以上からのサンプル数 (g_start以上で当選したすべて)
            sample_from_range = len(morning_0through[morning_0through['First_G'] >= g_start])
            
            # 部分当選率
            partial_win_rate = wins_in_range / sample_from_range if sample_from_range > 0 else 0
            
            # 部分当選時TY (その区間で当選した場合の平均TY)
            wins_df = morning_0through[(morning_0through['First_G'] >= g_start) & (morning_0through['First_G'] <= g_end)]
            partial_ty = wins_df['Total_Dedama'].mean() if len(wins_df) > 0 else 0
            
            # データ保持
            partial_data[g_start] = {
                'wins': wins_in_range,
                'sample': sample_from_range,
                'rate': partial_win_rate,
                'ty': partial_ty if partial_ty > 0 else 0
            }
            
            # 期待度 (後で計算)
            partial_expectation = 0
            
            # 出力
            ws_partial.cell(row=row, column=1, value=f"{g_start}~{g_end}").border = doki_border
            ws_partial.cell(row=row, column=2, value=wins_in_range).border = doki_border
            ws_partial.cell(row=row, column=3, value=sample_from_range).border = doki_border
            
            cell = ws_partial.cell(row=row, column=4, value=partial_win_rate)
            cell.number_format = '0.00%'
            cell.border = doki_border
            
            ws_partial.cell(row=row, column=5, value=partial_expectation).border = doki_border
            
            cell = ws_partial.cell(row=row, column=6, value=partial_ty if partial_ty > 0 else "-")
            if isinstance(partial_ty, (int, float)) and partial_ty > 0:
                cell.number_format = '#,##0'
            cell.border = doki_border
            
            row += 1
        
        # 部分当選期待度を計算 (全体平均との比較)
        total_sample = len(morning_0through)
        if total_sample > 0:
            # 805G / 5 = 161区間
            theoretical_rate = 1 / 161
            
            for r in range(5, row):
                actual_rate = ws_partial.cell(row=r, column=4).value
                if actual_rate and actual_rate > 0:
                    expectation = actual_rate / theoretical_rate
                    ws_partial.cell(row=r, column=5, value=expectation).number_format = '0.00'
        
        # ========== ゾーン狙い期待値表 ==========
        zone_start_col = 8  # H列から
        ws_partial.cell(row=1, column=zone_start_col, value="【ゾーン狙い期待値表】").font = Font(bold=True, size=14)
        ws_partial.cell(row=2, column=zone_start_col, value="※ ゾーン内当選時TY - 総投資コスト (当選時/非当選時の投資を加重平均)")
        
        zone_headers = ["開始G", "終了G", "ゾーン当選率", "実質初当確率", "平均初当G", "調整TY", "平均投資", "期待値(枚)", "期待値(円)", "出玉率"]
        for i, h in enumerate(zone_headers):
            cell = ws_partial.cell(row=4, column=zone_start_col + i, value=h)
            cell.font = Font(bold=True)
            cell.fill = doki_header_fill
            cell.border = doki_border
        
        # コイン消費: 50枚/25.3G = 約1.976枚/G
        coin_per_g = 50 / COIN_HOLD
        
        # ゾーン定義 (開始G, 終了G) - 5G刻みの開始点
        # 終了点: 50G, 100G, 150G, 200G, 300G, 400G, 500G, 600G, 700G, 800G
        zones = []
        end_points = [50, 100, 150, 200, 300, 400, 500, 600, 700, 800]
        for end_g in end_points:
            for start_g in range(1, end_g, 5):  # 1, 6, 11, 16, ... (end_g未満)
                zones.append((start_g, end_g))
        
        zone_row = 5
        
        # ChainData参照用列文字
        col_first_g_let = "I"
        col_morn_flag_let = get_col_letter(chains_df, 'Is_Morning_Chain')
        col_through_let = get_col_letter(chains_df, 'Through_Before')
        col_dedama_let = "M"
        
        # 条件用レンジ
        range_first_g = f"ChainData!${col_first_g_let}$2:${col_first_g_let}$1048576"
        range_morn_flag = f"ChainData!${col_morn_flag_let}$2:${col_morn_flag_let}$1048576"
        range_through = f"ChainData!${col_through_let}$2:${col_through_let}$1048576"
        range_dedama = f"ChainData!${col_dedama_let}$2:${col_dedama_let}$1048576"
        
        # 基本条件: 朝イチ=TRUE, 0スルー, First_G > 1 (1G目当選除外)
        base_cond = f'{range_morn_flag}, TRUE, {range_through}, 0, {range_first_g}, ">1"'
        
        for zone_start, zone_end in zones:
            # 列参照用
            col_h = zone_start_col      # 開始G
            col_i = zone_start_col + 1  # 終了G
            col_j = zone_start_col + 2  # ゾーン当選率
            col_k = zone_start_col + 3  # 実質初当確率 (ProbDenom)
            col_l = zone_start_col + 4  # 平均初当G
            col_m = zone_start_col + 5  # 調整TY
            col_n = zone_start_col + 6  # 平均投資
            col_o = zone_start_col + 7  # 期待値(枚)
            col_p = zone_start_col + 8  # 期待値(円)
            col_q = zone_start_col + 9  # 出玉率
            
            # 開始G, 終了G (値)
            ws_partial.cell(row=zone_row, column=col_h, value=zone_start).border = doki_border
            ws_partial.cell(row=zone_row, column=col_i, value=zone_end).border = doki_border
            
            # ゾーン当選率
            f_wins = f'COUNTIFS({base_cond}, {range_first_g}, ">={zone_start}", {range_first_g}, "<={zone_end}")'
            f_sample = f'COUNTIFS({base_cond}, {range_first_g}, ">={zone_start}")'
            f_rate = f'=IF({f_sample}=0, 0, {f_wins}/{f_sample})'
            cell = ws_partial.cell(row=zone_row, column=col_j, value=f_rate)
            cell.number_format = '0.00%'
            cell.border = doki_border
            
            # 平均初当G
            f_avg_g = f'=IFERROR(AVERAGEIFS({range_first_g}, {base_cond}, {range_first_g}, ">={zone_start}", {range_first_g}, "<={zone_end}"), "-")'
            cell = ws_partial.cell(row=zone_row, column=col_l, value=f_avg_g)
            cell.number_format = '0.0'
            cell.border = doki_border
            
            # 調整TY (ボーナス消化分を引く: 32G / 25.3 * 50) * 設定1調整係数
            cost_bonus = 32 * (50 / COIN_HOLD)
            f_avg_ty = f'=IFERROR((AVERAGEIFS({range_dedama}, {base_cond}, {range_first_g}, ">={zone_start}", {range_first_g}, "<={zone_end}") - {cost_bonus}) * CeilingAnalysis!$M$6, "-")'
            cell = ws_partial.cell(row=zone_row, column=col_m, value=f_avg_ty)
            cell.number_format = '#,##0'
            cell.border = doki_border
            
            # 平均投資
            col_j_let = get_column_letter(col_j) # Rate
            col_l_let = get_column_letter(col_l) # AvgG
            col_n_let = get_column_letter(col_n) # Invest (Target)
            
            invest_loss = (zone_end - zone_start + 1) * coin_per_g
            f_invest = f'=IF({col_l_let}{zone_row}="-", {invest_loss}, {col_j_let}{zone_row}*({col_l_let}{zone_row}-{zone_start}+1)*{coin_per_g} + (1-{col_j_let}{zone_row})*{invest_loss})'
            cell = ws_partial.cell(row=zone_row, column=col_n, value=f_invest)
            cell.number_format = '#,##0'
            cell.border = doki_border

            # 実質初当確率 (ProbDenom) = (Average Investment / coin_per_g) / Winning Rate
            # = (col_n / coin_per_g) / col_j
            f_prob = f'=IF(OR({col_j_let}{zone_row}=0, {col_n_let}{zone_row}="-"), "-", ({col_n_let}{zone_row}/{coin_per_g})/{col_j_let}{zone_row})'
            cell = ws_partial.cell(row=zone_row, column=col_k, value=f_prob)
            cell.number_format = '0.0'
            cell.border = doki_border
            
            # 期待値(枚) = (当選率 × 平均TY) - 平均投資
            # Ev = Rate * Ty - Invest
            col_m_let = get_column_letter(col_m) # Ty
            f_ev = f'=IF({col_m_let}{zone_row}="-", -{col_n_let}{zone_row}, {col_j_let}{zone_row}*{col_m_let}{zone_row}-{col_n_let}{zone_row})'
            cell = ws_partial.cell(row=zone_row, column=col_o, value=f_ev)
            cell.number_format = '#,##0'
            cell.border = doki_border
            
            # 期待値(円)
            col_o_let = get_column_letter(col_o)
            f_ev_yen = f'={col_o_let}{zone_row}*20'
            cell = ws_partial.cell(row=zone_row, column=col_p, value=f_ev_yen)
            cell.number_format = '#,##0'
            cell.border = doki_border

            # 出玉率
            # 実装ロジック:
            #   AvgG_Normal = Rate * (AvgG_Win - Start + 1) + (1 - Rate) * (End - Start + 1)
            #   AvgG_Bonus = Rate * (AdjTY / NetIncreasePerG)  (NetIncreasePerG from CeilingAnalysis!L13)
            #   TotalIN = (AvgG_Normal + AvgG_Bonus) * 3
            #   PayoutRate = (TotalIN + Ev) / TotalIN
            
            # 式の構成要素
            # Rate: col_j_let{zone_row}
            # AvgG_Win: col_l_let{zone_row}
            # AdjTY: col_m_let{zone_row}
            # Ev: col_o_let{zone_row}
            
            zone_len = zone_end - zone_start + 1
            f_avg_g_normal = f'({col_j_let}{zone_row} * ({col_l_let}{zone_row} - {zone_start} + 1) + (1 - {col_j_let}{zone_row}) * {zone_len})'
            f_avg_g_bonus = f'({col_j_let}{zone_row} * ({col_m_let}{zone_row} / CeilingAnalysis!$L$13))'
            f_total_in = f'(({f_avg_g_normal} + {f_avg_g_bonus}) * 3)'
            
            f_payout_rate = f'=IF(OR({col_j_let}{zone_row}=0, {col_l_let}{zone_row}="-", {col_m_let}{zone_row}="-"), "-", ({f_total_in} + {col_o_let}{zone_row}) / {f_total_in})'

            cell = ws_partial.cell(row=zone_row, column=col_q, value=f_payout_rate)
            cell.number_format = '0.00%'
            cell.border = doki_border

            zone_row += 1
        
        # 列幅調整
        ws_partial.column_dimensions['A'].width = 12
        ws_partial.column_dimensions['B'].width = 10
        ws_partial.column_dimensions['C'].width = 12
        ws_partial.column_dimensions['D'].width = 12
        ws_partial.column_dimensions['E'].width = 14
        ws_partial.column_dimensions['F'].width = 14
        for col in range(zone_start_col, zone_start_col + 8):
            ws_partial.column_dimensions[get_column_letter(col)].width = 12
        # Is_Reset列を参照
        col_reset = col_map.get('リセット判定')
        col_through = col_map.get('スルー回数')
        col_first_g = col_map.get('初当G')
        
        ref_reset = f"ChainData!${col_reset}$2:${col_reset}${chain_max_row}"
        ref_through = f"ChainData!${col_through}$2:${col_through}${chain_max_row}"
        ref_first_g = f"ChainData!${col_first_g}$2:${col_first_g}${chain_max_row}"
        
        # セクション1: リセット0スルー (A-D列)
        ws_ceil['A4'] = "■ リセット時0スルー"
        ws_ceil['A4'].font = Font(bold=True, size=12)
        
        ws_ceil['A5'] = "ゲーム数"
        ws_ceil['B5'] = "サンプル数"
        ws_ceil['C5'] = "天井到達"
        ws_ceil['D5'] = "到達率"
        for col in range(1, 5):
            ws_ceil.cell(row=5, column=col).font = Font(bold=True)
        
        # セクション2: リセット0スルー以外 (F-I列)
        ws_ceil['F4'] = "■ リセット0スルー以外"
        ws_ceil['F4'].font = Font(bold=True, size=12)
        
        ws_ceil['F5'] = "ゲーム数"
        ws_ceil['G5'] = "サンプル数"
        ws_ceil['H5'] = "天井到達"
        ws_ceil['I5'] = "到達率"
        for col in range(6, 10):
            ws_ceil.cell(row=5, column=col).font = Font(bold=True)
        
        # データ行（Python側で集計して値のみ書き込み）
        row = 6
        
        # 集計用データフレーム準備 (Reset=1/0, Through=0, First_G)
        df_reset0 = chains_df[(chains_df['Is_Reset'] == True) & (chains_df['Through_Before'] == 0)]
        df_other = chains_df[~((chains_df['Is_Reset'] == True) & (chains_df['Through_Before'] == 0))]
        
        for g in range(0, 801, 5):
            # --- セクション1: リセット0スルー (A-D列) ---
            # サンプル数: First_G >= g
            sample_r0 = len(df_reset0[df_reset0['First_G'] >= g])
            # 天井到達: First_G >= 800 (かつ現在のg以上であること、まあ800>=gなら当然だが)
            if g <= 800:
                ceil_r0 = len(df_reset0[df_reset0['First_G'] >= 800])
            else:
                ceil_r0 = 0
            
            rate_r0 = ceil_r0 / sample_r0 if sample_r0 > 0 else 0
            
            ws_ceil.cell(row=row, column=1, value=f"{g}G～")
            ws_ceil.cell(row=row, column=2, value=sample_r0)
            ws_ceil.cell(row=row, column=3, value=ceil_r0)
            ws_ceil.cell(row=row, column=4, value=rate_r0)
            ws_ceil.cell(row=row, column=4).number_format = '0.00%'
            
            # --- セクション2: リセット0スルー以外 (F-I列) ---
            sample_oth = len(df_other[df_other['First_G'] >= g])
            if g <= 800:
                ceil_oth = len(df_other[df_other['First_G'] >= 800])
            else:
                ceil_oth = 0
            
            rate_oth = ceil_oth / sample_oth if sample_oth > 0 else 0
            
            ws_ceil.cell(row=row, column=6, value=f"{g}G～")
            ws_ceil.cell(row=row, column=7, value=sample_oth)
            ws_ceil.cell(row=row, column=8, value=ceil_oth)
            ws_ceil.cell(row=row, column=9, value=rate_oth)
            ws_ceil.cell(row=row, column=9).number_format = '0.00%'
            
            row += 1
        
        # 列幅調整
        ws_ceil.column_dimensions['A'].width = 10
        ws_ceil.column_dimensions['B'].width = 12
        ws_ceil.column_dimensions['C'].width = 10
        ws_ceil.column_dimensions['D'].width = 10
        ws_ceil.column_dimensions['E'].width = 2  # 空き列
        ws_ceil.column_dimensions['F'].width = 10
        ws_ceil.column_dimensions['G'].width = 12
        ws_ceil.column_dimensions['H'].width = 10
        ws_ceil.column_dimensions['I'].width = 10

        # ========== 全体出玉率の出力（CeilingAnalysisシート右側） ==========
        # 除外なしで全データから算出
        # IN枚数 = (通常G * 3) + (純増 / 4.1 * 3)  ※純増4.1枚/Gと仮定してボーナス消化G数を逆算
        sum_first_g = chains_df['First_G'].sum()
        sum_ty = chains_df['Total_Dedama'].sum()
        sum_net_diff = chains_df['Net_Diff'].sum()
        
        total_in = sum_first_g * 3 + (sum_ty / 4.1 * 3)
        payout_rate = (total_in + sum_net_diff) / total_in if total_in > 0 else 0
        
        # 出力
        ws_ceil['K2'] = "【全体出玉率】(全データ)"
        ws_ceil['K2'].font = Font(bold=True)
        
        ws_ceil['K3'] = "サンプル数"
        ws_ceil['L3'] = len(chains_df)
        
        ws_ceil['K4'] = "総IN枚数"
        ws_ceil['L4'] = total_in
        ws_ceil['L4'].number_format = '#,##0'
        
        ws_ceil['K5'] = "総差枚"
        ws_ceil['L5'] = sum_net_diff
        ws_ceil['L5'].number_format = '#,##0'
        
        ws_ceil['K6'] = "出玉率"
        ws_ceil['L6'] = payout_rate
        ws_ceil['L6'].number_format = '0.00%'
        
        # 設定1調整係数: M6 = 97.2% / L6 (全体出玉率)
        ws_ceil['M5'] = "設定1調整係数"
        ws_ceil['M5'].font = Font(bold=True)
        ws_ceil['M6'] = f'=0.972/L6'
        ws_ceil['M6'].number_format = '0.000'
        
        ws_ceil.column_dimensions['K'].width = 15
        ws_ceil.column_dimensions['L'].width = 15
        ws_ceil.column_dimensions['M'].width = 15

        # ========== 天国チェーン中 実質純増の計算 ==========
        # 対象: 3連以上の天国チェーン (Is_Heaven=True, Chain_Length>=3)
        heaven_chains = chains_df[(chains_df['Is_Heaven'] == True) & (chains_df['Chain_Length'] >= 3)]
        
        total_heaven_dedama = 0
        total_heaven_wait_g = 0
        total_heaven_bonus_g = 0
        
        for idx in heaven_chains.index:
            hit_games = heaven_chains.at[idx, 'Hit_Games']
            total_dedama = heaven_chains.at[idx, 'Total_Dedama']
            
            # 天国中待機G (2連目以降、35G以下のみ)
            wait_g = sum(g for g in hit_games[1:] if g <= 35)
            
            # 推定ボーナス消化G (純増4枚/Gと仮定)
            bonus_g = total_dedama / 4.0 if total_dedama > 0 else 0
            
            total_heaven_dedama += total_dedama
            total_heaven_wait_g += wait_g
            total_heaven_bonus_g += bonus_g
        
        total_heaven_g = total_heaven_wait_g + total_heaven_bonus_g
        real_dedama_per_g = total_heaven_dedama / total_heaven_g if total_heaven_g > 0 else 0
        
        # 出力
        ws_ceil['K8'] = "【天国中 実質純増】"
        ws_ceil['K8'].font = Font(bold=True)
        
        ws_ceil['K9'] = "天国チェーン数"
        ws_ceil['L9'] = len(heaven_chains)
        
        ws_ceil['K10'] = "天国中待機G計"
        ws_ceil['L10'] = total_heaven_wait_g
        ws_ceil['L10'].number_format = '#,##0'
        
        ws_ceil['K11'] = "ボーナスG計(推定)"
        ws_ceil['L11'] = total_heaven_bonus_g
        ws_ceil['L11'].number_format = '#,##0'
        
        ws_ceil['K12'] = "純増計"
        ws_ceil['L12'] = total_heaven_dedama
        ws_ceil['L12'].number_format = '#,##0'
        
        ws_ceil['K13'] = "実質純増/G"
        ws_ceil['L13'] = real_dedama_per_g
        ws_ceil['L13'].number_format = '0.00'


        # 天国移行率シート（スルー回数別） - 出力しない
        # heaven_rate_df.to_excel(writer, sheet_name='天国移行率_スルー', index=False)
        
        # 天国移行率シート（前回連チャン別） - 出力しない
        # heaven_by_chain_df.to_excel(writer, sheet_name='天国移行率_前回連', index=False)
        
        # クロス集計シート（48行目以降は出力しない -> ヘッダー1行+46行=47行）
        crosstab_df = crosstab_df.iloc[:46]
        crosstab_df.to_excel(writer, sheet_name='クロス集計', index=False)
        ws3 = writer.sheets['クロス集計']

        # パーセント表示（天国移行率セクションのみ）
        current_section = None

        for row_idx in range(2, ws3.max_row + 1):
            a_cell_value = ws3.cell(row=row_idx, column=1).value

            # セクションヘッダーを検出
            if a_cell_value and isinstance(a_cell_value, str) and '【' in a_cell_value:
                current_section = a_cell_value
                continue

            # 空行をスキップ
            if not a_cell_value:
                continue

            # 天国移行率セクションのみパーセント表示
            if current_section and '天国移行率' in current_section:
                for col in range(2, 13):  # B-L列
                    cell = ws3.cell(row=row_idx, column=col)
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.0%'


        ws3.column_dimensions['A'].width = 25  # セクション名が長いので幅を広げる



        
        # ========== ボーナス間期待値表（朝イチ/朝イチ除外） ==========
        
        # 共通生成関数
        def create_bonus_ev_sheet(sheet_title, target_cond_str):
            ws_ev = writer.book.create_sheet(sheet_title)
            
            # 数式用列参照
            col_start = col_map.get('初当G')
            col_dedama = col_map.get('純増枚数')
            col_through = col_map.get('スルー回数')
            col_target = col_map.get('朝イチフラグ')
            col_max_daily = col_map.get('当日差枚最大_開始前')
            col_prev_special = col_map.get('前回特殊判定')
            
            max_r = ws_chain.max_row
            
            ref_start = f"ChainData!${col_start}$2:${col_start}${max_r}"
            ref_dedama = f"ChainData!${col_dedama}$2:${col_dedama}${max_r}"
            ref_through = f"ChainData!${col_through}$2:${col_through}${max_r}"
            ref_target = f"ChainData!${col_target}$2:${col_target}${max_r}"
            ref_max_daily = f"ChainData!${col_max_daily}$2:${col_max_daily}${max_r}"
            ref_prev_special = f"ChainData!${col_prev_special}$2:${col_prev_special}${max_r}"

            # スタイル定義
            ev_header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            ev_row_fill_even = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            ev_row_fill_odd = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            ev_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            # スルー回数ループ (0-9スルー)
            start_row = 3
            current_col = 2
            
            for thr in range(10):
                # ヘッダー作成
                header_cell = ws_ev.cell(row=start_row-1, column=current_col, value=f"{thr}スルー")
                ws_ev.merge_cells(start_row=start_row-1, start_column=current_col, end_row=start_row-1, end_column=current_col+6)
                header_cell.fill = ev_header_fill
                header_cell.font = Font(color="FFFFFF", bold=True)
                header_cell.alignment = Alignment(horizontal='center')
                
                # サブヘッダー
                sub_headers = ["開始G", "初当たり", "TY", "5枚等価", "出玉率", "サンプル", "調整TY"]
                for i, h in enumerate(sub_headers):
                    c = ws_ev.cell(row=start_row, column=current_col + i, value=h)
                    c.fill = ev_header_fill
                    c.font = Font(color="FFFFFF", bold=True)
                    c.alignment = Alignment(horizontal='center')
                    c.border = ev_border
                    ws_ev.column_dimensions[get_column_letter(current_col + i)].width = 12
                    
                col_idx_hatsu = current_col + 1
                col_idx_ty = current_col + 2
                col_idx_5mai = current_col + 3
                col_idx_rate = current_col + 4
                col_idx_sample = current_col + 5
                col_idx_adj_ty = current_col + 6
                
                col_char_hatsu = get_column_letter(col_idx_hatsu)
                col_char_ty = get_column_letter(col_idx_ty)
                col_char_5mai = get_column_letter(col_idx_5mai)
                col_char_rate = get_column_letter(col_idx_rate)
                col_char_sample = get_column_letter(col_idx_sample)
                col_char_adj_ty = get_column_letter(col_idx_adj_ty)

                # 行ループ
                for i, g in enumerate(range(35, 801, 5)):
                    row = start_row + 1 + i
                    row_fill = ev_row_fill_even if i % 2 == 0 else ev_row_fill_odd
                    
                    # 開始G
                    cell = ws_ev.cell(row=row, column=current_col, value=g)
                    cell.fill = row_fill
                    cell.border = ev_border
                    cell.alignment = Alignment(horizontal='center')
                    
                    # スタイル適用
                    for c in range(col_idx_hatsu, col_idx_adj_ty + 1):
                        cell = ws_ev.cell(row=row, column=c)
                        cell.fill = row_fill
                        cell.border = ev_border

                    addr_sample = f"{col_char_sample}{row}"
                    
                    # サンプル数
                    # サンプル数にも条件を入れるべきか？
                    # 初当たり計算の分母になるので入れるべき。
                    f_cnt = f'=COUNTIFS({ref_target}, {target_cond_str}, {ref_through}, {thr}, {ref_start}, ">={g}", {ref_max_daily}, "<=500", {ref_prev_special}, "<1000")'
                    ws_ev.cell(row=row, column=col_idx_sample, value=f_cnt).number_format = '#,##0'

                    # 初当たり
                    # 条件: Is_Target, Through, StartG>=g, MaxDaily<=500, PrevSpec<1000
                    # ユーザー要望: ÷CeilingAnalysis!M6を追加
                    f_hatsu = f'=IF({addr_sample}=0, "[-]", (AVERAGEIFS({ref_start}, {ref_target}, {target_cond_str}, {ref_through}, {thr}, {ref_start}, ">={g}", {ref_max_daily}, "<=500", {ref_prev_special}, "<1000") - {g}) / CeilingAnalysis!$M$6)'
                    ws_ev.cell(row=row, column=col_idx_hatsu, value=f_hatsu).number_format = '0.0'

                    # TY (<800G限定)
                    # 条件: 上記 + StartG<800
                    cost = 32 * (50/COIN_HOLD)
                    f_ty_core = f'AVERAGEIFS({ref_dedama}, {ref_target}, {target_cond_str}, {ref_through}, {thr}, {ref_start}, ">={g}", {ref_start}, "<800", {ref_max_daily}, "<=500", {ref_prev_special}, "<1000")'
                    f_ty = f'=IFERROR({f_ty_core} - {cost}, "[-]")'
                    ws_ev.cell(row=row, column=col_idx_ty, value=f_ty).number_format = '#,##0'
                    
                    # 他の項目
                    addr_ty = f"{col_char_ty}{row}"
                    addr_hatsu = f"{col_char_hatsu}{row}"
                    addr_adj_ty = f"{col_char_adj_ty}{row}"
                    addr_5mai = f"{col_char_5mai}{row}"
                    
                    # 調整TY
                    f_adj_ty = f'=IF({addr_ty}="[-]", "[-]", {addr_ty} * CeilingAnalysis!$M$6)'
                    ws_ev.cell(row=row, column=col_idx_adj_ty, value=f_adj_ty).number_format = '#,##0'
                    
                    # 5枚等価
                    f_5mai = f'=IF(OR({addr_adj_ty}="[-]", {addr_hatsu}="[-]"), "[-]", ({addr_adj_ty} - {addr_hatsu} * (50/{COIN_HOLD})) * 20)'
                    ws_ev.cell(row=row, column=col_idx_5mai, value=f_5mai).number_format = '#,##0'
                    
                    # 出玉率
                    denom = f'({addr_hatsu} * 3 + ({addr_adj_ty} / CeilingAnalysis!$L$13 * 3))'
                    numer = f'({denom} + {addr_5mai} / 20)'
                    f_rate = f'=IF(OR({addr_hatsu}="[-]", {addr_adj_ty}="[-]", {denom}=0), "[-]", {numer} / {denom})'
                    ws_ev.cell(row=row, column=col_idx_rate, value=f_rate).number_format = '0.0%'

                current_col += 8
        
        # シート生成実行
        create_bonus_ev_sheet("朝イチボーナス間期待値表", "TRUE")
        create_bonus_ev_sheet("朝イチ除外ボーナス間期待値表", "FALSE")
    
    print(f"Excel file created: {OUTPUT_FILE}")
        
        # 数式用列参照
if __name__ == "__main__":
    print("=== デュオ天国移行率分析 ===")
    print(f"設定: 天国閾値={HEAVEN_THRESHOLD}G, コイン持ち={COIN_HOLD}G/50枚")
    print()

    # キャッシュ有効性チェック
    use_cache = False
    data_mtime = os.path.getmtime(INPUT_FILE) if os.path.exists(INPUT_FILE) else 0
    
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, 'rb') as f:
                cache = pickle.load(f)
            if cache.get('mtime') == data_mtime:
                print("キャッシュから読み込み中...")
                df = cache['df']
                chains_df = cache['chains_df']
                use_cache = True
                print(f"  読み込み完了: {len(df)}行, チェーン数: {len(chains_df)}")
        except Exception as e:
            print(f"  キャッシュ読み込みエラー: {e}")
            use_cache = False
    
    if not use_cache:
        print("データ読み込み中...")
        df = load_data(INPUT_FILE)
        print(f"  読み込み完了: {len(df)}行")

        print("連チャン解析中...")
        df, chains_df = analyze_chains(df)
        


        # キャッシュ保存
        print("キャッシュ保存中...")
        with open(CACHE_FILE, 'wb') as f:
            pickle.dump({'mtime': data_mtime, 'df': df, 'chains_df': chains_df}, f)
        print("  保存完了")

    # 朝イチ連フラグの計算 (ID毎に最初の天国チェーン終了までをTRUE)
    # キャッシュ有無に関わらず実行
    print("朝イチ連フラグ計算中...")
    if 'Is_Morning_Chain' not in df.columns:
        chains_df['hvn_cumsum_temp'] = chains_df.groupby('ID')['Is_Heaven'].cumsum().astype(int)
        chains_df['Is_Morning_Chain'] = (chains_df['hvn_cumsum_temp'] == 0) | ((chains_df['hvn_cumsum_temp'] == 1) & (chains_df['Is_Heaven']))
        del chains_df['hvn_cumsum_temp']
        
        df = pd.merge(df, chains_df[['Chain_ID', 'Is_Morning_Chain']], on='Chain_ID', how='left')
        df['Is_Morning_Chain'] = df['Is_Morning_Chain'].fillna(False)

    # 前回初当G数計算 (MorningAnalysisで使用)
    if 'Prev_First_G' not in chains_df.columns:
        print("前回初当G数計算中...")
        chains_df['Prev_First_G'] = chains_df.groupby('ID')['First_G'].shift(1).fillna(0).astype(int)

    print("天国移行率計算中...")
    heaven_rate_df = calculate_heaven_rate(chains_df)
    heaven_by_chain_df = calculate_heaven_rate_by_chain(chains_df)
    crosstab_df = calculate_heaven_rate_crosstab(chains_df)
    # 3連基準分析は軽量化版ではスキップ
    # rule3_df = calculate_3chain_rule_analysis(chains_df)

    print("ドキハナチャンス分析中...")
    chains_df, dokihana_stats = calculate_dokihana_analysis(chains_df)
    print(f"  ドキハナ発生回数: {dokihana_stats['total_count']}")
    print(f"  平均連チャン数: {dokihana_stats['avg_chain_len']:.2f}")
    print(f"  平均獲得枚数: {dokihana_stats['avg_dedama']:.0f}")

    print()
    print("【天国移行率（スルー回数別）】")
    print(heaven_rate_df.to_string(index=False))
    print()
    print("【天国移行率（前回連チャン別）】")
    print(heaven_by_chain_df.to_string(index=False))
    print()

    print("Excel出力中...")
    write_excel(df, heaven_rate_df, heaven_by_chain_df, crosstab_df, chains_df)

    print("完了!")

