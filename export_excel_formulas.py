import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, Series
import traceback

# File path
INPUT_FILE = r'C:\Users\ilove\Desktop\解析\20251223_duo2.csv'
OUTPUT_FILE = r'C:\Users\ilove\Desktop\解析\analysis_result_v76.xlsx'

# Constants
COIN_HOLD = 25.3  # G per 50 coins
HEAVEN_THRESHOLD = 35
# Row limits removed - will use actual data size 

def load_data(filepath):
    try:
        df = pd.read_csv(filepath, encoding='utf-8-sig', header=0, names=[
            'Hall_Name', 'Hall_URL', 'Machine_No', 'Machine_Name_Used', 'Machine_URL', 
            'Count', 'Time', 'Start', 'Dedama', 'Status'
        ])
    except:
        df = pd.read_csv(filepath, encoding='cp932', header=0, names=[
            'Hall_Name', 'Hall_URL', 'Machine_No', 'Machine_Name_Used', 'Machine_URL', 
            'Count', 'Time', 'Start', 'Dedama', 'Status'
        ])
    
    df['Date'] = df['Machine_URL'].str.extract(r'target_date=(\d{4}-\d{2}-\d{2})')
    df['Start'] = pd.to_numeric(df['Start'], errors='coerce').fillna(0).astype(int)
    df['Dedama'] = pd.to_numeric(df['Dedama'], errors='coerce').fillna(0).astype(int)
    df['Count'] = pd.to_numeric(df['Count'], errors='coerce').fillna(0).astype(int)
    
    # Remove duplicate rows based on key columns
    before_count = len(df)
    df = df.drop_duplicates(subset=['Hall_Name', 'Date', 'Machine_No', 'Count'], keep='first')
    after_count = len(df)
    if before_count != after_count:
        print(f"Duplicates removed: {before_count - after_count} rows")
    
    df = df.sort_values(by=['Hall_Name', 'Date', 'Machine_No', 'Count'])
    return df

def process_logic(df):
    grouped = df.groupby(['Hall_Name', 'Date', 'Machine_No'])
    machine_stats = {} 
    
    # Initialize columns
    df['Chain_ID'] = 0
    df['Chain_Position'] = 0
    
    chain_counter = 0
    
    # Store per-chain RB through info for ChainData (chain_id -> {prev_thr_count, all_rb})
    chain_rb_info = {}
    
    for name, group in grouped:
        machine_hits = [] 
        for idx, row in group.iterrows():
            # Include Status for RB/BB detection
            machine_hits.append((idx, row['Start'], row.get('Status', '')))
        if not machine_hits: continue
        
        chains = [] 
        current_chain = [machine_hits[0]]
        for i in range(1, len(machine_hits)):
            h_idx, games, status = machine_hits[i]
            if games <= HEAVEN_THRESHOLD:
                current_chain.append(machine_hits[i])
            else:
                chains.append(current_chain)
                current_chain = [machine_hits[i]]
        chains.append(current_chain)
        
        # Assign Chain_ID and Position to each hit
        chain_ids_for_machine = []
        for chain in chains:
            chain_counter += 1
            chain_ids_for_machine.append(chain_counter)
            for pos, (hit_idx, _, _) in enumerate(chain, start=1):
                df.at[hit_idx, 'Chain_ID'] = chain_counter
                df.at[hit_idx, 'Chain_Position'] = pos
        
        # Analyze RB-only through patterns for this machine
        # Find heaven chains (len >= 2) and check if all previous throughs were RB
        through_count_so_far = 0
        through_statuses = []  # Status of each through's first hit
        through_chain_ids = []  # Chain IDs of throughs
        
        for i, chain in enumerate(chains):
            chain_len = len(chain)
            first_hit_status = chain[0][2]  # Status of first hit in chain
            current_chain_id = chain_ids_for_machine[i]
            
            if chain_len >= 2:
                # This is a heaven chain
                is_heaven = True
            else:
                is_heaven = False

            # Process through history if we had accumulated throughs
            if through_count_so_far > 0:
                # Determine if all throughs were RB
                final_all_rb = all('RB' in str(s).upper() for s in through_statuses)
            else:
                final_all_rb = False
            
            # Record chain info
            chain_rb_info[current_chain_id] = {
                'prev_thr_count': through_count_so_far,
                'all_rb': final_all_rb
            }
            
            if is_heaven:
                # Reset for next heaven search
                through_count_so_far = 0
                through_statuses = []
                through_chain_ids = []
            else:
                # This is a through (single hit chain)
                through_count_so_far += 1
                through_statuses.append(first_hit_status)
        
        morning_chain = chains[0]
        morning_games = morning_chain[0][1]
        
        through_count = 0
        for ch in chains:
            if len(ch) >= 2:
                break
            through_count += 1
        
        machine_stats[name] = {'MorningGames': morning_games, 'ThroughCount': through_count}
    
    # Store chain RB info in machine_stats for later use
    machine_stats['_chain_rb_info'] = chain_rb_info

    return df, machine_stats

def analyze_intervals(machine_stats):
    max_g = 0
    for k, s in machine_stats.items():
        if isinstance(k, str) and k.startswith('_'):
            continue  # Skip internal keys like _rb_through_stats
        if s['MorningGames'] > max_g: max_g = s['MorningGames']
    limit = ((max_g // 50) + 1) * 50
    ranges = []
    for i in range(0, limit, 50):
        ranges.append((i + 1, i + 50))
    bin_data = {r: {'count': 0, 'through_list': []} for r in ranges}
    for k, s in machine_stats.items():
        if isinstance(k, str) and k.startswith('_'):
            continue  # Skip internal keys
        mg = s['MorningGames']
        tc = s['ThroughCount']
        bin_idx = (mg - 1) // 50
        if bin_idx < 0: bin_idx = 0 
        r_key = (bin_idx * 50 + 1, bin_idx * 50 + 50)
        if r_key not in bin_data: bin_data[r_key] = {'count': 0, 'through_list': []}
        bin_data[r_key]['count'] += 1
        bin_data[r_key]['through_list'].append(tc)
    final_stats = []
    for k in sorted(bin_data.keys()):
        d = bin_data[k]
        cnt = d['count']
        avg_t = sum(d['through_list']) / cnt if cnt > 0 else 0
        final_stats.append({'Range': f"{k[0]}-{k[1]}G", 'Hit_Count': cnt, 'Avg_Through': avg_t})
    return pd.DataFrame(final_stats)

def write_excel(df, interval_df, machine_stats):
    print("Writing Excel v10...")
    # Extract chain RB info
    chain_rb_info = machine_stats.pop('_chain_rb_info', {})
    
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        # Include Chain_ID and Chain_Position calculated in Python
        output_df = df[['Hall_Name', 'Date', 'Machine_No', 'Count', 'Time', 'Start', 'Dedama', 'Status', 'Chain_ID', 'Chain_Position']].copy()
        
        jp_map_data = {
            'Hall_Name': 'ホール名', 'Date': '日付', 'Machine_No': '台番号', 'Count': '当たりNo',
            'Time': '当選時刻', 'Start': '当選G数', 'Dedama': '獲得枚数', 'Status': 'ボーナス種別',
            'Chain_ID': '連チャンID', 'Chain_Position': '連チャン位置'
        }
        output_df.rename(columns=jp_map_data, inplace=True)
        output_df['日付'] = output_df['日付'].astype(str)
        output_df.to_excel(writer, sheet_name='Data', index=False)
        
        workbook = writer.book
        ws_data = writer.sheets['Data']
        max_r = ws_data.max_row
        
        # I/J columns (連チャンID, 連チャン位置) are now direct values from Python in columns 9, 10
        # Add remaining calculated columns starting from K (column 11)
        headers = ['朝イチ連', '投資詳細(枚)', '差枚詳細(枚)', '天国フラグ', '当日差枚', '有利区間G数']
        start_calcs = 11  # Start at column K
        for i, h in enumerate(headers):
            ws_data.cell(row=1, column=start_calcs + i, value=h)
            
        for r in range(2, max_r + 1):
            prev_r = r - 1
            
            # K: 朝イチ連 (column 11)
            if r == 2:
                ws_data.cell(row=r, column=11, value=True)
            else:
                f_cxt = f"OR(A{r}<>A{prev_r}, B{r}<>B{prev_r}, C{r}<>C{prev_r})"
                f_morn = f"=IF(J{r}=1, IF({f_cxt}, TRUE, FALSE), K{prev_r})"
                ws_data.cell(row=r, column=11, value=f_morn)

            # L: 投資詳細(枚) (column 12)
            ws_data.cell(row=r, column=12, value=f"=F{r}/{COIN_HOLD}*50")
            # M: 差枚詳細(枚) (column 13)
            ws_data.cell(row=r, column=13, value=f"=G{r}-L{r}")
            # N: 天国フラグ (column 14)
            ws_data.cell(row=r, column=14, value=f"=IF(F{r}<={HEAVEN_THRESHOLD}, 1, 0)")
            
            # O: 当日差枚 (column 15) - 前回終了時点の累積差枚
            if r == 2:
                ws_data.cell(row=r, column=15, value=0)
            else:
                f_same_mach = f"AND(A{r}=A{prev_r}, B{r}=B{prev_r}, C{r}=C{prev_r})"
                f_daily = f"=IF({f_same_mach}, O{prev_r}+M{prev_r}, 0)"
                ws_data.cell(row=r, column=15, value=f_daily)
            
            # P: 有利区間G数 (column 16) - 当選時点の有利区間G数
            # BB=51G, RB=21Gを加算、同一台なら累積
            f_bonus_g = f'IF(ISNUMBER(SEARCH("BB", H{r})), 51, IF(ISNUMBER(SEARCH("RB", H{r})), 21, 0))'
            f_current_g = f"F{r}+{f_bonus_g}"
            if r == 2:
                ws_data.cell(row=r, column=16, value=f"={f_current_g}")
            else:
                f_same_mach = f"AND(A{r}=A{prev_r}, B{r}=B{prev_r}, C{r}=C{prev_r})"
                f_yuuri = f"=IF({f_same_mach}, P{prev_r}+{f_current_g}, {f_current_g})"
                ws_data.cell(row=r, column=16, value=f_yuuri)

        # ChainData
        ws_chain = workbook.create_sheet("ChainData")
        chain_headers = [
            '連チャンID', 'ホール名', '日付', '台番号', '朝イチ連', 
            '連チャン回数', '合計獲得', '合計G数', '純増差枚', 
            '天国フラグ', '行番号', '次回天国行', '次天国スルー', '次回天国連荘', '特殊判定', '2連目G', '34/35G履歴', '連荘中TY',
            '初当Status', '前Thr数', '前Thr全RB', '初当有利区間G', 'ドキハナ成功',
            '初当たりG', '当日差枚', '前回初当G', '前々回初当G', '初401Gハマリ',
            'ドキハナ開始位置', 'ドキハナ連チャン数', 'ドキハナ獲得枚数'
        ]
        for i, h in enumerate(chain_headers):
            ws_chain.cell(row=1, column=1 + i, value=h)

        # Use actual data size instead of fixed limits
        data_row_limit = max_r + 100  # Actual data rows + small buffer
        
        # Estimate chain count (roughly data rows / average chain length, assume ~5)
        # Add buffer for future data
        estimated_chains = (max_r // 3) + 500
        chain_row_limit = estimated_chains
        
        print(f"Data rows: {max_r}, Chain limit: {chain_row_limit}")
        
        # Data ranges with actual limits
        d_id_rng = f"Data!$I$2:$I${data_row_limit}"
        d_dedama_rng = f"Data!$G$2:$G${data_row_limit}"
        d_start_rng = f"Data!$F$2:$F${data_row_limit}"
        d_id_full = "Data!$I:$I"
        
        # Pre-compute chain info from DataFrame (now has Chain_ID)
        chain_info = {}
        for idx, row in df.iterrows():
            cid = row.get('Chain_ID', 0)
            pos = row.get('Chain_Position', 0)
            if cid == 0:
                continue
            if cid not in chain_info:
                chain_info[cid] = {
                    'Hall_Name': row['Hall_Name'],
                    'Date': str(row['Date']),
                    'Machine_No': row['Machine_No'],
                    'First_Start_G': row['Start'] if pos == 1 else 0
                }
            elif pos == 1:
                # Ensure First_Start_G is set for position 1
                chain_info[cid]['First_Start_G'] = row['Start']
        
        max_chain_id = max(chain_info.keys()) if chain_info else 0
        
        for r in range(2, chain_row_limit + 1):
            chain_id = r - 1  # Row 2 = ChainID 1
            
            if chain_id in chain_info:
                info = chain_info[chain_id]
                # A: Direct value
                ws_chain.cell(row=r, column=1, value=chain_id)
                # B: Direct value  
                ws_chain.cell(row=r, column=2, value=info['Hall_Name'])
                # C: Direct value
                ws_chain.cell(row=r, column=3, value=info['Date'])
                # D: Direct value
                ws_chain.cell(row=r, column=4, value=info['Machine_No'])
            else:
                # Skip rows beyond max chain
                continue
                
            chk_a = f"A{r}=\"\""
            crit = f"A{r}"
            
            def get_look(target_col):
                return f"INDEX(Data!${target_col}:${target_col}, MATCH({crit}, {d_id_full}, 0))"

            # E: formula (remains as formula)
            ws_chain.cell(row=r, column=5, value=f"=IF({chk_a}, \"\", {get_look('K')})")
            
            # F: Length using COUNTIF with correct range
            f_len = f"COUNTIF({d_id_rng}, {crit})"
            ws_chain.cell(row=r, column=6, value=f"=IF({chk_a}, \"\", {f_len})")
            
            # G: Dedama Sum
            ws_chain.cell(row=r, column=7, value=f"=IF({chk_a}, \"\", SUMIF({d_id_rng}, {crit}, {d_dedama_rng}))")
            # H: Start Sum
            ws_chain.cell(row=r, column=8, value=f"=IF({chk_a}, \"\", SUMIF({d_id_rng}, {crit}, {d_start_rng}))")
            
            ws_chain.cell(row=r, column=9, value=f"=IF({chk_a}, \"\", G{r}-(H{r}/{COIN_HOLD}*50))")
            ws_chain.cell(row=r, column=10, value=f"=IF({chk_a}, \"\", F{r}>=2)")
            ws_chain.cell(row=r, column=11, value=f"=IF({chk_a}, \"\", ROW())")
            
            # L: Next Heaven using MINIFS (standard function, no @ issue)
            c_B_rng = f"$B$2:$B${chain_row_limit}"
            c_C_rng = f"$C$2:$C${chain_row_limit}"
            c_D_rng = f"$D$2:$D${chain_row_limit}"
            c_F_rng = f"$F$2:$F${chain_row_limit}"
            c_K_rng = f"$K$2:$K${chain_row_limit}"
            
            # MINIFS with _xlfn. prefix to prevent Excel adding @ operator
            # Find smallest K where: same Hall, same Date, same Machine, F>=2, K>current
            f_min = (f"_xlfn.MINIFS({c_K_rng}, {c_B_rng}, B{r}, {c_C_rng}, C{r}, {c_D_rng}, D{r}, "
                     f"{c_F_rng}, \">=2\", {c_K_rng}, \">\"&K{r})")
            
            ws_chain.cell(row=r, column=12, value=f"=IF({chk_a}, \"\", IFERROR({f_min},0))")
            
            # Use F<2 instead of J=FALSE to avoid boolean comparison issues
            f_thr = f"=IF({chk_a}, \"\", IF(F{r}<2, \"\", IF(L{r}=0, \"終了\", L{r}-K{r}-1)))"
            ws_chain.cell(row=r, column=13, value=f_thr)
            
            # N: Next Heaven Chain Length (lookup F value at row L)
            c_F_col = f"$F$2:$F${chain_row_limit}"
            c_K_col = f"$K$2:$K${chain_row_limit}"
            # If L>0, lookup F at row L using INDEX/MATCH on K column
            f_next_len = f"INDEX({c_F_col}, MATCH(L{r}, {c_K_col}, 0))"
            ws_chain.cell(row=r, column=14, value=f"=IF({chk_a}, \"\", IF(L{r}=0, \"\", IFERROR({f_next_len}, \"\")))")
            
            # O: 特殊判定 = 純増差枚 - 最後の獲得枚数 + 初当たりGの投資分
            # 初当たりGの投資分を足し戻すことで、初当たりの投資を除外
            d_id_col = f"Data!$I$2:$I${data_row_limit}"
            d_pos_col = f"Data!$J$2:$J${data_row_limit}"
            d_dedama_col = f"Data!$G$2:$G${data_row_limit}"
            d_start_col = f"Data!$F$2:$F${data_row_limit}"
            
            # 最後の当たりの獲得枚数
            f_max_pos = f"_xlfn.MAXIFS({d_pos_col}, {d_id_col}, A{r})"
            f_last_dedama = f"SUMIFS({d_dedama_col}, {d_id_col}, A{r}, {d_pos_col}, {f_max_pos})"
            
            # 初当たりのゲーム数 (位置=1の行のStart値)
            f_first_start = f"SUMIFS({d_start_col}, {d_id_col}, A{r}, {d_pos_col}, 1)"
            
            # 特殊判定 = 純増差枚 - 最後Dedama + 初当たりGの投資分
            f_first_invest = f"({f_first_start}/{COIN_HOLD}*50)"
            f_special = f"I{r} - {f_last_dedama} + {f_first_invest}"
            ws_chain.cell(row=r, column=15, value=f"=IF({chk_a}, \"\", IFERROR({f_special}, \"\"))")
            
            # P: 2連目G（位置=2の行のStart値）- 2連以上の場合のみ
            f_second_start = f"SUMIFS({d_start_col}, {d_id_col}, A{r}, {d_pos_col}, 2)"
            ws_chain.cell(row=r, column=16, value=f"=IF({chk_a}, \"\", IF(F{r}>=2, {f_second_start}, \"\"))")
            
            # Q: 34/35G履歴 - 初当たりを除く当たりの中に34Gまたは35Gがあるか
            # SUMPRODUCTでチェック: 連チャンID合致 AND 位置>=2 AND (Start=34 OR Start=35)
            f_has_3435 = f"SUMPRODUCT(({d_id_col}=A{r})*({d_pos_col}>=2)*(({d_start_col}=34)+({d_start_col}=35)))>0"
            ws_chain.cell(row=r, column=17, value=f"=IF({chk_a}, \"\", IF(F{r}>=2, {f_has_3435}, FALSE))")
            
            # R: 連荘中TY - 初当たりを除く獲得枚数の合計 (Position >= 2)
            d_dedama_col = f"Data!$G$2:$G${data_row_limit}"
            f_chain_ty = f"SUMIFS({d_dedama_col}, {d_id_col}, A{r}, {d_pos_col}, \">=2\")"
            ws_chain.cell(row=r, column=18, value=f"=IF({chk_a}, \"\", IF(F{r}>=2, {f_chain_ty}, 0))")
            
            # S: 初当Status - このチェーンの初当たり(Position=1)のStatus (RB/BB)
            d_status_col = f"Data!$H$2:$H${data_row_limit}"
            f_first_status = f"INDEX({d_status_col}, MATCH(1, ({d_id_col}=A{r})*({d_pos_col}=1), 0))"
            ws_chain.cell(row=r, column=19, value=f"=IF({chk_a}, \"\", IFERROR({f_first_status}, \"\"))")
            
            # T: 前Thr数 - Python計算済みの直接値を使用
            # U: 前Thr全RB - Python計算済みの直接値を使用
            chain_id = r - 1  # Row 2 = ChainID 1
            if chain_id in chain_rb_info:
                info = chain_rb_info[chain_id]
                ws_chain.cell(row=r, column=20, value=info['prev_thr_count'])
                ws_chain.cell(row=r, column=21, value=info['all_rb'])
            else:
                ws_chain.cell(row=r, column=20, value="")
                ws_chain.cell(row=r, column=21, value="")
            
            # V: 初当有利区間G - Data!P列から初当たり(Position=1)の有利区間G数を取得
            d_yuuri_col = f"Data!$P$2:$P${data_row_limit}"
            f_first_yuuri = f"SUMIFS({d_yuuri_col}, {d_id_col}, A{r}, {d_pos_col}, 1)"
            ws_chain.cell(row=r, column=22, value=f"=IF({chk_a}, \"\", IFERROR({f_first_yuuri}, 0))")
            
            # W: ドキハナ成功 - 天国チェーンで初当以外に34/35Gがあり、その後も1回以上天国継続
            # 条件: F>=3 AND 34/35G履歴=TRUE (初当以外に34/35Gがあり、その後も継続)
            # F>=3が必要: 初当(1) + 34/35G(2) + その次(3) = 最低3連
            f_dokihana = f"AND(F{r}>=3, Q{r}=TRUE)"
            ws_chain.cell(row=r, column=23, value=f"=IF({chk_a}, \"\", {f_dokihana})")
            
            # X: 初当たりG - Data!F列から初当たり(Position=1)のG数を取得
            f_first_g = f"SUMIFS({d_start_col}, {d_id_col}, A{r}, {d_pos_col}, 1)"
            ws_chain.cell(row=r, column=24, value=f"=IF({chk_a}, \"\", IFERROR({f_first_g}, 0))")
            
            # Y: 当日差枚 - Data!O列から初当たり(Position=1)の当日差枚を取得
            d_daily_col = f"Data!$O$2:$O${data_row_limit}"
            f_first_daily = f"SUMIFS({d_daily_col}, {d_id_col}, A{r}, {d_pos_col}, 1)"
            ws_chain.cell(row=r, column=25, value=f"=IF({chk_a}, \"\", IFERROR({f_first_daily}, 0))")
            
            # Z: 前回初当G - 同一台の前のチェーンの初当たりG
            # 同一ホール・日付・台番号でK列(行番号)<現在の行の最大行を検索し、そのX列値を取得
            c_X_rng = f"$X$2:$X${chain_row_limit}"
            c_K_rng = f"$K$2:$K${chain_row_limit}"
            # MAXIFSで同一台かつK<現在Kの最大Kを取得、その行のX値をINDEX/MATCH
            f_prev_k = f"_xlfn.MAXIFS({c_K_rng}, {c_B_rng}, B{r}, {c_C_rng}, C{r}, {c_D_rng}, D{r}, {c_K_rng}, \"<\"&K{r})"
            f_prev_x = f"INDEX({c_X_rng}, MATCH({f_prev_k}, {c_K_rng}, 0))"
            ws_chain.cell(row=r, column=26, value=f"=IF({chk_a}, \"\", IFERROR({f_prev_x}, \"\"))")
            
            # AA: 前々回初当G - 同一台の前の前のチェーンの初当たりG
            # Z列を取得した行のさらに前の行を取得
            c_Z_rng = f"$Z$2:$Z${chain_row_limit}"
            f_prev_z = f"INDEX({c_Z_rng}, MATCH({f_prev_k}, {c_K_rng}, 0))"
            ws_chain.cell(row=r, column=27, value=f"=IF({chk_a}, \"\", IFERROR({f_prev_z}, \"\"))")
            
            # AB: 初401Gハマリ後 - その日その台で初めて401G以上ハマリの次のチェーン
            # 条件: E=FALSE(朝イチ連でない) AND Z>=401 AND 同一台で過去にZ>=401のチェーンが存在しない
            f_prev_z_401_count = f"COUNTIFS({c_B_rng}, B{r}, {c_C_rng}, C{r}, {c_D_rng}, D{r}, {c_K_rng}, \"<\"&K{r}, {c_Z_rng}, \">=401\")"
            f_first_401_after = f"AND(E{r}=FALSE, Z{r}>=401, {f_prev_z_401_count}=0)"
            ws_chain.cell(row=r, column=28, value=f"=IF({chk_a}, \"\", {f_first_401_after})")
            
            # AC: ドキハナ開始位置 - 34/35Gが発生した最初のPosition (Position>=2で探す)
            # SUMPRODUCTで34/35Gの最小Positionを取得
            f_dokihana_pos = f"_xlfn.MINIFS({d_pos_col}, {d_id_col}, A{r}, {d_pos_col}, \">=2\", {d_start_col}, 34)"
            f_dokihana_pos_35 = f"_xlfn.MINIFS({d_pos_col}, {d_id_col}, A{r}, {d_pos_col}, \">=2\", {d_start_col}, 35)"
            # 34と35の小さい方を取得（0を除外）
            f_min_dokihana = f"MIN(IF({f_dokihana_pos}=0, 999, {f_dokihana_pos}), IF({f_dokihana_pos_35}=0, 999, {f_dokihana_pos_35}))"
            f_dokihana_pos_result = f"IF({f_min_dokihana}=999, 0, {f_min_dokihana})"
            ws_chain.cell(row=r, column=29, value=f"=IF({chk_a}, \"\", IF(W{r}=FALSE, \"\", IFERROR({f_dokihana_pos_result}, \"\")))")
            
            # AD: ドキハナ連チャン数 = 連チャン回数(F) - ドキハナ開始位置(AC) + 1
            f_dokihana_len = f"F{r}-AC{r}+1"
            ws_chain.cell(row=r, column=30, value=f"=IF({chk_a}, \"\", IF(W{r}=FALSE, \"\", IF(AC{r}=\"\", \"\", {f_dokihana_len})))")
            
            # AE: ドキハナ獲得枚数 - ドキハナ開始位置以降のDedama合計
            f_dokihana_dedama = f"SUMIFS({d_dedama_col}, {d_id_col}, A{r}, {d_pos_col}, \">=\"&AC{r})"
            ws_chain.cell(row=r, column=31, value=f"=IF({chk_a}, \"\", IF(W{r}=FALSE, \"\", IF(AC{r}=\"\", \"\", IFERROR({f_dokihana_dedama}, 0))))")
            
        # GraphData
        jp_map_graph = {'Range': 'ゲーム数区間', 'Hit_Count': '当選回数', 'Avg_Through': '平均スルー回数'}
        interval_df_export = interval_df.rename(columns=jp_map_graph)
        interval_df_export.to_excel(writer, sheet_name='GraphData', index=False)
        ws_graph = writer.sheets['GraphData']
        
        chart1 = BarChart()
        chart1.type = "col"
        chart1.title = "朝イチ当選ゲーム数分布"
        chart1.y_axis.title = "台数"
        chart1.x_axis.title = "ゲーム数区間"
        data = Reference(ws_graph, min_col=2, min_row=1, max_row=ws_graph.max_row, max_col=2)
        cats = Reference(ws_graph, min_col=1, min_row=2, max_row=ws_graph.max_row)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.shape = 4
        ws_graph.add_chart(chart1, "E2")
        
        chart2 = BarChart()
        chart2.type = "col"
        chart2.title = "朝イチ当選G数別 平均スルー回数 (天国まで)"
        chart2.y_axis.title = "平均スルー回数"
        chart2.x_axis.title = "ゲーム数区間"
        data2 = Reference(ws_graph, min_col=3, min_row=1, max_row=ws_graph.max_row, max_col=3)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats)
        chart2.shape = 4
        ws_graph.add_chart(chart2, "E20")
        
        # PostHeavenAnalysis
        ws_ph = workbook.create_sheet("PostHeavenAnalysis")
        ws_ph['A1'] = "【天国連チャン終了後 次回天国までのスルー回数分布】"
        ws_ph['A1'].font = openpyxl.styles.Font(bold=True, size=14)
        ws_ph['A3'] = "■ 回数 (件数)"
        ws_ph['A3'].font = openpyxl.styles.Font(bold=True)
        
        ws_ph.cell(row=4, column=1, value="スルー回数")
        # Columns: 全体, 2連後, 3~9連後, 10連以上後, 3連後~9連後, 10連後~19連後, 20連以上後
        chain_lens = ['全体', 2, '3~9連', '10連以上', 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, '20連以上']
        max_through_disp = 10
        for i, cl in enumerate(chain_lens):
            if cl == '全体':
                header_text = '全体'
            elif cl == '3~9連':
                header_text = '3~9連後'
            elif cl == '10連以上':
                header_text = '10連以上後'
            elif cl == '20連以上':
                header_text = '20連以上後'
            else:
                header_text = f"{cl}連後"
            ws_ph.cell(row=4, column=2 + i, value=header_text)
            
        row_start = 5
        for t in range(max_through_disp + 1):
            ws_ph.cell(row=row_start + t, column=1, value=t)
        ws_ph.cell(row=row_start + max_through_disp + 1, column=1, value="11回以上")
        ws_ph.cell(row=row_start + max_through_disp + 2, column=1, value="終了")
        ws_ph.cell(row=row_start + max_through_disp + 3, column=1, value="合計")
        
        total_row_idx = row_start + max_through_disp + 3
        ref_thr = f"ChainData!$M$2:$M${chain_row_limit}"
        ref_len = f"ChainData!$F$2:$F${chain_row_limit}"
        
        for c_idx, cl in enumerate(chain_lens):
            col_idx = 2 + c_idx
            col_L = get_column_letter(col_idx)
            
            def get_f(cond_thr_val, is_str=False, is_gte=False):
                parts = []
                if is_gte: parts.append(f"{ref_thr}, \">=\"&{cond_thr_val}")
                elif is_str: parts.append(f"{ref_thr}, \"{cond_thr_val}\"")
                else: parts.append(f"{ref_thr}, {cond_thr_val}")
                # Add chain length condition
                if cl == '全体':
                    pass  # No additional condition
                elif cl == '3~9連':
                    parts.append(f"{ref_len}, \">=3\", {ref_len}, \"<=9\"")
                elif cl == '10連以上':
                    parts.append(f"{ref_len}, \">=10\"")
                elif cl == '20連以上':
                    parts.append(f"{ref_len}, \">=20\"")
                else:
                    parts.append(f"{ref_len}, {cl}")  # Exact match
                if len(parts) == 1: return f"=COUNTIF({parts[0]})"
                return f"=COUNTIFS({', '.join(parts)})"

            for t in range(max_through_disp + 1):
                ws_ph.cell(row=row_start + t, column=col_idx, value=get_f(t))
            ws_ph.cell(row=row_start + max_through_disp + 1, column=col_idx, value=get_f(11, is_gte=True))
            ws_ph.cell(row=row_start + max_through_disp + 2, column=col_idx, value=get_f("終了", is_str=True))
            # Sum only 0-10 and 11+ rows, excluding 終了
            end_sum_row = row_start + max_through_disp + 1  # 11回以上 row
            ws_ph.cell(row=total_row_idx, column=col_idx, value=f"=SUM({col_L}{row_start}:{col_L}{end_sum_row})")

        pct_base = total_row_idx + 3
        ws_ph.cell(row=pct_base-1, column=1, value="■ 天国移行率")
        ws_ph.cell(row=pct_base-1, column=1).font = openpyxl.styles.Font(bold=True)
        ws_ph.cell(row=pct_base, column=1, value="スルー回数")
        for i, cl in enumerate(chain_lens):
            if cl == '全体':
                header_text = '全体'
            elif cl == '3~9連':
                header_text = '3~9連後'
            elif cl == '10連以上':
                header_text = '10連以上後'
            elif cl == '20連以上':
                header_text = '20連以上後'
            else:
                header_text = f"{cl}連後"
            ws_ph.cell(row=pct_base, column=2 + i, value=header_text)
        
        # Helper function to create transition rate formula
        def get_trans_rate_f(thr_val, cl, is_gte=False):
            """
            Transition rate = (M=thr_val count) / (M>=thr_val count)
            = Those who hit heaven at exactly thr_val / Those who reached thr_val throughs
            """
            # Numerator: exact match (hit heaven at this through count)
            num_parts = []
            if is_gte:
                num_parts.append(f"{ref_thr}, \">=11\"")
            else:
                num_parts.append(f"{ref_thr}, {thr_val}")
            
            # Denominator: reached this through count (M >= thr_val)
            denom_parts = []
            denom_parts.append(f"{ref_thr}, \">=\"&{thr_val}")
            
            # Add chain length condition
            if cl == '全体':
                pass
            elif cl == '3~9連':
                num_parts.append(f"{ref_len}, \">=3\", {ref_len}, \"<=9\"")
                denom_parts.append(f"{ref_len}, \">=3\", {ref_len}, \"<=9\"")
            elif cl == '10連以上':
                num_parts.append(f"{ref_len}, \">=10\"")
                denom_parts.append(f"{ref_len}, \">=10\"")
            elif cl == '20連以上':
                num_parts.append(f"{ref_len}, \">=20\"")
                denom_parts.append(f"{ref_len}, \">=20\"")
            else:
                num_parts.append(f"{ref_len}, {cl}")
                denom_parts.append(f"{ref_len}, {cl}")
            
            if len(num_parts) == 1:
                num_f = f"COUNTIF({num_parts[0]})"
            else:
                num_f = f"COUNTIFS({', '.join(num_parts)})"
            
            if len(denom_parts) == 1:
                denom_f = f"COUNTIF({denom_parts[0]})"
            else:
                denom_f = f"COUNTIFS({', '.join(denom_parts)})"
            
            return f"=IFERROR({num_f}/{denom_f}, \"-\")"
            
        # Rows for 0-10 throughs
        for t in range(max_through_disp + 1):
            curr_pct_r = pct_base + 1 + t
            ws_ph.cell(row=curr_pct_r, column=1, value=t)
            for c_idx, cl in enumerate(chain_lens):
                ws_ph.cell(row=curr_pct_r, column=2+c_idx, value=get_trans_rate_f(t, cl))
                ws_ph.cell(row=curr_pct_r, column=2+c_idx).number_format = '0.0%'
        
        # 11回以上 row
        r_11plus = pct_base + 1 + max_through_disp + 1
        ws_ph.cell(row=r_11plus, column=1, value="11回以上")
        for c_idx, cl in enumerate(chain_lens):
            ws_ph.cell(row=r_11plus, column=2+c_idx, value=get_trans_rate_f(11, cl, is_gte=True))
            ws_ph.cell(row=r_11plus, column=2+c_idx).number_format = '0.0%'

        # Reference for 34/35G history (column Q)
        ref_3435 = f"ChainData!$Q$2:$Q${chain_row_limit}"
        
        # Section 2b: 34/35Gあり - 回数
        sec_3435_base = curr_pct_r + 4
        ws_ph.cell(row=sec_3435_base-1, column=1, value="■ 回数 (34/35Gあり)")
        ws_ph.cell(row=sec_3435_base-1, column=1).font = openpyxl.styles.Font(bold=True)
        
        ws_ph.cell(row=sec_3435_base, column=1, value="スルー回数")
        for i, cl in enumerate(chain_lens):
            if cl == '全体':
                header_text = '全体'
            elif cl == '3~9連':
                header_text = '3~9連後'
            elif cl == '10連以上':
                header_text = '10連以上後'
            elif cl == '20連以上':
                header_text = '20連以上後'
            else:
                header_text = f"{cl}連後"
            ws_ph.cell(row=sec_3435_base, column=2 + i, value=header_text)
        
        sec_3435_row_start = sec_3435_base + 1
        for t in range(max_through_disp + 1):
            ws_ph.cell(row=sec_3435_row_start + t, column=1, value=t)
        ws_ph.cell(row=sec_3435_row_start + max_through_disp + 1, column=1, value="11回以上")
        ws_ph.cell(row=sec_3435_row_start + max_through_disp + 2, column=1, value="終了")
        ws_ph.cell(row=sec_3435_row_start + max_through_disp + 3, column=1, value="合計")
        
        sec_3435_total = sec_3435_row_start + max_through_disp + 3
        
        for c_idx, cl in enumerate(chain_lens):
            col_idx = 2 + c_idx
            col_L = get_column_letter(col_idx)
            
            def get_f_3435(cond_thr_val, has_3435=True, is_str=False, is_gte=False):
                parts = []
                if is_gte: parts.append(f"{ref_thr}, \">=\"&{cond_thr_val}")
                elif is_str: parts.append(f'{ref_thr}, "{cond_thr_val}"')
                else: parts.append(f"{ref_thr}, {cond_thr_val}")
                # 34/35G condition
                parts.append(f"{ref_3435}, {has_3435}")
                # Chain length condition
                if cl == '全体':
                    pass
                elif cl == '3~9連':
                    parts.append(f'{ref_len}, ">=3", {ref_len}, "<=9"')
                elif cl == '10連以上':
                    parts.append(f'{ref_len}, ">=10"')
                elif cl == '20連以上':
                    parts.append(f'{ref_len}, ">=20"')
                else:
                    parts.append(f"{ref_len}, {cl}")
                return f"=COUNTIFS({', '.join(parts)})"
            
            for t in range(max_through_disp + 1):
                ws_ph.cell(row=sec_3435_row_start + t, column=col_idx, value=get_f_3435(t, True))
            ws_ph.cell(row=sec_3435_row_start + max_through_disp + 1, column=col_idx, value=get_f_3435(11, True, is_gte=True))
            ws_ph.cell(row=sec_3435_row_start + max_through_disp + 2, column=col_idx, value=get_f_3435("終了", True, is_str=True))
            end_sum = sec_3435_row_start + max_through_disp + 1
            ws_ph.cell(row=sec_3435_total, column=col_idx, value=f"=SUM({col_L}{sec_3435_row_start}:{col_L}{end_sum})")
        
        # Section 2c: 34/35Gなし - 回数
        sec_no3435_base = sec_3435_total + 3
        ws_ph.cell(row=sec_no3435_base-1, column=1, value="■ 回数 (34/35Gなし)")
        ws_ph.cell(row=sec_no3435_base-1, column=1).font = openpyxl.styles.Font(bold=True)
        
        ws_ph.cell(row=sec_no3435_base, column=1, value="スルー回数")
        for i, cl in enumerate(chain_lens):
            if cl == '全体':
                header_text = '全体'
            elif cl == '3~9連':
                header_text = '3~9連後'
            elif cl == '10連以上':
                header_text = '10連以上後'
            elif cl == '20連以上':
                header_text = '20連以上後'
            else:
                header_text = f"{cl}連後"
            ws_ph.cell(row=sec_no3435_base, column=2 + i, value=header_text)
        
        sec_no3435_row_start = sec_no3435_base + 1
        for t in range(max_through_disp + 1):
            ws_ph.cell(row=sec_no3435_row_start + t, column=1, value=t)
        ws_ph.cell(row=sec_no3435_row_start + max_through_disp + 1, column=1, value="11回以上")
        ws_ph.cell(row=sec_no3435_row_start + max_through_disp + 2, column=1, value="終了")
        ws_ph.cell(row=sec_no3435_row_start + max_through_disp + 3, column=1, value="合計")
        
        sec_no3435_total = sec_no3435_row_start + max_through_disp + 3
        
        for c_idx, cl in enumerate(chain_lens):
            col_idx = 2 + c_idx
            col_L = get_column_letter(col_idx)
            
            for t in range(max_through_disp + 1):
                ws_ph.cell(row=sec_no3435_row_start + t, column=col_idx, value=get_f_3435(t, False))
            ws_ph.cell(row=sec_no3435_row_start + max_through_disp + 1, column=col_idx, value=get_f_3435(11, False, is_gte=True))
            ws_ph.cell(row=sec_no3435_row_start + max_through_disp + 2, column=col_idx, value=get_f_3435("終了", False, is_str=True))
            end_sum = sec_no3435_row_start + max_through_disp + 1
            ws_ph.cell(row=sec_no3435_total, column=col_idx, value=f"=SUM({col_L}{sec_no3435_row_start}:{col_L}{end_sum})")
        
        # Update curr_pct_r for next section calculation
        curr_pct_r = sec_no3435_total
        
        # Section 2d: 34/35Gあり - 割合
        # Section 2d: 34/35Gあり - 天国移行率
        pct_3435_base = sec_no3435_total + 3
        ws_ph.cell(row=pct_3435_base-1, column=1, value="■ 天国移行率 (34/35Gあり)")
        ws_ph.cell(row=pct_3435_base-1, column=1).font = openpyxl.styles.Font(bold=True)
        
        ws_ph.cell(row=pct_3435_base, column=1, value="スルー回数")
        for i, cl in enumerate(chain_lens):
            if cl == '全体':
                header_text = '全体'
            elif cl == '3~9連':
                header_text = '3~9連後'
            elif cl == '10連以上':
                header_text = '10連以上後'
            elif cl == '20連以上':
                header_text = '20連以上後'
            else:
                header_text = f"{cl}連後"
            ws_ph.cell(row=pct_3435_base, column=2 + i, value=header_text)
        
        # Transition rate formula with 34/35G=TRUE condition
        def get_trans_rate_3435_f(thr_val, cl, has_3435=True, is_gte=False):
            # Numerator: M=thr_val AND Q=has_3435 
            num_parts = []
            if is_gte:
                num_parts.append(f"{ref_thr}, \">=11\"")
            else:
                num_parts.append(f"{ref_thr}, {thr_val}")
            num_parts.append(f"{ref_3435}, {has_3435}")
            
            # Denominator: M>=thr_val AND Q=has_3435
            denom_parts = []
            denom_parts.append(f"{ref_thr}, \">=\"&{thr_val}")
            denom_parts.append(f"{ref_3435}, {has_3435}")
            
            # Chain length condition
            if cl == '全体':
                pass
            elif cl == '3~9連':
                num_parts.append(f"{ref_len}, \">=3\", {ref_len}, \"<=9\"")
                denom_parts.append(f"{ref_len}, \">=3\", {ref_len}, \"<=9\"")
            elif cl == '10連以上':
                num_parts.append(f"{ref_len}, \">=10\"")
                denom_parts.append(f"{ref_len}, \">=10\"")
            elif cl == '20連以上':
                num_parts.append(f"{ref_len}, \">=20\"")
                denom_parts.append(f"{ref_len}, \">=20\"")
            else:
                num_parts.append(f"{ref_len}, {cl}")
                denom_parts.append(f"{ref_len}, {cl}")
            
            num_f = f"COUNTIFS({', '.join(num_parts)})"
            denom_f = f"COUNTIFS({', '.join(denom_parts)})"
            
            return f"=IFERROR({num_f}/{denom_f}, \"-\")"
        
        # Rows for 0-10 throughs (34/35Gあり)
        for t in range(max_through_disp + 1):
            curr_row = pct_3435_base + 1 + t
            ws_ph.cell(row=curr_row, column=1, value=t)
            for c_idx, cl in enumerate(chain_lens):
                ws_ph.cell(row=curr_row, column=2+c_idx, value=get_trans_rate_3435_f(t, cl, True))
                ws_ph.cell(row=curr_row, column=2+c_idx).number_format = '0.0%'
        
        # 11回以上 row
        r_11p_3435 = pct_3435_base + 1 + max_through_disp + 1
        ws_ph.cell(row=r_11p_3435, column=1, value="11回以上")
        for c_idx, cl in enumerate(chain_lens):
            ws_ph.cell(row=r_11p_3435, column=2+c_idx, value=get_trans_rate_3435_f(11, cl, True, is_gte=True))
            ws_ph.cell(row=r_11p_3435, column=2+c_idx).number_format = '0.0%'
        
        pct_3435_end = r_11p_3435
        
        # Section 2e: 34/35Gなし - 天国移行率
        pct_no3435_base = pct_3435_end + 3
        ws_ph.cell(row=pct_no3435_base-1, column=1, value="■ 天国移行率 (34/35Gなし)")
        ws_ph.cell(row=pct_no3435_base-1, column=1).font = openpyxl.styles.Font(bold=True)
        
        ws_ph.cell(row=pct_no3435_base, column=1, value="スルー回数")
        for i, cl in enumerate(chain_lens):
            if cl == '全体':
                header_text = '全体'
            elif cl == '3~9連':
                header_text = '3~9連後'
            elif cl == '10連以上':
                header_text = '10連以上後'
            elif cl == '20連以上':
                header_text = '20連以上後'
            else:
                header_text = f"{cl}連後"
            ws_ph.cell(row=pct_no3435_base, column=2 + i, value=header_text)
        
        # Rows for 0-10 throughs (34/35Gなし)
        for t in range(max_through_disp + 1):
            curr_row = pct_no3435_base + 1 + t
            ws_ph.cell(row=curr_row, column=1, value=t)
            for c_idx, cl in enumerate(chain_lens):
                ws_ph.cell(row=curr_row, column=2+c_idx, value=get_trans_rate_3435_f(t, cl, False))
                ws_ph.cell(row=curr_row, column=2+c_idx).number_format = '0.0%'
        
        # 11回以上 row
        r_11p_no3435 = pct_no3435_base + 1 + max_through_disp + 1
        ws_ph.cell(row=r_11p_no3435, column=1, value="11回以上")
        for c_idx, cl in enumerate(chain_lens):
            ws_ph.cell(row=r_11p_no3435, column=2+c_idx, value=get_trans_rate_3435_f(11, cl, False, is_gte=True))
            ws_ph.cell(row=r_11p_no3435, column=2+c_idx).number_format = '0.0%'
        
        curr_pct_r = r_11p_no3435

        # Section 3: Average Chain Length at Each Through Count
        avg_base = curr_pct_r + 3
        ws_ph.cell(row=avg_base-1, column=1, value="■ 平均連荘 (天国移行時)")
        ws_ph.cell(row=avg_base-1, column=1).font = openpyxl.styles.Font(bold=True)
        ws_ph.cell(row=avg_base, column=1, value="スルー回数")
        
        # Headers for average section
        for i, cl in enumerate(chain_lens):
            if cl == '全体':
                header_text = '全体'
            elif cl == '3~9連':
                header_text = '3~9連後'
            elif cl == '10連以上':
                header_text = '10連以上後'
            elif cl == '20連以上':
                header_text = '20連以上後'
            else:
                header_text = f"{cl}連後"
            ws_ph.cell(row=avg_base, column=2 + i, value=header_text)
        
        # Reference for next heaven chain length (column N)
        ref_next_len = f"ChainData!$N$2:$N${chain_row_limit}"
        
        # Row labels and AVERAGEIFS formulas
        avg_row_start = avg_base + 1
        for t in range(max_through_disp + 1):
            ws_ph.cell(row=avg_row_start + t, column=1, value=t)
        ws_ph.cell(row=avg_row_start + max_through_disp + 1, column=1, value="11回以上")
        
        # For each column (chain length category)
        for c_idx, cl in enumerate(chain_lens):
            col_idx = 2 + c_idx
            
            def get_avg_f(cond_thr_val, is_gte=False):
                # Build AVERAGEIFS: average of N where M=through_count and F condition
                parts_avg = [ref_next_len]  # Average range
                parts_crit = []
                
                # Through count condition
                if is_gte:
                    parts_crit.append(f"{ref_thr}, \">=\"&{cond_thr_val}")
                else:
                    parts_crit.append(f"{ref_thr}, {cond_thr_val}")
                
                # Chain length condition
                if cl == '全体':
                    pass
                elif cl == '3~9連':
                    parts_crit.append(f"{ref_len}, \">=3\", {ref_len}, \"<=9\"")
                elif cl == '10連以上':
                    parts_crit.append(f"{ref_len}, \">=10\"")
                elif cl == '20連以上':
                    parts_crit.append(f"{ref_len}, \">=20\"")
                else:
                    parts_crit.append(f"{ref_len}, {cl}")
                
                return f"=IFERROR(AVERAGEIFS({parts_avg[0]}, {', '.join(parts_crit)}), \"\")"
            
            # 0-10 through counts
            for t in range(max_through_disp + 1):
                ws_ph.cell(row=avg_row_start + t, column=col_idx, value=get_avg_f(t))
                ws_ph.cell(row=avg_row_start + t, column=col_idx).number_format = '0.00'
            
            # 11+ through count
            ws_ph.cell(row=avg_row_start + max_through_disp + 1, column=col_idx, value=get_avg_f(11, is_gte=True))
            ws_ph.cell(row=avg_row_start + max_through_disp + 1, column=col_idx).number_format = '0.00'

        # Section 4: 特殊判定別 スルー回数分布 (100枚刻み)
        # Same format as "天国連チャン終了後..." section
        spec_base = avg_row_start + max_through_disp + 4
        ws_ph.cell(row=spec_base-1, column=1, value="■ 特殊判定別 スルー回数分布 (100枚刻み)")
        ws_ph.cell(row=spec_base-1, column=1).font = openpyxl.styles.Font(bold=True)
        
        ref_special = f"ChainData!$O$2:$O${chain_row_limit}"
        
        # Column headers: Special judgment ranges (1~100, 101~200, ..., 2000+)
        spec_ranges = []
        for start in range(1, 2001, 100):
            end = start + 99
            spec_ranges.append((f"{start}～{end}", start, end))
        spec_ranges.append(("2001以上", 2001, None))
        
        ws_ph.cell(row=spec_base, column=1, value="スルー回数")
        for i, (label, low, high) in enumerate(spec_ranges):
            ws_ph.cell(row=spec_base, column=2 + i, value=label)
        
        # Row labels (same as main section)
        spec_row_start = spec_base + 1
        for t in range(max_through_disp + 1):
            ws_ph.cell(row=spec_row_start + t, column=1, value=t)
        ws_ph.cell(row=spec_row_start + max_through_disp + 1, column=1, value="11回以上")
        ws_ph.cell(row=spec_row_start + max_through_disp + 2, column=1, value="終了")
        ws_ph.cell(row=spec_row_start + max_through_disp + 3, column=1, value="合計")
        
        spec_total_row = spec_row_start + max_through_disp + 3
        
        # Fill formulas for each column (special judgment range)
        for c_idx, (label, low, high) in enumerate(spec_ranges):
            col_idx = 2 + c_idx
            col_L = get_column_letter(col_idx)
            
            def get_spec_f(thr_val, is_str=False, is_gte=False):
                # Through count condition
                if is_gte:
                    thr_cond = f'{ref_thr}, ">="&{thr_val}'
                elif is_str:
                    thr_cond = f'{ref_thr}, "{thr_val}"'
                else:
                    thr_cond = f'{ref_thr}, {thr_val}'
                
                # Special judgment condition  
                if high is None:  # 2001+
                    spec_cond = f'{ref_special}, ">=2001"'
                else:
                    spec_cond = f'{ref_special}, ">={low}", {ref_special}, "<={high}"'
                
                return f"=COUNTIFS({thr_cond}, {spec_cond})"
            
            # 0-10
            for t in range(max_through_disp + 1):
                ws_ph.cell(row=spec_row_start + t, column=col_idx, value=get_spec_f(t))
            # 11+
            ws_ph.cell(row=spec_row_start + max_through_disp + 1, column=col_idx, value=get_spec_f(11, is_gte=True))
            # 終了
            ws_ph.cell(row=spec_row_start + max_through_disp + 2, column=col_idx, value=get_spec_f("終了", is_str=True))
            # 合計 (0-11+, excluding 終了)
            end_sum = spec_row_start + max_through_disp + 1
            ws_ph.cell(row=spec_total_row, column=col_idx, value=f"=SUM({col_L}{spec_row_start}:{col_L}{end_sum})")
        
        # Percentage section for special judgment
        spec_pct_base = spec_total_row + 3
        ws_ph.cell(row=spec_pct_base-1, column=1, value="■ 特殊判定別 割合 (%)")
        ws_ph.cell(row=spec_pct_base-1, column=1).font = openpyxl.styles.Font(bold=True)
        
        ws_ph.cell(row=spec_pct_base, column=1, value="スルー回数")
        for i, (label, low, high) in enumerate(spec_ranges):
            ws_ph.cell(row=spec_pct_base, column=2 + i, value=label)
        
        # Copy row labels and calculate percentages
        spec_pct_rows = list(range(spec_row_start, spec_total_row + 1))
        for r_offset, src_row in enumerate(spec_pct_rows):
            curr_pct_r = spec_pct_base + 1 + r_offset
            label_val = ws_ph.cell(row=src_row, column=1).value
            ws_ph.cell(row=curr_pct_r, column=1, value=label_val)
            
            for c_idx, (_, _, _) in enumerate(spec_ranges):
                col_idx = 2 + c_idx
                col_L = get_column_letter(col_idx)
                f = f"=IF({col_L}{spec_total_row}>0, {col_L}{src_row}/{col_L}{spec_total_row}, 0)"
                ws_ph.cell(row=curr_pct_r, column=col_idx, value=f)
                ws_ph.cell(row=curr_pct_r, column=col_idx).number_format = '0.0%'
        
        # Average chain length section for special judgment
        spec_avg_base = spec_pct_base + len(spec_pct_rows) + 3
        ws_ph.cell(row=spec_avg_base-1, column=1, value="■ 特殊判定別 平均連荘 (天国移行時)")
        ws_ph.cell(row=spec_avg_base-1, column=1).font = openpyxl.styles.Font(bold=True)
        
        ws_ph.cell(row=spec_avg_base, column=1, value="スルー回数")
        for i, (label, low, high) in enumerate(spec_ranges):
            ws_ph.cell(row=spec_avg_base, column=2 + i, value=label)
        
        # Row labels for average
        spec_avg_row_start = spec_avg_base + 1
        for t in range(max_through_disp + 1):
            ws_ph.cell(row=spec_avg_row_start + t, column=1, value=t)
        ws_ph.cell(row=spec_avg_row_start + max_through_disp + 1, column=1, value="11回以上")
        
        # Fill average chain length formulas
        for c_idx, (label, low, high) in enumerate(spec_ranges):
            col_idx = 2 + c_idx
            
            def get_spec_avg(thr_val, is_gte=False):
                # Through count condition
                if is_gte:
                    thr_cond = f'{ref_thr}, ">="&{thr_val}'
                else:
                    thr_cond = f'{ref_thr}, {thr_val}'
                
                # Special judgment condition
                if high is None:
                    spec_cond = f'{ref_special}, ">=2001"'
                else:
                    spec_cond = f'{ref_special}, ">={low}", {ref_special}, "<={high}"'
                
                return f'=IFERROR(AVERAGEIFS({ref_next_len}, {thr_cond}, {spec_cond}), "")'
            
            for t in range(max_through_disp + 1):
                ws_ph.cell(row=spec_avg_row_start + t, column=col_idx, value=get_spec_avg(t))
                ws_ph.cell(row=spec_avg_row_start + t, column=col_idx).number_format = '0.00'
            
            ws_ph.cell(row=spec_avg_row_start + max_through_disp + 1, column=col_idx, value=get_spec_avg(11, is_gte=True))
            ws_ph.cell(row=spec_avg_row_start + max_through_disp + 1, column=col_idx).number_format = '0.00'

        # Section 5: 2連履歴詳細 - 2連目当選Gで分類
        detail_2_base = spec_avg_row_start + max_through_disp + 5
        ws_ph.cell(row=detail_2_base-1, column=1, value="■ 2連履歴詳細（2連目当選G別）")
        ws_ph.cell(row=detail_2_base-1, column=1).font = openpyxl.styles.Font(bold=True)
        
        # Reference for 2連目G (column P)
        ref_2nd_g = f"ChainData!$P$2:$P${chain_row_limit}"
        
        # Categories: 1G/2-10G/32G, 11-15G, 16-31G
        g2_categories = [
            ("1G,2-10G,32G", [("=", 1), (">=", 2), ("<=", 10), ("=", 32)], "complex1"),
            ("11-15G", 11, 15),
            ("16-31G", 16, 31),
        ]
        
        ws_ph.cell(row=detail_2_base, column=1, value="スルー回数")
        ws_ph.cell(row=detail_2_base, column=2, value="1G")
        ws_ph.cell(row=detail_2_base, column=3, value="2-10G,32G")
        ws_ph.cell(row=detail_2_base, column=4, value="11-15G")
        ws_ph.cell(row=detail_2_base, column=5, value="16-31G")
        ws_ph.cell(row=detail_2_base, column=6, value="33-35G")
        
        # Row labels
        d2_row_start = detail_2_base + 1
        for t in range(max_through_disp + 1):
            ws_ph.cell(row=d2_row_start + t, column=1, value=t)
        ws_ph.cell(row=d2_row_start + max_through_disp + 1, column=1, value="11回以上")
        ws_ph.cell(row=d2_row_start + max_through_disp + 2, column=1, value="終了")
        ws_ph.cell(row=d2_row_start + max_through_disp + 3, column=1, value="合計")
        
        d2_total_row = d2_row_start + max_through_disp + 3
        
        # Helper function for complex condition (2-10G, 32G)
        def get_d2_complex_f(thr_val, is_str=False, is_gte=False):
            if is_gte:
                thr_crit = f'({ref_thr}>={thr_val})*({ref_thr}<9999)'
            elif is_str:
                thr_crit = f'({ref_thr}="{thr_val}")'
            else:
                thr_crit = f'({ref_thr}={thr_val})'
            
            # F=2 AND ((P>=2 AND P<=10) OR P=32)
            g_crit = f'((({ref_2nd_g}>=2)*({ref_2nd_g}<=10))+({ref_2nd_g}=32))'
            f_crit = f'({ref_len}=2)'
            
            return f'=SUMPRODUCT({thr_crit}*{f_crit}*{g_crit})'
        
        def get_d2_range_f(thr_val, low, high, is_str=False, is_gte=False):
            if is_gte:
                thr_cond = f'{ref_thr}, ">="&{thr_val}'
            elif is_str:
                thr_cond = f'{ref_thr}, "{thr_val}"'
            else:
                thr_cond = f'{ref_thr}, {thr_val}'
            
            return f'=COUNTIFS({thr_cond}, {ref_len}, 2, {ref_2nd_g}, ">={low}", {ref_2nd_g}, "<={high}")'
        
        def get_d2_exact_f(thr_val, exact_val, is_str=False, is_gte=False):
            if is_gte:
                thr_cond = f'{ref_thr}, ">="&{thr_val}'
            elif is_str:
                thr_cond = f'{ref_thr}, "{thr_val}"'
            else:
                thr_cond = f'{ref_thr}, {thr_val}'
            
            return f'=COUNTIFS({thr_cond}, {ref_len}, 2, {ref_2nd_g}, {exact_val})'
        
        # Column 2: 1G
        for t in range(max_through_disp + 1):
            ws_ph.cell(row=d2_row_start + t, column=2, value=get_d2_exact_f(t, 1))
        ws_ph.cell(row=d2_row_start + max_through_disp + 1, column=2, value=get_d2_exact_f(11, 1, is_gte=True))
        ws_ph.cell(row=d2_row_start + max_through_disp + 2, column=2, value=get_d2_exact_f("終了", 1, is_str=True))
        ws_ph.cell(row=d2_total_row, column=2, value=f"=SUM(B{d2_row_start}:B{d2_row_start + max_through_disp + 1})")
        
        # Column 3: 2-10G,32G (complex)
        for t in range(max_through_disp + 1):
            ws_ph.cell(row=d2_row_start + t, column=3, value=get_d2_complex_f(t))
        ws_ph.cell(row=d2_row_start + max_through_disp + 1, column=3, value=get_d2_complex_f(11, is_gte=True))
        ws_ph.cell(row=d2_row_start + max_through_disp + 2, column=3, value=get_d2_complex_f("終了", is_str=True))
        ws_ph.cell(row=d2_total_row, column=3, value=f"=SUM(C{d2_row_start}:C{d2_row_start + max_through_disp + 1})")
        
        # Column 4: 11-15G
        for t in range(max_through_disp + 1):
            ws_ph.cell(row=d2_row_start + t, column=4, value=get_d2_range_f(t, 11, 15))
        ws_ph.cell(row=d2_row_start + max_through_disp + 1, column=4, value=get_d2_range_f(11, 11, 15, is_gte=True))
        ws_ph.cell(row=d2_row_start + max_through_disp + 2, column=4, value=get_d2_range_f("終了", 11, 15, is_str=True))
        ws_ph.cell(row=d2_total_row, column=4, value=f"=SUM(D{d2_row_start}:D{d2_row_start + max_through_disp + 1})")
        
        # Column 5: 16-31G
        for t in range(max_through_disp + 1):
            ws_ph.cell(row=d2_row_start + t, column=5, value=get_d2_range_f(t, 16, 31))
        ws_ph.cell(row=d2_row_start + max_through_disp + 1, column=5, value=get_d2_range_f(11, 16, 31, is_gte=True))
        ws_ph.cell(row=d2_row_start + max_through_disp + 2, column=5, value=get_d2_range_f("終了", 16, 31, is_str=True))
        ws_ph.cell(row=d2_total_row, column=5, value=f"=SUM(E{d2_row_start}:E{d2_row_start + max_through_disp + 1})")
        
        # Column 6: 33-35G
        for t in range(max_through_disp + 1):
            ws_ph.cell(row=d2_row_start + t, column=6, value=get_d2_range_f(t, 33, 35))
        ws_ph.cell(row=d2_row_start + max_through_disp + 1, column=6, value=get_d2_range_f(11, 33, 35, is_gte=True))
        ws_ph.cell(row=d2_row_start + max_through_disp + 2, column=6, value=get_d2_range_f("終了", 33, 35, is_str=True))
        ws_ph.cell(row=d2_total_row, column=6, value=f"=SUM(F{d2_row_start}:F{d2_row_start + max_through_disp + 1})")
        
        # Transition rate section for 2連詳細
        d2_pct_base = d2_total_row + 3
        ws_ph.cell(row=d2_pct_base-1, column=1, value="■ 2連履歴詳細 天国移行率")
        ws_ph.cell(row=d2_pct_base-1, column=1).font = openpyxl.styles.Font(bold=True)
        
        ws_ph.cell(row=d2_pct_base, column=1, value="スルー回数")
        ws_ph.cell(row=d2_pct_base, column=2, value="1G")
        ws_ph.cell(row=d2_pct_base, column=3, value="2-10G,32G")
        ws_ph.cell(row=d2_pct_base, column=4, value="11-15G")
        ws_ph.cell(row=d2_pct_base, column=5, value="16-31G")
        ws_ph.cell(row=d2_pct_base, column=6, value="33-35G")
        
        # Helper function for 2連詳細 transition rate
        def get_d2_trans_exact(thr_val, exact_val, is_gte=False):
            """Transition rate for exact 2nd G value"""
            if is_gte:
                num_thr = f'{ref_thr}, ">=11"'
                denom_thr = f'{ref_thr}, ">="&{thr_val}'
            else:
                num_thr = f'{ref_thr}, {thr_val}'
                denom_thr = f'{ref_thr}, ">="&{thr_val}'
            
            # Numerator: M=thr, F=2, P=exact_val
            num_f = f'COUNTIFS({num_thr}, {ref_len}, 2, {ref_2nd_g}, {exact_val})'
            # Denominator: M>=thr, F=2, P=exact_val
            denom_f = f'COUNTIFS({denom_thr}, {ref_len}, 2, {ref_2nd_g}, {exact_val})'
            
            return f'=IFERROR({num_f}/{denom_f}, "-")'
        
        def get_d2_trans_range(thr_val, low, high, is_gte=False):
            """Transition rate for range of 2nd G values"""
            if is_gte:
                num_thr = f'{ref_thr}, ">=11"'
                denom_thr = f'{ref_thr}, ">="&{thr_val}'
            else:
                num_thr = f'{ref_thr}, {thr_val}'
                denom_thr = f'{ref_thr}, ">="&{thr_val}'
            
            # Numerator: M=thr, F=2, P>=low, P<=high
            num_f = f'COUNTIFS({num_thr}, {ref_len}, 2, {ref_2nd_g}, ">={low}", {ref_2nd_g}, "<={high}")'
            # Denominator: M>=thr, F=2, P>=low, P<=high
            denom_f = f'COUNTIFS({denom_thr}, {ref_len}, 2, {ref_2nd_g}, ">={low}", {ref_2nd_g}, "<={high}")'
            
            return f'=IFERROR({num_f}/{denom_f}, "-")'
        
        def get_d2_trans_complex(thr_val, is_gte=False):
            """Transition rate for 2-10G or 32G pattern"""
            if is_gte:
                num_thr = f'{ref_thr}, ">=11"'
                denom_thr = f'{ref_thr}, ">="&{thr_val}'
            else:
                num_thr = f'{ref_thr}, {thr_val}'
                denom_thr = f'{ref_thr}, ">="&{thr_val}'
            
            # Numerator: count for 2-10 range + 32 exact
            num_2_10 = f'COUNTIFS({num_thr}, {ref_len}, 2, {ref_2nd_g}, ">=2", {ref_2nd_g}, "<=10")'
            num_32 = f'COUNTIFS({num_thr}, {ref_len}, 2, {ref_2nd_g}, 32)'
            # Denominator: same conditions with M>=thr
            denom_2_10 = f'COUNTIFS({denom_thr}, {ref_len}, 2, {ref_2nd_g}, ">=2", {ref_2nd_g}, "<=10")'
            denom_32 = f'COUNTIFS({denom_thr}, {ref_len}, 2, {ref_2nd_g}, 32)'
            
            return f'=IFERROR(({num_2_10}+{num_32})/({denom_2_10}+{denom_32}), "-")'
        
        # Rows for 0-10 throughs
        for t in range(max_through_disp + 1):
            pct_row = d2_pct_base + 1 + t
            ws_ph.cell(row=pct_row, column=1, value=t)
            ws_ph.cell(row=pct_row, column=2, value=get_d2_trans_exact(t, 1))
            ws_ph.cell(row=pct_row, column=3, value=get_d2_trans_complex(t))
            ws_ph.cell(row=pct_row, column=4, value=get_d2_trans_range(t, 11, 15))
            ws_ph.cell(row=pct_row, column=5, value=get_d2_trans_range(t, 16, 31))
            ws_ph.cell(row=pct_row, column=6, value=get_d2_trans_range(t, 33, 35))
            for col in [2, 3, 4, 5, 6]:
                ws_ph.cell(row=pct_row, column=col).number_format = '0.0%'
        
        # 11回以上 row
        r_11p_d2 = d2_pct_base + 1 + max_through_disp + 1
        ws_ph.cell(row=r_11p_d2, column=1, value="11回以上")
        ws_ph.cell(row=r_11p_d2, column=2, value=get_d2_trans_exact(11, 1, is_gte=True))
        ws_ph.cell(row=r_11p_d2, column=3, value=get_d2_trans_complex(11, is_gte=True))
        ws_ph.cell(row=r_11p_d2, column=4, value=get_d2_trans_range(11, 11, 15, is_gte=True))
        ws_ph.cell(row=r_11p_d2, column=5, value=get_d2_trans_range(11, 16, 31, is_gte=True))
        ws_ph.cell(row=r_11p_d2, column=6, value=get_d2_trans_range(11, 33, 35, is_gte=True))
        for col in [2, 3, 4, 5, 6]:
            ws_ph.cell(row=r_11p_d2, column=col).number_format = '0.0%'

        # Section: ドキハナ成功 有利区間別分布 (0-3000G in 50G intervals)
        dokihana_base = r_11p_d2 + 4
        ws_ph.cell(row=dokihana_base-1, column=1, value="■ ドキハナ成功 有利区間別分布")
        ws_ph.cell(row=dokihana_base-1, column=1).font = openpyxl.styles.Font(bold=True)
        
        ws_ph.cell(row=dokihana_base, column=1, value="有利区間G数")
        ws_ph.cell(row=dokihana_base, column=2, value="ドキハナ成功件数")
        ws_ph.cell(row=dokihana_base, column=3, value="全チェーン数")
        ws_ph.cell(row=dokihana_base, column=4, value="発生率")
        
        # References for ChainData V (初当有利区間G) and W (ドキハナ成功) columns
        ref_yuuri = f"ChainData!$V$2:$V${chain_row_limit}"
        ref_dokihana = f"ChainData!$W$2:$W${chain_row_limit}"
        
        # Generate 100G interval rows (1-100, 101-200, ..., up to 2901-3000)
        dokihana_row = dokihana_base + 1
        for start in range(0, 3001, 100):
            end = start + 100
            if start == 0:
                label = f"1 ~ 100"
                yuuri_cond_low = 1
            else:
                label = f"{start+1} ~ {end}"
                yuuri_cond_low = start + 1
            yuuri_cond_high = end
            
            ws_ph.cell(row=dokihana_row, column=1, value=label)
            # ドキハナ成功件数: COUNTIFS(V>=low, V<=high, W=TRUE)
            f_success = f'=COUNTIFS({ref_yuuri}, ">={yuuri_cond_low}", {ref_yuuri}, "<={yuuri_cond_high}", {ref_dokihana}, TRUE)'
            ws_ph.cell(row=dokihana_row, column=2, value=f_success)
            # 全チェーン数: COUNTIFS(V>=low, V<=high)
            f_total = f'=COUNTIFS({ref_yuuri}, ">={yuuri_cond_low}", {ref_yuuri}, "<={yuuri_cond_high}")'
            ws_ph.cell(row=dokihana_row, column=3, value=f_total)
            # 発生率
            ws_ph.cell(row=dokihana_row, column=4, value=f"=IFERROR(B{dokihana_row}/C{dokihana_row}, 0)")
            ws_ph.cell(row=dokihana_row, column=4).number_format = '0.00%'
            
            dokihana_row += 1
        
        # Total row
        ws_ph.cell(row=dokihana_row, column=1, value="合計")
        ws_ph.cell(row=dokihana_row, column=2, value=f"=COUNTIF({ref_dokihana}, TRUE)")
        ws_ph.cell(row=dokihana_row, column=3, value=f"=COUNTA({ref_yuuri})")
        ws_ph.cell(row=dokihana_row, column=4, value=f"=IFERROR(B{dokihana_row}/C{dokihana_row}, 0)")
        ws_ph.cell(row=dokihana_row, column=4).number_format = '0.00%'

        # Section: 前回401G+ハマり & 差枚<1000 初当たり分布 (50G intervals)
        hamar_base = dokihana_row + 4
        ws_ph.cell(row=hamar_base-1, column=1, value="■ 前回401G+ハマり & 差枚<1000 初当たり分布")
        ws_ph.cell(row=hamar_base-1, column=1).font = openpyxl.styles.Font(bold=True)
        
        ws_ph.cell(row=hamar_base, column=1, value="初当たりG数")
        ws_ph.cell(row=hamar_base, column=2, value="件数")
        ws_ph.cell(row=hamar_base, column=3, value="全件数")
        ws_ph.cell(row=hamar_base, column=4, value="割合")
        
        # References for ChainData X (初当たりG), Y (当日差枚), Z (前回初当G) columns
        ref_first_g = f"ChainData!$X$2:$X${chain_row_limit}"
        ref_daily_medal = f"ChainData!$Y$2:$Y${chain_row_limit}"
        ref_prev_g = f"ChainData!$Z$2:$Z${chain_row_limit}"
        
        # Generate 50G interval rows (1-50, 51-100, ..., up to 801+)
        hamar_row = hamar_base + 1
        for start in range(0, 801, 50):
            end = start + 50
            if start == 0:
                label = f"1 ~ 50"
                g_cond_low = 1
            else:
                label = f"{start+1} ~ {end}"
                g_cond_low = start + 1
            g_cond_high = end
            
            ws_ph.cell(row=hamar_row, column=1, value=label)
            # 件数: COUNTIFS(X>=low, X<=high, Z>=401, Y<1000)
            f_count = f'=COUNTIFS({ref_first_g}, ">={g_cond_low}", {ref_first_g}, "<={g_cond_high}", {ref_prev_g}, ">=401", {ref_daily_medal}, "<1000")'
            ws_ph.cell(row=hamar_row, column=2, value=f_count)
            # 全件数: COUNTIFS(Z>=401, Y<1000)
            f_total = f'=COUNTIFS({ref_prev_g}, ">=401", {ref_daily_medal}, "<1000")'
            ws_ph.cell(row=hamar_row, column=3, value=f_total)
            # 割合
            ws_ph.cell(row=hamar_row, column=4, value=f"=IFERROR(B{hamar_row}/C{hamar_row}, 0)")
            ws_ph.cell(row=hamar_row, column=4).number_format = '0.00%'
            
            hamar_row += 1
        
        # 801G+ row
        ws_ph.cell(row=hamar_row, column=1, value="801 ~")
        f_count_801 = f'=COUNTIFS({ref_first_g}, ">=801", {ref_prev_g}, ">=401", {ref_daily_medal}, "<1000")'
        ws_ph.cell(row=hamar_row, column=2, value=f_count_801)
        ws_ph.cell(row=hamar_row, column=3, value=f'=COUNTIFS({ref_prev_g}, ">=401", {ref_daily_medal}, "<1000")')
        ws_ph.cell(row=hamar_row, column=4, value=f"=IFERROR(B{hamar_row}/C{hamar_row}, 0)")
        ws_ph.cell(row=hamar_row, column=4).number_format = '0.00%'
        hamar_row += 1
        
        # Total row
        ws_ph.cell(row=hamar_row, column=1, value="合計")
        ws_ph.cell(row=hamar_row, column=2, value=f'=COUNTIFS({ref_prev_g}, ">=401", {ref_daily_medal}, "<1000")')
        ws_ph.cell(row=hamar_row, column=3, value=f'=COUNTIFS({ref_prev_g}, ">=401", {ref_daily_medal}, "<1000")')
        ws_ph.cell(row=hamar_row, column=4, value="100.00%")

        # Section: 前回はじめての401G+ハマり & 差枚<1000 初当たり分布 (50G intervals)
        first401_base = hamar_row + 4
        ws_ph.cell(row=first401_base-1, column=1, value="■ 前回はじめての401G+ハマり & 差枚<1000 初当たり分布")
        ws_ph.cell(row=first401_base-1, column=1).font = openpyxl.styles.Font(bold=True)
        
        ws_ph.cell(row=first401_base, column=1, value="初当たりG数")
        ws_ph.cell(row=first401_base, column=2, value="件数")
        ws_ph.cell(row=first401_base, column=3, value="全件数")
        ws_ph.cell(row=first401_base, column=4, value="割合")
        
        # Reference for ChainData AB (初401Gハマリ) column
        ref_first_401 = f"ChainData!$AB$2:$AB${chain_row_limit}"
        
        # Generate 50G interval rows (1-50, 51-100, ..., up to 801+)
        first401_row = first401_base + 1
        for start in range(0, 801, 50):
            end = start + 50
            if start == 0:
                label = f"1 ~ 50"
                g_cond_low = 1
            else:
                label = f"{start+1} ~ {end}"
                g_cond_low = start + 1
            g_cond_high = end
            
            ws_ph.cell(row=first401_row, column=1, value=label)
            # 件数: COUNTIFS(X>=low, X<=high, AB=TRUE, Y<1000)
            f_count = f'=COUNTIFS({ref_first_g}, ">={g_cond_low}", {ref_first_g}, "<={g_cond_high}", {ref_first_401}, TRUE, {ref_daily_medal}, "<1000")'
            ws_ph.cell(row=first401_row, column=2, value=f_count)
            # 全件数: COUNTIFS(AB=TRUE, Y<1000)
            f_total = f'=COUNTIFS({ref_first_401}, TRUE, {ref_daily_medal}, "<1000")'
            ws_ph.cell(row=first401_row, column=3, value=f_total)
            # 割合
            ws_ph.cell(row=first401_row, column=4, value=f"=IFERROR(B{first401_row}/C{first401_row}, 0)")
            ws_ph.cell(row=first401_row, column=4).number_format = '0.00%'
            
            first401_row += 1
        
        # 801G+ row
        ws_ph.cell(row=first401_row, column=1, value="801 ~")
        f_count_801_first = f'=COUNTIFS({ref_first_g}, ">=801", {ref_first_401}, TRUE, {ref_daily_medal}, "<1000")'
        ws_ph.cell(row=first401_row, column=2, value=f_count_801_first)
        ws_ph.cell(row=first401_row, column=3, value=f'=COUNTIFS({ref_first_401}, TRUE, {ref_daily_medal}, "<1000")')
        ws_ph.cell(row=first401_row, column=4, value=f"=IFERROR(B{first401_row}/C{first401_row}, 0)")
        ws_ph.cell(row=first401_row, column=4).number_format = '0.00%'
        first401_row += 1
        
        # Total row
        ws_ph.cell(row=first401_row, column=1, value="合計")
        ws_ph.cell(row=first401_row, column=2, value=f'=COUNTIFS({ref_first_401}, TRUE, {ref_daily_medal}, "<1000")')
        ws_ph.cell(row=first401_row, column=3, value=f'=COUNTIFS({ref_first_401}, TRUE, {ref_daily_medal}, "<1000")')
        ws_ph.cell(row=first401_row, column=4, value="100.00%")

        # Summary
        ws_summary = workbook.create_sheet("Summary")
        ws_summary['A1'] = "【朝イチ・天国分析】"
        ws_summary['A1'].font = openpyxl.styles.Font(bold=True, size=14)
        
        max_r = ws_data.max_row
        range_start = f"Data!$F$2:$F${max_r}"
        range_pos = f"Data!$J$2:$J${max_r}"
        range_morn = f"Data!$K$2:$K${max_r}"
        
        ws_summary['A3'] = "朝イチ初当たり平均G数"
        ws_summary['B3'] = f"=AVERAGEIFS({range_start}, {range_pos}, 1, {range_morn}, TRUE)"
        ws_summary['A4'] = "朝イチ天国移行率"
        
        ref_c_morn = f"ChainData!$E$2:$E${chain_row_limit}"
        ref_c_len = f"ChainData!$F$2:$F${chain_row_limit}"
        ref_c_net = f"ChainData!$I$2:$I${chain_row_limit}"
        
        ws_summary['B4'] = f"=COUNTIFS({ref_c_len}, \">=2\", {ref_c_morn}, TRUE) / COUNTIF({ref_c_morn}, TRUE)"
        ws_summary['B4'].number_format = '0.0%'
        
        ws_summary['A5'] = "朝イチ天国平均連チャン数"
        ws_summary['B5'] = f"=AVERAGEIFS({ref_c_len}, {ref_c_len}, \">=2\", {ref_c_morn}, TRUE)"
        ws_summary['B5'].number_format = '0.00'
        ws_summary['A6'] = "朝イチ天国平均差枚数"
        ws_summary['B6'] = f"=AVERAGEIFS({ref_c_net}, {ref_c_len}, \">=2\", {ref_c_morn}, TRUE)"
        ws_summary['B6'].number_format = '#,##0'

        ws_summary['A9'] = "【連チャン継続率】 (全期間)"
        ws_summary['A9'].font = openpyxl.styles.Font(bold=True)
        ws_summary['A10'] = "遷移"
        ws_summary['B10'] = "開始回数"
        ws_summary['C10'] = "継続回数"
        ws_summary['D10'] = "継続率"
        for i in range(1, 21):
            r = 10 + i
            if i == 1:
                label = "初当たり→2連"
            else:
                label = f"{i}連→{i+1}連"
            ws_summary.cell(row=r, column=1, value=label)
            ws_summary.cell(row=r, column=2, value=f"=COUNTIF({ref_c_len}, \">={i}\")")
            ws_summary.cell(row=r, column=3, value=f"=COUNTIF({ref_c_len}, \">={i+1}\")")
            ws_summary.cell(row=r, column=4, value=f"=IF(B{r}>0, C{r}/B{r}, 0)")
            ws_summary.cell(row=r, column=4).number_format = '0.0%'
            
        ws_summary['F9'] = "【天国終了後移行挙動】"
        ws_summary['F9'].font = openpyxl.styles.Font(bold=True)
        ws_summary['F10'] = "詳細分布へ移動"
        ws_summary['G10'] = '=HYPERLINK("#PostHeavenAnalysis!A1", "クリックして詳細を見る")'
        
        # Second pattern: Using Data!J column (hit position)
        ws_summary['A34'] = "【連チャン継続率】 (Data!J列参照)"
        ws_summary['A34'].font = openpyxl.styles.Font(bold=True)
        ws_summary['A35'] = "遷移"
        ws_summary['B35'] = "開始回数"
        ws_summary['C35'] = "継続回数"
        ws_summary['D35'] = "継続率"
        
        # Use Data!J column for hit position-based calculation
        for i in range(1, 21):
            r = 35 + i
            if i == 1:
                label = "初当たり→2連"
            else:
                label = f"{i}連→{i+1}連"
            ws_summary.cell(row=r, column=1, value=label)
            # 開始: J=i のカウント（その位置に到達した数）
            # 継続: J=i+1 のカウント（次の位置に到達した数）
            ws_summary.cell(row=r, column=2, value=f'=COUNTIF({range_pos}, {i})')
            ws_summary.cell(row=r, column=3, value=f'=COUNTIF({range_pos}, {i+1})')
            ws_summary.cell(row=r, column=4, value=f"=IF(B{r}>0, C{r}/B{r}, 0)")
            ws_summary.cell(row=r, column=4).number_format = '0.0%'
        
        # Morning First Hit Distribution Section (50G increments)
        morn_dist_base = 58
        ws_summary.cell(row=morn_dist_base, column=1, value="【朝一初当たりゲーム数分布】")
        ws_summary.cell(row=morn_dist_base, column=1).font = openpyxl.styles.Font(bold=True)
        
        ws_summary.cell(row=morn_dist_base+1, column=1, value="ボナ間G数")
        ws_summary.cell(row=morn_dist_base+1, column=2, value="件数")
        ws_summary.cell(row=morn_dist_base+1, column=3, value="振分")
        ws_summary.cell(row=morn_dist_base+1, column=4, value="当選率")
        
        # Reference for morning first hits: Data where J=1 (first hit) AND K=TRUE (morning)
        range_start_morn = f"Data!$F$2:$F${max_r}"
        range_pos_morn = f"Data!$J$2:$J${max_r}"
        range_morn_flag = f"Data!$K$2:$K${max_r}"
        
        # Total morning first hits
        morn_total_row = morn_dist_base + 22  # Row for total
        morn_total_formula = f"=COUNTIFS({range_pos_morn}, 1, {range_morn_flag}, TRUE)"
        
        # Generate 50G interval rows (1-50, 51-100, ..., up to 1000+)
        intervals = [(1, 50)]
        for start in range(51, 801, 50):
            intervals.append((start, start + 49))
        intervals.append(("801+", None))
        
        for i, (low, high) in enumerate(intervals):
            r = morn_dist_base + 2 + i
            
            if high is None:
                # 801G+
                label = "801 ~ "
                count_f = f"=COUNTIFS({range_pos_morn}, 1, {range_morn_flag}, TRUE, {range_start_morn}, \">=801\")"
                cumul_f = f"COUNTIFS({range_pos_morn}, 1, {range_morn_flag}, TRUE)"
            else:
                label = f"{low} ~ {high}"
                count_f = f'=COUNTIFS({range_pos_morn}, 1, {range_morn_flag}, TRUE, {range_start_morn}, ">={low}", {range_start_morn}, "<={high}")'
                # Cumulative: count where Start <= high (no = prefix for embedding)
                cumul_f = f'COUNTIFS({range_pos_morn}, 1, {range_morn_flag}, TRUE, {range_start_morn}, "<={high}")'
            
            ws_summary.cell(row=r, column=1, value=label)
            ws_summary.cell(row=r, column=2, value=count_f)
            # Distribution %
            ws_summary.cell(row=r, column=3, value=f"=IF($B${morn_total_row}>0, B{r}/$B${morn_total_row}, 0)")
            ws_summary.cell(row=r, column=3).number_format = '0.0%'
            # Cumulative rate
            if high is None:
                ws_summary.cell(row=r, column=4, value="100.0%")
            else:
                ws_summary.cell(row=r, column=4, value=f"=IF($B${morn_total_row}>0, {cumul_f}/$B${morn_total_row}, 0)")
            ws_summary.cell(row=r, column=4).number_format = '0.0%'
        
        # Total row
        ws_summary.cell(row=morn_total_row, column=1, value="合計")
        ws_summary.cell(row=morn_total_row, column=2, value=morn_total_formula)
        ws_summary.cell(row=morn_total_row, column=3, value="100.0%")
        ws_summary.cell(row=morn_total_row, column=4, value="-")
        
        # ========== ドキハナチャンス分析シート ==========
        ws_doki = workbook.create_sheet("DokihanaAnalysis")
        ws_doki['A1'] = "【ドキハナチャンス分析】"
        ws_doki['A1'].font = openpyxl.styles.Font(bold=True, size=14)
        ws_doki['A2'] = "34Gまたは35Gで発生する特別な天国連チャンの分析"
        
        # ChainDataの範囲参照
        c_dokihana_success = f"ChainData!$W$2:$W${chain_row_limit}"  # ドキハナ成功フラグ
        c_dokihana_len = f"ChainData!$AD$2:$AD${chain_row_limit}"    # ドキハナ連チャン数
        c_dokihana_dedama = f"ChainData!$AE$2:$AE${chain_row_limit}" # ドキハナ獲得枚数
        c_chain_len = f"ChainData!$F$2:$F${chain_row_limit}"         # 元の連チャン回数
        
        # 基本統計
        ws_doki['A4'] = "■ 基本統計"
        ws_doki['A4'].font = openpyxl.styles.Font(bold=True)
        
        ws_doki['A5'] = "ドキハナ発生回数"
        ws_doki['B5'] = f"=COUNTIF({c_dokihana_success}, TRUE)"
        
        ws_doki['A6'] = "平均ドキハナ連チャン数"
        ws_doki['B6'] = f"=IFERROR(AVERAGEIF({c_dokihana_success}, TRUE, {c_dokihana_len}), \"-\")"
        ws_doki['B6'].number_format = '0.00'
        
        ws_doki['A7'] = "平均ドキハナ獲得枚数"
        ws_doki['B7'] = f"=IFERROR(AVERAGEIF({c_dokihana_success}, TRUE, {c_dokihana_dedama}), \"-\")"
        ws_doki['B7'].number_format = '#,##0'
        
        ws_doki['A8'] = "合計ドキハナ獲得枚数"
        ws_doki['B8'] = f"=SUMIF({c_dokihana_success}, TRUE, {c_dokihana_dedama})"
        ws_doki['B8'].number_format = '#,##0'
        
        # 連チャン数分布
        ws_doki['A10'] = "■ ドキハナ連チャン数分布"
        ws_doki['A10'].font = openpyxl.styles.Font(bold=True)
        
        ws_doki['A11'] = "連チャン数"
        ws_doki['B11'] = "発生回数"
        ws_doki['C11'] = "割合"
        ws_doki['D11'] = "平均獲得枚数"
        
        for chain in range(2, 21):  # 2連〜20連
            r = 12 + (chain - 2)
            ws_doki.cell(row=r, column=1, value=f"{chain}連")
            # 発生回数
            ws_doki.cell(row=r, column=2, value=f"=COUNTIFS({c_dokihana_success}, TRUE, {c_dokihana_len}, {chain})")
            # 割合
            ws_doki.cell(row=r, column=3, value=f"=IFERROR(B{r}/$B$5, 0)")
            ws_doki.cell(row=r, column=3).number_format = '0.0%'
            # 平均獲得枚数
            ws_doki.cell(row=r, column=4, value=f"=IFERROR(AVERAGEIFS({c_dokihana_dedama}, {c_dokihana_success}, TRUE, {c_dokihana_len}, {chain}), \"-\")")
            ws_doki.cell(row=r, column=4).number_format = '#,##0'
        
        # 21連以上
        r_21plus = 12 + 19
        ws_doki.cell(row=r_21plus, column=1, value="21連以上")
        ws_doki.cell(row=r_21plus, column=2, value=f"=COUNTIFS({c_dokihana_success}, TRUE, {c_dokihana_len}, \">=21\")")
        ws_doki.cell(row=r_21plus, column=3, value=f"=IFERROR(B{r_21plus}/$B$5, 0)")
        ws_doki.cell(row=r_21plus, column=3).number_format = '0.0%'
        ws_doki.cell(row=r_21plus, column=4, value=f"=IFERROR(AVERAGEIFS({c_dokihana_dedama}, {c_dokihana_success}, TRUE, {c_dokihana_len}, \">=21\"), \"-\")")
        ws_doki.cell(row=r_21plus, column=4).number_format = '#,##0'
        
        # 元の連チャン数別ドキハナ発生率
        ws_doki['A35'] = "■ 元の連チャン数別 ドキハナ発生率"
        ws_doki['A35'].font = openpyxl.styles.Font(bold=True)
        
        ws_doki['A36'] = "元の連チャン数"
        ws_doki['B36'] = "天国回数"
        ws_doki['C36'] = "ドキハナ回数"
        ws_doki['D36'] = "ドキハナ発生率"
        ws_doki['E36'] = "平均ドキハナ連数"
        
        c_heaven_flag = f"ChainData!$J$2:$J${chain_row_limit}"  # 天国フラグ
        
        for chain in range(3, 21):  # 3連〜20連（ドキハナは最低3連必要）
            r = 37 + (chain - 3)
            ws_doki.cell(row=r, column=1, value=f"{chain}連")
            # 天国回数
            ws_doki.cell(row=r, column=2, value=f"=COUNTIF({c_chain_len}, {chain})")
            # ドキハナ回数
            ws_doki.cell(row=r, column=3, value=f"=COUNTIFS({c_dokihana_success}, TRUE, {c_chain_len}, {chain})")
            # ドキハナ発生率
            ws_doki.cell(row=r, column=4, value=f"=IFERROR(C{r}/B{r}, 0)")
            ws_doki.cell(row=r, column=4).number_format = '0.0%'
            # 平均ドキハナ連数
            ws_doki.cell(row=r, column=5, value=f"=IFERROR(AVERAGEIFS({c_dokihana_len}, {c_dokihana_success}, TRUE, {c_chain_len}, {chain}), \"-\")")
            ws_doki.cell(row=r, column=5).number_format = '0.00'
        
        # 21連以上
        r_21plus_rate = 37 + 18
        ws_doki.cell(row=r_21plus_rate, column=1, value="21連以上")
        ws_doki.cell(row=r_21plus_rate, column=2, value=f"=COUNTIF({c_chain_len}, \">=21\")")
        ws_doki.cell(row=r_21plus_rate, column=3, value=f"=COUNTIFS({c_dokihana_success}, TRUE, {c_chain_len}, \">=21\")")
        ws_doki.cell(row=r_21plus_rate, column=4, value=f"=IFERROR(C{r_21plus_rate}/B{r_21plus_rate}, 0)")
        ws_doki.cell(row=r_21plus_rate, column=4).number_format = '0.0%'
        ws_doki.cell(row=r_21plus_rate, column=5, value=f"=IFERROR(AVERAGEIFS({c_dokihana_len}, {c_dokihana_success}, TRUE, {c_chain_len}, \">=21\"), \"-\")")
        ws_doki.cell(row=r_21plus_rate, column=5).number_format = '0.00'


    print(f"Excel file created: {OUTPUT_FILE}")

if __name__ == "__main__":
    print("Start...")
    try:
        df = load_data(INPUT_FILE)
        print("Data Loaded")
        df, stats = process_logic(df)
        print("Logic Processed")
        interval_df = analyze_intervals(stats)
        print("Intervals Analyzed")
        write_excel(df, interval_df, stats)
    except Exception as e:
        print(f"ERROR: {e}")
        traceback.print_exc()
